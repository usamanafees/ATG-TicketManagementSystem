from django.conf import settings
from django.core.exceptions import PermissionDenied
from django.shortcuts import render,reverse, redirect, get_object_or_404, render_to_response
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponseRedirect, HttpResponse, Http404, JsonResponse, HttpResponseBadRequest
from itrak.models import *
from django_datatables_view.base_datatable_view import BaseDatatableView
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
import pytz
from django.core.mail import EmailMessage
from django.core.mail import EmailMultiAlternatives
import os
import time
import smtplib
import mimetypes
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.message import Message
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.text import MIMEText
from django.utils.crypto import get_random_string
from django.conf.urls import url
from django.template.loader import render_to_string, get_template
from itrak.views.Load import *
from itrak.views.Email import *
from django.db.models.query import QuerySet
from django.core import signing
from datetime import datetime, timezone, timedelta
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
import json
import operator
from functools import reduce
from django.db.models import Q, Count, F, Func, Sum
from django.apps import apps
from django.db import transaction, IntegrityError
from copy import deepcopy
from django.db.models import Case, F, FloatField, IntegerField, Sum, When
from django.db.models.functions import Cast
from django.db.models.expressions import RawSQL
from operator import itemgetter
import collections
from django.utils.html import escape
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
# Create your views here.


# Custom Decorator Start#

user_login_required = user_passes_test(lambda user: user.is_active,
                                       login_url='/')  # Here user_passes_test decorator designates the user is active.


# def active_user_required(view_func):
#     decorated_view_func = login_required(user_login_required(view_func))
#     return decorated_view_func
from functools import wraps
def active_user_required(view_func):
    @wraps(view_func)
    def wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/')
        else:
            if request.user.user_org.org_is_delete==False:
                return view_func(request, *args, **kwargs)
            else:
                return redirect('signout')
    decorated_view_func = login_required(user_login_required(view_func))
    return wraps(decorated_view_func)(wrapped)

# Custom Decorator End#


# Get Saved Search Reports Start#

@active_user_required
def savedSearches(request):
    load_sidebar = get_sidebar(request)
    org_id = request.user.user_org_id
    savedSearches = TicketSavedSearch.objects.filter(org_id=org_id).filter(Q(save_created_by_id=request.user.id)|Q(is_share=1))
    context = {
        'sidebar': load_sidebar,
        'savedSearches': savedSearches,
    }
    return render(request, 'itrak/Reports/saved_searches.html', context)


# Get Saved Search Reports End#


# Get Saved Search Process Start#

@active_user_required
def saveSearchProcess(request):
    user_id = request.user.id
    org_id = request.user.user_org_id
    save_id = request.GET.get('savedSearch')
    obj = TicketSavedSearch.objects.get(pk=save_id)
    print('1223')
    print(obj.note_entered_by)
    #User-Account Mapping Check Start
    accountsList = getAccountIDsOfCurrentUser(request)

    if len(accountsList) > 0:
        accountFilter = {'account_id__in':accountsList}
    else: 
        accountFilter = {}
    
    user_type = userType(request) 
    accountsList = getMappedUserIDsWithCurrentUer(request)
    if user_type == 'manager':
        userFilter = Q(ticket_assign_to_id__in = accountsList) | Q(ticket_caller_id__in = accountsList) | Q(ticket_next_action_id__in=accountsList) | Q(ticket_created_by_id__in=accountsList)
    elif user_type == 'enduser':
        userFilter = Q(ticket_assign_to_id=user_id) | Q(ticket_caller_id=user_id) | Q(ticket_next_action_id=user_id) | Q(ticket_created_by_id=user_id)
    else:
        userFilter = Q(pk__isnull=False)

    #User-Account Mapping Check End

    kwargs = {
        '{0}__{1}'.format('ticket_is_delete', 'iexact'): 0,
        '{0}__{1}'.format('ticket_is_active', 'iexact'): 1,
    }
    ticket_status_dict = {0: "Opened", 1: "Closed"}
    labor_hours_dict = {"0": "Less Than", "1": "More Than", "2": "Equal"}
    fielddict = {}
    # fielddict = {"1": ''}
    if obj.ticket_status != None:
        ticket_status = obj.ticket_status
        kwargs.setdefault('ticket_status__iexact', ticket_status)
        fielddict.update({'Ticket Status': ticket_status_dict[ticket_status]})
    if obj.ticket_sub_status_id != None:
        ticket_sub_status = obj.ticket_sub_status_id
        kwargs.setdefault('ticket_sub_status_id', ticket_sub_status)
        result = SubStatus.objects.only('sub_status_text').get(pk=ticket_sub_status).sub_status_text
        if obj.ticket_status != None:
            fielddict.update({'Ticket Status': ticket_status_dict[ticket_status] + ' - ' + result})
        else:
            fielddict.update({'Ticket Status': ' - ' + result})
    if obj.priority_id != None:
        priority = obj.priority_id
        kwargs.setdefault('priority_id', priority)
        result = Priority.objects.only('priority_name').get(pk=priority).priority_name
        fielddict.update({'Priority': result})
    if obj.ticket_type_id != None:
        ticket_type = obj.ticket_type_id
        kwargs.setdefault('ticket_type_id', ticket_type)
        result = TicketType.objects.only('ttype_name').get(pk=ticket_type).ttype_name
        fielddict.update({'Ticket Type': result})
    if obj.ticket_subtype1_id != None:
        subtype1 = obj.ticket_subtype1_id
        kwargs.setdefault('ticket_subtype1_id', subtype1)
        result = TicketType.objects.only('ttype_name').get(pk=subtype1).ttype_name
        fielddict.update({'Subtype 1': result})
    if obj.ticket_subtype2_id != None:
        subtype2 = obj.ticket_subtype2_id
        kwargs.setdefault('ticket_subtype2_id', subtype2)
        result = TicketType.objects.only('ttype_name').get(pk=subtype2).ttype_name
        fielddict.update({'Subtype 2': result})
    if obj.ticket_subtype3_id != None:
        subtype3 = obj.ticket_subtype3_id
        kwargs.setdefault('ticket_subtype3_id', subtype3)
        result = TicketType.objects.only('ttype_name').get(pk=subtype3).ttype_name
        fielddict.update({'Subtype 3': result})
    if obj.ticket_subtype4_id != None:
        subtype4 = obj.ticket_subtype4_id
        kwargs.setdefault('ticket_subtype4_id', subtype4)
        result = TicketType.objects.only('ttype_name').get(pk=subtype4).ttype_name
        fielddict.update({'Subtype 4': result})
    if obj.subject != '':
        subject = obj.subject
        kwargs.setdefault('subject__icontains', subject)
        fielddict.update({'Subject': subject})
    if obj.ticket_note != '':
        # ticket_note = request.POST.get('ticket_note')
        ticket_note = obj.ticket_note
        kwargs.setdefault('ticketNote__note_detail__icontains', ticket_note)
        fielddict.update({'Notes': ticket_note})
    if obj.all_three != '':
        # all_three = request.POST.get('all_three')
        all_three = obj.all_three
        args = Q(subject__icontains=all_three) | Q(ticketNote__note_detail__icontains=all_three)
        fielddict.update({'Search All Three': all_three})
    if obj.ticket_record_locator != '':
        record_locator = obj.ticket_record_locator
        kwargs.setdefault('ticket_record_locator__startswith', record_locator)
        fielddict.update({'Record Locator': record_locator})
    if obj.ticket_caller_name != '':
        caller_name = obj.ticket_caller_name
        kwargs.setdefault('ticket_caller_name__startswith', caller_name)
        fielddict.update({'Caller Name': caller_name})
    if obj.ticket_caller_phone != '':
        caller_phone = obj.ticket_caller_phone
        kwargs.setdefault('ticket_caller_phone__startswith', caller_phone)
        fielddict.update({'Caller Phone': caller_phone})
    if obj.ticket_caller_email != '':
        caller_email = obj.ticket_caller_email
        kwargs.setdefault('ticket_caller_email__startswith', caller_email)
        fielddict.update({'Caller Email': caller_email})
    if obj.ticket_passenger_name != '':
        # passenger_name = request.POST.get('passenger_name')
        passenger_name = obj.ticket_passenger_name
        print(passenger_name)
        kwargs.setdefault('ticket_passenger_name__startswith', passenger_name)
        fielddict.update({'Passenger Name': passenger_name})
    if obj.note_entered_by_id != None:
        note_entered_by = obj.note_entered_by_id
        print(note_entered_by)
        kwargs.setdefault('ticketNote__note_created_by_id', note_entered_by)
        result = User.objects.values('display_name').get(pk=note_entered_by).get('display_name')
        print('result')
        print(result)
        fielddict.update({'Note Entered By': result})
    if obj.submitted_by_id != None:
        submitted_by = obj.submitted_by_id
        kwargs.setdefault('ticket_caller_id', submitted_by)
        result = User.objects.values('display_name').get(pk=submitted_by).get('display_name')
        fielddict.update({'Submitted By': result})
    if obj.entered_by_id != None:
        # entered_by = request.POST.get('entered_by')
        entered_by = obj.entered_by_id
        kwargs.setdefault('ticket_created_by_id', entered_by)
        result = User.objects.values('display_name').get(pk=entered_by).get('display_name')
        fielddict.update({'Entered By': result})
    if obj.task_assigned_to_id != None and obj.ever_assigned == 0:
        assigned_to = obj.task_assigned_to_id
        kwargs.setdefault('ticket_assign_to_id', assigned_to)
        result = User.objects.values('display_name').get(pk=assigned_to).get('display_name')
        fielddict.update({'Assigned To': result})
    if obj.task_assigned_to_id != None and obj.ever_assigned == 1:
        assigned_to = obj.task_assigned_to_id
        kwargs.setdefault('ticketURLog__urlog_event', 1)
        kwargs.setdefault('ticketURLog__urlog_user_id', assigned_to)
        result = User.objects.values('display_name').get(pk=assigned_to).get('display_name')
        fielddict.update({'Assigned To': result})
    if obj.assigned_by_id != None:
        assigned_by = obj.assigned_by_id
        kwargs.setdefault('ticket_assign_by_id', assigned_by)
        result = User.objects.values('display_name').get(pk=assigned_by).get('display_name')
        fielddict.update({'Assigned By': result})
    if obj.next_action_id != None and obj.ever_next_action == 0:
        next_action = obj.next_action_id
        kwargs.setdefault('ticket_next_action_id', next_action)
        result = User.objects.values('display_name').get(pk=next_action).get('display_name')
        fielddict.update({'Next Action': result})
    if obj.next_action_id != None and obj.ever_next_action == 1:
        next_action = obj.next_action_id
        kwargs.setdefault('ticketURLog__urlog_event', 2)
        kwargs.setdefault('ticketURLog__urlog_user_id', next_action)
        result = User.objects.values('display_name').get(pk=next_action).get('display_name')
        fielddict.update({'Next Action': result})
    if obj.closed_by_id != None:
        closed_by = obj.closed_by_id
        kwargs.setdefault('ticket_closed_by_id', closed_by)
        result = User.objects.values('display_name').get(pk=closed_by).get('display_name')
        fielddict.update({'Closed By': result})
    if obj.org_id != None:
        org_id = obj.org_id
        kwargs.setdefault('ticket_org_id', org_id)
        result = Organization.objects.values('org_name').get(pk=org_id).get('display_name')
        fielddict.update({'Organization': result})
    if obj.account_id != None:
        account_id = obj.account_id
        kwargs.setdefault('account_id', account_id)
        # result = Client.objects.only('client_name').get(pk=client_id).client_name
        # fielddict.update({'Client': result})
    if obj.date_opened != '':
        date_opened = obj.date_opened
        dopen_start = datetime.strptime(obj.date_opened.split(' - ')[0], '%m/%d/%Y').strftime('%Y-%m-%d')
        dopen_end = datetime.strptime(obj.date_opened.split(' - ')[1], '%m/%d/%Y').strftime('%Y-%m-%d')
        kwargs.setdefault('ticket_created_at__gte', dopen_start)
        kwargs.setdefault('ticket_created_at__lte', dopen_end)
        fielddict.update({'Date Opened': date_opened})
    if obj.date_closed != '':
        date_closed = obj.date_closed
        dclose_start = datetime.strptime(obj.date_closed.split(' - ')[0], '%m/%d/%Y').strftime('%Y-%m-%d')
        dclose_end = datetime.strptime(obj.date_closed.split(' - ')[1], '%m/%d/%Y').strftime('%Y-%m-%d')
        kwargs.setdefault('ticket_closed_at__gte', dclose_start)
        kwargs.setdefault('ticket_closed_at__lte', dclose_end)
        fielddict.update({'Date Closed': date_closed})
    if obj.labour_hours_val != '':
        labour_hours_val = obj.labour_hours_val
        labour_hours_val += ':00:00'
        labour_hours = obj.labour_hours
        if labour_hours == '0':
            condition = 'lte'
        elif labour_hours == '1':
            condition = 'gte'
        else:
            condition = 'startswith'
        kwargs.setdefault('labour_hours__' + condition, str(labour_hours_val))
        fielddict.update({'Labor Hours - ' + labor_hours_dict[labour_hours]: labour_hours})
    if obj.task_description != '':
        task_description = obj.task_description
        kwargs.setdefault('ticketManager__tmgr_task__task_description__icontains', task_description)
        fielddict.update({'Task Description': task_description})
    if obj.task_assigned_to_id != None:
        task_assigned_to = obj.task_assigned_to_id
        kwargs.setdefault('ticketManager__task_assigned_to_id', task_assigned_to)
        result = User.objects.only('display_name').get(pk=task_assigned_to).display_name
        fielddict.update({'Task Assigned To': result})
    if obj.task_completion_date != '':
        task_completion_date = obj.task_completion_date
        dtask_start = datetime.strptime(obj.task_completion_date.split(' - ')[0], '%m/%d/%Y').strftime('%Y-%m-%d')
        dtask_end = datetime.strptime(obj.task_completion_date.split(' - ')[1], '%m/%d/%Y').strftime('%Y-%m-%d')
        kwargs.setdefault('ticketManager__task_due_date__gte', dtask_start)
        kwargs.setdefault('ticketManager__task_due_date__lte', dtask_end)
        fielddict.update({'Task Completion Date': task_completion_date})
    if obj.search_title != '':
        search_title = obj.search_title
    else:
        search_title = 'Search Results'

    if obj.search_title != '':
        search_title = obj.search_title
    if obj.output_view != '':
        output_view = obj.output_view
    sortargs = []
    sortresponse = []
    sortdict = {
        "ticket_assign_to__display_name": "Assigned To",
        "ticket_closed_at": "Date Closed",
        "ticket_created_at": "Date Opened",
        "ticket_created_by__display_name": "Entered By",
        "ticket_id": "Ticket Nbr",
        "ticket_subtype1__ttype_name": "Ticket Subtype",
        "ticket_type__ttype_name": "Ticket Type",
        "ticket_modified_at": "Last Activity",
        "ticket_next_action__display_name": "Next Action",
        "ticket_org__org_name": "Organization",
        "priority__p_display_order": "Priority",
        "ticket_status": "Status",
        "ticket_caller__display_name": "Submitted By"
        # "ticket_client__client_name": "Client"
    }
    sortorder_dict = {
        "0": "Asc",
        "1": "Desc"
    }
    if obj.sort_column1 != '':
        sort_column1 = obj.sort_column1
        sort_order1 = obj.sort_order1
        if sort_order1 == '0':
            sort_value = sort_column1
        else:
            sort_value = '-' + sort_column1

        sortargs.append(sort_value)
        sortresponse.append(sortdict[sort_column1] + " , " + sortorder_dict[sort_order1])

    if obj.sort_column2 != '':
        sort_column2 = obj.sort_column2
        sort_order2 = obj.sort_order2
        if sort_order2 == '0':
            sort_value = sort_column2
        else:
            sort_value = '-' + str(sort_column2)

        sortargs.append(sort_value)
        sortresponse.append(sortdict[sort_column2] + " , " + sortorder_dict[sort_order2])

    if obj.sort_column3 != '':
        sort_column3 = obj.sort_column3
        sort_order3 = obj.sort_order3
        if sort_order3 == '0':
            sort_value = sort_column3
        else:
            sort_value = '-' + str(sort_column3)

        sortargs.append(sort_value)
        sortresponse.append(sortdict[sort_column3] + " , " + sortorder_dict[sort_order3])

    if not sortargs:
        if obj.all_three != '':
            tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(**kwargs).filter(args).filter(**accountFilter).filter(userFilter).distinct()
            ticketid_list = Ticket.objects.filter(ticket_org_id=org_id).filter(**kwargs).filter(args).filter(**accountFilter).filter(userFilter).distinct().values_list('ticket_id', flat=True)
        else:
            tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(**kwargs).filter(**accountFilter).filter(userFilter).distinct()
            ticketid_list = Ticket.objects.filter(ticket_org_id=org_id).filter(**kwargs).filter(**accountFilter).filter(userFilter).distinct().values_list('ticket_id', flat=True)
    else:
        if obj.all_three != '':
            tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(**kwargs).filter(args).filter(**accountFilter).filter(userFilter).distinct().order_by(*sortargs)
            ticketid_list = Ticket.objects.filter(ticket_org_id=org_id).filter(**kwargs).filter(args).filter(**accountFilter).filter(userFilter).distinct().values_list('ticket_id',
                                                                                                flat=True).order_by(
                *sortargs)
        else:
            tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(**kwargs).filter(**accountFilter).filter(userFilter).distinct().order_by(*sortargs)
            ticketid_list = Ticket.objects.filter(ticket_org_id=org_id).filter(**kwargs).filter(**accountFilter).filter(userFilter).distinct().values_list('ticket_id', flat=True).order_by(
                *sortargs)

    # Total Time Open Case Start #
    if obj.total_time_open_val != '':
        total_time_open_val = obj.total_time_open_val
        total_time_open = obj.total_time_open

        ticket_list = []
        if tickets:
            for ticket in tickets:
                if ticket.ticket_status == 0:
                    delta = datetime.now(timezone.utc) - ticket.ticket_created_at
                else:
                    delta = ticket.ticket_closed_at - ticket.ticket_created_at
                if total_time_open == '0':
                    if int(delta.days) <= int(total_time_open_val):
                        ticket_list.append(ticket.ticket_id)
                elif total_time_open == '1':
                    if int(delta.days) >= int(total_time_open_val):
                        ticket_list.append(ticket.ticket_id)
                else:
                    if int(delta.days) == int(total_time_open_val):
                        ticket_list.append(ticket.ticket_id)

            if not sortargs and ticket_list:
                tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_id__in=ticket_list).filter(ticket_is_delete=0).filter(
                    ticket_is_active=1).filter(**accountFilter).filter(userFilter)
                ticketid_list = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_id__in=ticket_list).filter(
                    ticket_is_delete=0).filter(ticket_is_active=1).filter(**accountFilter).filter(userFilter).values_list('ticket_id', flat=True)

            elif sortargs and ticket_list:
                tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_id__in=ticket_list).filter(ticket_is_delete=0).filter(
                    ticket_is_active=1).filter(**accountFilter).filter(userFilter).order_by(*sortargs)
                ticketid_list = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_id__in=ticket_list).filter(
                    ticket_is_delete=0).filter(ticket_is_active=1).filter(**accountFilter).filter(userFilter).values_list('ticket_id', flat=True).order_by(
                    *sortargs)
            else:
                tickets = []
        fielddict.update({'Total Time Open - ' + labor_hours_dict[total_time_open]: total_time_open_val})
    # Total Time Open Case End #

    if obj.show_criteria != 1:
        fielddict = {"1": ''}

    # for key, value in kwargs.items():
    #     print("The value of {} is {}".format(key, value))
    # return HttpResponse('OK')

    context = {
        'tickets': tickets,
        'search_title': search_title,
        'sortresponses': sortresponse,
        'fielddict': fielddict,
        'output_view': output_view,
        'ticketid_list': list(ticketid_list)
    }

    if output_view == 'OutputToExcel' or output_view == 'DetailOutputToExcel':
        if output_view == 'OutputToExcel':
            return get_xls_from_tickets(list(ticketid_list), '0')
        else:
            return get_xls_from_tickets(list(ticketid_list), '1')
    else:
        return render(request, 'itrak/Ticket/tickets_search_list.html', context)


# Get Saved Search Process End#


@csrf_exempt
# Delete Saved Search Against ID Start#
def savedSearchRemove(request):
    response_data = {}
    if request.is_ajax() and request.method == 'POST':
        save_id = request.POST.get('save_id')
        # probably you want to add a regex check if the username value is valid here
        if save_id:
            print(save_id)
            result = TicketSavedSearch.objects.filter(saved_search_id=save_id).delete()
            try:
                response_data['response'] = 'Success'
            except:
                response_data['response'] = 'Error'
        else:
            response_data['response'] = 'Error'
    else:
        response_data['response'] = 'Error'
    messages.success(request, 'Request Succeed! Report  successfully deleted.')
    return JsonResponse(response_data)

    # Delete Saved Search Against ID End#
@csrf_exempt
# Delete Saved Search Against ID Start#
def savedSearchRemove2(request):
    response_data = {}
    if request.is_ajax() and request.method == 'POST':
        save_id = request.POST.get('save_id')
        # probably you want to add a regex check if the username value is valid here
        if save_id:
            print(save_id)
            result = SavedRBReports.objects.get(pk=save_id)
            result.rb_report_is_delete = 1
            result.rb_modified_by_id = request.user.id
            result.rb_modified_at = datetime.now()
            result.save()
            try:
                response_data['response'] = 'Success'
            except:
                response_data['response'] = 'Error'
        else:
            response_data['response'] = 'Error'
    else:
        response_data['response'] = 'Error'
    messages.success(request, 'Request Succeed! Report  successfully deleted.')
    return JsonResponse(response_data)

    # Delete Saved Search Against ID End#    


# Get Saved Search Reports Start#

@active_user_required
def summaryReports(request):
    load_sidebar = get_sidebar(request)
    savedSearches = TicketSavedSearch.objects.filter()
    context = {
        'sidebar': load_sidebar,
        'savedSearches': savedSearches,
    }
    return render(request, 'itrak/Reports/summary_reports.html', context)


# Get Saved Search Reports End#


# Get Date Range For Summry Report  Start#

@active_user_required
def getReportDateRange(request):
    report_id = request.GET.get('summaryReport')
    load_sidebar = get_sidebar(request)

    # for key, value in kwargs.items():
    #     print("The value of {} is {}".format(key, value))
    # return HttpResponse('OK')

    context = {
        'report_id': report_id,
        'sidebar': load_sidebar,
    }

    return render(request, 'itrak/Reports/get_report_daterange.html', context)


# Get Date Range For Summry Report End#


# Get Summary Report Result Start#

@active_user_required
def summaryReportResults(request):
    if request.method == 'POST':
        org_id  = request.user.user_org_id
        report_id = request.POST.get('report_id')
        date_range = request.POST.get('date_range')
        start_date = datetime.strptime(request.POST.get('date_range').split(' - ')[0], '%m/%d/%Y').strftime('%Y-%m-%d')
        end_date = datetime.strptime(request.POST.get('date_range').split(' - ')[1] + " 23:59:59",
                                     '%m/%d/%Y %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
        sort_by = request.POST.get('sort_by')
        load_sidebar = get_sidebar(request)

        # Instead of having an error on your server,
        # your user will get a 404 meaning that he tries to access a non existing resource.

        reportdict = {
            "OrgContractTime": "Contract Time Used by Organization - Select Date Range",
            "OrgContractExp": "Contracts Expiring by Organization - Select Date Range",
            "OrgContractTimeOverage": "Contract Overage Time by Organization - Select Date Range",
            "IssuesPriority": "Tickets by Priority - Select Date Range",
            "IssuesPriorityByDay": "Tickets by Priority (Day) - Select Date Range",
            "IssuesPriorityByMonth": "Tickets by Priority (Month) - Select Date Range",
            "AssignedTo": "Tickets by Assigned To - Select Date Range",
            "Inactive": "Tickets by Inactive User - Results",
            "IssueSubTypes": "Tickets by Ticket Subtype - Select Date Range",
            "IssueTypes": "Tickets by Ticket Type - Select Date Range",
            "Locations": "Tickets by Account - Select Date Range",
            "NextActionBy": "Tickets by Next Action - Select Date Range",
            "Organizations": "Tickets by Organization - Select Date Range",
            "SubmittedBy": "Tickets by Submitter - Select Date Range",
            "DepartmentsSubmit": "Tickets by Submitting Department - Select Date Range",
            "IncidentsByRep": "Tickets by Rep - Select Date Range",
            "OrgTimes": "Labor Hours by Organization - Select Date Range",
            "UserTimes": "Labor Hours by User - Select Date Range",
            "OrgTimeOpen": "Total Time Open by Organization - Select Date Range",
            "UserTimeOpen": "Total Time Open by User - Select Date getReportDateRange"
        }
        #User-Account Mapping Check Start
        accountsList = getAccountIDsOfCurrentUser(request)
        # return HttpResponse(accountsList)

        if len(accountsList) > 0:
            accountFilter = {'account_id__in':accountsList}
        else: 
            accountFilter = {}
        
        user_id = request.user.id
        user_type = userType(request) 
        accountsList = getMappedUserIDsWithCurrentUer(request)
        if user_type == 'manager':
            userFilter = Q(ticket_assign_to_id__in = accountsList) | Q(ticket_caller_id__in = accountsList) | Q(ticket_next_action_id__in=accountsList) | Q(ticket_created_by_id__in=accountsList)
        elif user_type == 'enduser':
            userFilter = Q(ticket_assign_to_id=user_id) | Q(ticket_caller_id=user_id) | Q(ticket_next_action_id=user_id) | Q(ticket_created_by_id=user_id)
        else:
            userFilter = Q(pk__isnull=False)
        #User-Account Mapping Check End
        if report_id in reportdict:
            if report_id == 'IssuesPriority':
                try:
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                        priority_id__isnull=False).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(tcount=Count('ticket_id'))
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                try:
                    priorities = Priority.objects.filter(user_org_id=org_id).values('priority_id', 'priority_name').filter(prior_is_delete=0)
                except Priority.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))

                tickets_open_before = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_id', 'priority__priority_name').filter(
                    priority_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (
                                Q(ticket_status=0) | Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                            ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_before = sorted(tickets_open_before, key=itemgetter('tcount'), reverse=True)
                tickets_open_before_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_open_in = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_id', 'priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(tcount=Count('ticket_id'))
                tickets_open_in = sorted(tickets_open_in, key=itemgetter('tcount'), reverse=True)
                tickets_open_in_list = list(Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(tcount=Count('ticket_id')))

                tickets_closed = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_id', 'priority__priority_name').filter(
                    priority_id__isnull=False).filter(
                    ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                    ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_closed = sorted(tickets_closed, key=itemgetter('tcount'), reverse=True)
                tickets_closed_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(
                    ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                    ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))

                tickets_left_opened = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_id', 'priority__priority_name').filter(
                    priority_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(tcount=Count('ticket_id'))
                tickets_left_opened = sorted(tickets_left_opened, key=itemgetter('tcount'), reverse=True)
                tickets_left_opened_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(Q(ticket_created_at__lt=start_date) | (
                            Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))

                if tickets_open_before or tickets_open_in or tickets_closed or tickets_left_opened:

                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'tickets_open_before': tickets_open_before,
                        'tickets_open_before_list': tickets_open_before_list,
                        'tickets_open_in': tickets_open_in,
                        'tickets_open_in_list': tickets_open_in_list,
                        'tickets_closed': tickets_closed,
                        'tickets_closed_list': tickets_closed_list,
                        'tickets_left_opened': tickets_left_opened,
                        'tickets_left_opened_list': tickets_left_opened_list,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'priorities': priorities,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_priority_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'OrgContractExp':
                organizationServiceContract = OrginaztionContract.objects.filter(oc_org_id=org_id).filter(
                    org_contract_begin_date__gte=start_date).filter(org_contract_begin_date__lte=end_date).filter(
                    org_contract_is_delete=0)
                if organizationServiceContract:
                    context = {
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'organizationServiceContract': organizationServiceContract,
                        'sidebar': load_sidebar,
                    }
                    return render(request, 'itrak/Reports/summary_Contracts_Expiring_by_Organization.html', context)
                else:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'IssuesPriorityByDay':
                try:
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                        priority_id__isnull=False).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(tcount=Count('ticket_id'))
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                try:
                    priorities = Priority.objects.filter(user_org_id=org_id).values('priority_id', 'priority_name').filter(prior_is_delete=0)
                except Priority.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                total_tickets = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(tcount=Count('ticket_id'))
                total_tickets_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(tcount=Count('ticket_id'))
                tickets_sunday = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__week_day=1).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_sunday_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__week_day=1).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_monday = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__week_day=2).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_monday_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__week_day=2).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_tuesday = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__week_day=3).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_tuesday_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__week_day=3).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_wednesday = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__week_day=4).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_wednesday_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__week_day=4).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_thursday = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__week_day=5).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_thursday_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__week_day=5).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_friday = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__week_day=6).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_friday_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__week_day=6).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_saturday = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__week_day=7).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_saturday_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__week_day=7).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(
                    tcount=Count('ticket_id'))
                total_sunday = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__week_day=1).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter)
                total_monday = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__week_day=2).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter)
                total_tuesday = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__week_day=3).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter)
                total_wednesday = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__week_day=4).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter)
                total_thursday = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__week_day=5).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter)
                total_friday = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__week_day=6).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter)
                total_saturday = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__week_day=7).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter)
                grand_total = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(**accountFilter).filter(userFilter)

                if tickets_sunday or tickets_monday or tickets_tuesday or tickets_wednesday or tickets_thursday or tickets_friday or tickets_saturday:

                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'total_tickets': total_tickets,
                        'total_tickets_list': total_tickets_list,
                        'tickets_sunday': tickets_sunday,
                        'tickets_sunday_list': tickets_sunday_list,
                        'tickets_monday': tickets_monday,
                        'tickets_monday_list': tickets_monday_list,
                        'tickets_tuesday': tickets_tuesday,
                        'tickets_tuesday_list': tickets_tuesday_list,
                        'tickets_wednesday': tickets_wednesday,
                        'tickets_wednesday_list': tickets_wednesday_list,
                        'tickets_thursday': tickets_thursday,
                        'tickets_thursday_list': tickets_thursday_list,
                        'tickets_friday': tickets_friday,
                        'tickets_friday_list': tickets_friday_list,
                        'tickets_saturday': tickets_saturday,
                        'tickets_saturday_list': tickets_saturday_list,
                        'total_sunday': total_sunday,
                        'total_monday': total_monday,
                        'total_tuesday': total_tuesday,
                        'total_wednesday': total_wednesday,
                        'total_thursday': total_thursday,
                        'total_friday': total_friday,
                        'total_saturday': total_saturday,
                        'grand_total': grand_total,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'priorities': priorities,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_priority_day_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'IssuesPriorityByMonth':
                try:
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                        priority_id__isnull=False).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(tcount=Count('ticket_id'))
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                try:
                    priorities = Priority.objects.filter(user_org_id=org_id).values('priority_id', 'priority_name').filter(prior_is_delete=0)
                except Priority.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                total_tickets = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(tcount=Count('ticket_id'))
                total_tickets_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).annotate(tcount=Count('ticket_id'))
                tickets_january = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=1).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_february = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=2).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_march = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=3).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_april = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=4).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_may = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(priority_id__isnull=False).filter(
                    ticket_created_at__month=5).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(tcount=Count('ticket_id'))
                tickets_june = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=6).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_july = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=7).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_august = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=8).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_september = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=9).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_october = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=10).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_november = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=11).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_december = Ticket.objects.filter(ticket_org_id=org_id).values('priority__priority_name').filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=12).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_january_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=1).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_february_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=2).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_march_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=3).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_april_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=4).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_may_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=5).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_june_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=6).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_july_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=7).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_august_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=8).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_september_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=9).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_october_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=10).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_november_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=11).filter(
                    ticket_created_at__gte=start_date).filter(userFilter).filter(**accountFilter).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).annotate(
                    tcount=Count('ticket_id'))
                tickets_december_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('priority__priority_name', flat=True).filter(
                    priority_id__isnull=False).filter(ticket_created_at__month=12).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                total_january = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__month=1).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter)
                total_february = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__month=2).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter)
                total_march = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__month=3).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter)
                total_april = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__month=4).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter)
                total_may = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(ticket_created_at__month=5).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter)
                total_june = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(ticket_created_at__month=6).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter)
                total_july = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(ticket_created_at__month=7).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter)
                total_august = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__month=8).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter)
                total_september = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__month=9).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter)
                total_october = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__month=10).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter)
                total_november = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__month=11).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter)
                total_december = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__month=12).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter)
                grand_total = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter)

                if total_january or total_february or total_march or total_april or total_may or total_june or total_june or total_july or total_august or total_september or total_october or total_november or total_december:

                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'total_tickets': total_tickets,
                        'total_tickets_list': total_tickets_list,
                        'tickets_january': tickets_january,
                        'tickets_january_list': tickets_january_list,
                        'tickets_february': tickets_february,
                        'tickets_february_list': tickets_february_list,
                        'tickets_march': tickets_march,
                        'tickets_march_list': tickets_march_list,
                        'tickets_april': tickets_april,
                        'tickets_april_list': tickets_april_list,
                        'tickets_may': tickets_may,
                        'tickets_may_list': tickets_may_list,
                        'tickets_june': tickets_june,
                        'tickets_june_list': tickets_june_list,
                        'tickets_july': tickets_july,
                        'tickets_july_list': tickets_july_list,
                        'tickets_august': tickets_august,
                        'tickets_august_list': tickets_august_list,
                        'tickets_september': tickets_september,
                        'tickets_september_list': tickets_september_list,
                        'tickets_october': tickets_october,
                        'tickets_october_list': tickets_october_list,
                        'tickets_november': tickets_november,
                        'tickets_november_list': tickets_november_list,
                        'tickets_december': tickets_december,
                        'tickets_december_list': tickets_december_list,
                        'total_january': total_january,
                        'total_february': total_february,
                        'total_march': total_march,
                        'total_april': total_april,
                        'total_may': total_may,
                        'total_june': total_june,
                        'total_july': total_july,
                        'total_august': total_august,
                        'total_september': total_september,
                        'total_october': total_october,
                        'total_november': total_november,
                        'total_december': total_december,
                        'grand_total': grand_total,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'priorities': priorities,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_priority_month_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'AssignedTo':
                try:
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_assign_to_id', 'ticket_assign_to__display_name').filter(
                        ticket_assign_to__isnull=False).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                        tcount=Count('ticket_id')).order_by('ticket_assign_to__display_name')
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                tickets_open_before = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_assign_to_id',
                                                            'ticket_assign_to__display_name').filter(
                    ticket_assign_to_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_before = sorted(tickets_open_before, key=itemgetter('tcount'), reverse=True)
                tickets_open_before_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_assign_to_id', flat=True).filter(
                    ticket_assign_to_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_open_in = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_assign_to_id', 'ticket_assign_to__display_name').filter(
                    ticket_assign_to_id__isnull=False).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_in = sorted(tickets_open_in, key=itemgetter('tcount'), reverse=True)
                tickets_open_in_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_assign_to_id', flat=True).filter(
                    ticket_assign_to_id__isnull=False).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_closed = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_assign_to_id', 'ticket_assign_to__display_name').filter(
                    ticket_assign_to_id__isnull=False).filter(ticket_status=1).filter(
                    ticket_closed_at__gte=start_date).filter(ticket_closed_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_closed = sorted(tickets_closed, key=itemgetter('tcount'), reverse=True)
                tickets_closed_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_assign_to_id', flat=True).filter(
                    ticket_assign_to_id__isnull=False).filter(ticket_status=1).filter(
                    ticket_closed_at__gte=start_date).filter(ticket_closed_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_left_opened = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_assign_to_id',
                                                            'ticket_assign_to__display_name').filter(
                    ticket_assign_to_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_left_opened = sorted(tickets_left_opened, key=itemgetter('tcount'), reverse=True)
                tickets_left_opened_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_assign_to_id', flat=True).filter(
                    ticket_assign_to_id__isnull=False).filter(Q(ticket_created_at__lt=start_date) | (
                        Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                if tickets_open_before or tickets_open_in or tickets_closed or tickets_left_opened:

                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'tickets_open_before': tickets_open_before,
                        'tickets_open_before_list': tickets_open_before_list,
                        'tickets_open_in': tickets_open_in,
                        'tickets_open_in_list': tickets_open_in_list,
                        'tickets_closed': tickets_closed,
                        'tickets_closed_list': tickets_closed_list,
                        'tickets_left_opened': tickets_left_opened,
                        'tickets_left_opened_list': tickets_left_opened_list,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_assign_to_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'Inactive':
                try:
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_assign_to_id', 'ticket_assign_to__display_name').filter(
                        ticket_assign_to__isnull=False).filter(ticket_assign_to__is_active=False).filter(
                        ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                        tcount=Count('ticket_id')
                    )
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                tickets_open_before = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_assign_to_id',
                                                            'ticket_assign_to__display_name').filter(
                    ticket_assign_to_id__isnull=False).filter(ticket_assign_to__is_active=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_before = sorted(tickets_open_before, key=itemgetter('tcount'), reverse=True)
                tickets_open_before_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_assign_to_id', flat=True).filter(
                    ticket_assign_to_id__isnull=False).filter(ticket_assign_to__is_active=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_open_in = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_assign_to_id', 'ticket_assign_to__first_name',
                                                        'ticket_assign_to__last_name').filter(
                    ticket_assign_to_id__isnull=False).filter(ticket_assign_to__is_active=False).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))
                tickets_open_in = sorted(tickets_open_in, key=itemgetter('tcount'), reverse=True)
                tickets_open_in_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_assign_to_id', flat=True).filter(
                    ticket_assign_to_id__isnull=False).filter(ticket_assign_to__is_active=False).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))

                tickets_closed = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_assign_to_id', 'ticket_assign_to__display_name').filter(
                    ticket_assign_to_id__isnull=False).filter(ticket_assign_to__is_active=False).filter(
                    ticket_status=1).filter(
                    ticket_closed_at__gte=start_date).filter(ticket_closed_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_closed = sorted(tickets_closed, key=itemgetter('tcount'), reverse=True)
                tickets_closed_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_assign_to_id', flat=True).filter(
                    ticket_assign_to_id__isnull=False).filter(ticket_assign_to__is_active=False).filter(
                    ticket_status=1).filter(
                    ticket_closed_at__gte=start_date).filter(ticket_closed_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_left_opened = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_assign_to_id',
                                                            'ticket_assign_to__display_name').filter(
                    ticket_assign_to_id__isnull=False).filter(ticket_assign_to__is_active=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_left_opened = sorted(tickets_left_opened, key=itemgetter('tcount'), reverse=True)
                tickets_left_opened_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_assign_to_id', flat=True).filter(
                    ticket_assign_to_id__isnull=False).filter(ticket_assign_to__is_active=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                if tickets_open_before or tickets_open_in or tickets_closed or tickets_left_opened:

                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'tickets_open_before': tickets_open_before,
                        'tickets_open_before_list': tickets_open_before_list,
                        'tickets_open_in': tickets_open_in,
                        'tickets_open_in_list': tickets_open_in_list,
                        'tickets_closed': tickets_closed,
                        'tickets_closed_list': tickets_closed_list,
                        'tickets_left_opened': tickets_left_opened,
                        'tickets_left_opened_list': tickets_left_opened_list,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_inactive_user_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'IssueSubTypes':
                try:
                    ttypes = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_subtype1__parent_id', 'ticket_subtype1_id',
                                                   'ticket_subtype1__ttype_name').filter(
                        ticket_subtype1_id__isnull=False).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                        tcount=Count('ticket_id')).order_by('ticket_type__ttype_name', 'ticket_subtype1__ttype_name')
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                try:
                    parent_ttypes = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_type_id', 'ticket_type__ttype_name').filter(
                        ticket_subtype1_id__isnull=False).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(tcount=Count('ticket_id'))
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                tickets_open_before = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_subtype1_id', 'ticket_subtype1__ttype_name').filter(
                    ticket_subtype1_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_before = sorted(tickets_open_before, key=itemgetter('tcount'), reverse=True)
                tickets_open_before_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_subtype1_id', flat=True).filter(
                    ticket_subtype1_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_open_in = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_subtype1_id', 'ticket_subtype1__ttype_name').filter(
                    ticket_subtype1_id__isnull=False).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_in = sorted(tickets_open_in, key=itemgetter('tcount'), reverse=True)
                tickets_open_in_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_subtype1_id', flat=True).filter(
                    ticket_subtype1_id__isnull=False).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_closed = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_subtype1_id', 'ticket_subtype1__ttype_name').filter(
                    ticket_subtype1_id__isnull=False).filter(
                    ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                    ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_closed = sorted(tickets_closed, key=itemgetter('tcount'), reverse=True)
                tickets_closed_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_subtype1_id', flat=True).filter(
                    ticket_subtype1_id__isnull=False).filter(
                    ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                    ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_left_opened = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_subtype1_id', 'ticket_subtype1__ttype_name').filter(
                    ticket_subtype1_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_left_opened = sorted(tickets_left_opened, key=itemgetter('tcount'), reverse=True)
                tickets_left_opened_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_subtype1_id', flat=True).filter(
                    ticket_subtype1_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                parent_open_before = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_subtype1__parent_id').filter(
                    ticket_subtype1_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))

                parent_open_in = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_subtype1__parent_id').filter(
                    ticket_subtype1_id__isnull=False).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                parent_closed = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_subtype1__parent_id').filter(
                    ticket_subtype1_id__isnull=False).filter(
                    ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                    ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                parent_left_opened = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_subtype1__parent_id').filter(
                    ticket_subtype1_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                if tickets_open_before or tickets_open_in or tickets_closed or tickets_left_opened:

                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'tickets_open_before': tickets_open_before,
                        'tickets_open_before_list': tickets_open_before_list,
                        'tickets_open_in': tickets_open_in,
                        'tickets_open_in_list': tickets_open_in_list,
                        'tickets_closed': tickets_closed,
                        'tickets_closed_list': tickets_closed_list,
                        'tickets_left_opened': tickets_left_opened,
                        'tickets_left_opened_list': tickets_left_opened_list,
                        'parent_open_before': parent_open_before,
                        'parent_open_in': parent_open_in,
                        'parent_closed': parent_closed,
                        'parent_left_opened': parent_left_opened,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'ttypes': ttypes,
                        'parent_ttypes': parent_ttypes,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_ticket_subtype_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'IssueTypes':
                try:
                    ttypes = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_type_id', 'ticket_type__ttype_name').filter(
                        ticket_type_id__isnull=False).filter(
                        ticket_type__has_parent=0).filter(ticket_type__ttype_is_delete=0).filter(
                        ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(tcount=Count('ticket_id')).order_by('ticket_type__ttype_name')
                    ttypes_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_type_id', flat=True).filter(
                        ticket_type_id__isnull=False).filter(ticket_type__has_parent=0).filter(
                        ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                        tcount=Count('ticket_id')
                    )
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                tickets_open_before = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_type_id', 'ticket_type__ttype_name').filter(
                    ticket_type_id__isnull=False).filter(ticket_type__ttype_is_delete=0).filter(
                    ticket_type__has_parent=0).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_before = sorted(tickets_open_before, key=itemgetter('tcount'), reverse=True)
                tickets_open_before_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_type_id', flat=True).filter(
                    ticket_type_id__isnull=False).filter(ticket_type__ttype_is_delete=0).filter(
                    ticket_type__has_parent=0).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_open_in = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_type_id', 'ticket_type__ttype_name').filter(
                    ticket_type_id__isnull=False).filter(ticket_type__ttype_is_delete=0).filter(
                    ticket_type__has_parent=0).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_in = sorted(tickets_open_in, key=itemgetter('tcount'), reverse=True)
                tickets_open_in_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_type_id', flat=True).filter(
                    ticket_type_id__isnull=False).filter(ticket_type__ttype_is_delete=0).filter(
                    ticket_type__has_parent=0).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_closed = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_type_id', 'ticket_type__ttype_name').filter(
                    ticket_type_id__isnull=False).filter(
                    ticket_type__ttype_is_delete=0).filter(ticket_status=1).filter(ticket_type__has_parent=0).filter(
                    ticket_closed_at__gte=start_date).filter(
                    ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_closed = sorted(tickets_closed, key=itemgetter('tcount'), reverse=True)
                tickets_closed_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_type_id', flat=True).filter(
                    ticket_type_id__isnull=False).filter(
                    ticket_type__ttype_is_delete=0).filter(ticket_type__has_parent=0).filter(ticket_status=1).filter(
                    ticket_closed_at__gte=start_date).filter(ticket_closed_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_left_opened = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_type_id', 'ticket_type__ttype_name').filter(
                    ticket_type_id__isnull=False).filter(ticket_type__ttype_is_delete=0).filter(
                    ticket_type__has_parent=0).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_left_opened = sorted(tickets_left_opened, key=itemgetter('tcount'), reverse=True)
                tickets_left_opened_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_type_id', flat=True).filter(
                    ticket_type_id__isnull=False).filter(ticket_type__ttype_is_delete=0).filter(
                    ticket_type__has_parent=0).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id'))

                # ttype_tickets_closed = TicketType.objects.values('ttype_id', 'ttype_name').filter(has_parent=0).filter(
                #     ttype_is_delete=0).filter(ticketType__ticket_closed_at__range=(start_date, end_date)).annotate(
                #     tcount=Count('ticketType__ticket_id')).order_by('-tcount')
                #
                #
                # for ttype in ttypes_list:
                #     if ttype not in tickets_closed_list:
                #         print(ttype)
                #         tickets_closed_list.append(ttype)

                # print(ttypes_list)
                # print(tickets_open_before)
                # print(tickets_open_before_list)
                # print(tickets_open_in)
                # print(tickets_open_in_list)
                # print(tickets_closed)
                # print(tickets_closed_list)
                # print(tickets_left_opened)
                # print(tickets_left_opened_list)

                if tickets_open_before or tickets_open_in or tickets_closed or tickets_left_opened:

                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'tickets_open_before': tickets_open_before,
                        'tickets_open_before_list': tickets_open_before_list,
                        'tickets_open_in': tickets_open_in,
                        'tickets_open_in_list': tickets_open_in_list,
                        'tickets_closed': tickets_closed,
                        'tickets_closed_list': tickets_closed_list,
                        'tickets_left_opened': tickets_left_opened,
                        'tickets_left_opened_list': tickets_left_opened_list,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'ttypes': ttypes,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_ticket_type_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'Locations':
                
                try:
                    accounts = Ticket.objects.filter(ticket_org_id=org_id).values('account_id','account__acc_name').filter(
                        account_id__isnull=False).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                        tcount=Count('ticket_id')).order_by('account__acc_name')
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                tickets_open_before = Ticket.objects.filter(ticket_org_id=org_id).values('account_id', 'account__acc_name').filter(
                    account_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (
                            Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                
                tickets_open_before = sorted(tickets_open_before, key=itemgetter('tcount'), reverse=True)
                tickets_open_before_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('account_id', flat=True).filter(
                    account_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_open_in = Ticket.objects.filter(ticket_org_id=org_id).values('account_id', 'account__acc_name').filter(
                    account_id__isnull=False).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_in = sorted(tickets_open_in, key=itemgetter('tcount'), reverse=True)
                tickets_open_in_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('account_id', flat=True).filter(
                    account_id__isnull=False).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_closed = Ticket.objects.filter(ticket_org_id=org_id).values('account_id', 'account__acc_name').filter(
                    account_id__isnull=False).filter(
                    ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                    ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_closed = sorted(tickets_closed, key=itemgetter('tcount'), reverse=True)
                tickets_closed_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('account_id', flat=True).filter(
                    account_id__isnull=False).filter(
                    ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                    ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_left_opened = Ticket.objects.filter(ticket_org_id=org_id).values('account_id', 'account__acc_name').filter(
                    account_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_left_opened = sorted(tickets_left_opened, key=itemgetter('tcount'), reverse=True)
                tickets_left_opened_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('account_id', flat=True).filter(
                    account_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                
                if tickets_open_before or tickets_open_in or tickets_closed or tickets_left_opened:
                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'tickets_open_before': tickets_open_before,
                        'tickets_open_before_list': tickets_open_before_list,
                        'tickets_open_in': tickets_open_in,
                        'tickets_open_in_list': tickets_open_in_list,
                        'tickets_closed': tickets_closed,
                        'tickets_closed_list': tickets_closed_list,
                        'tickets_left_opened': tickets_left_opened,
                        'tickets_left_opened_list': tickets_left_opened_list,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'accounts': accounts,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_account_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'NextActionBy':
                try:
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_next_action_id', 'ticket_next_action__display_name').filter(
                        ticket_next_action_id__isnull=False).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                        tcount=Count('ticket_id')).order_by('ticket_next_action__display_name')
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                tickets_open_before = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_next_action_id',
                                                            'ticket_next_action__display_name').filter(
                    ticket_next_action_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_before = sorted(tickets_open_before, key=itemgetter('tcount'), reverse=True)
                tickets_open_before_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_next_action_id', flat=True).filter(
                    ticket_next_action_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_open_in = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_next_action_id',
                                                        'ticket_next_action__display_name').filter(
                    ticket_next_action_id__isnull=False).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_in = sorted(tickets_open_in, key=itemgetter('tcount'), reverse=True)
                tickets_open_in_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_next_action_id', flat=True).filter(
                    ticket_next_action_id__isnull=False).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_closed = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_next_action_id',
                                                       'ticket_next_action__display_name').filter(
                    ticket_next_action_id__isnull=False).filter(ticket_status=1).filter(
                    ticket_closed_at__gte=start_date).filter(
                    ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_closed = sorted(tickets_closed, key=itemgetter('tcount'), reverse=True)
                tickets_closed_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_next_action_id', flat=True).filter(
                    ticket_next_action_id__isnull=False).filter(ticket_status=1).filter(
                    ticket_closed_at__gte=start_date).filter(
                    ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_left_opened = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_next_action_id',
                                                            'ticket_next_action__display_name').filter(
                    ticket_next_action_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_left_opened = sorted(tickets_left_opened, key=itemgetter('tcount'), reverse=True)
                tickets_left_opened_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_next_action_id', flat=True).filter(
                    ticket_next_action_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                if tickets_open_before or tickets_open_in or tickets_closed or tickets_left_opened:

                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'tickets_open_before': tickets_open_before,
                        'tickets_open_before_list': tickets_open_before_list,
                        'tickets_open_in': tickets_open_in,
                        'tickets_open_in_list': tickets_open_in_list,
                        'tickets_closed': tickets_closed,
                        'tickets_closed_list': tickets_closed_list,
                        'tickets_left_opened': tickets_left_opened,
                        'tickets_left_opened_list': tickets_left_opened_list,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_next_action_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'Organizations':
                try:
                    organizations = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_org_id', 'ticket_org__org_name').filter(
                        ticket_org_id__isnull=False).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                        tcount=Count('ticket_id')).order_by('ticket_org__org_name')
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                tickets_open_before = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_org_id', 'ticket_org__org_name').filter(
                    ticket_org_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_before = sorted(tickets_open_before, key=itemgetter('tcount'), reverse=True)
                tickets_open_before_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_org_id', flat=True).filter(
                    ticket_org_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_open_in = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_org_id', 'ticket_org__org_name').filter(
                    ticket_org_id__isnull=False).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_in = sorted(tickets_open_in, key=itemgetter('tcount'), reverse=True)
                tickets_open_in_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_org_id', flat=True).filter(
                    ticket_org_id__isnull=False).filter(
                    ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_closed = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_org_id', 'ticket_org__org_name').filter(
                    ticket_org_id__isnull=False).filter(
                    ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                    ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_closed = sorted(tickets_closed, key=itemgetter('tcount'), reverse=True)
                tickets_closed_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_org_id', flat=True).filter(
                    ticket_org_id__isnull=False).filter(
                    ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                    ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_left_opened = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_org_id', 'ticket_org__org_name').filter(
                    ticket_org_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_left_opened = sorted(tickets_left_opened, key=itemgetter('tcount'), reverse=True)
                tickets_left_opened_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_org_id', flat=True).filter(
                    ticket_org_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                if tickets_open_before or tickets_open_in or tickets_closed or tickets_left_opened:

                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'tickets_open_before': tickets_open_before,
                        'tickets_open_before_list': tickets_open_before_list,
                        'tickets_open_in': tickets_open_in,
                        'tickets_open_in_list': tickets_open_in_list,
                        'tickets_closed': tickets_closed,
                        'tickets_closed_list': tickets_closed_list,
                        'tickets_left_opened': tickets_left_opened,
                        'tickets_left_opened_list': tickets_left_opened_list,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'organizations': organizations,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_organization_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'SubmittedBy':
                try:
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_caller_id', 'ticket_caller__display_name').filter(
                        ticket_caller_id__isnull=False).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                        tcount=Count('ticket_id')).order_by('ticket_caller__display_name')
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))

                tickets_open_before = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_caller_id', 'ticket_caller__display_name').filter(
                    ticket_caller_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_before = sorted(tickets_open_before, key=itemgetter('tcount'), reverse=True)
                tickets_open_before_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_caller_id', flat=True).filter(
                    ticket_caller_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_open_in = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_caller_id', 'ticket_caller__display_name').filter(
                    ticket_caller_id__isnull=False).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_in = sorted(tickets_open_in, key=itemgetter('tcount'), reverse=True)
                tickets_open_in_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_caller_id', flat=True).filter(
                    ticket_caller_id__isnull=False).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_closed = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_caller_id', 'ticket_caller__display_name').filter(
                    ticket_caller_id__isnull=False).filter(
                    ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                    ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_closed = sorted(tickets_closed, key=itemgetter('tcount'), reverse=True)
                tickets_closed_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_caller_id', flat=True).filter(
                    ticket_caller_id__isnull=False).filter(
                    ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                    ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_left_opened = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_caller_id', 'ticket_caller__display_name').filter(
                    ticket_caller_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_left_opened = sorted(tickets_left_opened, key=itemgetter('tcount'), reverse=True)
                tickets_left_opened_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_caller_id', flat=True).filter(
                    ticket_caller_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                if tickets_open_before or tickets_open_in or tickets_closed or tickets_left_opened:

                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'tickets_open_before': tickets_open_before,
                        'tickets_open_before_list': tickets_open_before_list,
                        'tickets_open_in': tickets_open_in,
                        'tickets_open_in_list': tickets_open_in_list,
                        'tickets_closed': tickets_closed,
                        'tickets_closed_list': tickets_closed_list,
                        'tickets_left_opened': tickets_left_opened,
                        'tickets_left_opened_list': tickets_left_opened_list,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_caller_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'DepartmentsSubmit':
                try:
                    departments = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_caller__user_dep_id',
                                                        'ticket_caller__user_dep__dep_name').filter(
                        ticket_caller__isnull=False).filter(ticket_caller__user_dep__isnull=False).filter(
                        ticket_is_delete=0).annotate(
                        tcount=Count('ticket_id')).order_by('ticket_caller__user_dep__dep_name')
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                tickets_open_before = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_caller__user_dep_id',
                                                            'ticket_caller__user_dep__dep_name').filter(
                    ticket_caller__user_dep_id__isnull=False).filter(ticket_created_at__lt=start_date).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_before = sorted(tickets_open_before, key=itemgetter('tcount'), reverse=True)
                tickets_open_before_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_caller__user_dep_id', flat=True).filter(
                    ticket_caller__user_dep_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) & (Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                        ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_open_in = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_caller__user_dep_id',
                                                        'ticket_caller__user_dep__dep_name').filter(
                    ticket_caller__user_dep_id__isnull=False).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open_in = sorted(tickets_open_in, key=itemgetter('tcount'), reverse=True)
                tickets_open_in_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_caller__user_dep_id', flat=True).filter(
                    ticket_caller__user_dep_id__isnull=False).filter(ticket_created_at__gte=start_date).filter(
                    ticket_created_at__lte=end_date).filter(ticket_is_delete=0).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_closed = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_caller__user_dep_id',
                                                       'ticket_caller__user_dep__dep_name').filter(
                    ticket_caller__user_dep_id__isnull=False).filter(ticket_status=1).filter(
                    ticket_closed_at__gte=start_date).filter(ticket_closed_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_closed = sorted(tickets_closed, key=itemgetter('tcount'), reverse=True)
                tickets_closed_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_caller__user_dep_id', flat=True).filter(
                    ticket_caller__user_dep_id__isnull=False).filter(ticket_status=1).filter(
                    ticket_closed_at__gte=start_date).filter(ticket_closed_at__lte=end_date).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )

                tickets_left_opened = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_caller__user_dep_id',
                                                            'ticket_caller__user_dep__dep_name').filter(
                    ticket_caller__user_dep_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_left_opened = sorted(tickets_left_opened, key=itemgetter('tcount'), reverse=True)
                tickets_left_opened_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_caller__user_dep_id', flat=True).filter(
                    ticket_caller__user_dep_id__isnull=False).filter(
                    Q(ticket_created_at__lt=start_date) | (
                                Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                    Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(ticket_closed_at__gt=end_date)).filter(
                    ticket_is_delete=0).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                if tickets_open_before or tickets_open_in or tickets_closed or tickets_left_opened:

                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'tickets_open_before': tickets_open_before,
                        'tickets_open_before_list': tickets_open_before_list,
                        'tickets_open_in': tickets_open_in,
                        'tickets_open_in_list': tickets_open_in_list,
                        'tickets_closed': tickets_closed,
                        'tickets_closed_list': tickets_closed_list,
                        'tickets_left_opened': tickets_left_opened,
                        'tickets_left_opened_list': tickets_left_opened_list,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'departments': departments,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_submit_department_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'IncidentsByRep':
                try:
                    total_tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_created_by__isnull=False).filter(
                        ticket_is_delete=0).filter(
                        Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date)).filter(userFilter).filter(**accountFilter).count()
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_created_by_id', 'ticket_created_by__display_name').filter(
                        ticket_created_by__isnull=False).filter(ticket_is_delete=0).filter(
                        Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date)).filter(userFilter).filter(**accountFilter).annotate(
                        tcount=Count('ticket_id')).order_by(
                        'ticket_created_by__display_name')
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                tickets_open = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_created_by_id', 'ticket_created_by__display_name').filter(
                    ticket_created_by__isnull=False).filter(ticket_is_delete=0).filter(ticket_status=0).filter(
                    Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date)).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                tickets_open = sorted(tickets_open, key=itemgetter('tcount'), reverse=True)
                tickets_open_list = Ticket.objects.filter(ticket_org_id=org_id).values_list('ticket_created_by_id', flat=True).filter(
                    ticket_created_by__isnull=False).filter(ticket_is_delete=0).filter(ticket_status=0).filter(
                    Q(ticket_created_at__gte=start_date) | Q(ticket_created_at__lte=end_date)).filter(userFilter).filter(**accountFilter).annotate(
                    tcount=Count('ticket_id')
                )
                if tickets:

                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'tickets_open': tickets_open,
                        'tickets_open_list': tickets_open_list,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'total_tickets': total_tickets,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_support_rep_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'OrgTimes':

                try:
                    org_labour_hours = TicketNote.objects.values('note_ticket__ticket_org_id',
                                                                 'note_ticket__ticket_org__org_name').filter(
                        note_is_delete=0).filter(note_ticket__ticket_org_id__isnull=False).filter(
                        note_created_at__gte=start_date).filter(note_created_at__lte=end_date).annotate(
                        thours=Sum(Cast('tnote_laborhour_hours', IntegerField())) + Sum(
                            Cast('tnote_laborhour_minutes', IntegerField())) / 60,
                        tminutes=Sum(Cast('tnote_laborhour_minutes', IntegerField())) % 60).order_by(
                        'note_ticket__ticket_org__org_name')
                except TicketNote.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                try:
                    total_labour_hours = TicketNote.objects.filter(note_is_delete=0).filter(
                        note_ticket__ticket_org_id__isnull=False).filter(
                        note_created_at__gte=start_date).filter(note_created_at__lte=end_date).aggregate(
                        thours=Sum(Cast('tnote_laborhour_hours', IntegerField())) + Sum(
                            Cast('tnote_laborhour_minutes', IntegerField())) / 60,
                        tminutes=Sum(Cast('tnote_laborhour_minutes', IntegerField())) % 60
                    )
                except TicketNote.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                if org_labour_hours:
                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'org_labour_hours': org_labour_hours,
                        'total_hours': total_labour_hours['thours'],
                        'total_minutes': total_labour_hours['tminutes'],
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_org_labourhour_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'UserTimes':
                try:
                    # user_labour_hours = TicketNote.objects.values('note_ticket__ticket_caller_id',
                    #                                               'note_ticket__ticket_caller__display_name',
                    #                                               'note_ticket__ticket_caller__first_name',
                    #                                               'note_ticket__ticket_caller__last_name',
                    #                                               'tnote_laborhour_hours', 'tnote_laborhour_minutes',
                    #                                               'labour_hours', 'note_created_at').filter(
                    #     note_is_delete=0).filter(note_ticket__ticket_caller_id__isnull=False).filter(
                    #     note_created_at__gte=start_date).filter(note_created_at__lte=end_date).order_by(
                    #     'note_ticket__ticket_caller_id')
                    user_labour_hours = TicketNote.objects.values('note_created_by_id',
                                                                  'note_created_by__display_name').filter(
                        note_is_delete=0).filter(note_created_by_id__isnull=False).filter(
                        note_created_at__gte=start_date).filter(note_created_at__lte=end_date).annotate(
                        thours=Sum(Cast('tnote_laborhour_hours', IntegerField())) + Sum(
                            Cast('tnote_laborhour_minutes', IntegerField())) / 60,
                        tminutes=Sum(Cast('tnote_laborhour_minutes', IntegerField())) % 60).order_by(
                        'note_created_by__display_name')
                except TicketNote.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                try:
                    total_labour_hours = TicketNote.objects.filter(note_is_delete=0).filter(
                        note_created_by_id__isnull=False).filter(
                        note_created_at__gte=start_date).filter(note_created_at__lte=end_date).aggregate(
                        thours=Sum(Cast('tnote_laborhour_hours', IntegerField())) + Sum(
                            Cast('tnote_laborhour_minutes', IntegerField())) / 60,
                        tminutes=Sum(Cast('tnote_laborhour_minutes', IntegerField())) % 60
                    )
                except TicketNote.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                if user_labour_hours:
                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'user_labour_hours': user_labour_hours,
                        'total_hours': total_labour_hours['thours'],
                        'total_minutes': total_labour_hours['tminutes'],
                        'sort_by': sort_by
                    }
                    return render(request, 'itrak/Reports/summary_report_user_labourhour_result.html', context)
                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'OrgTimeOpen':
                try:
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_org_id', 'ticket_org__org_name', 'ticket_status',
                                                    'ticket_created_at',
                                                    'ticket_id', 'ticket_closed_at').filter(ticket_is_delete=0).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_org_id')
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                last_id = 0
                tot_hours = {}
                g_total_hours = 0
                if tickets:
                    for ticket in tickets:

                        if ticket['ticket_org_id'] != last_id:
                            if ticket['ticket_status'] == 0:
                                delta = datetime.now(timezone.utc) - ticket['ticket_created_at']
                                days, seconds = delta.days, delta.seconds
                                hours = days * 24 + seconds // 3600
                                minutes = (seconds % 3600) // 60
                            else:
                                delta = ticket['ticket_closed_at'] - ticket['ticket_created_at']
                                days, seconds = delta.days, delta.seconds
                                hours = days * 24 + seconds // 3600
                                minutes = (seconds % 3600) // 60

                            total_hours = round((hours) + (minutes / 100), 2)
                            tot_hours.update({ticket['ticket_org_id']: total_hours})
                        else:
                            if ticket['ticket_status'] == 0:
                                delta = datetime.now(timezone.utc) - ticket['ticket_created_at']
                                days, seconds = delta.days, delta.seconds
                                hours = days * 24 + seconds // 3600
                                minutes = (seconds % 3600) // 60
                            else:
                                delta = ticket['ticket_closed_at'] - ticket['ticket_created_at']
                                days, seconds = delta.days, delta.seconds
                                hours = days * 24 + seconds // 3600
                                minutes = (seconds % 3600) // 60

                            total_hours = round(tot_hours[ticket['ticket_org_id']] + (hours) + (minutes / 100), 2)
                            tot_hours.update({ticket['ticket_org_id']: total_hours})

                        g_total_hours += round(((hours) + (minutes / 100)), 2)
                        last_id = ticket['ticket_org_id']

                if tot_hours:
                    g_total_hours = round(g_total_hours, 2)
                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tot_hours': tot_hours,
                        'g_total_hours': g_total_hours,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_org_totaltime_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'UserTimeOpen':
                try:
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_caller_id', 'ticket_caller__display_name', 'ticket_status',
                                                    'ticket_created_at',
                                                    'ticket_id', 'ticket_closed_at').filter(ticket_is_delete=0).filter(
                        ticket_caller_id__isnull=False).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_caller_id')
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                last_id = 0
                tot_hours, userDict = {}, {}
                g_total_hours = 0
                if tickets:
                    for ticket in tickets:

                        if ticket['ticket_caller_id'] != last_id:
                            if ticket['ticket_status'] == 0:
                                delta = datetime.now(timezone.utc) - ticket['ticket_created_at']
                                days, seconds = delta.days, delta.seconds
                                hours = days * 24 + seconds // 3600
                                minutes = (seconds % 3600) // 60
                            else:
                                delta = ticket['ticket_closed_at'] - ticket['ticket_created_at']
                                days, seconds = delta.days, delta.seconds
                                hours = days * 24 + seconds // 3600
                                minutes = (seconds % 3600) // 60

                            total_hours = round((hours) + (minutes / 100), 2)
                            tot_hours.update({ticket['ticket_caller_id']: total_hours})
                        else:
                            if ticket['ticket_status'] == 0:
                                delta = datetime.now(timezone.utc) - ticket['ticket_created_at']
                                days, seconds = delta.days, delta.seconds
                                hours = days * 24 + seconds // 3600
                                minutes = (seconds % 3600) // 60
                            else:
                                delta = ticket['ticket_closed_at'] - ticket['ticket_created_at']
                                days, seconds = delta.days, delta.seconds
                                hours = days * 24 + seconds // 3600
                                minutes = (seconds % 3600) // 60

                            total_hours = round(tot_hours[ticket['ticket_caller_id']] + (hours) + (minutes / 100), 2)
                            tot_hours.update({ticket['ticket_caller_id']: total_hours})
                        full_name = ticket['ticket_caller_id']

                        g_total_hours += round(((hours) + (minutes / 100)), 2)
                        last_id = ticket['ticket_caller_id']
                if tot_hours:
                    g_total_hours = round(g_total_hours, 2)
                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tot_hours': tot_hours,
                        'g_total_hours': g_total_hours,
                        'sort_by': sort_by
                    }

                    return render(request, 'itrak/Reports/summary_report_user_totaltime_result.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
        else:
            return render_to_response('itrak/page-404.html')


# Get Summary Report Result End#


# Get Summary Report Ticket List Start#

@active_user_required
def getSummaryReportTicketList(request):
    if request.method == 'GET':
        org_id = request.user.user_org_id
        report_id = request.GET.get('reportId')
        start_date = datetime.strptime(request.GET.get('startDate'), '%m/%d/%Y').strftime('%Y-%m-%d')
        end_date = datetime.strptime(request.GET.get('endDate') + " 23:59:59", '%m/%d/%Y %H:%M:%S').strftime(
            '%Y-%m-%d %H:%M:%S')
        # Instead of having an error on your server,
        # your user will get a 404 meaning that he tries to access a non existing resource.

        reportdict = {
            "OrgContractTime": "Contract Time Used by Organization - Select Date Range",
            "OrgContractExp": "Contracts Expiring by Organization - Select Date Range",
            "OrgContractTimeOverage": "Contract Overage Time by Organization - Select Date Range",
            "IssuesPriority": "Tickets by Priority - Select Date Range",
            "IssuesPriorityByDay": "Tickets by Priority (Day) - Select Date Range",
            "IssuesPriorityByMonth": "Tickets by Priority (Month) - Select Date Range",
            "AssignedTo": "Tickets by Assigned To - Select Date Range",
            "Inactive": "Tickets by Inactive User - Results",
            "IssueSubTypes": "Tickets by Ticket Subtype - Select Date Range",
            "IssueTypes": "Tickets by Ticket Type - Select Date Range",
            "Locations": "Tickets by Account - Select Date Range",
            "NextActionBy": "Tickets by Next Action - Select Date Range",
            "Organizations": "Tickets by Organization - Select Date Range",
            "SubmittedBy": "Tickets by Submitter - Select Date Range",
            "DepartmentsSubmit": "Tickets by Submitting Department - Select Date Range",
            "OrgTimes": "Tickets by Rep - Select Date Range",
            "IncidentsByRep": "Labor Hours by Organization - Select Date Range",
            "UserTimes": "Labor Hours by User - Select Date Range",
            "OrgTimeOpen": "Total Time Open by Organization - Select Date Range",
            "UserTimeOpen": "Total Time Open by User - Select Date getReportDateRange"
        }

        #User-Account Mapping Check Start
        accountsList = getAccountIDsOfCurrentUser(request)

        if len(accountsList) > 0:
            accountFilter = {'account_id__in':accountsList}
        else: 
            accountFilter = {}

        user_id = request.user.id
        user_type = userType(request) 
        accountsList = getMappedUserIDsWithCurrentUer(request)
        if user_type == 'manager':
            userFilter = Q(ticket_assign_to_id__in = accountsList) | Q(ticket_caller_id__in = accountsList) | Q(ticket_next_action_id__in=accountsList) | Q(ticket_created_by_id__in=accountsList)
        elif user_type == 'enduser':
            userFilter = Q(ticket_assign_to_id=user_id) | Q(ticket_caller_id=user_id) | Q(ticket_next_action_id=user_id) | Q(ticket_created_by_id=user_id)
        else:
            userFilter = Q(pk__isnull=False)

        #User-Account Mapping Check End

        if report_id in reportdict:
            if report_id == 'IssuesPriority' or report_id == 'IssuesPriorityByDay' or report_id == 'IssuesPriorityByMonth' or report_id == 'AssignedTo' or report_id == 'Inactive' or report_id == 'IssueSubTypes' or report_id == 'IssueTypes' or report_id == 'Locations' or report_id == 'NextActionBy' or report_id == 'Organizations' or report_id == 'SubmittedBy' or report_id == 'DepartmentsSubmit' or report_id == 'IncidentsByRep' or report_id == 'OrgTimes' or report_id == 'UserTimes' or report_id == 'OrgTimeOpen' or report_id == 'UserTimeOpen':
                if 'rowId' in request.GET:
                    rowId = request.GET.get('rowId')
                if 'columnId' in request.GET:
                    columnId = request.GET.get('columnId')
            if report_id == 'IssuesPriority':
                try:
                    reportTitle = Priority.objects.filter(user_org_id=org_id).values_list('priority_name', flat=True).get(pk=rowId)
                except Priority.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                if columnId == 'OB':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) & (
                                    Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                                ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )
                if columnId == 'OI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )
                if columnId == 'CI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(ticket_status=1).filter(
                        ticket_closed_at__gte=start_date).filter(ticket_closed_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )
                if columnId == 'LO':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) | (
                                    Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                        Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(
                            ticket_closed_at__gt=end_date)).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if tickets:
                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'reportTitle': reportTitle,
                    }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'IssuesPriorityByDay':
                if 'rowId' in locals():
                    try:
                        reportTitle = Priority.objects.filter(user_org_id=org_id).values_list('priority_name', flat=True).get(pk=rowId)
                    except Priority.DoesNotExist:
                        messages.error(request, 'Request Failed! No Record Found.')
                        # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                        return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                else:
                    reportTitle = 'Totals'

                if columnId == 'TOT':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'SUN':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__week_day=1).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'MON':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__week_day=2).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TUE':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__week_day=3).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'WED':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__week_day=4).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'THU':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__week_day=5).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'FRI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__week_day=6).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'SAT':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__week_day=7).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TSUN':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__week_day=1).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TMON':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__week_day=2).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TTUE':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__week_day=3).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TWED':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__week_day=4).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TTHU':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__week_day=5).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TFRI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__week_day=6).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TSAT':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__week_day=7).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TTOT':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if tickets:

                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'reportTitle': reportTitle,
                    }
                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'IssuesPriorityByMonth':
                
                if 'rowId' in locals():
                    try:
                        reportTitle = Priority.objects.filter(user_org_id=org_id).values_list('priority_name', flat=True).get(pk=rowId)
                    except Priority.DoesNotExist:
                        messages.error(request, 'Request Failed! No Record Found.')
                        # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                        return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                else:
                    reportTitle = 'Totals'
                
                # TIcket Month Wises and Priority Wise Calculation Start
                if columnId == 'TOT':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'JAN':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__month=1).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )
                if columnId == 'FEB':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__month=2).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'MAR':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__month=3).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'APR':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__month=4).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'MAY':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__month=5).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'JUN':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__month=6).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'JUL':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__month=7).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'AUG':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__month=8).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'SEP':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__month=9).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'OCT':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__month=10).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'NOV':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__month=11).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'DEC':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id=rowId).filter(
                        ticket_created_at__month=12).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                # TIcket Month Wises and Priority Wise Calculation ENd

                # TIcket Total Month Wises Calculation Start

                if columnId == 'TJAN':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__month=1).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TFEB':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__month=2).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TMAR':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__month=3).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TAPR':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__month=4).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TMAY':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__month=5).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TJUN':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__month=6).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TJUL':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__month=7).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TAUG':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__month=8).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TSEP':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__month=9).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TOCT':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__month=10).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TNOV':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__month=11).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TDEC':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__month=12).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TTOT':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(priority_id__isnull=False).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if tickets:
                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'reportTitle': reportTitle,
                    }
                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'AssignedTo':
                try:
                    reportTitle = User.objects.filter(user_org_id=org_id).values_list('display_name', flat=True).get(pk=rowId)
                except User.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                if columnId == 'OB':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_assign_to_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) & (
                                    Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                                ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )
                if columnId == 'OI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_assign_to_id=rowId).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )
                if columnId == 'CI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_assign_to_id=rowId).filter(ticket_status=1).filter(
                        ticket_closed_at__gte=start_date).filter(ticket_closed_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )
                if columnId == 'LO':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_assign_to_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) | (
                                    Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                        Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(
                            ticket_closed_at__gt=end_date)).filter(ticket_is_delete=0).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if tickets:
                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'reportTitle': reportTitle,
                    }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'Inactive':
                try:
                    reportTitle = User.objects.filter(user_org_id=org_id).values_list('display_name', flat=True).get(pk=rowId)
                except User.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))

                if columnId == 'OB':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_assign_to_id=rowId).filter(
                        ticket_assign_to__is_active=False).filter(
                        Q(ticket_created_at__lt=start_date) & (
                                    Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                                ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'OI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_assign_to_id=rowId).filter(
                        ticket_assign_to__is_active=False).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(**accountFilter).filter(userFilter).filter(
                        ticket_is_delete=0).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'CI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_assign_to_id=rowId).filter(ticket_status=1).filter(
                        ticket_assign_to__is_active=False).filter(ticket_closed_at__gte=start_date).filter(
                        ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'LO':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_assign_to_id=rowId).filter(
                        ticket_assign_to__is_active=False).filter(
                        Q(ticket_created_at__lt=start_date) | (
                                    Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                        Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(
                            ticket_closed_at__gt=end_date)).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if tickets:
                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'reportTitle': reportTitle,
                    }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'IssueSubTypes':
                if 'rowId' in locals():
                    try:
                        reportTitle = TicketType.objects.filter(user_org_id=org_id).values_list('ttype_name', flat=True).get(pk=rowId)
                    except TicketType.DoesNotExist:
                        messages.error(request, 'Request Failed! No Record Found.')
                        # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                        return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                else:
                    reportTitle = 'Totals'

                if columnId == 'OB':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_subtype1_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) & (
                                    Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                                ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'OI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_subtype1_id=rowId).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'CI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_subtype1_id=rowId).filter(
                        ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                        ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'LO':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_subtype1_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) | (
                                    Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                        Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(
                            ticket_closed_at__gt=end_date)).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TOB':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_subtype1__parent_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) & (
                                    Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                                ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TOI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_subtype1__parent_id=rowId).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TCI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_subtype1__parent_id=rowId).filter(
                        ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                        ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'TLO':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_subtype1__parent_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) | (
                                    Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                        Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(
                            ticket_closed_at__gt=end_date)).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if tickets:
                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'reportTitle': reportTitle,
                    }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'IssueTypes':
                try:
                    reportTitle = TicketType.objects.filter(user_org_id=org_id).values_list('ttype_name', flat=True).get(pk=rowId)
                except TicketType.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                if columnId == 'OB':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_type_id=rowId).filter(ticket_type__has_parent=0).filter(
                        Q(ticket_created_at__lt=start_date) & (
                                    Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                                ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )
                if columnId == 'OI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_type_id=rowId).filter(ticket_type__has_parent=0).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )
                if columnId == 'CI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_type_id=rowId).filter(ticket_type__has_parent=0).filter(
                        ticket_status=1).filter(
                        ticket_closed_at__gte=start_date).filter(ticket_closed_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )
                if columnId == 'LO':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_type_id=rowId).filter(ticket_type__has_parent=0).filter(
                        Q(ticket_created_at__lt=start_date) | (
                                    Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                        Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(
                            ticket_closed_at__gt=end_date)).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if tickets:
                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'reportTitle': reportTitle,
                    }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'Locations':
                try:
                    reportTitle = GlobalACCTS.objects.values_list('acc_name', flat=True).get(pk=rowId)
                except Client.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                if columnId == 'OB':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(account_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) & (
                                    Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                                ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )
                if columnId == 'OI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(account_id=rowId).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )
                if columnId == 'CI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(account_id=rowId).filter(
                        ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                        ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )
                if columnId == 'LO':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(account_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) | (
                                    Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                        Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(
                            ticket_closed_at__gt=end_date)).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if tickets:
                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'reportTitle': reportTitle,
                    }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'NextActionBy':
                try:
                    reportTitle = User.objects.filter(user_org_id=org_id).values_list('display_name', flat=True).get(pk=rowId)
                except User.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                if columnId == 'OB':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_next_action_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) & (
                                    Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                                ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'OI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_next_action_id=rowId).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'CI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_next_action_id=rowId).filter(
                        ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                        ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'LO':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_next_action_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) | (
                                    Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                        Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(
                            ticket_closed_at__gt=end_date)).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if tickets:
                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'reportTitle': reportTitle,
                    }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'Organizations':
                try:
                    reportTitle = Organization.objects.filter(org_id=org_id).values_list('org_name', flat=True).get(pk=rowId)
                except Organization.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                if columnId == 'OB':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_org_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) & (
                                    Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                                ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'OI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_org_id=rowId).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'CI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_org_id=rowId).filter(
                        ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                        ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'LO':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_org_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) | (
                                    Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                        Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(
                            ticket_closed_at__gt=end_date)).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if tickets:
                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'reportTitle': reportTitle,
                    }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'SubmittedBy':
                try:
                    reportTitle = User.objects.filter(user_org_id=org_id).values_list('display_name', flat=True).get(pk=rowId)
                except User.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                if columnId == 'OB':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_caller_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) & (
                                    Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                                ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'OI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_caller_id=rowId).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'CI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_caller_id=rowId).filter(
                        ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                        ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'LO':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_caller_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) | (
                                    Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                        Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(
                            ticket_closed_at__gt=end_date)).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if tickets:
                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'reportTitle': reportTitle,
                    }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'DepartmentsSubmit':
                try:
                    reportTitle = Department.objects.filter(user_org_id=org_id).values_list('dep_name', flat=True).get(pk=rowId)
                except Department.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                if columnId == 'OB':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_caller__user_dep_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) & (
                                Q(ticket_status=0) | Q(ticket_closed_at__isnull=True) | Q(
                            ticket_closed_at__gte=start_date))).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'OI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_caller__user_dep_id=rowId).filter(
                        ticket_created_at__gte=start_date).filter(ticket_created_at__lte=end_date).filter(
                        ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'CI':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_caller__user_dep_id=rowId).filter(
                        ticket_status=1).filter(ticket_closed_at__gte=start_date).filter(
                        ticket_closed_at__lte=end_date).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'LO':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_caller__user_dep_id=rowId).filter(
                        Q(ticket_created_at__lt=start_date) | (
                                    Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date))).filter(
                        Q(ticket_closed_at__isnull=True) | Q(ticket_status=0) | Q(
                            ticket_closed_at__gt=end_date)).filter(ticket_is_delete=0).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if tickets:
                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'reportTitle': reportTitle,
                    }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'IncidentsByRep':
                try:
                    reportTitle = User.objects.filter(user_org_id=org_id).values_list('display_name', flat=True).get(pk=rowId)
                except User.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                if columnId == 'TOT':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_created_by_id=rowId).filter(ticket_is_delete=0).filter(
                        Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date)).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if columnId == 'OB':
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).filter(ticket_created_by_id=rowId).filter(ticket_is_delete=0).filter(
                        ticket_status=0).filter(
                        Q(ticket_created_at__gte=start_date) & Q(ticket_created_at__lte=end_date)).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_status', 'ticket_sub_status__sub_status_text', 'priority__priority_name',
                        'ticket_created_at'
                    )

                if tickets:
                    load_sidebar = get_sidebar(request)
                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'reportTitle': reportTitle,
                    }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'OrgTimes':
                try:
                    reportTitle = Organization.objects.filter(org_id=org_id).values_list('org_name', flat=True).get(pk=rowId)
                except Organization.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                try:
                    tickets = TicketNote.objects.values('note_ticket__ticket_org__org_name', 'note_ticket_id',
                                                        'note_ticket__subject',
                                                        'note_ticket__ticket_status',
                                                        'note_ticket__ticket_sub_status__sub_status_text',
                                                        'note_created_at', ).filter(
                        note_is_delete=0).filter(note_ticket_id__isnull=False).filter(
                        note_ticket__ticket_org_id=rowId).filter(
                        note_created_at__gte=start_date).filter(note_created_at__lte=end_date).annotate(
                        hours=Sum(Cast('tnote_laborhour_hours', IntegerField())) + Sum(
                            Cast('tnote_laborhour_minutes', IntegerField())) / 60,
                        minutes=Sum(Cast('tnote_laborhour_minutes', IntegerField())) % 60).exclude(
                        Q(hours=0) & Q(minutes=0)).order_by(
                        'note_ticket__ticket_org__org_name', 'note_ticket_id', 'note_created_by'
                    )
                except TicketNote.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                try:
                    total_labour_hours = TicketNote.objects.filter(note_is_delete=0).filter(
                        note_ticket_id__isnull=False).filter(note_ticket__ticket_org_id=rowId).filter(
                        note_created_at__gte=start_date).filter(note_created_at__lte=end_date).aggregate(
                        thours=Sum(Cast('tnote_laborhour_hours', IntegerField())) + Sum(
                            Cast('tnote_laborhour_minutes', IntegerField())) / 60,
                        tminutes=Sum(Cast('tnote_laborhour_minutes', IntegerField())) % 60
                    )
                except TicketNote.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                if tickets:
                    load_sidebar = get_sidebar(request)

                    if tickets:
                        load_sidebar = get_sidebar(request)
                        context = {
                            'report_id': report_id,
                            'sidebar': load_sidebar,
                            'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                            'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                            'tickets': tickets,
                            'total_hours': total_labour_hours['thours'],
                            'total_minutes': total_labour_hours['tminutes'],
                            'reportTitle': reportTitle,
                        }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)
                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'UserTimes':
                try:
                    reportTitle = User.objects.filter(user_org_id=org_id).values_list('display_name', flat=True).get(pk=rowId)
                except User.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                try:
                    tickets = TicketNote.objects.values('note_ticket__ticket_org__org_name', 'note_ticket_id',
                                                        'note_ticket__subject',
                                                        'note_ticket__ticket_status',
                                                        'note_ticket__ticket_sub_status__sub_status_text',
                                                        'note_created_at',
                                                        hours=(Cast('tnote_laborhour_hours', IntegerField()) + Cast(
                                                            'tnote_laborhour_minutes', IntegerField()) / 60),
                                                        minutes=(Cast('tnote_laborhour_minutes',
                                                                      IntegerField()) % 60)).filter(
                        note_is_delete=0).filter(
                        note_ticket_id__isnull=False).filter(note_created_by_id=rowId).filter(
                        note_created_at__gte=start_date).filter(
                        note_created_at__lte=end_date).exclude(Q(hours=0) & Q(minutes=0)).order_by(
                        'note_ticket__ticket_org__org_name', 'note_ticket_id', 'note_created_by'
                    )
                except TicketNote.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                try:
                    total_labour_hours = TicketNote.objects.filter(note_is_delete=0).filter(
                        note_ticket_id__isnull=False).filter(
                        note_created_by_id=rowId).filter(note_created_at__gte=start_date).filter(
                        note_created_at__lte=end_date).aggregate(
                        thours=Sum(Cast('tnote_laborhour_hours', IntegerField())) + Sum(
                            Cast('tnote_laborhour_minutes', IntegerField())) / 60,
                        tminutes=Sum(Cast('tnote_laborhour_minutes', IntegerField())) % 60
                    )
                except TicketNote.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                if tickets:
                    load_sidebar = get_sidebar(request)

                    if tickets:
                        load_sidebar = get_sidebar(request)
                        context = {
                            'report_id': report_id,
                            'sidebar': load_sidebar,
                            'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                            'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                            'tickets': tickets,
                            'total_hours': total_labour_hours['thours'],
                            'total_minutes': total_labour_hours['tminutes'],
                            'reportTitle': reportTitle,
                        }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)
                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'OrgTimeOpen':
                try:
                    reportTitle = Organization.objects.filter(org_id=org_id).values_list('org_name', flat=True).get(pk=rowId)
                except Organization.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                try:
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_id', 'ticket_status', 'submitted_at', 'ticket_created_at',
                                                    'ticket_closed_at', 'ticket_org__org_name',
                                                    'subject', 'ticket_sub_status__sub_status_text').filter(
                        ticket_org_id=rowId).filter(
                        ticket_is_delete=0).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_org__org_name', 'ticket_id', 'submitted_at'
                    )
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
                g_total_hours = 0
                if tickets:
                    for ticket in tickets:
                        if ticket['ticket_status'] == 0:
                            delta = datetime.now(timezone.utc) - ticket['ticket_created_at']
                            days, seconds = delta.days, delta.seconds
                            hours = days * 24 + seconds // 3600
                            minutes = (seconds % 3600) // 60
                        else:
                            delta = ticket['ticket_closed_at'] - ticket['ticket_created_at']
                            days, seconds = delta.days, delta.seconds
                            hours = days * 24 + seconds // 3600
                            minutes = (seconds % 3600) // 60

                        total_hours = round((hours) + (minutes / 100), 2)
                        ticket['thours'] = total_hours
                        g_total_hours += round(((hours) + (minutes / 100)), 2)

                if tickets:
                    g_total_hours = round(g_total_hours, 2)
                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'g_total_hours': g_total_hours,
                        'reportTitle': reportTitle,
                    }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))
            elif report_id == 'UserTimeOpen':
                try:
                    reportTitle = User.objects.filter(user_org_id=org_id).values_list('display_name', flat=True).get(pk=rowId)
                except User.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))

                try:
                    tickets = Ticket.objects.filter(ticket_org_id=org_id).values('ticket_id', 'ticket_status', 'submitted_at', 'ticket_created_at',
                                                    'ticket_closed_at',
                                                    'ticket_org__org_name', 'subject', 'ticket_status',
                                                    'ticket_sub_status__sub_status_text').filter(
                        ticket_caller_id=rowId).filter(
                        ticket_is_delete=0).filter(ticket_created_at__gte=start_date).filter(
                        ticket_created_at__lte=end_date).filter(**accountFilter).filter(userFilter).order_by(
                        'ticket_org__org_name', 'ticket_id', 'submitted_at'
                    )
                except Ticket.DoesNotExist:
                    messages.error(request, 'Request Failed! No Record Found.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))

                g_total_hours = 0
                if tickets:
                    for ticket in tickets:
                        if ticket['ticket_status'] == 0:
                            delta = datetime.now(timezone.utc) - ticket['ticket_created_at']
                            days, seconds = delta.days, delta.seconds
                            hours = days * 24 + seconds // 3600
                            minutes = (seconds % 3600) // 60
                        else:
                            delta = ticket['ticket_closed_at'] - ticket['ticket_created_at']
                            days, seconds = delta.days, delta.seconds
                            hours = days * 24 + seconds // 3600
                            minutes = (seconds % 3600) // 60

                        total_hours = round((hours) + (minutes / 100), 2)
                        ticket['thours'] = total_hours
                        g_total_hours += round(((hours) + (minutes / 100)), 2)

                if tickets:
                    g_total_hours = round(g_total_hours, 2)
                    load_sidebar = get_sidebar(request)

                    context = {
                        'report_id': report_id,
                        'sidebar': load_sidebar,
                        'start_date': datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y'),
                        'end_date': datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y'),
                        'tickets': tickets,
                        'g_total_hours': g_total_hours,
                        'reportTitle': reportTitle,
                    }

                    return render(request, 'itrak/Reports/report_defaultIssue_list.html', context)

                else:
                    messages.error(request, 'Request Failed! No Record Found.Please try another Date Range.')
                    # return redirect("/Home_getReportDateRange?summaryReport=" + str(report_id))
                    return redirect(reverse('getReportDateRange') + '?summaryReport=' + str(report_id))

        else:
            return render_to_response('itrak/page-404.html')

    else:
        return render_to_response('itrak/page-404.html')


# Get Summary Report Ticket List End#


# Get Summary Report Result End#

# Get Report Writer Start #

@active_user_required
def reportWriterQueries(request):
    load_sidebar = get_sidebar(request)
    request.session['pair'] = ''
    request.session['selected_fields_array'] = ''
    request.session['unselected_fields_array'] = ''
    request.session['filter_statement'] = ''
    request.session['expressions'] = ''
    request.session['filter_expression_array'] = ''
    request.session['expression_content_array'] = ''
    qb_queries = SavedQBQuries.objects.filter(qb_query_is_delete=0).filter(Q(qbQueryID__qb_query_share_with_id=request.user.id) | Q(qb_created_by_id=request.user.id)).distinct()
    context = {
        'sidebar': load_sidebar,
        'qb_queries': qb_queries
    }
    return render(request, 'itrak/Reports/report_writer.html', context)


# Get Report Writer End #

# New Query Writer Start #

@active_user_required
def newQuery(request):
    load_sidebar = get_sidebar(request)
    Pairs = DataSetsPair.objects.filter(ds_is_deleted=0)
    context = {
        'sidebar': load_sidebar,
        'pairs': Pairs,
    }
    if request.session.get('pair'):
        context['pair_set'] = int(request.session['pair'])
    else:
        request.session['pair'] = ''
        request.session['selected_fields_array'] = ''
        request.session['unselected_fields_array'] = ''
        request.session['filter_statement'] = ''
        request.session['expression'] = ''
        request.session['filter_expression_array'] = ''
        request.session['expression_content_array'] = ''

    return render(request, 'itrak/Reports/new_query.html', context)


# New Query Writer End #

# Get pair Fields Start #

@csrf_exempt
def getPairFields(request):
    if request.is_ajax() and request.method == 'POST':
        fields = DataSetsPairFields.objects.filter(df_pair_id=request.POST.get('data_pair')).exclude(df_actual_table_name = 'Client').order_by('df_order')
        # return HttpResponse(fields.query)
        request.session['pair'] = request.POST.get('data_pair')
        return render(request, 'itrak/Reports/data_pair_fields.html', {'fields': fields})


# Get pair Fields End #
#
#
# Get pair Selected Fields Start #

@csrf_exempt
def getPairSelectedFields(request):
    if request.is_ajax() and request.method == 'POST':
        selectedFields = request.session['selected_fields_array']
        unselectedFields = request.session['unselected_fields_array']
        return render(request, 'itrak/Reports/data_selected_pair_fields.html',
                      {'selecetdFields': selectedFields, 'unselecetdFields': unselectedFields})


# Get pair Selected Fields End #

# Second Query Writer Start #

@active_user_required
def secondQuery(request):
    load_sidebar = get_sidebar(request)
    fields = DataSetsPairFields.objects.filter(df_pair_id=request.session['pair']).exclude(df_actual_table_name = 'Client')
    context = {}
    if request.session['filter_statement']:
        context['filter_statement'] = request.session['filter_statement']
    if request.session['expressions']:
        filter_expression_array = request.session['expressions']
        context['filter_expression_array'] = filter_expression_array
    if request.session['expression_content_array']:
        expression_content_array = request.session['expression_content_array']
        context['expression_content_array'] = zip(expression_content_array, filter_expression_array)
    context['sidebar'] = load_sidebar
    context['fields'] = fields
    return render(request, 'itrak/Reports/new_query2.html', context)


# second Query Writer End #

#  Get FieldsConditions Start #

@csrf_exempt
def getFieldsConditions(request):
    if request.is_ajax() and request.method == 'POST':
        field_type = DataSetsPairFields.objects.values('df_condition_type').filter(
            df_actual_column_name=request.POST.get('field_name')).first()
        conditions = DataSetsFieldsConditions.objects.filter(dc_type=field_type['df_condition_type'])
        options = ['<option value="None"></option>']
        for condition in conditions:
            if 'condition' in request.POST and condition.dc_name == request.POST['condition']:
                options.append(
                    '<option value="' + condition.dc_name + '" data-id="' + condition.dc_type_name + '" selected>' + condition.dc_name + '</option>')
            else:
                options.append(
                    '<option value="' + condition.dc_name + '" data-id="' + condition.dc_type_name + '">' + condition.dc_name + '</option>')
        return HttpResponse(options)


# Get Fields Conditions End #

#  Set Selected Fields Start #

@csrf_exempt
def setSelectedFields(request):
    if request.is_ajax() and request.method == 'POST':
        request.session['unselected_fields_array'] = request.POST.get('unselected_fields').split(",")
        request.session['selected_fields_array'] = request.POST.get('selected_fields').split(",")
        request.session['filter_statement'] = ''
        request.session['expressions'] = ''
        request.session['filter_expression_array'] = ''
        request.session['expression_content_array'] = ''
        return HttpResponse(request.session['selected_fields_array'])


# Set Selected Fields End #

# Third Query Writer Start #

@active_user_required
def thirdQuery(request):
    load_sidebar = get_sidebar(request)
    org_id = request.user.user_org_id
    users = User.objects.filter(user_org_id = org_id).values('id', 'first_name', 'last_name').filter(is_delete=0)
    context = {
        'sidebar': load_sidebar,
        'users': users,
    }
    return render(request, 'itrak/Reports/new_query3.html', context)


# Third Query Writer End #

# Set Filters Start #

@csrf_exempt
def setFilters(request):
    if request.is_ajax() and request.method == 'POST':
        if request.POST.get('expression'):
            filter_expression = request.POST.get('expression').split("^")
            expression_content = request.POST.get('expression_content').split("|")
            expressions = []
            for exp in filter_expression:
                temp = exp.split("del\xa0\xa0\xa0")
                expressions.append(temp[1])
            request.session['expressions'] = expressions
            request.session['expression_content_array'] = expression_content
            request.session['filter_statement'] = request.POST.get('statement')
            filter_expression_array = {}
            for filter_exp in expression_content:
                filter_exp_dict = {}
                split_expression = filter_exp.split("',")
                field = split_expression[0].replace("'", "")
                filter_exp_dict['modelName'] = field.split(",")[1]
                filter_exp_dict['modelAttr'] = field.split(",")[0]
                filter_exp_dict['condition'] = split_expression[1].replace("'", "")
                filter_exp_dict['fieldValue'] = split_expression[2].replace("'", "")
                try:
                    fieldType = DataSetsPairFields.objects.values_list('df_condition_type', flat=True).filter(
                        df_actual_column_name=filter_exp_dict['modelAttr']).first()
                except:
                    return render_to_response('itrak/page-404.html')
                # if (fieldType == '2' or fieldType == '4') and (filter_exp_dict['condition'] != 'In' or filter_exp_dict['condition'] != 'Not In'):
                #     filter_exp_dict['fieldValue'] = split_expression[2]+"'"
                if (filter_exp_dict['condition'] == 'In' or filter_exp_dict['condition'] == 'Not In'):
                    field_values_in = split_expression[2].replace("'", "")
                    filter_exp_dict['fieldValue'] = field_values_in.split(',')
                    # filter_exp_dict['fieldValue'] = (','.join("'" + item + "'" for item in filter_exp_dict['fieldValue']))
                    # filter_exp_dict['fieldValue'] = list(filter_exp_dict['fieldValue'].split(","))
                if (fieldType == '2') and (
                        filter_exp_dict['condition'] == 'Is Blank' or filter_exp_dict['condition'] == 'Is Not Blank'):
                    filter_exp_dict['fieldValue'] = ''
                if (fieldType == '1' or fieldType == '3' or fieldType == '4') and (
                        filter_exp_dict['condition'] == 'Is Blank' or filter_exp_dict['condition'] == 'Is Not Blank' or
                        filter_exp_dict['condition'] == 'Is True' or filter_exp_dict['condition'] == 'Is False'):
                    if (filter_exp_dict['condition'] == 'Is Blank' or filter_exp_dict['condition'] == 'Is Not Blank'):
                        filter_exp_dict['condition'] = 'Is Null' if filter_exp_dict[
                                                                        'condition'] == 'Is Blank' else 'Is Not Null'
                    filter_exp_dict['fieldValue'] = 'True' if (
                                filter_exp_dict['condition'] == 'Is Blank' or filter_exp_dict[
                            'condition'] == 'Is Not Blank') else 1
                if (fieldType == '4') and (filter_exp_dict['condition'] != 'Is Blank' and filter_exp_dict[
                    'condition'] != 'Is Not Blank' and filter_exp_dict['condition'] != 'Is Null' and filter_exp_dict[
                                               'condition'] != 'Is Not Null'):
                    if filter_exp_dict['condition'] == '=' or filter_exp_dict['condition'] == '<>' or filter_exp_dict[
                        'condition'] == '<' or filter_exp_dict['condition'] == '>' or filter_exp_dict[
                        'condition'] == '>=' or filter_exp_dict['condition'] == '<=':
                        filter_exp_dict['fieldValue'] = datetime.strptime(filter_exp_dict['fieldValue'],
                                                                          '%m/%d/%Y').strftime('%Y-%m-%d')
                    else:
                        start_date = datetime.strptime(filter_exp_dict['fieldValue'].split(' - ')[0],
                                                       '%m/%d/%Y').strftime('%Y-%m-%d')
                        end_date = datetime.strptime(filter_exp_dict['fieldValue'].split(' - ')[1] + " 23:59:59",
                                                     '%m/%d/%Y %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
                        filter_exp_dict['fieldValue'] = start_date + ' - ' + end_date
                        filter_exp_dict['fieldValue'] = filter_exp_dict['fieldValue'].split(' - ')
                    if filter_exp_dict['condition'] == '=':
                        filter_exp_dict['condition'] = 'Starts with'
                    elif filter_exp_dict['condition'] == '<>':
                        filter_exp_dict['condition'] = 'Starts Not with'
                filter_exp_dict['fieldLabel'] = split_expression[3].split(",")[1]
                filter_expression_array[filter_exp_dict['fieldLabel'].replace("'", "")] = filter_exp_dict

            request.session['filter_expression_array'] = filter_expression_array
            return HttpResponse(expressions)
        return HttpResponse(1)


# Set Filters End #


# Get Compound Filter For Query Builder Start #

@csrf_exempt
def getCompoundFilterForQuery(filter_statement, filter_expression_array, conditiondict):
    argument_list = []
    counter = 0
    filter_operator = 'and'
    query_commit = 0
    query_not_commit = 0
    compound_filter = ''
    tempcompound_filter = []
    # remain_filter_statement = []  #filter_statement.copy() can also be  used but restrict to Python 3

    for key in range(len(filter_statement)):
        if key == len(filter_statement):
            return compound_filter
        print(filter_statement)
        print(key)
        print(filter_statement[key])
        try:
            if len(filter_statement[key]) == 1:
                if filter_statement[key] == '(':
                    remain_filter_statement = []
                    print(argument_list)
                    print(compound_filter)
                    while True:
                        print('Hello')
                        print(filter_statement)
                        print(filter_statement[key + 1])
                        remain_filter_statement.append(filter_statement.pop(key + 1))
                        print(remain_filter_statement)
                        if filter_statement[key + 1] == ')':
                            remain_filter_statement.append(filter_statement.pop(key + 1))
                            break
                    tempcompound_filter.append(compound_filter)
                    argument_list.append(
                        getCompoundFilterForQuery(remain_filter_statement, filter_expression_array, conditiondict))
                    print('Compound Filter Mill Gaya')
                    print(compound_filter)
                    print(argument_list[0])
                    print(filter_statement)
                    if compound_filter:
                        if filter_operator == 'or' and query_commit == 1:
                            print('Idhr aya ha')
                            argument_list.insert(0, compound_filter)
                            print(argument_list[0])
                            print(argument_list[1])
                            compound_filter = (argument_list[0].__or__(argument_list[1]))
                            argument_list.clear()
                        elif filter_operator == 'and' and query_commit == 1:
                            argument_list.insert(0, compound_filter)
                            compound_filter = argument_list[0].__and__(argument_list[1])
                            argument_list.clear()
                        return compound_filter
                    else:
                        compound_filter = argument_list[0]
                        argument_list.clear()
                        counter = counter + 1
                        if key + 1 == len(filter_statement):
                            print('Here it oomes')
                            return compound_filter

                elif filter_statement[key] == ')':
                    print(argument_list)
                    print(compound_filter)
                    return compound_filter
                else:
                    # print(filter_expression_array[filter_statement[key]])
                    expression = filter_expression_array[filter_statement[key]]
                    modelName = expression['modelName']
                    modelAttr = expression['modelAttr']
                    condition = expression['condition']
                    fieldValue = expression['fieldValue']
                    fieldLabel = expression['fieldLabel']

                    print(modelName)
                    print(modelAttr)
                    print(condition)
                    print(fieldValue)
                    print(fieldLabel)
                    print(compound_filter)
                    print(query_commit)

                    # Condition Implementation For NOT Case AND Condition BASED
                    if query_not_commit == 0:
                        if condition == '<>' or condition == 'Not In' or condition == 'Not Like' or condition == 'Is Not Blank' or condition == 'Is Not Null' or condition == 'Is False' or condition == 'Starts Not with':
                            argument_list.append(~Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                        else:
                            argument_list.append(Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                    elif query_not_commit == 1:
                        query_not_commit = 0
                        if condition == '<>' or condition == 'Not In' or condition == 'Not Like' or condition == 'Is Not Blank' or condition == 'Is Not Null' or condition == 'Is False' or condition == 'Starts Not with':
                            argument_list.append(Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                        else:
                            argument_list.append(~Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                    print('Test')
                    print(argument_list)
                    print('Test')
                    # First Iteration of Filter Expression
                    if counter == 0:
                        compound_filter = argument_list[0]
                        argument_list.clear()
                        counter = counter + 1

                    # First Increment in Filter Expression
                    if filter_operator == 'or' and query_commit == 1:
                        print(compound_filter)
                        argument_list.insert(0, compound_filter)
                        print(argument_list)
                        print('Hello')
                        print(argument_list[0])
                        print(argument_list[1])
                        compound_filter = (argument_list[0].__or__(argument_list[1]))
                        argument_list.clear()
                    elif filter_operator == 'and' and query_commit == 1:
                        argument_list.insert(0, compound_filter)
                        compound_filter = argument_list[0].__and__(argument_list[1])
                        argument_list.clear()
            elif filter_statement[key].lower() == 'and' or filter_statement[key].lower() == 'or':
                filter_operator = filter_statement[key].lower()
                query_commit = 1
            elif filter_statement[key].lower() == 'not':
                query_not_commit = 1
            else:
                return render_to_response('itrak/page-404.html')
        except:
            return render_to_response('itrak/page-404.html')
    return compound_filter


# Get Compound Filter For QueryBuilder End #


# Save Final Query From Query Builder Start #

@active_user_required
def saveFinalNewQuery(request):
    with transaction.atomic():
        if request.method == 'POST':
            if request.GET.get('IsRun') == 'True':
                IsRun = 1
            else:
                IsRun = 0
            if 'query_name' in request.POST:
                query_name = request.POST.get('query_name')
            if 'description' in request.POST and request.POST['description']:
                description = request.POST.get('description')
            else:
                description = ''
            if 'createReport' in request.POST and request.POST['createReport']:
                createReport = request.POST.get('createReport')
            else:
                createReport = 0
            if 'share' in request.POST and request.POST['share']:
                share = request.POST.get('share')
                share_with = request.POST.getlist('share_with')
                share_with = list(filter(lambda x: x != 'multiselect-all', share_with))
            else:
                share = 0
            pair_id = int(request.session['pair'])
            expression_content_array = request.session['expression_content_array']
            filter_expressions = request.session['expressions']
            filter_expression_array = request.session['filter_expression_array']
            filter_statement = request.session['filter_statement']
            selected_fields = request.session['selected_fields_array']
            unselected_fields = request.session['unselected_fields_array']

            obj = SavedQBQuries(
                qb_query_name=query_name,
                qb_query_description=description,
                qb_query_pair_id=pair_id,
                qb_query_create_report=createReport,
                qb_query_is_share=share,
                qb_filter_expressions=json.dumps(filter_expressions),
                qb_filter_expression_array=json.dumps(filter_expression_array),
                qb_expression_content_array=json.dumps(expression_content_array),
                qb_filter_statement=filter_statement,
                qb_selected_fields=json.dumps(selected_fields),
                qb_unselected_fields=json.dumps(unselected_fields),
                qb_created_by_id=request.user.id
            )
            obj.save()
            insert_id = SavedQBQuries.objects.latest('pk').qb_query_id

            if 'share' in request.POST:
                for userID in share_with:
                    obj1 = SavedQBQuriesShareWith(qb_query_id=insert_id, qb_query_share_with_id=userID)
                    obj1.save()

            if IsRun == 0:
                messages.success(request, 'Request Succeed! Query created.')
                return redirect('reportWriterQueries')
            else:
                # return redirect("/Home_qbQueryProcess?qbQuery=" + str(insert_id))
                return redirect(reverse('qbQueryProcess') + '?qbQuery=' + str(insert_id))
        else:
            messages.error(request, 'Request Failed! Query cannot be created.Please try again.')
            return redirect('reportWriterQueries')


# Save Final Query From Query Builder StartEnd #


# Get Query Builder Query Process Start#

@active_user_required
def qbQueryProcess(request, queryId=None, call_reportProcessFun=False):
    org_id = request.user.user_org_id
    SQL  = """
                select distinct id
                from AT_USERS a
                where 1=1
                AND a.user_org_id = '"""+str(org_id)+"""'
            """
        
    cursor = connection.cursor()
    cursor.execute(SQL)
    users = dictfetchall(cursor)
    usersList = []
    for account in users:
        usersList.append(account['id'])
    if len(usersList) > 0:
        createdFilter = {'qb_created_by_id__in':usersList}
    else: 
        createdFilter = {}
    if queryId:
        query_id = queryId
    else:
        query_id = request.GET.get('qbQuery')

    # Get List Of Associated Accounts
    accountsList = getAccountIDsOfCurrentUser(request)
    userAccountList = getAccountIDsOfCurrentUser(request)
    if len(accountsList) > 0:
        accountFilter = {'account_id__in':accountsList}
    else: 
        accountFilter = {}
    user_id = request.user.id
    user_type = userType(request) 
    accountsList = getMappedUserIDsWithCurrentUer(request)
    if user_type == 'manager':
        userFilter = Q(ticket_assign_to_id__in = accountsList) | Q(ticket_caller_id__in = accountsList) | Q(ticket_next_action_id__in=accountsList) | Q(ticket_created_by_id__in=accountsList)
    elif user_type == 'enduser':
        userFilter = Q(ticket_assign_to_id=user_id) | Q(ticket_caller_id=user_id) | Q(ticket_next_action_id=user_id) | Q(ticket_created_by_id=user_id)
    else:
        userFilter = Q(pk__isnull=False)

    context = {}
    obj = SavedQBQuries.objects.get(pk=query_id)
    
    query_name = obj.qb_query_name
    query_description = obj.qb_query_description
    qb_query_pair_id = obj.qb_query_pair_id
    

    filter_expression_array = json.loads(obj.qb_filter_expression_array)
    filter_statement = (obj.qb_filter_statement).split(" ") if obj.qb_filter_statement else ''
    selected_fields = json.loads(obj.qb_selected_fields)
    
    selected_columns = []
    argument_list = []
    conditiondict = {
        "=": "exact",
        "<>": "exact",
        "In": "in",
        "Not In": "in",
        "Like": "contains",
        "Not Like": "contains",
        "Is Blank": "exact",
        "Is Not Blank": "exact",
        "Is Null": "isnull",
        "Is Not Null": "isnull",
        "Is True": "exact",
        "Is False": "exact",
        ">": "gt",
        "<": "lt",
        ">=": "gte",
        "<=": "lte",
        "Starts with": "startswith",
        "Starts Not with": "startswith",
        "Today": "range",
        "Yesterday": "range",
        "Tomorrow": "range",
        "This Week": "range",
        "Last Week": "range",
        "Next Week": "range",
        "This Month": "range",
        "Last Month": "range",
        "Next Month": "range",
        "This Year": "range",
        "Last Year": "range",
        "Next Year": "range",
        "Custom Date Range": "range"
    }

    for field in selected_fields:
        try:
            actual_field = \
            DataSetsPairFields.objects.values('df_actual_column_name').filter(df_pair_id=qb_query_pair_id).filter(
                df_name=field)[0]['df_actual_column_name']
        except:
            return render_to_response('itrak/page-404.html')
        selected_columns.append(actual_field)
    
    counter = 0
    filter_operator = 'and'
    query_commit = 0
    query_not_commit = 0
    compound_filter = ''
    if filter_statement:
        if len(filter_statement) == 1:
            for key in filter_statement:
                # try:
                if len(key) == 1:
                    expression = filter_expression_array[key]
                    modelName = expression['modelName']
                    modelAttr = expression['modelAttr']
                    condition = expression['condition']
                    fieldValue = expression['fieldValue']
                    fieldLabel = expression['fieldLabel']

                    # print(expression)
                    # print(modelName)
                    # print(modelAttr)
                    # print(condition)
                    # print(fieldValue)
                    # print(fieldLabel)

                    modelInstance = apps.get_model('itrak', modelName)
                    print(modelInstance)
                    # print(selected_columns)
                    if modelName == 'Ticket':
                        finalquery = modelInstance.objects.filter(ticket_is_delete=0).filter(ticket_org_id=org_id).values(*tuple(selected_columns)).filter().filter(**accountFilter).filter(userFilter)
                    elif modelName == 'Organization':
                        finalquery = modelInstance.objects.filter(org_is_delete=0).filter(ticketOrg__ticket_is_delete=0).filter(org_id=org_id).values(*tuple(selected_columns)).filter()
                    elif modelName == 'Group':
                        finalquery = modelInstance.objects.filter(gp_is_delete=0).filter(group_org_id=org_id).values(*tuple(selected_columns)).filter()
                    elif modelName == 'SavedQBQuries':
                        finalquery = modelInstance.objects.filter(qb_query_is_delete=0).filter(**createdFilter).values(*tuple(selected_columns)).filter()
                    elif modelName == 'User':
                        finalquery = modelInstance.objects.filter(is_delete=0).filter(user_org_id=org_id).values(*tuple(selected_columns)).filter()
                    elif modelName == 'GlobalACCTS':
                        finalquery = modelInstance.objects.filter(id__in = userAccountList).values(*tuple(selected_columns)).filter()
                    else:
                        finalquery = modelInstance.objects.values(*tuple(selected_columns)).filter()
                    # print(modelInstance)
                    if condition == '<>' or condition == 'Not In' or condition == 'Not Like' or condition == 'Is Not Blank' or condition == 'Is Not Null' or condition == 'Is False' or condition == 'Starts Not with':
                        argument_list.append(~Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                    else:
                        argument_list.append(Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                        finalquery = finalquery.filter(reduce(operator.and_, argument_list))
                else:
                    return render_to_response('itrak/page-404.html')
            # except:
            #     return render_to_response('itrak/page-404.html')
        else:
            if filter_statement.__contains__('(') or filter_statement.__contains__(')'):
                try:
                    modelName = DataSetsPairFields.objects.only('df_primary_table_name').get(
                        df_actual_column_name=selected_columns[0]).df_primary_table_name
                except:
                    return render_to_response('itrak/page-404.html')
                modelInstance = apps.get_model('itrak', modelName)
                if modelName == 'Ticket':
                    finalquery = modelInstance.objects.filter(ticket_is_delete=0).filter(ticket_org_id=org_id).values(*tuple(selected_columns)).filter().filter(**accountFilter).filter(userFilter)
                elif modelName == 'Organization':
                    finalquery = modelInstance.objects.filter(org_is_delete=0).filter(ticketOrg__ticket_is_delete=0).filter(org_id=org_id).values(*tuple(selected_columns)).filter()
                elif modelName == 'Group':
                    finalquery = modelInstance.objects.filter(gp_is_delete=0).filter(group_org_id=org_id).values(*tuple(selected_columns)).filter()
                elif modelName == 'SavedQBQuries':
                    finalquery = modelInstance.objects.filter(qb_query_is_delete=0).filter(**createdFilter).values(*tuple(selected_columns)).filter()
                elif modelName == 'User':
                    finalquery = modelInstance.objects.filter(is_delete=0).filter(user_org_id=org_id).values(*tuple(selected_columns)).filter()
                elif modelName == 'GlobalACCTS':
                    finalquery = modelInstance.objects.filter(id__in = userAccountList).values(*tuple(selected_columns)).filter()
                else:
                    finalquery = modelInstance.objects.values(*tuple(selected_columns)).filter()
                compound_filter = getCompoundFilterForQuery(filter_statement, filter_expression_array, conditiondict)
                print(compound_filter)
            else:
                for key in filter_statement:
                    # print(key)
                    # print(filter_operator)
                    try:
                        if len(key) == 1:
                            key = key.upper()
                            # print(key)
                            # print(filter_expression_array[key])
                            expression = filter_expression_array[key]
                            modelName = expression['modelName']
                            modelAttr = expression['modelAttr']
                            condition = expression['condition']
                            fieldValue = expression['fieldValue']
                            fieldLabel = expression['fieldLabel']

                            # print(expression)
                            # print(modelName)
                            # print(modelAttr)
                            # print(condition)
                            # print(fieldValue)
                            # print(fieldLabel)

                            modelInstance = apps.get_model('itrak', modelName)
                            # print(modelInstance)
                            # print(selected_columns)

                            # Condition Implementation For NOT Case AND Condition BASED
                            if query_not_commit == 0:
                                if condition == '<>' or condition == 'Not In' or condition == 'Not Like' or condition == 'Is Not Blank' or condition == 'Is Not Null' or condition == 'Is False' or condition == 'Starts Not with':
                                    argument_list.append(
                                        ~Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                                else:
                                    argument_list.append(Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                            elif query_not_commit == 1:
                                query_not_commit = 0
                                if condition == '<>' or condition == 'Not In' or condition == 'Not Like' or condition == 'Is Not Blank' or condition == 'Is Not Null' or condition == 'Is False' or condition == 'Starts Not with':
                                    argument_list.append(Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                                else:
                                    argument_list.append(Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                                    print(argument_list)

                            # First Iteration of Filter Expression
                            if counter == 0:
                                if modelName == 'Ticket':
                                    finalquery = modelInstance.objects.filter(ticket_is_delete=0).filter(ticket_org_id=org_id).values(*tuple(selected_columns)).filter().filter(**accountFilter).filter(userFilter)
                                elif modelName == 'Organization':
                                    finalquery = modelInstance.objects.filter(org_is_delete=0).filter(ticketOrg__ticket_is_delete=0).filter(org_id=org_id).values(*tuple(selected_columns)).filter()
                                elif modelName == 'Group':
                                    finalquery = modelInstance.objects.filter(gp_is_delete=0).filter(group_org_id=org_id).values(*tuple(selected_columns)).filter()
                                elif modelName == 'SavedQBQuries':
                                    finalquery = modelInstance.objects.filter(qb_query_is_delete=0).filter(**createdFilter).values(*tuple(selected_columns)).filter()
                                elif modelName == 'User':
                                    finalquery = modelInstance.objects.filter(is_delete=0).filter(user_org_id=org_id).values(*tuple(selected_columns)).filter()
                                elif modelName == 'GlobalACCTS':
                                    finalquery = modelInstance.objects.filter(id__in = userAccountList).values(*tuple(selected_columns)).filter()
                                else:
                                    finalquery = modelInstance.objects.values(*tuple(selected_columns)).filter()
                                compound_filter = argument_list[0]
                                argument_list.clear()

                            counter = counter + 1
                            # First Increment in Filter Expression
                            if filter_operator == 'or' and query_commit == 1:
                                argument_list.insert(0, compound_filter)
                                compound_filter = argument_list[0].__or__(argument_list[1])
                                # finalquery = finalquery.filter(reduce(operator.or_, argument_list))
                                argument_list.clear()
                            elif filter_operator == 'and' and query_commit == 1:
                                argument_list.insert(0, compound_filter)
                                compound_filter = argument_list[0].__and__(argument_list[1])
                                print(compound_filter)
                                # finalquery = finalquery.filter(reduce(operator.and_, argument_list))
                                argument_list.clear()
                        elif key.lower() == 'and' or key.lower() == 'or':
                            filter_operator = key.lower()
                            query_commit = 1
                        elif key.lower() == 'not':
                            query_not_commit = 1
                        else:
                            return render_to_response('itrak/page-404.html')
                    except:
                        return render_to_response('itrak/page-404.html')
            if modelName == 'Ticket':
                finalquery = finalquery.filter(ticket_is_delete=0).filter(compound_filter).filter(**accountFilter).filter(userFilter)
            elif modelName == 'Organization':
                finalquery = finalquery.filter(org_is_delete=0).filter(ticketOrg__ticket_is_delete=0).filter(org_id=org_id).values(*tuple(selected_columns)).filter()
            elif modelName == 'Group':
                finalquery = finalquery.filter(gp_is_delete=0).filter(group_org_id=org_id).values(*tuple(selected_columns)).filter()
            elif modelName == 'SavedQBQuries':
                finalquery = finalquery.filter(qb_query_is_delete=0).filter(**createdFilter).values(*tuple(selected_columns)).filter()
            elif modelName == 'User':
                finalquery = finalquery.filter(is_delete=0).filter(user_org_id=org_id).values(*tuple(selected_columns)).filter()
            elif modelName == 'GlobalACCTS':
                finalquery = finalquery.filter(id__in = userAccountList).values(*tuple(selected_columns)).filter()
            else:
                finalquery = finalquery.filter(compound_filter)
    else:
        try:
            modelName = \
            DataSetsPairFields.objects.values('df_primary_table_name').filter(df_pair_id=qb_query_pair_id).filter(
                df_actual_column_name=selected_columns[0])[0]['df_primary_table_name']
        except:
            return render_to_response('itrak/page-404.html')
        modelInstance = apps.get_model('itrak', modelName)
        # print(modelInstance)
        # print(modelName)
        # return HttpResponse(modelName)
        if modelName == 'Ticket':
            finalquery = modelInstance.objects.filter(ticket_is_delete=0).filter(ticket_org_id=org_id).values(*tuple(selected_columns)).filter().filter(**accountFilter).filter(userFilter)
            # print(modelInstance)
            # print(modelName)
            # return HttpResponse(finalquery.query)
        elif modelName == 'Organization':
            finalquery = modelInstance.objects.filter(org_is_delete=0).filter(ticketOrg__ticket_is_delete=0).filter(org_id=org_id).values(*tuple(selected_columns)).filter()
            # print(modelInstance)
            # print(modelName)
            # return HttpResponse(finalquery.query)
        elif modelName == 'Group':
            finalquery = modelInstance.objects.filter(gp_is_delete=0).filter(group_org_id=org_id).values(*tuple(selected_columns)).filter()
            # print(modelInstance)
            # print(modelName)
            # return HttpResponse(finalquery.query)
        elif modelName == 'SavedQBQuries':
            finalquery = modelInstance.objects.filter(qb_query_is_delete=0).filter(**createdFilter).values(*tuple(selected_columns)).filter()
        elif modelName == 'User':
            finalquery = modelInstance.objects.filter(is_delete=0).filter(user_org_id=org_id).values(*tuple(selected_columns)).filter()
            # print(modelInstance)
            # print(modelName)
            # return HttpResponse(finalquery.query)
        elif modelName == 'GlobalACCTS':
            finalquery = modelInstance.objects.filter(id__in = userAccountList).values(*tuple(selected_columns)).filter()
            # return HttpResponse(finalquery.query)
        else:
            finalquery = modelInstance.objects.values(*tuple(selected_columns)).filter()
        
    if call_reportProcessFun:
        return finalquery
    
    timezone = MySettings.objects.filter(m_user_id=request.user.id).first()
    eastern = pytz.timezone(timezone.m_time_zone)
    time = datetime.now(eastern)
    time = time.strftime('%m/%d/%Y %I:%M %p')
    context = {
        'query_name': query_name,
        'query_description': query_description,
        'selected_fields': selected_fields,
        'time':time,
        'query_id':query_id,
        'records': finalquery
    }

    return render(request, 'itrak/Reports/qb_query_result_list.html', context)


# Get Query Builder Query Process End#


# Edit Query Builder Query Step 1 Start #

@active_user_required
def editQueryStep1(request):
    query_id = request.GET.get('qbQuery')
    load_sidebar = get_sidebar(request)
    Pairs = DataSetsPair.objects.all()
    queryObj = SavedQBQuries.objects.get(pk=query_id)
    context = {
        'sidebar': load_sidebar,
        'pairs': Pairs,
    }
    request.session['pair'] = queryObj.qb_query_pair_id
    request.session['selected_fields_array'] = json.loads(queryObj.qb_selected_fields)
    request.session['unselected_fields_array'] = json.loads(queryObj.qb_unselected_fields)

    if request.session['pair']:
        context['pair_set'] = int(request.session['pair'])
    else:
        request.session['pair'] = ''
        request.session['selected_fields_array'] = ''
        request.session['unselected_fields_array'] = ''
        request.session['filter_statement'] = ''
        request.session['expression'] = ''
        request.session['filter_expression_array'] = ''
        request.session['expression_content_array'] = ''
    context['query_id'] = query_id

    return render(request, 'itrak/Reports/edit_query.html', context)


# Edit Query Builder Query Step 1 End #


# Edit Query Builder Query Step 2 Start #

@active_user_required
def editQueryStep2(request):
    query_id = request.GET.get('qbQuery')
    load_sidebar = get_sidebar(request)
    fields = DataSetsPairFields.objects.filter(df_pair_id=request.session['pair'])
    queryObj = SavedQBQuries.objects.get(pk=query_id)

    if not request.session['expressions']:
        request.session['filter_statement'] = queryObj.qb_filter_statement
        request.session['expressions'] = json.loads(queryObj.qb_filter_expressions)
        request.session['expression_content_array'] = json.loads(queryObj.qb_expression_content_array)

    context = {}
    if request.session['filter_statement']:
        context['filter_statement'] = request.session['filter_statement']
    if request.session['expressions']:
        filter_expression_array = request.session['expressions']
        context['filter_expression_array'] = filter_expression_array
    if request.session['expression_content_array']:
        expression_content_array = request.session['expression_content_array']
        context['expression_content_array'] = zip(expression_content_array, filter_expression_array)
    context['sidebar'] = load_sidebar
    context['fields'] = fields
    context['query_id'] = query_id
    # print(context['filter_statement'])
    return render(request, 'itrak/Reports/edit_query2.html', context)


# Edit Query Builder Query Step 2 End #


# Edit Query Builder Query Step 3 Start #

@active_user_required
def editQueryStep3(request):
    org_id = request.user.user_org_id
    query_id = request.GET.get('qbQuery')
    load_sidebar = get_sidebar(request)
    users = User.objects.filter(user_org_id=org_id).values('id', 'first_name', 'last_name').filter(is_delete=0)
    queryObj = SavedQBQuries.objects.get(pk=query_id)
    # share_lists = queryObj.qb_filter_share_with

    if 'clone_id' in request.session and request.session['clone_id'] != '':
        share_lists = SavedQBQuriesShareWith.objects.values_list('qb_query_share_with_id', flat=True).filter(
            qb_query_id=request.session['clone_id'])
    else:
        share_lists = SavedQBQuriesShareWith.objects.values_list('qb_query_share_with_id', flat=True).filter(
            qb_query_id=query_id)

    context = {
        'sidebar': load_sidebar,
        'users': users,
        'query_id': query_id,
        'queryObj': queryObj,
        'share_lists': share_lists
    }
    return render(request, 'itrak/Reports/edit_query3.html', context)


# Edit Query Builder Query Step 3 End #


# Update Final Query From Query Builder Start #

@active_user_required
def updateFinalNewQuery(request):
    with transaction.atomic():
        request.session['clone_id'] = ''
        if request.method == 'POST':
            if request.GET.get('IsRun') == 'True':
                IsRun = 1
            else:
                IsRun = 0
            if 'query_id' in request.POST:
                query_id = request.POST.get('query_id')
                try:
                    queryObj = SavedQBQuries.objects.get(pk=query_id)
                except:
                    return render_to_response('itrak/page-404.html')
            if 'query_name' in request.POST:
                query_name = request.POST.get('query_name')
            if 'description' in request.POST and request.POST['description']:
                description = request.POST.get('description')
            else:
                description = ''
            if 'createReport' in request.POST and request.POST['createReport']:
                createReport = request.POST.get('createReport')
            else:
                createReport = 0
            if 'share' in request.POST and request.POST['share']:
                share = request.POST.get('share')
                share_with = request.POST.getlist('share_with')
                share_with = list(filter(lambda x: x != 'multiselect-all',share_with))

            else:
                share = 0
                share_with = None

            pair_id = int(request.session['pair'])
            expression_content_array = request.session['expression_content_array']
            filter_expressions = request.session['expressions']
            filter_expression_array = request.session['filter_expression_array']
            filter_statement = request.session['filter_statement']
            selected_fields = request.session['selected_fields_array']
            unselected_fields = request.session['unselected_fields_array']

            queryObj.qb_query_name = query_name
            queryObj.qb_query_description = description
            queryObj.qb_query_pair_id = pair_id
            queryObj.qb_query_create_report = createReport
            queryObj.qb_query_is_share = share
            queryObj.qb_filter_expressions = json.dumps(filter_expressions)
            queryObj.qb_filter_expression_array = json.dumps(filter_expression_array)
            queryObj.qb_expression_content_array = json.dumps(expression_content_array)
            queryObj.qb_filter_statement = filter_statement
            queryObj.qb_selected_fields = json.dumps(selected_fields)
            queryObj.qb_unselected_fields = json.dumps(unselected_fields)
            queryObj.qb_modified_by_id = request.user.id
            queryObj.note_modified_at = datetime.now()

            queryObj.save()
            insert_id = queryObj.qb_query_id
            queryShareWithObj = SavedQBQuriesShareWith.objects.filter(qb_query_id=insert_id)
            queryShareWithObj.delete()

            if share_with:
                for userID in share_with:
                    obj1 = SavedQBQuriesShareWith(
                        qb_query_id=insert_id,
                        qb_query_share_with_id=userID
                    )
                    obj1.save()

            if IsRun == 0:
                messages.success(request, 'Request Succeed! Query updated.')
                return redirect('reportWriterQueries')
            else:
                # return redirect("/Home_qbQueryProcess?qbQuery=" + str(query_id))
                return redirect(reverse('qbQueryProcess') + '?qbQuery=' + str(query_id))
        else:
            messages.error(request, 'Request Failed! Query cannot be updated.Please try again.')
            return redirect('reportWriterQueries')


# UpdateFinal Query From Query Builder StartEnd #


# Ticket Search Save Updating End#


@csrf_exempt
# Delete Saved Search Against ID Start#
def deleteQBQuery(request):
    with transaction.atomic():
        if request.method == 'GET':
            query_id = request.GET.get('qbQuery')
            # probably you want to add a regex check if the username value is valid here
            if query_id:
                try:
                    queryObj = SavedQBQuries.objects.get(pk=query_id)
                except:
                    return render_to_response('itrak/page-404.html')

                queryObj.qb_query_is_delete = 1
                queryObj.qb_modified_by_id = request.user.id
                queryObj.note_modified_at = datetime.now()
                queryObj.save()
                messages.success(request, 'Request Succeed! Query deleted.')
                return redirect('reportWriterQueries')
            else:
                messages.error(request, 'Request Failed! Query cannot be delete.Please try again.')
                return redirect('reportWriterQueries')


# Delete Saved Search Against ID End#

@csrf_exempt
# Delete Saved Search Against ID Start#
def deleteRBQuery(request):
    print('hello')
    with transaction.atomic():
        if request.method == 'GET':
            report_id = request.GET.get('rbQuery')
            # probably you want to add a regex check if the username value is valid here
            if report_id:
                try:
                    queryObj = SavedRBReports.objects.get(pk=report_id)
                except:
                    return render_to_response('itrak/page-404.html')

                queryObj.rb_report_is_delete = 1
                queryObj.rb_modified_by_id = request.user.id
                queryObj.rb_modified_at = datetime.now()
                queryObj.save()
                messages.success(request, 'Request Succeed! Query deleted.')
                return redirect('reportWriterReports')
            else:
                messages.error(request, 'Request Failed! Query cannot be delete.Please try again.')
                return redirect('reportWriterReports')


# Delete Saved Search Against ID End#


# Clone Query From Query Builder Start #

@active_user_required
def cloneQBQuery(request):
    with transaction.atomic():
        if request.method == 'GET':
            if 'qbQuery' in request.GET:
                query_id = request.GET.get('qbQuery')
                request.session['clone_id'] = query_id
                try:
                    queryObj = SavedQBQuries.objects.get(pk=query_id)
                except:
                    return render_to_response('itrak/page-404.html')

                newQueryObj = deepcopy(queryObj)
                newQueryObj.qb_query_id = None
                newQueryObj.qb_query_name = newQueryObj.qb_query_name + '(copy)'
                newQueryObj.save()
                insert_id = SavedQBQuries.objects.latest('pk').qb_query_id

            messages.success(request, 'Request Succeed! Query Clone created!.')
            # return redirect("/Home_EditQueryStep1?qbQuery=" + str(insert_id))
            return redirect(reverse('editQueryStep1') + '?qbQuery=' + str(insert_id))
        else:
            messages.error(request, 'Request Failed! Query Clone cannot be created.Please try again.')
            return redirect('reportWriterQueries')


# Clone Query From Query Builder End #

# Clone RBQuery From Query Builder Start #

@active_user_required
def cloneRBQuery(request):
    with transaction.atomic():
        if request.method == 'GET':
            if 'rbReport' in request.GET:
                report_id = request.GET.get('rbReport')
                request.session['rb_clone_id'] = report_id
                try:
                    queryObj = SavedRBReports.objects.get(pk=report_id)
                except:
                    return render_to_response('itrak/page-404.html')

                newQueryObj = deepcopy(queryObj)
                newQueryObj.rb_report_id = None
                newQueryObj.rb_report_name = newQueryObj.rb_report_name + '(copy)'
                newQueryObj.save()
                insert_id = SavedRBReports.objects.latest('pk').rb_report_id

            messages.success(request, 'Request Succeed! Query Clone created!.')
            # return redirect("/Home_editReportStep1?rbReport=" + str(insert_id))
            return redirect(reverse('editReportStep1') + '?rbReport=' + str(insert_id))
        else:
            messages.error(request, 'Request Failed! Query Clone cannot be created.Please try again.')
            return redirect('reportWriterReports')


# Clone RBQuery From Query Builder End #



# Report Writter Settings Start #

@active_user_required
def reportWriterSettings(request):
    user_id = request.user.id
    load_sidebar = get_sidebar(request)
    data_types = ReportSettingDataType.objects.filter(data_type_is_delete=0)
    justifications = ReportSettingJustification.objects.filter(justify_is_delete=0)
    # saved_settings = SaveReportSetting.objects.first()
    saved_settings = SaveReportSetting.objects.filter(rpt_setting_created_by_id= user_id).first()
    print(saved_settings)
    context = {
        'sidebar': load_sidebar,
        'data_types': data_types,
        'justifications': justifications,
        'saved_settings': saved_settings
    }
    return render(request, 'itrak/Reports/report_writer_settings.html', context)


# Report Writter Settings End #


# Update Report Setting Start #

@active_user_required
def updateReportSettings(request):
    with transaction.atomic():
        user_id = request.user.id
        org_id = request.user.user_org_id
        if request.method == 'POST':

            rpt_setting_dataType = request.POST.getlist('dataType')
            rpt_setting_width = request.POST.getlist('width')
            rpt_setting_maxRecordsReturn = request.POST.get('max_no_records')
            rpt_setting_maxRecordsPrint = request.POST.get('no_records_print')
            rpt_setting_maxRecordsDisplay = request.POST.get('no_records_display')
            rpt_setting_format = request.POST.getlist('format')
            rpt_setting_justification = request.POST.getlist('justicifation')

            # reportSettings = SaveReportSetting.objects.all()
            reportSettings = SaveReportSetting.objects.filter(rpt_setting_created_by_id= user_id)
            reportSettings.delete()

            for i, data_type in enumerate(rpt_setting_dataType):
                dt = ReportSettingDataType.objects.filter(rpt_data_type_name=rpt_setting_dataType[i]).first()
                if rpt_setting_width[i] == '':
                    rpt_setting_width[i] = 0
                obj = SaveReportSetting(rpt_setting_dataType_id=dt.rpt_data_type_id,
                                        rpt_setting_width=rpt_setting_width[i],
                                        max_records_return=rpt_setting_maxRecordsReturn,
                                        max_records_print=rpt_setting_maxRecordsPrint,
                                        max_records_display=rpt_setting_maxRecordsDisplay,
                                        rpt_setting_is_delete=False,
                                        rpt_setting_format_id=rpt_setting_format[i],
                                        rpt_setting_justification_id=rpt_setting_justification[i],
                                        rpt_setting_created_by_id=request.user.id)
                obj.save()
            return redirect('reportWriterSettings')
        else:
            messages.error(request, 'Request Failed! Report Settings cannot be updated.Please try again.')
            return redirect('reportWriterSettings')


# Update Report Settings End #


# Get Modal By Pair/Data set id Start #

@csrf_exempt
def getModalReportDataSet(request):
    if request.is_ajax() and request.method == 'POST':
        pairs = DataSetsPair.objects.filter(ds_is_deleted=0)
        context = {
            'pairs': pairs,
        }
        return render(request, 'itrak/Reports/data_set_field_list_modal.html', context)


# Get Modal By Pair/Data set id End #


# Get Pair/Data set Fields and Description Start #

@csrf_exempt
def getQueryDescriptionById(request):
    if request.is_ajax() and request.method == 'POST':
        reportDesc = SavedQBQuries.objects.get(pk=request.POST.get('reportID'))
        return HttpResponse(reportDesc.qb_query_description)


# Get Pair/Data set Fields and Description End #


# Get Query Definition By Id Start #

@csrf_exempt
def getModalQueryDef(request):
    if request.is_ajax() and request.method == 'POST':
        queryDetails = SavedQBQuries.objects.get(pk=request.POST.get('queryID'))
        querySharedWith = SavedQBQuriesShareWith.objects.filter(qb_query_id=request.POST.get('queryID'))
        context = {
            'queryDetails': queryDetails,
            'querySharedWith': querySharedWith,
        }
        return render(request, 'itrak/Reports/query_definition_modal.html', context)


# Get Query Definition By Id End #

# Get Query Definition By Id Start #

@csrf_exempt
def getModalReportDef(request):
    if request.is_ajax() and request.method == 'POST':
        reportDetails = SavedRBReports.objects.get(pk=request.POST.get('reportID'))
        reportSharedWith = SavedRBReportsShareWith.objects.filter(rb_report_id=request.POST.get('reportID'))
        context = {
            'queryDetails': reportDetails,
            'reportSharedWith': reportSharedWith,
        }
        return render(request, 'itrak/Reports/report_definition_modal.html', context)


# Get Query Definition By Id End #


# Clear All Session and Get Report Writer Create Report Page Start #

@active_user_required
def reportWriterReports(request):
    load_sidebar = get_sidebar(request)
    request.session['query'] = ''
    request.session['selected_query_fields_array'] = ''
    request.session['unselected_query_fields_array'] = ''
    request.session['selected_group_fields_array'] = ''
    request.session['unselected_group_fields_array'] = ''
    request.session['selected_order_fields_array'] = ''
    request.session['selected_group_sorting'] = ''
    request.session['selected_sort_expressions'] = ''
    request.session['selected_format_fields_array'] = ''

    rb_reports = SavedRBReports.objects.filter(rb_report_is_delete=0).filter(Q(rbReportID__rb_report_share_with_id=request.user.id)|Q(rb_created_by_id=request.user.id)).distinct()
    context = {
        'sidebar': load_sidebar,
        'rb_reports': rb_reports
    }
    return render(request, 'itrak/Reports/report_writer_reports.html', context)


# Clear All Session and Get Report Writer Create Report Page End #


# Create Report Through Report Writer Step 1 Start #

@active_user_required
def newReport(request):
    load_sidebar = get_sidebar(request)
    qb_queries = SavedQBQuries.objects.filter(qb_query_is_delete=0).filter(Q(qbQueryID__qb_query_share_with_id=request.user.id) | Q(qb_created_by_id=request.user.id)).distinct()

    context = {
        'sidebar': load_sidebar,
        'qb_queries': qb_queries
    }
    if request.session['query']:
        context['query_set'] = int(request.session['query'])
    else:
        request.session['query'] = ''
        request.session['selected_query_fields_array'] = ''
        request.session['unselected_query_fields_array'] = ''
        request.session['selected_group_fields_array'] = ''
        request.session['unselected_group_fields_array'] = ''
        request.session['selected_order_fields_array'] = ''
        request.session['selected_group_sorting'] = ''
        request.session['selected_sort_expressions'] = ''
        request.session['selected_fields_formating'] = ''
        request.session['selected_format_fields_array'] = ''

    return render(request, 'itrak/Reports/new_report.html', context)


# Create Report Through Report Writer Step 1 End #


# Get Selected/Available Fields on Query Select in Report Writer Step1 Start #

@csrf_exempt
def getQueryFields(request):
    if request.is_ajax() and request.method == 'POST':
        request.session['query'] = request.POST.get('query')
        queryDetails = SavedQBQuries.objects.get(pk=request.POST.get('query'))
        return render(request, 'itrak/Reports/data_query_fields.html', {'queryDetails': queryDetails})


# Get Selected/Available Fields on Query Select in Report Writer Step1 End #


# Get Session Selected/Available Fields on Query Select in Report Writer Step1 Start #

@csrf_exempt
def getQuerySelectedFields(request):
    if request.is_ajax() and request.method == 'POST':
        step = request.POST.get('step')
        if step == '1':
            selectedFields = request.session['selected_query_fields_array']
            unselectedFields = request.session['unselected_query_fields_array']
            return render(request, 'itrak/Reports/data_selected_query_fields.html',
                          {'selecetdFields': selectedFields, 'unselecetdFields': unselectedFields})
        elif step == '2':
            selectedFields = request.session['selected_group_fields_array']
            unselectedFields = request.session['unselected_group_fields_array']
            return render(request, 'itrak/Reports/data_selected_query_fields.html',
                          {'selecetdFields': selectedFields, 'unselecetdFields': unselectedFields})
        elif step == '3':
            if request.session['selected_group_sorting'] and len(request.session['selected_group_sorting']) > 0:
                selectedSortingFields = request.session['selected_group_sorting']
                return render(request, 'itrak/Reports/data_selected_group_fields.html',
                              {'selectedSortingFields': selectedSortingFields})
            else:
                selectedFields = request.session['selected_order_fields_array']
                return render(request, 'itrak/Reports/data_selected_group_fields.html',
                              {'selectedFields': selectedFields})
        elif step == '4':
            selectedSortExpressions = request.session['selected_sort_expressions']
            unselectedFields = request.session['unselected_group_fields_array']
            return render(request, 'itrak/Reports/data_selected_sort_fields.html',
                          {'selectedSortExpressions': selectedSortExpressions, 'unselectedFields': unselectedFields})
        elif step == '5':
            if request.session['selected_fields_formating'] and len(request.session['selected_fields_formating']) > 0:
                selectedFormatingFields = request.session['selected_fields_formating']
                selectedGroupFields = request.session['selected_group_fields_array']
                selectedFields = request.session['selected_format_fields_array']
                query_id = request.session['query']
                return render(request, 'itrak/Reports/data_selected_formating_fields.html',
                              {'selectedFormatingFields': selectedFormatingFields,
                               'selectedGroupFields': selectedGroupFields, 'query_id': query_id,
                               'selectedFields': selectedFields})
            else:
                selectedFields = request.session['selected_format_fields_array']
                selectedGroupFields = request.session['selected_group_fields_array']
                query_id = request.session['query']
                return render(request, 'itrak/Reports/data_selected_formating_fields.html',
                              {'query_id': query_id, 'selectedGroupFields': selectedGroupFields,
                               'selectedFields': selectedFields})


# Get Session Selected/Available Fields on Query Select in Report Writer Step1 End#


# Create Report Through Report Writer Step 2 Start  #

@active_user_required
def secondReport(request):
    load_sidebar = get_sidebar(request)
    context = {}
    context['sidebar'] = load_sidebar
    return render(request, 'itrak/Reports/new_report2.html', context)


# Create Report Through Report Writer Step 2 End #


# Set Selected/Available fields of Step 1 in Session Start  #

@csrf_exempt
def setReportQuerySelectedFields(request):
    if request.is_ajax() and request.method == 'POST':
        request.session['unselected_query_fields_array'] = request.POST.get('unselected_fields').split(",")
        request.session['selected_query_fields_array'] = request.POST.get('selected_fields').split(",")
        if request.session['unselected_group_fields_array'] and request.session['selected_group_fields_array']:
            joinedlist = request.session['unselected_group_fields_array'] + request.session[
                'selected_group_fields_array']

            if collections.Counter(list(set(request.session['selected_query_fields_array']).intersection(
                    joinedlist))) == collections.Counter(request.session['selected_query_fields_array']):
                request.session['unselected_group_fields_array'] = request.session['unselected_group_fields_array']
                request.session['selected_group_fields_array'] = request.session['selected_group_fields_array']
            else:
                request.session['unselected_group_fields_array'] = request.session['selected_query_fields_array']
                request.session['selected_group_fields_array'] = ''
                request.session['selected_group_sorting'] = []
                request.session['selected_sort_expressions'] = ''
        else:
            request.session['unselected_group_fields_array'] = request.session['selected_query_fields_array']
            request.session['selected_group_fields_array'] = ''
            request.session['selected_group_sorting'] = []
            request.session['selected_sort_expressions'] = ''

        if request.session['selected_format_fields_array']:
            if collections.Counter(request.session['selected_format_fields_array']) == collections.Counter(
                    request.session['selected_query_fields_array']):
                request.session['selected_format_fields_array'] = request.session['selected_format_fields_array']
            else:
                request.session['selected_format_fields_array'] = request.session['selected_query_fields_array']
                request.session['selected_fields_formating'] = []
        else:
            request.session['selected_format_fields_array'] = request.session['selected_query_fields_array']
            request.session['selected_fields_formating'] = []

        return HttpResponse(request.session['selected_query_fields_array'])


# Set Selected/Available fields of Step 1 in Session End #


# Set Selected/Available fields of Step 2 in Session Start #

@csrf_exempt
def setReportGroupSelectedFields(request):
    if request.is_ajax() and request.method == 'POST':
        request.session['unselected_group_fields_array'] = request.POST.get('unselected_fields').split(",")
        request.session['selected_group_fields_array'] = request.POST.get('selected_fields').split(",")
        if request.session['selected_order_fields_array']:
            if collections.Counter(request.session['selected_group_fields_array']) == collections.Counter(
                    request.session['selected_order_fields_array']):
                request.session['selected_order_fields_array'] = request.session['selected_order_fields_array']
            else:
                request.session['selected_order_fields_array'] = request.session['selected_group_fields_array']
                request.session['selected_group_sorting'] = []
                request.session['selected_sort_expressions'] = ''
        else:
            request.session['selected_order_fields_array'] = request.session['selected_group_fields_array']
            request.session['selected_group_sorting'] = []
        return HttpResponse(request.session['selected_group_fields_array'])


# Set Selected/Available fields of Step 1 in Session End #


# Create Report Through Report Writer Step 3 Start  #

@active_user_required
def thirdReport(request):
    load_sidebar = get_sidebar(request)
    context = {
        'sidebar': load_sidebar,
    }
    return render(request, 'itrak/Reports/new_report3.html', context)


# Create Report Through Report Writer Step 3 End #


# Set Selected fields Header/Footer Check and Sort Order of Step 3 in Session Start #

@csrf_exempt
def setReportSortOrderVal(request):
    if request.is_ajax() and request.method == 'POST':
        selected_fields = request.session['selected_group_fields_array']
        sort_orders = request.POST.get('sort_order').split(",")
        show_headers = request.POST.get('sort_header').split(",")
        show_footers = request.POST.get('sort_footer').split(",")

        request.session['selected_group_sorting'] = []
        # iterates over 2 lists and excutes
        # 2 times as len(value)= 2 which is the
        count = 1
        for (field, order) in zip(selected_fields, sort_orders):
            sortDict = {}
            sortDict['field_name'] = field
            sortDict['sort_order'] = order
            if str(count) in show_headers:
                sortDict['header'] = 1
            else:
                sortDict['header'] = 0
            if str(count) in show_footers:
                sortDict['footer'] = 1
            else:
                sortDict['footer'] = 0
            count = count + 1
            request.session['selected_group_sorting'].append(sortDict)
        return HttpResponse(request.session['selected_group_sorting'])


#  Set Selected fields Header/Footer Check and Sort Order of Step 3 in Session Start #


# Create Report Through Report Writer Step 4 Start  #

@active_user_required
def fourthReport(request):
    load_sidebar = get_sidebar(request)
    context = {
        'sidebar': load_sidebar,
    }
    return render(request, 'itrak/Reports/new_report4.html', context)


# Create Report Through Report Writer Step 4 End #


# Set Sort Expression for Step 4 in Session Start #

@csrf_exempt
def setSavedSortExpression(request):
    if request.is_ajax() and request.method == 'POST':
        request.session['selected_sort_expressions'] = request.POST.get('sort_expressions').split(",")
        return HttpResponse(request.session['selected_sort_expressions'])


#  Set Sort Expression for Step 4 in Session Start #


# Create Report Through Report Writer Step 5 Start  #

@active_user_required
def fifthReport(request):
    load_sidebar = get_sidebar(request)
    context = {
        'sidebar': load_sidebar,
    }
    return render(request, 'itrak/Reports/new_report5.html', context)


# Create Report Through Report Writer Step 5 End #


# Set Format Expression for Step 5 in Session Start #

@csrf_exempt
def setSavedFormatExpression(request):
    if request.is_ajax() and request.method == 'POST':
        selected_fields = request.session['selected_fields_formating']
        actual_column_names = request.POST.get('actual_column_names').split(",")
        column_names = request.POST.get('column_names').split(",")
        formats = request.POST.get('formats').split(",")
        subtotals = request.POST.get('subtotals').split(",")
        justifications = request.POST.get('justifications').split(",")
        column_widths = request.POST.get('column_widths').split(",")

        request.session['selected_fields_formating'] = []
        for (actual_column_name, column_name, format, subtotal, justification, column_width) in zip(actual_column_names,
                                                                                                    column_names,
                                                                                                    formats, subtotals,
                                                                                                    justifications,
                                                                                                    column_widths):
            formatDict = {}
            formatDict['actual_column_name'] = actual_column_name
            formatDict['column_name'] = column_name
            formatDict['format'] = format
            formatDict['subtotal'] = subtotal
            formatDict['justification'] = justification
            formatDict['column_width'] = column_width
            request.session['selected_fields_formating'].append(formatDict)
        return HttpResponse(request.session['selected_fields_formating'])


#  Set Format Expression for Step 5 in Session Start #


# Create Report Through Report Writer Step 6 Start  #

@active_user_required
def sixthReport(request):
    org_id = request.user.user_org_id
    load_sidebar = get_sidebar(request)
    users = User.objects.filter(user_org_id = org_id).values('id', 'first_name', 'last_name').filter(is_delete=0)
    context = {
        'sidebar': load_sidebar,
        'users': users,
    }
    return render(request, 'itrak/Reports/new_report6.html', context)


# Create Report Through Report Writer Step 6 End #


# Save Final Query From Query Builder Start #

@active_user_required
def saveFinalNewReport(request):
    with transaction.atomic():
        if request.method == 'POST':
            if request.GET.get('IsRun') == 'True':
                IsRun = 1
            else:
                IsRun = 0
            if 'report_name' in request.POST:
                report_name = request.POST.get('report_name')
            if 'report_title' in request.POST:
                report_title = request.POST.get('report_title')
            if 'description' in request.POST and request.POST['description']:
                description = request.POST.get('description')
            else:
                description = ''
            if 'share' in request.POST and request.POST['share']:
                share = request.POST.get('share')
                share_with = request.POST.getlist('share_with')
                share_with = list(filter(lambda x: x != 'multiselect-all', share_with))
            else:
                share = 0
            query_id = int(request.session['query'])
            selected_query_fields_array = json.dumps(request.session['selected_query_fields_array'])
            unselected_query_fields_array = json.dumps(request.session['unselected_query_fields_array'])
            selected_group_fields_array = json.dumps(request.session['selected_group_fields_array'])
            unselected_group_fields_array = json.dumps(request.session['unselected_group_fields_array'])
            selected_order_fields_array = json.dumps(request.session['selected_order_fields_array'])
            selected_group_sorting = json.dumps(request.session['selected_group_sorting'])
            selected_sort_expressions = json.dumps(request.session['selected_sort_expressions'])
            selected_fields_formating = json.dumps(request.session['selected_fields_formating'])
            selected_format_fields_array = json.dumps(request.session['selected_format_fields_array'])

            obj = SavedRBReports(
                rb_report_name=report_name,
                rb_report_title=report_title,
                rb_report_description=description,
                rb_report_query_id=query_id,
                rb_report_is_share=share,
                rb_selected_query_fields_array=selected_query_fields_array,
                rb_unselected_query_fields_array=unselected_query_fields_array,
                rb_selected_group_fields_array=selected_group_fields_array,
                rb_unselected_group_fields_array=unselected_group_fields_array,
                rb_selected_order_fields_array=selected_order_fields_array,
                rb_selected_group_sorting=selected_group_sorting,
                rb_selected_sort_expressions=selected_sort_expressions,
                rb_selected_fields_formating=selected_fields_formating,
                rb_selected_format_fields_array=selected_format_fields_array,
                rb_created_by_id=request.user.id
            )
            obj.save()
            insert_id = SavedRBReports.objects.latest('pk').rb_report_id

            if 'share' in request.POST:
                for userID in share_with:
                    obj1 = SavedRBReportsShareWith(rb_report_id=insert_id, rb_report_share_with_id=userID)
                    obj1.save()

            if IsRun == 0:
                messages.success(request, 'Request Succeed! Report created.')
                return redirect('reportWriterReports')
            else:
                # return redirect("/Home_rbReportProcess?rbReport=" + str(insert_id))
                return redirect(reverse('rbReportProcess') + '?rbReport=' + str(insert_id))
        else:
            messages.error(request, 'Request Failed! Report cannot be created.Please try again.')
            return redirect('reportWriterReports')


# Save Final Query From Query Builder StartEnd #


# Edit Report Builder Report Step 1 Start #

@active_user_required
def editReportStep1(request):
    report_id = request.GET.get('rbReport')
    load_sidebar = get_sidebar(request)
    # qb_queries = SavedQBQuries.objects.filter().filter(qb_query_is_delete=0)
    qb_queries = SavedQBQuries.objects.filter(qb_query_is_delete=0).filter(Q(qbQueryID__qb_query_share_with_id=request.user.id) | Q(qb_created_by_id=request.user.id)).distinct()
    context = {
        'sidebar': load_sidebar,
        'qb_queries': qb_queries,
        'report_id': report_id
    }

    reportObj = SavedRBReports.objects.get(pk=report_id)

    request.session['query'] = reportObj.rb_report_query_id
    request.session['selected_query_fields_array'] = json.loads(reportObj.rb_selected_query_fields_array)
    request.session['unselected_query_fields_array'] = json.loads(reportObj.rb_unselected_query_fields_array)
    request.session['selected_group_fields_array'] = json.loads(reportObj.rb_selected_group_fields_array)
    request.session['unselected_group_fields_array'] = json.loads(reportObj.rb_unselected_group_fields_array)
    request.session['selected_order_fields_array'] = json.loads(reportObj.rb_selected_order_fields_array)
    request.session['selected_group_sorting'] = json.loads(reportObj.rb_selected_group_sorting)
    request.session['selected_sort_expressions'] = json.loads(reportObj.rb_selected_sort_expressions)
    request.session['selected_fields_formating'] = json.loads(reportObj.rb_selected_fields_formating)
    request.session['selected_format_fields_array'] = json.loads(reportObj.rb_selected_format_fields_array)

    if request.session['query']:
        context['query_set'] = int(request.session['query'])
    else:
        request.session['query'] = ''
        request.session['selected_query_fields_array'] = ''
        request.session['unselected_query_fields_array'] = ''
        request.session['selected_group_fields_array'] = ''
        request.session['unselected_group_fields_array'] = ''
        request.session['selected_order_fields_array'] = ''
        request.session['selected_group_sorting'] = ''
        request.session['selected_sort_expressions'] = ''
        request.session['selected_fields_formating'] = ''
        request.session['selected_format_fields_array'] = ''
    return render(request, 'itrak/Reports/edit_report.html', context)


# Edit Report Builder Report Step 1 End #


# Edit Report Through Report Writer Step 2 Start  #

@active_user_required
def editReportStep2(request):
    report_id = request.GET.get('rbReport')
    load_sidebar = get_sidebar(request)
    context = {
        'sidebar': load_sidebar,
        'report_id': report_id
    }
    return render(request, 'itrak/Reports/edit_report2.html', context)


# Edit Report Through Report Writer Step 2 End #


# Edit Report Through Report Writer Step 3 Start  #

@active_user_required
def editReportStep3(request):
    report_id = request.GET.get('rbReport')
    load_sidebar = get_sidebar(request)
    context = {
        'sidebar': load_sidebar,
        'report_id': report_id
    }
    return render(request, 'itrak/Reports/edit_report3.html', context)


# Edit Report Through Report Writer Step 3 End #


# Edit Report Through Report Writer Step 4 Start  #

@active_user_required
def editReportStep4(request):
    report_id = request.GET.get('rbReport')
    load_sidebar = get_sidebar(request)
    context = {
        'sidebar': load_sidebar,
        'report_id': report_id
    }
    return render(request, 'itrak/Reports/edit_report4.html', context)


# Edit Report Through Report Writer Step 4 End #


# Create Report Through Report Writer Step 5 Start  #

@active_user_required
def editReportStep5(request):
    report_id = request.GET.get('rbReport')
    load_sidebar = get_sidebar(request)
    context = {
        'sidebar': load_sidebar,
        'report_id': report_id
    }
    return render(request, 'itrak/Reports/edit_report5.html', context)


# Create Report Through Report Writer Step 5 End #


# Create Report Through Report Writer Step 6 Start  #

@active_user_required
def editReportStep6(request):
    report_id = request.GET.get('rbReport')
    load_sidebar = get_sidebar(request)
    org_id = request.user.user_org_id
    users = User.objects.filter(user_org_id = org_id).values('id', 'first_name', 'last_name').filter(is_delete=0)
    reportObj = SavedRBReports.objects.get(pk=report_id)
    share_lists = SavedRBReportsShareWith.objects.values_list('rb_report_share_with_id', flat=True).filter(
        rb_report_id=report_id)
    if 'rb_clone_id' in request.session and request.session['rb_clone_id'] != '':
        share_lists = SavedRBReportsShareWith.objects.values_list('rb_report_share_with_id', flat=True).filter(
            rb_report_id=request.session['rb_clone_id'])
    else:
        share_lists = SavedRBReportsShareWith.objects.values_list('rb_report_share_with_id', flat=True).filter(
            rb_report_id=report_id)

    context = {
        'sidebar': load_sidebar,
        'users': users,
        'share_lists': share_lists,
        'reportObj': reportObj,
        'report_id': report_id
    }
    return render(request, 'itrak/Reports/edit_report6.html', context)


# Create Report Through Report Writer Step 6 End #


# Update FinalReport From Report Builder Start #

@active_user_required
def updateFinalReport(request):
    with transaction.atomic():
        request.session['rb_clone_id'] = ''
        if request.method == 'POST':
            if request.GET.get('IsRun') == 'True':
                IsRun = 1
            else:
                IsRun = 0
            if 'report_id' in request.POST:
                report_id = request.POST.get('report_id')
                try:
                    reportObj = SavedRBReports.objects.get(pk=report_id)
                except:
                    return render_to_response('itrak/page-404.html')
            if 'report_name' in request.POST:
                report_name = request.POST.get('report_name')
            if 'report_title' in request.POST:
                report_title = request.POST.get('report_title')
            else:
                report_title = ''
            if 'description' in request.POST and request.POST['description']:
                description = request.POST.get('description')
            else:
                description = ''
            if 'share' in request.POST and request.POST['share']:
                share = request.POST.get('share')
                share_with = request.POST.getlist('share_with')
                share_with = list(filter(lambda x: x != 'multiselect-all', share_with))
            else:
                share = 0
                share_with = None

            query_id = int(request.session['query'])
            selected_query_fields_array = json.dumps(request.session['selected_query_fields_array'])
            unselected_query_fields_array = json.dumps(request.session['unselected_query_fields_array'])
            selected_group_fields_array = json.dumps(request.session['selected_group_fields_array'])
            unselected_group_fields_array = json.dumps(request.session['unselected_group_fields_array'])
            selected_order_fields_array = json.dumps(request.session['selected_order_fields_array'])
            selected_group_sorting = json.dumps(request.session['selected_group_sorting'])
            selected_sort_expressions = json.dumps(request.session['selected_sort_expressions'])
            selected_fields_formating = json.dumps(request.session['selected_fields_formating'])
            selected_format_fields_array = json.dumps(request.session['selected_format_fields_array'])

            reportObj.rb_report_name = report_name
            reportObj.rb_report_title = report_title
            reportObj.rb_report_description = description
            reportObj.rb_report_query_id = query_id
            reportObj.rb_report_is_share = share
            reportObj.rb_selected_query_fields_array = selected_query_fields_array
            reportObj.rb_unselected_query_fields_array = unselected_query_fields_array
            reportObj.rb_selected_group_fields_array = selected_group_fields_array
            reportObj.rb_unselected_group_fields_array = unselected_group_fields_array
            reportObj.rb_selected_order_fields_array = selected_order_fields_array
            reportObj.rb_selected_group_sorting = selected_group_sorting
            reportObj.rb_selected_sort_expressions = selected_sort_expressions
            reportObj.rb_selected_fields_formating = selected_fields_formating
            reportObj.rb_selected_format_fields_array = selected_format_fields_array
            reportObj.rb_modifed_by_id = request.user.id
            reportObj.rb_modified_at = datetime.now()

            reportObj.save()
            updated_id = reportObj.rb_report_id

            reportShareWithObj = SavedRBReportsShareWith.objects.filter(rb_report_id=updated_id).delete()

            if share_with:
                for userID in share_with:
                    obj1 = SavedRBReportsShareWith(
                        rb_report_id=updated_id,
                        rb_report_share_with_id=userID
                    )
                    obj1.save()

            if IsRun == 0:
                messages.success(request, 'Request Succeed! Report updated.')
                return redirect('reportWriterReports')
            else:
                # return redirect("/Home_rbReportProcess?rbReport=" + str(report_id))
                return redirect(reverse('rbReportProcess') + '?rbReport=' + str(report_id))
        else:
            messages.error(request, 'Request Failed! Report cannot be updated.Please try again.')
            return redirect('reportWriterReports')


# UpdateFinal Report From Report Builder StartEnd #


# Get Query Builder Query Process Start#

@active_user_required
def rbReportProcess(request):
    report_id = request.GET.get('rbReport')
    obj = SavedRBReports.objects.get(pk=report_id)

    report_name = obj.rb_report_name
    report_title = obj.rb_report_title
    report_description = obj.rb_report_description
    rb_report_query_id = obj.rb_report_query_id

    selected_fields = selected_query_fields = json.loads(obj.rb_selected_query_fields_array)
    unselected_query_fields = json.loads(obj.rb_unselected_query_fields_array)
    selected_group_fields = json.loads(obj.rb_selected_group_fields_array)
    selected_group_fields = list(filter(lambda x: x != '',selected_group_fields))
    # test = filter(None,selected_group_fields)
    unselected_group_fields = json.loads(obj.rb_unselected_group_fields_array)
    selected_order_fields = json.loads(obj.rb_selected_order_fields_array)
    selected_group_sorting = json.loads(obj.rb_selected_group_sorting)
    selected_group_sorting = list(filter(lambda x: x['field_name'] != '',selected_group_sorting))
    selected_sort_expressions = json.loads(obj.rb_selected_sort_expressions)
    selected_sort_expressions = list(filter(lambda x: x != '', selected_sort_expressions))
    report_headers = selected_fields_formating = json.loads(obj.rb_selected_fields_formating)
    selected_format_fields = json.loads(obj.rb_selected_format_fields_array)
    query_result = qbQueryProcess(request, rb_report_query_id, True)

    queryObj = SavedQBQuries.objects.get(pk=rb_report_query_id)
    qb_query_pair_id = queryObj.qb_query_pair_id

    if selected_group_fields:
        group_field_formating = []
        for ele in selected_group_fields:
            keyValList = ele
            selected_fields.remove(ele)
            # Setting Group Fields At Start of Report Header
            group_index = list(filter(lambda d: d['actual_column_name'] in keyValList, report_headers))[0]
            report_headers.remove(group_index)
            group_field_formating.append(group_index)
        selected_fields[0:0] = selected_group_fields
        report_headers[0:0] = group_field_formating


    selected_columns = []
    sorted_columns = []

    for field in selected_fields:
        try:
            actual_field = DataSetsPairFields.objects.values('df_actual_column_name').filter(df_pair_id=qb_query_pair_id).filter(
            df_name=field)[0]['df_actual_column_name']
        except:
            return render_to_response('itrak/page-404.html')
        selected_columns.append(actual_field)
    # Set the Sorting If Sorting Expression Exist

    for sort_expression in selected_sort_expressions:
        sort_string = sort_expression
        if sort_string.find("(Asc)") == -1:
            column_name = sort_string.split(' (Desc)')[0]
        else:
            column_name = sort_string.split(' (Asc)')[0]
        try:
            actual_field = DataSetsPairFields.objects.values('df_actual_column_name').filter(df_pair_id=qb_query_pair_id).filter(
                df_name=column_name)[0]['df_actual_column_name']
        except:
            return render_to_response('itrak/page-404.html')
        if sort_string.find("(Asc)") == -1:
            sort_value = '-' + actual_field
        else:
            sort_value = actual_field
        sorted_columns.append(sort_value)

    if selected_sort_expressions:
        finalquery = query_result.values(*tuple(selected_columns)).order_by(*sorted_columns)
        print(finalquery)
    else:
        finalquery = query_result.values(*tuple(selected_columns))
        print(finalquery)
    grouped_records = []
    if selected_group_fields:
        for field in selected_group_sorting:
            group_selected_columns = []
            sortargs = []
            try:
                actual_field = DataSetsPairFields.objects.values('df_actual_column_name').filter(df_pair_id=qb_query_pair_id).filter(
                    df_name=field['field_name'])[0]['df_actual_column_name']
            except:
                return render_to_response('itrak/page-404.html')
            group_selected_columns.append(actual_field)

            # Setting Sort Order Asc and Desc For Group Sort
            if field['sort_order'] == '0':
                sort_value = actual_field
            else:
                sort_value = '-' + actual_field
            sortargs.append(sort_value)

            grouped_records = query_result.values(*tuple(group_selected_columns)).filter().annotate(tcount=Count(actual_field)).order_by(*sortargs)

            # Setting Justification, Column Width, Format
            keyValList = field['field_name']
            group_format_array = list(filter(lambda d: d['actual_column_name'] in keyValList, selected_fields_formating))[0]
            # Setting Check For Header/Footer for Display
            for record in grouped_records:
                record['group_record_name'] = list(record.items())[0][1]
                record['group_record_column'] = list(record.items())[0][0]
                record['header'] = field['header']
                record['footer'] = field['footer']
                record['colspan'] = len(selected_fields) - 1
                if group_format_array:
                    record['justification'] = group_format_array['justification']
                    record['column_width'] = group_format_array['column_width']
                    record['format'] = group_format_array['format']
            break

    report_records = zip(finalquery, selected_fields_formating)

    timezone = MySettings.objects.filter(m_user_id=request.user.id).first()
    eastern = pytz.timezone(timezone.m_time_zone)
    time = datetime.now(eastern)
    time = time.strftime('%m/%d/%Y %I:%M %p')
    # print(report_records)
    # print(selected_group_fields)
    # print(selected_group_sorting)

    context = {
        'report_name': report_name,
        'report_id':report_id,
        'report_title': report_title,
        'report_description': report_description,
        'report_query_id': rb_report_query_id,
        'report_query_pair_id': qb_query_pair_id,
        'selected_fields': selected_fields,
        'selected_columns': selected_columns,
        'report_headers': report_headers,
        'records': finalquery,
        'fields_formating' : selected_fields_formating,
        'report_records': report_records,
        'grouped_records': grouped_records,
        'groups': selected_group_fields,
        'ungroups':selected_sort_expressions,
        'groups_sorting': selected_group_sorting,
        'time':time,
    }

    return render(request, 'itrak/Reports/rb_report_result_list.html', context)


# Get Query Builder Query Process End#


# Get Report set Fields and Description Start #

@csrf_exempt
def getReportDescriptionById(request):
    if request.is_ajax() and request.method == 'POST':
        reportDesc = SavedRBReports.objects.get(pk=request.POST.get('reportID'))
        return HttpResponse(reportDesc.rb_report_description)


# Get Report set Fields and Description End #


# Add Schedule Report Start
@active_user_required
@csrf_exempt
def addScheduledReport(request):
    org_id = request.user.user_org_id
    load_sidebar = get_sidebar(request)
    users = User.objects.filter(user_org_id = org_id).filter(is_delete=0)
    schedule_report_span = ScheduleReportSpan.objects.filter()
    context = {
        'sidebar': load_sidebar,
        'schedule_report_span':schedule_report_span,
        'users': users,
    }
    return render(request, 'itrak/Reports/add_report_schedule.html', context)


@active_user_required
# Add Scheduled Report End

# get Reports By Report Type SaveSearch|ReportWrier Start
@csrf_exempt
def getReportsByReportType(request):
    if request.method == 'POST' and request.is_ajax():
        reptype = request.POST.get('rp_type')
        if reptype == '0':
            result = TicketSavedSearch.objects.filter(save_created_by_id=request.user.id)
            print(result)

        if reptype == '1':
            result = SavedRBReports.objects.filter(rb_report_is_delete=0).filter(Q(rbReportID__rb_report_share_with_id=request.user.id)|Q(rb_created_by_id=request.user.id)).distinct()
            print(result)
        response_data = {}
        try:
            response_data['response'] = serializers.serialize('json', result)
        except:
            response_data['response'] = 'No Record Found'
        return JsonResponse(response_data)


# get Reports By Report Type SaveSearch|ReportWrier End

# save schedule report Start
@csrf_exempt
def saveScheduleReport(request):
    if request.method == 'POST':
        if 'sch_name' in request.POST:
            sch_name = request.POST.get('sch_name')
        if 'is_active' in request.POST:
            is_active = request.POST.get('is_active')
        else:
            is_active =0
        if 'rp_type' in request.POST:
            rp_type = request.POST.get('rp_type')
        if 'rep_name' in request.POST:
            rep_name = request.POST.get('rep_name')
        if 'out_type' in request.POST:
            out_type = request.POST.get('out_type')
        if 'schedule' in request.POST:
            schedule = request.POST.get('schedule')
        if 'end_sch_date' in request.POST and request.POST.get('end_sch_date') != '':
            end_sch_date = datetime.strptime(request.POST.get('end_sch_date'), '%m-%d-%Y').strftime('%Y-%m-%d')
        else:
            end_sch_date = None
        if 'notify_error' in request.POST:
            notify_error = request.POST.get('notify_error')
        if 'comment' in request.POST:
            comment = request.POST.get('comment')
        if rp_type == '0':
            print('report_name')
            report_name = TicketSavedSearch.objects.get(pk=rep_name)
            print('report_name')
        print(type(rp_type))
        if rp_type ==  '1':
            print('report_name')
            report_name = SavedRBReports.objects.get(pk=rep_name)
            print(report_name)
        # obj = ScheduledReport(
        #     sch_rpt_name=sch_name,
        #     is_active=is_active,
        #     rep_type=rp_type,
        #     rep_name=rep_name,
        #     out_type=out_type,
        #     schedule=schedule,
        #     end_sch_rpt_date=end_sch_date,
        #     notify_error_id=notify_error,
        #     comment=comment,
        #     sch_rpt_created_by_id=request.user.id,
        #     sch_rpt_org_id=request.user.user_org_id
        # )
        # obj.save()
        # insert_id = ScheduledReport.objects.latest('pk').sch_rpt_id
        # if 'schedule' in request.POST:
        #     schedule = request.POST.get('schedule')
        #     if schedule == '1':
        #         obj1 = ScheduleReportSchedulebyFilters(
        #             sch_rep_span_id_id = schedule,
        #             sch_rep_id_id= insert_id
        #         )
        #         obj1.save()
        #     if schedule == '2':
        #         if 'week_days' in request.POST and request.POST['week_days']:
        #             week_days = request.POST.getlist('week_days')                       
        #         obj2 = ScheduleReportSchedulebyFilters(
        #             sch_rep_span_id_id = schedule,
        #             sch_rep_id_id= insert_id,
        #             week_days= week_days
        #         )
        #         obj2.save()
        #     if schedule == '3':
        #         if 'is_MonthDay_Specific' in request.POST:
        #             is_MonthDay_Specific = request.POST.get('is_MonthDay_Specific')
        #             if is_MonthDay_Specific == '0':
        #                 if 'Day_value' in request.POST:
        #                     Day_value = request.POST.get('Day_value')
        #                 if 'Day_every_month' in request.POST:
        #                     Day_every_month = request.POST.get('Day_every_month')  
        #                 obj3 = ScheduleReportSchedulebyFilters(
        #                     sch_rep_span_id_id = schedule,
        #                     sch_rep_id_id= insert_id,
        #                     days= Day_value,
        #                     is_MonthDay_Specific= is_MonthDay_Specific,
        #                     everyMonth = Day_every_month
        #                     )
        #                 obj3.save()
        #             else:
        #                 if 'the_first' in request.POST:
        #                     the_first = request.POST.get('the_first')
        #                 if 'the_days' in request.POST:
        #                     the_days = request.POST.get('the_days')
        #                 if 'the_every_month' in request.POST:
        #                     the_every_month = request.POST.get('the_every_month')       
        #                 obj4 = ScheduleReportSchedulebyFilters(
        #                     sch_rep_span_id_id = schedule,
        #                     sch_rep_id_id= insert_id,
        #                     the_first= the_first,
        #                     single_WeakDay=the_days,
        #                     is_MonthDay_Specific= is_MonthDay_Specific,
        #                     everyMonth = the_every_month
        #                     )
        #                 obj4.save()
        #     if schedule == '4':
        #         if 'quartely_Beginning_Date' in request.POST:
        #             quartely_Beginning_Date = datetime.strptime(request.POST.get('quartely_Beginning_Date'), '%m-%d-%Y').strftime('%Y-%m-%d')
        #         obj5 = ScheduleReportSchedulebyFilters(
        #                     sch_rep_span_id_id = schedule,
        #                     sch_rep_id_id= insert_id,
        #                     quartely_Beginning_Date=quartely_Beginning_Date
        #                     )
        #         obj5.save()
        #     if schedule == '5':
        #         if 'biannually_start_date' in request.POST:
        #             biannually_start_date = datetime.strptime(request.POST.get('biannually_start_date'), '%m-%d-%Y').strftime('%Y-%m-%d')
        #         if 'biannually_end_date' in request.POST:
        #             biannually_end_date = datetime.strptime(request.POST.get('biannually_end_date'), '%m-%d-%Y').strftime('%Y-%m-%d')    
        #         obj6 = ScheduleReportSchedulebyFilters(
        #                     sch_rep_span_id_id = schedule,
        #                     sch_rep_id_id= insert_id,
        #                     biannually_start_date=biannually_start_date,
        #                     biannually_end_date=biannually_end_date
        #                     )
        #         obj6.save() 
        #     if schedule =='6':
        #         if 'annually_date' in request.POST:
        #             annually_date = datetime.strptime(request.POST.get('annually_date'), '%m-%d-%Y').strftime('%Y-%m-%d')
        #         obj7 = ScheduleReportSchedulebyFilters(
        #                     sch_rep_span_id_id = schedule,
        #                     sch_rep_id_id= insert_id,
        #                     annually_date=annually_date
        #                     )
        #         obj7.save()
        #     if schedule == '7':
        #         if 'onetime_date' in request.POST:
        #             onetime_date = datetime.strptime(request.POST.get('onetime_date'), '%m-%d-%Y').strftime('%Y-%m-%d')
        #         obj8 = ScheduleReportSchedulebyFilters(
        #                     sch_rep_span_id_id = schedule,
        #                     sch_rep_id_id= insert_id,
        #                     onetime_date=onetime_date
        #                     )
        #         obj8.save()    

        messages.success(request, 'Request Succeed! Scheduled Report added.')
        return redirect('lsitScheduleReport')

    # save schedule report end


# list schedule Report start
@active_user_required
@csrf_exempt
def lsitScheduleReport(request):
    org_id = request.user.id
    load_sidebar = get_sidebar(request)
    users = User.objects.filter(user_org_id = org_id).filter(is_delete=0)
    context = {
        'sidebar': load_sidebar,
        'users': users,
    }
    return render(request, 'itrak/Reports/list_report_schedule.html', context)


# list schedule Report end

# Schedule Report Delete Request Start#

@active_user_required
def DelScheduleReport(request):
    # Instead of having an error on your server,
    # your user will get a 404 meaning that he tries to access a non existing resource.
    # data = get_object_or_404(Organization , pk = id)
    id = request.GET.get('schrepID')
    try:
        obj = ScheduledReport.objects.get(pk=id)
    except ScheduledReport.DoesNotExist:
        return render_to_response('itrak/page-404.html')
    # If Object Response is Empty
    if not obj:
        messages.success(request, 'Request Failed! No Record Found.')
        return redirect('lsitScheduleReport')
    else:
        obj.sch_rpt_is_delete = 1
        obj.save()
        messages.success(request, 'Request Succeed! Scheduled Report deleted.')
        return redirect('lsitScheduleReport')

# Schedule Report Delete Request End#

# Schedule Report  Edit Request Start#

@active_user_required
def editScheduleReport(request):
    # Instead of having an error on your server,
    # your user will get a 404 meaning that he tries to access a non existing resource.
    # data = get_object_or_404(Organization , pk = id)
    id = request.GET.get('schrepID')
    try:
        sch_id = signing.loads(id)
        data = ScheduledReport.objects.get(pk=sch_id)
        data1 = ScheduleReportSchedulebyFilters.objects.get(sch_rep_id_id=sch_id)
        week_days = ScheduleReportSchedulebyFilters.objects.values_list('week_days', flat=True).filter(
        sch_rep_id_id=sch_id)[0]
    except ScheduledReport.DoesNotExist:
        return render_to_response('itrak/page-404.html')
    # If Object Response is Empty
    if not data:
        messages.success(request, 'Request Failed! No Record Found.')
        return redirect('lsitScheduleReport')
    else:
        load_sidebar = get_sidebar(request)
        users = User.objects.filter(user_org_id = request.user.user_org_id).filter(is_delete=0)
        schrep = ScheduleReportResp.objects.filter(sch_rep_id_id=sch_id)
        context = {
            'sidebar': load_sidebar,
            'data': data,
            'data1':data1,
            'week_days':week_days,
            'schrep':schrep,
            'users':users
        }

        return render(request, 'itrak/Reports/edit_report_schedule.html', context)

# Schedule Report  Edit Request End#

#Schedule Report Update Request Start
@active_user_required
def updateScheduledReport(request):
    if request.method == 'POST':
        id = request.POST.get('sch_rpt_id')
        # Instead of having an error on your server,
        # your user will get a 404 meaning that he tries to access a non existing resource.
        # data = get_object_or_404(Organization , pk = id)
        print(id)
        try:
            obj = ScheduledReport.objects.get(pk=id)
            SchRepSchbyFilter = ScheduleReportSchedulebyFilters.objects.get(sch_rep_id_id=id)
        except ScheduledReport.DoesNotExist:
            return render_to_response('itrak/page-404.html')
        # If Object Response is Empty
        if not obj:
            messages.success(request, 'Request Failed! No Record Found.')
            return redirect('lsitScheduleReport')
        else:
            if 'sch_name' in request.POST:
                obj.sch_rpt_name = request.POST.get('sch_name')
            if 'is_active' in request.POST:
                obj.is_active = request.POST.get('is_active')
            else:
                obj.is_active =0
            if 'rp_type' in request.POST:
                obj.rp_type = request.POST.get('rp_type')
            if 'rep_name' in request.POST:
                obj.rep_name = request.POST.get('rep_name')
            if 'out_type' in request.POST:
                obj.out_type = request.POST.get('out_type')
            if 'schedule' in request.POST:
                obj.schedule = request.POST.get('schedule')
            if 'end_sch_date' in request.POST and request.POST.get('end_sch_date') != '':
                obj.end_sch_rpt_date = datetime.strptime(request.POST.get('end_sch_date'), '%m-%d-%Y').strftime('%Y-%m-%d')
            else:
                obj.end_sch_rpt_date = None
            if 'notify_error_id' in request.POST:
                obj.notify_error_id = request.POST.get('notify_error_id')
            if 'comment' in request.POST:
                obj.comment = request.POST.get('comment')

            obj.save()
            SchRepSchbyFilter.delete()
            if 'schedule' in request.POST:
                schedule = request.POST.get('schedule')
            if schedule == '1':
                obj1 = ScheduleReportSchedulebyFilters(
                    sch_rep_span_id_id = schedule,
                    sch_rep_id_id= id 
                )
                obj1.save()
            if schedule == '2':
                if 'week_days' in request.POST and request.POST['week_days']:
                    week_days = request.POST.getlist('week_days')                       
                obj2 = ScheduleReportSchedulebyFilters(
                    sch_rep_span_id_id = schedule,
                    sch_rep_id_id= id ,
                    week_days= week_days
                )
                obj2.save()
            if schedule == '3':
                if 'is_MonthDay_Specific' in request.POST:
                    is_MonthDay_Specific = request.POST.get('is_MonthDay_Specific')
                    if is_MonthDay_Specific == '0':
                        if 'Day_value' in request.POST:
                            Day_value = request.POST.get('Day_value')
                        if 'Day_every_month' in request.POST:
                            Day_every_month = request.POST.get('Day_every_month')  
                        obj3 = ScheduleReportSchedulebyFilters(
                            sch_rep_span_id_id = schedule,
                            sch_rep_id_id= id ,
                            days= Day_value,
                            is_MonthDay_Specific= is_MonthDay_Specific,
                            everyMonth = Day_every_month
                            )
                        obj3.save()
                    else:
                        if 'the_first' in request.POST:
                            the_first = request.POST.get('the_first')
                        if 'the_days' in request.POST:
                            the_days = request.POST.get('the_days')
                        if 'the_every_month' in request.POST:
                            the_every_month = request.POST.get('the_every_month')       
                        obj4 = ScheduleReportSchedulebyFilters(
                            sch_rep_span_id_id = schedule,
                            sch_rep_id_id= id ,
                            the_first= the_first,
                            single_WeakDay=the_days,
                            is_MonthDay_Specific= is_MonthDay_Specific,
                            everyMonth = the_every_month
                            )
                        obj4.save()
            if schedule == '4':
                if 'quartely_Beginning_Date' in request.POST:
                    quartely_Beginning_Date = datetime.strptime(request.POST.get('quartely_Beginning_Date'), '%m-%d-%Y').strftime('%Y-%m-%d')
                obj5 = ScheduleReportSchedulebyFilters(
                            sch_rep_span_id_id = schedule,
                            sch_rep_id_id= id,
                            quartely_Beginning_Date=quartely_Beginning_Date
                            )
                obj5.save()
            if schedule == '5':
                if 'biannually_start_date' in request.POST:
                    biannually_start_date = datetime.strptime(request.POST.get('biannually_start_date'), '%m-%d-%Y').strftime('%Y-%m-%d')
                if 'biannually_end_date' in request.POST:
                    biannually_end_date = datetime.strptime(request.POST.get('biannually_end_date'), '%m-%d-%Y').strftime('%Y-%m-%d')    
                obj6 = ScheduleReportSchedulebyFilters(
                            sch_rep_span_id_id = schedule,
                            sch_rep_id_id= id,
                            biannually_start_date=biannually_start_date,
                            biannually_end_date=biannually_end_date
                            )
                obj6.save() 
            if schedule =='6':
                if 'annually_date' in request.POST:
                    annually_date = datetime.strptime(request.POST.get('annually_date'), '%m-%d-%Y').strftime('%Y-%m-%d')
                obj7 = ScheduleReportSchedulebyFilters(
                            sch_rep_span_id_id = schedule,
                            sch_rep_id_id= id,
                            annually_date=annually_date
                            )
                obj7.save()
            if schedule == '7':
                if 'onetime_date' in request.POST:
                    onetime_date = datetime.strptime(request.POST.get('onetime_date'), '%m-%d-%Y').strftime('%Y-%m-%d')
                obj8 = ScheduleReportSchedulebyFilters(
                            sch_rep_span_id_id = schedule,
                            sch_rep_id_id= id,
                            onetime_date=onetime_date
                            )
                obj8.save()

            # return HttpResponse('Success')
            messages.success(request, 'Request Succeed! Schedule Report updated.')
            return redirect('lsitScheduleReport')
    else:
        messages.error(request, 'Request Failed! Organization cannot be updated.Please try again.')
        return redirect('lsitScheduleReport')

# Schedule Report Update Request End#
# save schedule report Recipient Start
@csrf_exempt
def addScheduledReportRecpient(request):
    if request.method == 'POST':
        response_data1 = {}
        try:
            a = request.POST.get('sr_resp_recipt_user')
            b = request.POST.get('sr_resp_recipt_user_email')
            if a:
                n = ScheduleReportResp.objects.get(sr_resp_recipt_user_id=a)
                return JsonResponse(1,safe=False)
            else:
                n = ScheduleReportResp.objects.get(sr_resp_recipt_user_id=b)
                return JsonResponse(1, safe=False)

        except ScheduleReportResp.DoesNotExist:
            if 'sr_resp_recipt_user' in request.POST and request.POST.get('sr_resp_recipt_user') != '':
                sr_resp_recipt_user = request.POST.get('sr_resp_recipt_user')
                if 'sch_rep_id' in request.POST:
                  sch_rep_id = request.POST.get('sch_rep_id')
                obj = ScheduleReportResp(
                    sr_resp_recipt_user_id=sr_resp_recipt_user,
                    sch_rep_id_id=sch_rep_id,
                    sr_resp_created_at=request.user.id
                )
                obj.save()
                response_data = list(ScheduleReportResp.objects.values().filter(sch_rep_id=sch_rep_id))
                return JsonResponse(response_data, safe=False)
            if 'sr_resp_recipt_user_email' in request.POST:
                sr_resp_recipt_user_email = request.POST.get('sr_resp_recipt_user_email')
                if 'sch_rep_id' in request.POST:
                  sch_rep_id = request.POST.get('sch_rep_id')
                obj = ScheduleReportResp(
                    sr_resp_recipt_user_id=sr_resp_recipt_user_email,
                    sch_rep_id_id=sch_rep_id,
                    sr_resp_created_at=request.user.id
                )
                obj.save()
                response_data = list(ScheduleReportResp.objects.values().filter(sch_rep_id=sch_rep_id))
                return JsonResponse(response_data, safe=False)

    # save schedule report Recipient end
# schedule report Recipient Delete Request Start#

@active_user_required
def delScheduledReportRecpient(request):
    # Instead of having an error on your server,
    # your user will get a 404 meaning that he tries to access a non existing resource.
    # data = get_object_or_404(Organization , pk = id)
    response_data1 = {}
    if request.method == 'POST':
        sr_resp_id =  request.POST.get('sr_resp_id')
        sch_rep_id =  request.POST.get('sch_rep_id')
        if sr_resp_id:
            serviceContact = ScheduleReportResp.objects.get(sr_resp_id=sr_resp_id)
            serviceContact.delete()
            response_data1['response'] = 'Success'
        else:
            response_data1['response'] = 'No Record Found'
        response_data = list(ScheduleReportResp.objects.values().filter(sch_rep_id=sch_rep_id))
        return JsonResponse(response_data, safe=False)

# schedule report Recipient Delete Request End#    

# Datatable Code Start Here#
class ScheduleRptListJson(BaseDatatableView):
    # The model we're going to show
    model = ScheduledReport

    # define the columns that will be returned
    # columns = ['action', 'org_id', 'org_name', 'is_internal', 'display']

    # define column names that will be used in sorting
    # order is important and should be same as order of columns
    # displayed by datatables. For non sortable columns use empty
    # value like ''
    order_columns = ['', 'sch_rpt_name', 'sch_rpt_created_by__display_name', 'schedule', 'end_sch_rpt_date', 'sch_rpt_created_at', 'comment',
                     'is_active']

    # set max limit of records returned, this is used to protect our site if someone tries to attack our site
    # and make it return huge amount of data
    max_display_length = 500

    def render_column(self, row, column):
        # We want to render user as a custom column
        # rid = base64.urlsafe_b64encode(bytes(str(row.org_id), 'ascii'))
        # signer = Signer(salt='extra')
        # original = signer.sign(rid)
        # value = signer.unsign(original)
        # new = base64.b64decode(value).decode('ascii')
        # # return HttpResponse(original)
        # print(type(new))
        rid = signing.dumps(row.sch_rpt_id)

        if column == 'action':
            # escape HTML for security reasons
            # return escape('{0}'.format(row.site_title))
            return '<a href="Home_ScheduledReportEdit?schrepID=' + str(
                rid) + '"><i class="fa fa-pencil"></i></a> | <a href="#" data-href="Home_ScheduledReportDel?schrepID=' + str(
                row.sch_rpt_id) + '" data-toggle="modal" data-target="#confirm-delete"><i class="fa fa-trash-o"></a>'
            # return '<a href="{% url 'editOrganization' %}">Edit</a> | <a href="deleteOrg/' + str(row.org_id) + '">Delete</a>'
        elif column == 'is_active':
            if row.is_active == True:
                return '''<img alt="Completed" src="static/itrak/images/check.gif" border="0" style="">'''
            else:
                return escape('{0}'.format(''))
        elif column == 'schedule':
            if row.schedule == 1:
                return escape('{0}'.format('Daily'))
            elif row.schedule == 2:
                return escape('{0}'.format('Weekly'))
            elif row.schedule == 3:
                return escape('{0}'.format('Monthly'))
            elif row.schedule == 4:
                return escape('{0}'.format('Quarterly'))
            elif row.schedule == 5:
                return escape('{0}'.format('Biannually'))
            elif row.schedule == 6:
                return escape('{0}'.format('Annually'))                    
            else:
                return escape('{0}'.format('One Time Only'))
        elif column == 'end_sch_rpt_date':
            if row.end_sch_rpt_date != None:
                print(row.end_sch_rpt_date)
                return datetime.strptime(str(row.end_sch_rpt_date), '%Y-%m-%d %H:%M:%S%z').strftime('%d/%m/%Y')
            else:
                return ''
            # elif column == 'end_sch_rpt_date':
            #     print(row.end_sch_rpt_date)
            #     return datetime.strptime(str(row.end_sch_rpt_date),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y')
        else:
            return super(ScheduleRptListJson, self).render_column(row, column)

    def get_initial_queryset(self):
        # return queryset used as base for futher sorting/filtering
        # these are simply objects displayed in datatable
        # You should not filter data returned here by any filter values entered by user. This is because
        # we need some base queryset to count total number of records.
        user_id = self.request.user.id
        org_id = self.request.user.user_org_id
        return ScheduledReport.objects.filter(sch_rpt_org_id = org_id).filter(sch_rpt_is_delete=0)
        # return Organization.objects.filter(org_is_active=0, org_is_delete=1)

    def filter_queryset(self, qs):
        # use parameters passed in GET request to filter queryset

        # simple example:
        search = self.request.GET.get('search[value]', None)
        if search:
            qs = qs.filter(Q(sch_rpt_id__icontains=search) | Q(sch_rpt_name__icontains=search))
            # qs = qs.filter(name__istartswith=search)

        # more advanced example using extra parameters
        # filter_customer = self.request.GET.get('customer', None)
        #
        # if filter_customer:
        #     customer_parts = filter_customer.split(' ')
        #     qs_params = None
        #     for part in customer_parts:
        #         q = Q(customer_firstname__istartswith=part)|Q(customer_lastname__istartswith=part)
        #         qs_params = qs_params | q if qs_params else q
        #     qs = qs.filter(qs_params)
        return qs

        # def prepare_results(self, qs):
        #     # prepare list with output column data
        #     # queryset is already paginated here
        #     json_data = []
        #     for item in qs:
        #         json_data.append([
        #             escape("{0}".format('OK')),
        #             escape(item.org_id),  # escape HTML for security reasons
        #             escape(item.org_name),  # escape HTML for security reasons
        #             escape(item.is_internal),  # escape HTML for security reasons
        #             escape("{0}".format('OK'))
        #             # item.get_state_display(),
        #             # item.created.strftime("%Y-%m-%d %H:%M:%S"),
        #
        #         ])
        #     return json_data

# Datatable Code End Here#

@csrf_exempt
def export_rb_query_Report_xls(request):
    if request.method == 'POST':
 
        queryheader = request.POST.get('queryheader')

        records = qbQueryProcess(request, queryheader, True)
        saveQBProcess = SavedQBQuries.objects.get(pk=queryheader)
        selected_fields = json.loads(saveQBProcess.qb_selected_fields)
        # print('abc')
        # print(query_result)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-TicketSearchList.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Delete the default worksheet
        workbook.remove(workbook.active)

        # Define some styles and formatting that will be later used for cells
        header_font = Font(size=7, name='Segoe UI', bold=True, color='FFFFFF')
        centered_alignment = Alignment(horizontal='left')
        border_bottom = Border(
            bottom=Side(border_style='medium', color='21316f'),
        )
        wrapped_alignment = Alignment(
            vertical='top',
            wrap_text=True
        )

        # Define the column titles and widths
        columns = []

        # Create a worksheet/tab with the title of the category
        worksheet = workbook.create_sheet(
            title='TicketSearchEcxel',
            index=0,
        )
        # Define the background color of the header cells
        fill = PatternFill(
            start_color='21316f',
            end_color='21316f',
            fill_type='solid',
        )
        row_num = 1

        # Assign values, styles, and formatting for each cell in the header
        for col_num, (column_title) in enumerate(selected_fields, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.border = border_bottom
            cell.alignment = centered_alignment
            cell.fill = fill
            # set column width
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 20

        # Iterate through all movies of a category
        for record in records:
            row_num += 1
            if int(row_num) % 2 == 0:
                row_fill = PatternFill(
                    start_color='FFEFD5',
                    end_color='FFEFD5',
                    fill_type='solid',
                )
            else:
                row_fill = PatternFill(
                    start_color='FFFFFF',
                    end_color='FFFFFF',
                    fill_type='solid',
                )


            row_font = Font(size=8, name='Segoe UI', bold=False, color='000000')
            # Assign values, styles, and formatting for each cell in the row


            recordList = []
            for key,value in record.items():
                print(key)
                if key == 'created_at'  and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'client_created_at' and value is not None:
                    value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                            '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticket_is_reopen_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticket_closed_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticketNote__note_modified_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticketNote__note_created_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticketOrg__ticket_assign_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticketOrg__ticket_next_action_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'qb_created_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'qb_modified_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'rbQuerySetsQuery__rb_created_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'rbQuerySetsQuery__rb_modified_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticketManager__tmgr_completion_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticket_modified_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'submitted_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt), '%Y-%m-%d %H:%M:%S%z').strftime('%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'UserOrgId__created_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'UserOrgId__last_login' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                else:
                    recordList.append(value)
                
            print(recordList)
            for col_num, (cell_value) in enumerate(recordList, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.style = 'Normal'
                # if cell_format == 'Currency':
                #     cell.number_format = '#,##0.00 €'
                # if col_num == 8:
                #     cell.number_format = '[h]:mm;@'
                cell.alignment = wrapped_alignment
                cell.fill = row_fill
                cell.font = row_font

        # freeze the first row
        worksheet.freeze_panes = worksheet['A2']

        # set tab color
        worksheet.sheet_properties.tabColor = 'FFFFFF'

        workbook.save(response)
        return response
@csrf_exempt
def export_rb_report_Report_xls(request):
    if request.method == 'POST':

        reportheader = request.POST.get('reportheader')

        saveRBProcessheader = SavedRBReports.objects.get(pk=reportheader)
        print(saveRBProcessheader)
        # saveQBProcess = SavedQBQuries.objects.get(pk=reportheader)
        selected_fields = json.loads(saveRBProcessheader.rb_selected_query_fields_array)
        # # print('abc')
        # # print(query_result)
        response = HttpResponse(
             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
         )
        response['Content-Disposition'] = 'attachment; filename={date}-ReportResult.xlsx'.format(
             date=datetime.now().strftime('%Y-%m-%d'),
         )
        workbook = Workbook()

        # Delete the default worksheet
        workbook.remove(workbook.active)
        #
        # Define some styles and formatting that will be later used for cells
        header_font = Font(size=7, name='Segoe UI', bold=True, color='FFFFFF')
        centered_alignment = Alignment(horizontal='left')
        border_bottom = Border(
             bottom=Side(border_style='medium', color='21316f'),
         )
        wrapped_alignment = Alignment(
             vertical='top',
             wrap_text=True
         )
        #
        # Define the column titles and widths
        columns = []
        #
        # Create a worksheet/tab with the title of the category
        worksheet = workbook.create_sheet(
             title='ReportResultEcxel',
             index=0,
         )
        # Define the background color of the header cells
        fill = PatternFill(
             start_color='21316f',
             end_color='21316f',
             fill_type='solid',
         )
        row_num = 1
        #
        # # Assign values, styles, and formatting for each cell in the header
        for col_num, (column_title) in enumerate(selected_fields, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.border = border_bottom
            cell.alignment = centered_alignment
            cell.fill = fill
            # set column width
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 20
        #
        obj = SavedRBReports.objects.get(pk=reportheader)

        report_name = obj.rb_report_name
        report_title = obj.rb_report_title
        report_description = obj.rb_report_description
        rb_report_query_id = obj.rb_report_query_id

        selected_fields = selected_query_fields = json.loads(obj.rb_selected_query_fields_array)
        unselected_query_fields = json.loads(obj.rb_unselected_query_fields_array)
        selected_group_fields = json.loads(obj.rb_selected_group_fields_array)
        selected_group_fields = list(filter(lambda x: x != '', selected_group_fields))
        # test = filter(None,selected_group_fields)
        unselected_group_fields = json.loads(obj.rb_unselected_group_fields_array)
        selected_order_fields = json.loads(obj.rb_selected_order_fields_array)
        selected_group_sorting = json.loads(obj.rb_selected_group_sorting)
        selected_group_sorting = list(filter(lambda x: x['field_name'] != '', selected_group_sorting))
        selected_sort_expressions = json.loads(obj.rb_selected_sort_expressions)
        selected_sort_expressions = list(filter(lambda x: x != '', selected_sort_expressions))
        report_headers = selected_fields_formating = json.loads(obj.rb_selected_fields_formating)
        selected_format_fields = json.loads(obj.rb_selected_format_fields_array)
        query_result = qbQueryProcess(request, rb_report_query_id, True)

        queryObj = SavedQBQuries.objects.get(pk=rb_report_query_id)
        qb_query_pair_id = queryObj.qb_query_pair_id

        if selected_group_fields:
            group_field_formating = []
            for ele in selected_group_fields:
                keyValList = ele
                selected_fields.remove(ele)
                # Setting Group Fields At Start of Report Header
                group_index = list(filter(lambda d: d['actual_column_name'] in keyValList, report_headers))[0]
                report_headers.remove(group_index)
                group_field_formating.append(group_index)
            selected_fields[0:0] = selected_group_fields
            report_headers[0:0] = group_field_formating

        selected_columns = []
        sorted_columns = []

        for field in selected_fields:
            try:
                actual_field = \
                DataSetsPairFields.objects.values('df_actual_column_name').filter(df_pair_id=qb_query_pair_id).filter(
                    df_name=field)[0]['df_actual_column_name']
            except:
                return render_to_response('itrak/page-404.html')
            selected_columns.append(actual_field)
        # Set the Sorting If Sorting Expression Exist

        for sort_expression in selected_sort_expressions:
            sort_string = sort_expression
            if sort_string.find("(Asc)") == -1:
                column_name = sort_string.split(' (Desc)')[0]
            else:
                column_name = sort_string.split(' (Asc)')[0]
            try:
                actual_field = \
                DataSetsPairFields.objects.values('df_actual_column_name').filter(df_pair_id=qb_query_pair_id).filter(
                    df_name=column_name)[0]['df_actual_column_name']
            except:
                return render_to_response('itrak/page-404.html')
            if sort_string.find("(Asc)") == -1:
                sort_value = '-' + actual_field
            else:
                sort_value = actual_field
            sorted_columns.append(sort_value)

        if selected_sort_expressions:
            finalquery = query_result.values(*tuple(selected_columns)).order_by(*sorted_columns)
            print(finalquery)
        else:
            finalquery = query_result.values(*tuple(selected_columns))
            print(finalquery)




        # # Iterate through all movies of a category
        for record in finalquery:
            row_num += 1
            if int(row_num) % 2 == 0:
                row_fill = PatternFill(
                    start_color='FFEFD5',
                    end_color='FFEFD5',
                    fill_type='solid',
                )
            else:
                row_fill = PatternFill(
                    start_color='FFFFFF',
                    end_color='FFFFFF',
                    fill_type='solid',
                )

            row_font = Font(size=8, name='Segoe UI', bold=False, color='000000')
        #     # Assign values, styles, and formatting for each cell in the row
        #
            recordList = []
            for key, value in record.items():
                print(key)
                if key == 'created_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'client_created_at' and value is not None:
                    value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                            '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticket_created_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticket_is_reopen_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticket_closed_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticketNote__note_modified_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticketNote__note_created_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticketOrg__ticket_assign_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticketOrg__ticket_next_action_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'qb_created_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'qb_modified_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'rbQuerySetsQuery__rb_created_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'rbQuerySetsQuery__rb_modified_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticketManager__tmgr_completion_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticket_modified_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                        '%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'submitted_at' and value is not None:
                    uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                    local_dt = value.astimezone(pytz.timezone(uTimeZone))
                    value1 = datetime.strptime(str(local_dt), '%Y-%m-%d %H:%M:%S%z').strftime('%m/%d/%Y %I:%M %p')
                    recordList.append(value1)
                elif key == 'ticketAttach__attach_created_at' and value is not None:
                    if value != None:
                        value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                            '%m/%d/%Y %I:%M %p')
                        recordList.append(value1)
                elif key == 'client_created_at' and value is not None:
                    if value != None:
                        value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                            '%m/%d/%Y %I:%M %p')
                        recordList.append(value1)
                elif key == 'ticket_assign_at' and value is not None:
                    if value != None:
                        value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                            '%m/%d/%Y %I:%M %p')
                        recordList.append(value1)
                elif key == 'ticket_next_action_at' and value is not None:
                    if value != None:
                        value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                            '%m/%d/%Y %I:%M %p')
                        recordList.append(value1)
                elif key == 'submitted_at' and value is not None:
                    if value != None:
                        value1 = datetime.strptime(str(value), '%Y-%m-%d %H:%M:%S%z').strftime('%m/%d/%Y %I:%M %p')
                        recordList.append(value1)
                elif key == 'ticket_created_at' and value is not None:
                    if value != None:
                        value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                            '%m/%d/%Y %I:%M %p')
                        recordList.append(value1)
                else:
                    recordList.append(value)
            for col_num, (cell_value) in enumerate(recordList, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.style = 'Normal'
                # if cell_format == 'Currency':
                #     cell.number_format = '#,##0.00 €'
                # if col_num == 8:
                #     cell.number_format = '[h]:mm;@'
                cell.alignment = wrapped_alignment
                cell.fill = row_fill
                cell.font = row_font
        #
        # # freeze the first row
        worksheet.freeze_panes = worksheet['A2']
        #
        # # set tab color
        worksheet.sheet_properties.tabColor = 'FFFFFF'
        #
        workbook.save(response)
        return response

#CHECK SAVE SEARCH IS SHAREABLE OR NOT.
@csrf_exempt
def isSaveSearchShareable(request):
    if request.is_ajax() and request.method == 'POST':
        savedSearchID = request.POST.get('saved_search_id')
        savedSearches = TicketSavedSearch.objects.get(pk= savedSearchID)
        isShare = savedSearches.is_share 
        return HttpResponse(json.dumps(isShare), content_type="application/json")
    else:
        return HttpResponse('fail')



def runDailyScheduleJob(request):
    scheduledReports = ScheduledReport.objects.filter(sch_rpt_saved_search=1018).filter(sch_rpt_is_delete= 0).filter(is_active= 1).filter(schedule = 1)
    res = 1
    if scheduledReports:
        print('here1')
        for scheduledReport in scheduledReports:
            print('here2')
            # For Saved Search
            if scheduledReport.sch_rpt_saved_search:
                obj = TicketSavedSearch.objects.get(pk=1018)
                print('here3')
                kwargs = {
                    '{0}__{1}'.format('ticket_is_delete', 'iexact'): 0,
                    '{0}__{1}'.format('ticket_is_active', 'iexact'): 1,
                }
                ticket_status_dict = {0: "Opened", 1: "Closed"}
                labor_hours_dict = {"0": "Less Than", "1": "More Than", "2": "Equal"}
                fielddict = {}
                # fielddict = {"1": ''}
                if obj.ticket_status != None:
                    ticket_status = obj.ticket_status
                    kwargs.setdefault('ticket_status__iexact', ticket_status)
                    fielddict.update({'Ticket Status': ticket_status_dict[ticket_status]})
                if obj.ticket_sub_status_id != None:
                    ticket_sub_status = obj.ticket_sub_status_id
                    kwargs.setdefault('ticket_sub_status_id', ticket_sub_status)
                    result = SubStatus.objects.only('sub_status_text').get(pk=ticket_sub_status).sub_status_text
                    if obj.ticket_status != None:
                        fielddict.update({'Ticket Status': ticket_status_dict[ticket_status] + ' - ' + result})
                    else:
                        fielddict.update({'Ticket Status': ' - ' + result})
                if obj.priority_id != None:
                    priority = obj.priority_id
                    kwargs.setdefault('priority_id', priority)
                    result = Priority.objects.only('priority_name').get(pk=priority).priority_name
                    fielddict.update({'Priority': result})
                if obj.ticket_type_id != None:
                    ticket_type = obj.ticket_type_id
                    kwargs.setdefault('ticket_type_id', ticket_type)
                    result = TicketType.objects.only('ttype_name').get(pk=ticket_type).ttype_name
                    fielddict.update({'Ticket Type': result})
                if obj.ticket_subtype1_id != None:
                    subtype1 = obj.ticket_subtype1_id
                    kwargs.setdefault('ticket_subtype1_id', subtype1)
                    result = TicketType.objects.only('ttype_name').get(pk=subtype1).ttype_name
                    fielddict.update({'Subtype 1': result})
                if obj.ticket_subtype2_id != None:
                    subtype2 = obj.ticket_subtype2_id
                    kwargs.setdefault('ticket_subtype2_id', subtype2)
                    result = TicketType.objects.only('ttype_name').get(pk=subtype2).ttype_name
                    fielddict.update({'Subtype 2': result})
                if obj.ticket_subtype3_id != None:
                    subtype3 = obj.ticket_subtype3_id
                    kwargs.setdefault('ticket_subtype3_id', subtype3)
                    result = TicketType.objects.only('ttype_name').get(pk=subtype3).ttype_name
                    fielddict.update({'Subtype 3': result})
                if obj.ticket_subtype4_id != None:
                    subtype4 = obj.ticket_subtype4_id
                    kwargs.setdefault('ticket_subtype4_id', subtype4)
                    result = TicketType.objects.only('ttype_name').get(pk=subtype4).ttype_name
                    fielddict.update({'Subtype 4': result})
                if obj.subject != '':
                    subject = obj.subject
                    kwargs.setdefault('subject__icontains', subject)
                    fielddict.update({'Subject': subject})
                if obj.ticket_note != '':
                    ticket_note = request.POST.get('ticket_note')
                    kwargs.setdefault('ticketNote__note_detail__icontains', ticket_note)
                    fielddict.update({'Notes': ticket_note})
                if obj.all_three != '':
                    all_three = request.POST.get('all_three')
                    args = Q(subject__icontains=all_three) | Q(ticketNote__note_detail__icontains=all_three)
                    fielddict.update({'Search All Three': all_three})
                if obj.ticket_record_locator != '':
                    record_locator = obj.ticket_record_locator
                    kwargs.setdefault('ticket_record_locator__startswith', record_locator)
                    fielddict.update({'Record Locator': record_locator})
                if obj.ticket_caller_name != '':
                    caller_name = obj.ticket_caller_name
                    kwargs.setdefault('ticket_caller_name__startswith', caller_name)
                    fielddict.update({'Caller Name': caller_name})
                if obj.ticket_caller_phone != '':
                    caller_phone = obj.ticket_caller_phone
                    kwargs.setdefault('ticket_caller_phone__startswith', caller_phone)
                    fielddict.update({'Caller Phone': caller_phone})
                if obj.ticket_caller_email != '':
                    caller_email = obj.ticket_caller_email
                    kwargs.setdefault('ticket_caller_email__startswith', caller_email)
                    fielddict.update({'Caller Email': caller_email})
                if obj.ticket_passenger_name != '':
                    passenger_name = request.POST.get('passenger_name')
                    kwargs.setdefault('ticket_passenger_name__startswith', passenger_name)
                    fielddict.update({'Passenger Name': passenger_name})
                if obj.note_entered_by_id != None:
                    note_entered_by = request.POST.get('note_entered_by')
                    kwargs.setdefault('ticketNote__note_created_by_id', note_entered_by)
                    result = User.objects.only('display_name').get(pk=note_entered_by).display_name
                    fielddict.update({'Note Entered By': result})
                if obj.submitted_by_id != None:
                    submitted_by = request.POST.get('submitted_by')
                    kwargs.setdefault('ticket_caller_id', submitted_by)
                    result = User.objects.only('display_name').get(pk=submitted_by).display_name
                    fielddict.update({'Submitted By': result})
                if obj.entered_by_id != None:
                    entered_by = request.POST.get('entered_by')
                    kwargs.setdefault('ticket_created_by_id', entered_by)
                    result = User.objects.only('display_name').get(pk=entered_by).display_name
                    fielddict.update({'Entered By': result})
                if obj.task_assigned_to_id != None and obj.ever_assigned == 0:
                    assigned_to = obj.task_assigned_to_id
                    kwargs.setdefault('ticket_assign_to_id', assigned_to)
                    result = User.objects.only('display_name').get(pk=assigned_to).display_name
                    fielddict.update({'Assigned To': result})
                if obj.task_assigned_to_id != None and obj.ever_assigned == 1:
                    assigned_to = obj.task_assigned_to_id
                    kwargs.setdefault('ticketURLog__urlog_event', 1)
                    kwargs.setdefault('ticketURLog__urlog_user_id', assigned_to)
                    result = User.objects.only('display_name').get(pk=assigned_to).display_name
                    fielddict.update({'Assigned To': result})
                if obj.assigned_by_id != None:
                    assigned_by = obj.assigned_by_id
                    kwargs.setdefault('ticket_assign_by_id', assigned_by)
                    result = User.objects.only('display_name').get(pk=assigned_by).display_name
                    fielddict.update({'Assigned By': result})
                if obj.next_action_id != None and obj.ever_next_action == 0:
                    next_action = obj.next_action_id
                    kwargs.setdefault('ticket_next_action_id', next_action)
                    result = User.objects.only('display_name').get(pk=next_action).display_name
                    fielddict.update({'Next Action': result})
                if obj.next_action_id != None and obj.ever_next_action == 1:
                    next_action = obj.next_action_id
                    kwargs.setdefault('ticketURLog__urlog_event', 2)
                    kwargs.setdefault('ticketURLog__urlog_user_id', next_action)
                    result = User.objects.only('display_name').get(pk=next_action).display_name
                    fielddict.update({'Next Action': result})
                if obj.closed_by_id != None:
                    closed_by = obj.closed_by_id
                    kwargs.setdefault('ticket_closed_by_id', closed_by)
                    result = User.objects.only('display_name').get(pk=closed_by).display_name
                    fielddict.update({'Closed By': result})
                if obj.org_id != None:
                    org_id = obj.org_id
                    kwargs.setdefault('ticket_org_id', org_id)
                    result = Organization.objects.only('org_name').get(pk=org_id).org_name
                    fielddict.update({'Organization': result})
                # if obj.client_id != None:
                #     client_id = obj.client_id
                #     kwargs.setdefault('ticket_client_id', client_id)
                #     result = Client.objects.only('client_name').get(pk=client_id).client_name
                #     fielddict.update({'Client': result})
                if obj.date_opened != '':
                    date_opened = obj.date_opened
                    dopen_start = datetime.strptime(obj.date_opened.split(' - ')[0], '%m/%d/%Y').strftime('%Y-%m-%d')
                    dopen_end = datetime.strptime(obj.date_opened.split(' - ')[1], '%m/%d/%Y').strftime('%Y-%m-%d')
                    kwargs.setdefault('ticket_created_at__gte', dopen_start)
                    kwargs.setdefault('ticket_created_at__lte', dopen_end)
                    fielddict.update({'Date Opened': date_opened})
                if obj.date_closed != '':
                    date_closed = obj.date_closed
                    dclose_start = datetime.strptime(obj.date_closed.split(' - ')[0], '%m/%d/%Y').strftime('%Y-%m-%d')
                    dclose_end = datetime.strptime(obj.date_closed.split(' - ')[1], '%m/%d/%Y').strftime('%Y-%m-%d')
                    kwargs.setdefault('ticket_closed_at__gte', dclose_start)
                    kwargs.setdefault('ticket_closed_at__lte', dclose_end)
                    fielddict.update({'Date Closed': date_closed})
                if obj.labour_hours_val != '':
                    labour_hours_val = obj.labour_hours_val
                    labour_hours_val += ':00:00'
                    labour_hours = obj.labour_hours
                    if labour_hours == '0':
                        condition = 'lte'
                    elif labour_hours == '1':
                        condition = 'gte'
                    else:
                        condition = 'startswith'
                    kwargs.setdefault('labour_hours__' + condition, str(labour_hours_val))
                    fielddict.update({'Labor Hours - ' + labor_hours_dict[labour_hours]: labour_hours})
                if obj.task_description != '':
                    task_description = obj.task_description
                    kwargs.setdefault('ticketManager__tmgr_task__task_description__icontains', task_description)
                    fielddict.update({'Task Description': task_description})
                if obj.task_assigned_to_id != None:
                    task_assigned_to = obj.task_assigned_to_id
                    kwargs.setdefault('ticketManager__task_assigned_to_id', task_assigned_to)
                    result = User.objects.only('display_name').get(pk=task_assigned_to).display_name
                    fielddict.update({'Task Assigned To': result})
                if obj.task_completion_date != '':
                    task_completion_date = obj.task_completion_date
                    dtask_start = datetime.strptime(obj.task_completion_date.split(' - ')[0], '%m/%d/%Y').strftime('%Y-%m-%d')
                    dtask_end = datetime.strptime(obj.task_completion_date.split(' - ')[1], '%m/%d/%Y').strftime('%Y-%m-%d')
                    kwargs.setdefault('ticketManager__task_due_date__gte', dtask_start)
                    kwargs.setdefault('ticketManager__task_due_date__lte', dtask_end)
                    fielddict.update({'Task Completion Date': task_completion_date})
                if obj.search_title != '':
                    search_title = obj.search_title
                else:
                    search_title = 'Search Results'

                if obj.search_title != '':
                    search_title = obj.search_title
                if obj.output_view != '':
                    output_view = obj.output_view 
                # Query To get all ticket ids
                if obj.all_three != '':
                    tickets = Ticket.objects.filter(**kwargs).filter(args).distinct()
                    ticketid_list = Ticket.objects.filter(**kwargs).filter(args).distinct().values_list('ticket_id', flat=True)
                else:
                    tickets = Ticket.objects.filter(**kwargs).distinct()
                    ticketid_list = Ticket.objects.filter(**kwargs).distinct().values_list('ticket_id', flat=True)

                print(list(ticketid_list))

                # Saving Excel in Specific Directory
                get_xls_from_scheduled_saved_search(request,list(ticketid_list), '0', scheduledReport.sch_rpt_created_by_id)

                # Send Email Request with Attachment Start#
                smtp = getSMTPSettings(1)
                smtpserver = smtplib.SMTP(smtp['email_server'], smtp['port'])
                # print(smtpserver.ehlo())
                smtpserver.starttls()
                # print(smtpserver.ehlo)
                smtpserver.login(smtp["user_name"], smtp["password"]) 
                print(smtp)
                emails_to = []
                if scheduledReport.notify_error_id:
                    to = User.objects.values_list('email',flat=True).get(pk=scheduledReport.notify_error_id)
                    emails_to.append(to)
                print(emails_to)
                # # recipients = ScheduleReportResp.objects.values_list('sr_resp_recipt_user_id',flat=True).filter(sch_rep_id_id = 9)
                # # print(recipients)
                # to_recipients = User.objects.values_list('email',flat=True).filter(pk=2)
                # print(to_recipients)
                # for to_recipient in to_recipients:
                #     emails_to.append(to_recipient) 
                print(emails_to)
                for email_to in emails_to:
                    msg = EmailMessage('Scheduled Report', 'Report For Saved Search', smtp["email_sender_name"], [email_to])
                    msg.content_subtype = "html"  
                    dirname = os.path.dirname(__file__)
                    filename = os.path.join(dirname, 'Attachments\Report.xlsx')
                    msg.attach_file(filename)
                    msg.send()
                # Send Email Request with Attachment End#

                # return HttpResponse(request,ticketid_list)
            # time.sleep(1)
            # For Report Writer Search
            if res == 1:
                print('elif')
                saveRBProcessheader = SavedRBReports.objects.get(pk=2060)
                print(saveRBProcessheader)
                # saveQBProcess = SavedQBQuries.objects.get(pk=reportheader)
                selected_fields = json.loads(saveRBProcessheader.rb_selected_query_fields_array)
                # print('abc')
                # print(query_result)
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = 'attachment; filename={date}-ReportResult.xlsx'.format(
                    date=datetime.now().strftime('%Y-%m-%d'),
                )
                # workbook = Workbook()
                wb = Workbook()
                ws1 = wb.active

                # Delete the default worksheet
                # workbook.remove(workbook.active)
                #
                # Define some styles and formatting that will be later used for cells
                header_font = Font(size=7, name='Segoe UI', bold=True, color='FFFFFF')
                centered_alignment = Alignment(horizontal='left')
                border_bottom = Border(
                    bottom=Side(border_style='medium', color='21316f'),
                )
                wrapped_alignment = Alignment(
                    vertical='top',
                    wrap_text=True
                )
                #
                # Define the column titles and widths
                columns = []
                #
                # Create a worksheet/tab with the title of the category
                worksheet = wb.create_sheet(
                    title='ReportResultEcxel',
                    index=0,
                )
                # Define the background color of the header cells
                fill = PatternFill(
                    start_color='21316f',
                    end_color='21316f',
                    fill_type='solid',
                )
                row_num = 1
                #
                # # Assign values, styles, and formatting for each cell in the header
                for col_num, (column_title) in enumerate(selected_fields, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title
                    cell.font = header_font
                    cell.border = border_bottom
                    cell.alignment = centered_alignment
                    cell.fill = fill
                    # set column width
                    column_letter = get_column_letter(col_num)
                    column_dimensions = worksheet.column_dimensions[column_letter]
                    column_dimensions.width = 20
                #
                obj = SavedRBReports.objects.get(pk=2060)

                report_name = obj.rb_report_name
                report_title = obj.rb_report_title
                report_description = obj.rb_report_description
                rb_report_query_id = obj.rb_report_query_id

                selected_fields = selected_query_fields = json.loads(obj.rb_selected_query_fields_array)
                unselected_query_fields = json.loads(obj.rb_unselected_query_fields_array)
                selected_group_fields = json.loads(obj.rb_selected_group_fields_array)
                selected_group_fields = list(filter(lambda x: x != '', selected_group_fields))
                # test = filter(None,selected_group_fields)
                unselected_group_fields = json.loads(obj.rb_unselected_group_fields_array)
                selected_order_fields = json.loads(obj.rb_selected_order_fields_array)
                selected_group_sorting = json.loads(obj.rb_selected_group_sorting)
                selected_group_sorting = list(filter(lambda x: x['field_name'] != '', selected_group_sorting))
                selected_sort_expressions = json.loads(obj.rb_selected_sort_expressions)
                selected_sort_expressions = list(filter(lambda x: x != '', selected_sort_expressions))
                report_headers = selected_fields_formating = json.loads(obj.rb_selected_fields_formating)
                selected_format_fields = json.loads(obj.rb_selected_format_fields_array)
                query_result = qbQueryProcess(request, rb_report_query_id, True)

                queryObj = SavedQBQuries.objects.get(pk=rb_report_query_id)
                qb_query_pair_id = queryObj.qb_query_pair_id

                if selected_group_fields:
                    group_field_formating = []
                    for ele in selected_group_fields:
                        keyValList = ele
                        selected_fields.remove(ele)
                        # Setting Group Fields At Start of Report Header
                        group_index = list(filter(lambda d: d['actual_column_name'] in keyValList, report_headers))[0]
                        report_headers.remove(group_index)
                        group_field_formating.append(group_index)
                    selected_fields[0:0] = selected_group_fields
                    report_headers[0:0] = group_field_formating

                selected_columns = []
                sorted_columns = []

                for field in selected_fields:
                    try:
                        actual_field = \
                        DataSetsPairFields.objects.values('df_actual_column_name').filter(df_pair_id=qb_query_pair_id).filter(
                            df_name=field)[0]['df_actual_column_name']
                    except:
                        return render_to_response('itrak/page-404.html')
                    selected_columns.append(actual_field)
                # Set the Sorting If Sorting Expression Exist

                for sort_expression in selected_sort_expressions:
                    sort_string = sort_expression
                    if sort_string.find("(Asc)") == -1:
                        column_name = sort_string.split(' (Desc)')[0]
                    else:
                        column_name = sort_string.split(' (Asc)')[0]
                    try:
                        actual_field = \
                        DataSetsPairFields.objects.values('df_actual_column_name').filter(df_pair_id=qb_query_pair_id).filter(
                            df_name=column_name)[0]['df_actual_column_name']
                    except:
                        return render_to_response('itrak/page-404.html')
                    if sort_string.find("(Asc)") == -1:
                        sort_value = '-' + actual_field
                    else:
                        sort_value = actual_field
                    sorted_columns.append(sort_value)

                if selected_sort_expressions:
                    finalquery = query_result.values(*tuple(selected_columns)).order_by(*sorted_columns)
                    print(finalquery)
                else:
                    finalquery = query_result.values(*tuple(selected_columns))
                    print(finalquery)

                # # Iterate through all movies of a category
                for record in finalquery:
                    row_num += 1
                    if int(row_num) % 2 == 0:
                        row_fill = PatternFill(
                            start_color='FFEFD5',
                            end_color='FFEFD5',
                            fill_type='solid',
                        )
                    else:
                        row_fill = PatternFill(
                            start_color='FFFFFF',
                            end_color='FFFFFF',
                            fill_type='solid',
                        )

                    row_font = Font(size=8, name='Segoe UI', bold=False, color='000000')
                #     # Assign values, styles, and formatting for each cell in the row
                #
                    recordList = []
                    for key, value in record.items():
                        print(key)
                        if key == 'created_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'client_created_at' and value is not None:
                            value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                    '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticket_created_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticket_is_reopen_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticket_closed_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticketNote__note_modified_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticketNote__note_created_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticketOrg__ticket_assign_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticketOrg__ticket_next_action_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'qb_created_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'qb_modified_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'rbQuerySetsQuery__rb_created_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'rbQuerySetsQuery__rb_modified_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticketManager__tmgr_completion_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticket_modified_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'submitted_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt), '%Y-%m-%d %H:%M:%S%z').strftime('%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticketAttach__attach_created_at' and value is not None:
                            if value != None:
                                value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                    '%m/%d/%Y %I:%M %p')
                                recordList.append(value1)
                        elif key == 'client_created_at' and value is not None:
                            if value != None:
                                value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                    '%m/%d/%Y %I:%M %p')
                                recordList.append(value1)
                        elif key == 'ticket_assign_at' and value is not None:
                            if value != None:
                                value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                    '%m/%d/%Y %I:%M %p')
                                recordList.append(value1)
                        elif key == 'ticket_next_action_at' and value is not None:
                            if value != None:
                                value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                    '%m/%d/%Y %I:%M %p')
                                recordList.append(value1)
                        elif key == 'submitted_at' and value is not None:
                            if value != None:
                                value1 = datetime.strptime(str(value), '%Y-%m-%d %H:%M:%S%z').strftime('%m/%d/%Y %I:%M %p')
                                recordList.append(value1)
                        elif key == 'ticket_created_at' and value is not None:
                            if value != None:
                                value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                    '%m/%d/%Y %I:%M %p')
                                recordList.append(value1)
                        else:
                            recordList.append(value)
                    for col_num, (cell_value) in enumerate(recordList, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.style = 'Normal'
                        # if cell_format == 'Currency':
                        #     cell.number_format = '#,##0.00 €'
                        # if col_num == 8:
                        #     cell.number_format = '[h]:mm;@'
                        cell.alignment = wrapped_alignment
                        cell.fill = row_fill
                        cell.font = row_font
                #
                # # freeze the first row
                worksheet.freeze_panes = worksheet['A2']
                #
                # # set tab color
                worksheet.sheet_properties.tabColor = 'FFFFFF'
                #
                # workbook.save(response)
                # return response
                dirname = os.path.dirname(__file__)
                filename = os.path.join(dirname, 'Attachments\Report.xlsx')
                # return HttpResponse(dirname)
                wb.save(filename)

                # Send Email Request with Attachment Start#
                smtp = getSMTPSettings(1)
                smtpserver = smtplib.SMTP(smtp['email_server'], smtp['port'])
                smtpserver.ehlo()
                smtpserver.starttls()
                smtpserver.ehlo
                smtpserver.login(smtp["user_name"], smtp["password"]) 

                emails_to = []
                if scheduledReport.notify_error_id:
                    to = User.objects.values_list('email',flat=True).get(pk=scheduledReport.notify_error_id)
                    emails_to.append(to)
                
                recipients = ScheduleReportResp.objects.values_list('sr_resp_recipt_user_id',flat=True).filter(sch_rep_id_id = 2)
                to_recipients = User.objects.values_list('email',flat=True).filter(pk__in=recipients)
                
                for to_recipient in to_recipients:
                    emails_to.append(to_recipient) 
                print(emails_to)
                # return HttpResponse(emails_to)
                
                for email_to in emails_to:
                    msg = EmailMessage('Scheduled Report', 'Report For Report Writer', smtp["email_sender_name"], [email_to])
                    msg.content_subtype = "html"  
                    dirname = os.path.dirname(__file__)
                    filename = os.path.join(dirname, 'Attachments\Report.xlsx')
                    msg.attach_file(filename)
                    msg.send()
                # Send Email Request with Attachment End#
            return HttpResponse('passed')
    else:
        return HttpResponse('failed')



@csrf_exempt
def get_xls_from_scheduled_saved_search(request, tickets, type, user_id):
    print('export')
    """
    Downloads all movies as Excel file with a worksheet for each movie category
    """
    tickets = tickets
    xlsType = type
    # category_queryset = Ticket.objects.all()
    ticket_queryset = Ticket.objects.filter(ticket_id__in=json.loads(str(tickets))).filter(ticket_is_delete=0).filter(ticket_is_active=1)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    dirname = os.path.dirname(__file__)
    response['Content-Disposition'] = 'attachment; filename={date}-TicketSearchList.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    # workbook = Workbook()
    wb = Workbook()
    ws1 = wb.active
    # ws1.title = "1st Hour"
    # wb.save('D:/atg-extra/FileName11.xlsx')

    # workbook = xlsxwriter.workbook('demo.xlsx')
    # workbook= workbook.create_sheet('TicketSearchList.xlsx', idx)
    
    # save_path = os.path.join(BASE_DIR, 'backup')
    # writer = ExcelWriter('D:/atg-extra/media/PythonExportt.xlsx'.format(save_path))
    # writer.save()
    # workbook= openpyxl.load_workbook('myfile.xlsx')
    # path='D:/atg-extra/workbook.xlsx'
    # workbook=openpyxl.load_workbook(path)

    # Define some styles and formatting that will be later used for cells
    header_font = Font(size=7, name='Segoe UI', bold=True, color='FFFFFF')
    centered_alignment = Alignment(horizontal='left')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='21316f'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    if xlsType == '0':
        # Define the column titles and widths
        columns = [
            ('Ticket #', 7),
            ('Status', 7),
            ('SubStatus', 15),
            ('Subject', 40),
            ('Ticket Type', 15),
            ('Subtype', 20),
            ('Priority', 7),
            ('Submitted Date', 15),
            ('Submitted By Name', 15),
            ('Entered By Name', 15),
            ('Assigned To Name', 15),
            ('Assigned Date', 15),
            ('Next Action Name', 15),
            ('Closed Date', 15),
            ('Last Activity', 15),
            ('Total Time Open', 15),
            ('Time Open (Days)', 15),
            ('Adj Time Open', 15),
            ('Adj Time Open (Days)', 15),
        ]
    else:
        # Define the column titles and widths
        columns = [
            ('Ticket #', 7),
            ('Status', 7),
            ('SubStatus', 15),
            ('Subject', 40),
            ('Full Description', 40),
            ('Ticket Type', 15),
            ('Subtype', 20),
            ('Subtype 2', 20),
            ('Subtype 3', 20),
            ('Subtype 4', 20),
            ('Priority', 7),
            ('Organization', 10),
            ('Client ID', 7),
            ('Client', 15),
            ('Record Locator', 10),
            ('Caller Name', 15),
            ('Caller Phone', 15),
            ('Caller Email', 20),
            ('Passenger Name', 15),
            ('Address', 40),
            ('Agent Error or Goodwill', 20),
            ('Agent Responsible (if applic)', 20),
            ('Airline Ticket #', 15),
            ('Amount of Payout', 15),
            ('Amount Saved', 10),
            ('Attention', 15),
            ('Check Number', 15),
            ('Company', 15),
            ('Correction / Containment Actions', 40),
            ('Corrective Action', 40),
            ('Is Traveler a VIP?', 10),
            ('Notes to print on check', 40),
            ('Pay to the order of', 30),
            ('Payout Required?', 15),
            ('Resp Vendor\'s City (if applic)', 40),
            ('Root Cause', 40),
            ('Vendor Name Resp (if applic)', 30),
            ('Who Approved', 20),
            ('Submitted Date', 15),
            ('Submitted By Name', 15),
            ('Submitted By Dept', 20),
            ('Submitted By Phone', 20),
            ('Assigned To', 15),
            ('Assigned To Name', 15),
            ('Assigned Date', 15),
            ('Assigned To Phone', 15),
            ('Next Action', 15),
            ('Next Action Name', 15),
            ('Next Action Date', 15),
            ('Next Action Phone', 15),
            ('Closed By', 15),
            ('Closed By Name', 15),
            ('Closed Date', 15),
            ('Closed By Phone', 15),
            ('ReOpened By', 15),
            ('ReOpened By Name', 15),
            ('ReOpened Date', 15),
            ('ReOpened By Phone', 15),
            ('Last Activity', 15),
            ('Total Time Open', 15),
            ('Time Open (Days)', 15),
            ('Adj Time Open', 15),
            ('Adj Time Open (Days)', 15),
        ]

    # Create a worksheet/tab with the title of the category
    worksheet = wb.create_sheet(
        title='TicketSearchEcxel',
        index=0,
        
    )
    # Define the background color of the header cells
    fill = PatternFill(
        start_color='21316f',
        end_color='21316f',
        fill_type='solid',
    )
    row_num = 1

    # Assign values, styles, and formatting for each cell in the header
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = border_bottom
        cell.alignment = centered_alignment
        cell.fill = fill
        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width

    # Iterate through all movies of a category
    for ticket in ticket_queryset:
        row_num += 1

        if ticket.ticket_status == 0:
            delta = datetime.now(timezone.utc) - ticket.ticket_created_at
        else:
            delta = ticket.ticket_closed_at - ticket.ticket_created_at

        days, seconds = delta.days, delta.seconds
        hours = days * 24 + seconds
        minutes = (seconds % 3600)
        seconds = seconds % 60

        if ticket.description:
            s = MLStripper()
            s.feed(ticket.description)
            plain_description = s.get_data()
        else:
            plain_description = ''

        if xlsType == '0':
            # Define data and formats for each cell in the row
            row = [
                (ticket.ticket_id, 'Normal'),
                ('Open' if ticket.ticket_status == 0 else 'Close', 'Normal'),
                (ticket.ticket_sub_status.sub_status_text if ticket.ticket_sub_status_id else '', 'Normal'),
                (ticket.subject, 'Normal'),
                # (timedelta(minutes=movie.ticket_created_at), 'Normal'),
                (ticket.ticket_type.ttype_name if ticket.ticket_type_id else '', 'Normal'),
                (ticket.ticket_subtype1.ttype_name if ticket.ticket_subtype1_id else '', 'Normal'),
                (ticket.priority.priority_name if ticket.priority_id else '', 'Normal'),
                (datetime.strptime(str(ticket.submitted_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S%z').strftime('%d/%m/%Y %I:%M %p') if ticket.submitted_at else '', 'Normal'),
                # (datetime.strptime(str(
                    # datetime.strptime(str(ticket.submitted_date) + ' ' + str(ticket.submitted_time) + '.000001+00:00','%Y-%m-%d %H:%M:%S.%f%z').astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %H:%M %p') if ticket.submitted_date else '', 'Normal'),
                # (datetime.strptime(str(ticket.submitted_date), '%Y-%m-%d').strftime('%d/%m/%Y') if ticket.submitted_date else '', 'Normal'),
                (ticket.ticket_caller.display_name if ticket.ticket_caller_id else '', 'Normal'),
                (ticket.ticket_created_by.display_name if ticket.ticket_created_by_id else '', 'Normal'),
                (ticket.ticket_assign_to.display_name if ticket.ticket_assign_to_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_assign_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_assign_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_next_action.display_name if ticket.ticket_next_action_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_closed_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_closed_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_closed_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_closed_at else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_modified_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_modified_at else '', 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),

            ]
        else:

            #Data Building of Condition Fields Start#
            if ticket.agent_error_goodwill:
                if ticket.agent_error_goodwill == '0':
                    agent_error_goodwill = 'Agent Error'
                else:
                    agent_error_goodwill = 'Goodwill'
            else:
                agent_error_goodwill = ''

            if ticket.is_traveler_vip:
                if ticket.is_traveler_vip == '0':
                    is_traveler_vip = 'Yes'
                else:
                    is_traveler_vip = 'No'
            else:
                is_traveler_vip = ''

            if ticket.is_payout_required:
                if ticket.is_payout_required == '0':
                    is_payout_required = 'Yes'
                else:
                    is_payout_required = 'No'
            else:
                is_payout_required = ''

            #Data Building of Condition Fields End#


            # Define data and formats for each cell in the row
            row = [
                (ticket.ticket_id, 'Normal'),
                ('Open' if ticket.ticket_status == 0 else 'Close', 'Normal'),
                (ticket.ticket_sub_status.sub_status_text if ticket.ticket_sub_status_id else '', 'Normal'),
                (ticket.subject, 'Normal'),
                (plain_description, 'Normal'),
                (ticket.ticket_type.ttype_name if ticket.ticket_type_id else '', 'Normal'),
                (ticket.ticket_subtype1.ttype_name if ticket.ticket_subtype1_id else '', 'Normal'),
                (ticket.ticket_subtype2.ttype_name if ticket.ticket_subtype2_id else '', 'Normal'),
                (ticket.ticket_subtype3.ttype_name if ticket.ticket_subtype3_id else '', 'Normal'),
                (ticket.ticket_subtype4.ttype_name if ticket.ticket_subtype4_id else '', 'Normal'),
                (ticket.priority.priority_name if ticket.priority_id else '', 'Normal'),
                (ticket.ticket_org.org_name if ticket.ticket_org_id else '', 'Normal'),
                (ticket.ticket_client.client_cus_id if ticket.ticket_client_id else '', 'Normal'),
                (ticket.ticket_client.client_name if ticket.ticket_client_id else '', 'Normal'),
                (ticket.ticket_record_locator if ticket.ticket_record_locator else '', 'Normal'),
                (ticket.ticket_caller_name if ticket.ticket_caller_name else '', 'Normal'),
                (ticket.ticket_caller_phone if ticket.ticket_caller_phone else '', 'Normal'),
                (ticket.ticket_caller_email if ticket.ticket_caller_email else '', 'Normal'),
                (ticket.ticket_passenger_name if ticket.ticket_passenger_name else '', 'Normal'),
                (ticket.ticket_address if ticket.ticket_address else '', 'Normal'),
                (agent_error_goodwill, 'Normal'),
                (ticket.agent_responsible if ticket.agent_responsible else '', 'Normal'),
                (ticket.airline_ticket_no if ticket.airline_ticket_no else '', 'Normal'),
                (ticket.ticket_payout_amount if ticket.ticket_payout_amount else '', 'Normal'),
                (ticket.amount_saved if ticket.amount_saved else '', 'Normal'),
                (ticket.ticket_attention if ticket.ticket_attention else '', 'Normal'),
                (ticket.check_number if ticket.check_number else '', 'Normal'),
                (ticket.ticket_company if ticket.ticket_company else '', 'Normal'),
                (ticket.corr_cont_actions if ticket.corr_cont_actions else '', 'Normal'),
                (ticket.corrective_action if ticket.corrective_action else '', 'Normal'),
                (is_traveler_vip, 'Normal'),
                (ticket.notes_on_check if ticket.notes_on_check else '', 'Normal'),
                (ticket.ticket_order_of_pay if ticket.ticket_order_of_pay else '', 'Normal'),
                (is_payout_required, 'Normal'),
                (ticket.vresponsible_city if ticket.vresponsible_city else '', 'Normal'),
                (ticket.ticket_root_cause if ticket.ticket_root_cause else '', 'Normal'),
                (ticket.vendor_responsible if ticket.vendor_responsible else '', 'Normal'),
                (ticket.check_approved_by if ticket.check_approved_by else '', 'Normal'),
                (datetime.strptime(str(ticket.submitted_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S%z').strftime('%d/%m/%Y %I:%M %p') if ticket.submitted_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.submitted_date), '%Y-%m-%d').strftime('%d/%m/%Y') if ticket.submitted_date else '', 'Normal'),
                (ticket.ticket_caller.display_name if ticket.ticket_caller_id else '', 'Normal'),
                (ticket.ticket_caller.user_dep.dep_name if ticket.ticket_caller_id and ticket.ticket_caller.user_dep_id else '', 'Normal'),
                (ticket.ticket_caller.phone_no if ticket.ticket_caller_id and ticket.ticket_caller_id else '', 'Normal'),
                (ticket.ticket_assign_to.display_name if ticket.ticket_assign_to_id else '', 'Normal'),
                (ticket.ticket_assign_to.first_name if ticket.ticket_assign_to_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_assign_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_assign_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_assign_to.phone_no if ticket.ticket_assign_to_id else '', 'Normal'),
                (ticket.ticket_next_action.display_name if ticket.ticket_next_action_id else '', 'Normal'),
                (ticket.ticket_next_action.first_name if ticket.ticket_next_action_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_next_action_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_next_action_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_next_action.phone_no if ticket.ticket_next_action_id else '', 'Normal'),
                (ticket.ticket_closed_by.display_name if ticket.ticket_closed_by_id else '', 'Normal'),
                (ticket.ticket_closed_by.first_name if ticket.ticket_closed_by_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_closed_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_closed_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_closed_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_closed_at else '', 'Normal'),
                (ticket.ticket_closed_by.phone_no if ticket.ticket_closed_by_id else '', 'Normal'),
                (ticket.ticket_is_reopen_by.display_name if ticket.ticket_is_reopen_by_id else '', 'Normal'),
                (ticket.ticket_is_reopen_by.first_name if ticket.ticket_is_reopen_by_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_is_reopen_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_is_reopen_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_is_reopen_by.phone_no if ticket.ticket_is_reopen_by_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_modified_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_modified_at else '', 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),

            ]

        if int(row_num) % 2 == 0:
            row_fill = PatternFill(
                start_color='FFEFD5',
                end_color='FFEFD5',
                fill_type='solid',
            )
        else:
            row_fill = PatternFill(
                start_color='FFFFFF',
                end_color='FFFFFF',
                fill_type='solid',
            )
        print(row_num)

        row_font = Font(size=8, name='Segoe UI', bold=False, color='000000')
        # Assign values, styles, and formatting for each cell in the row
        for col_num, (cell_value, cell_format) in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.style = cell_format
            # if cell_format == 'Currency':
            #     cell.number_format = '#,##0.00 €'
            # if col_num == 8:
            #     cell.number_format = '[h]:mm;@'
            cell.alignment = wrapped_alignment
            cell.fill = row_fill
            cell.font = row_font

    # freeze the first row
    worksheet.freeze_panes = worksheet['A2']

    # set tab color
    worksheet.sheet_properties.tabColor = 'FFFFFF'
    dirname = os.path.dirname(__file__)
    filename = os.path.join(dirname, 'Attachments\Report.xlsx')
    # return HttpResponse(dirname)
    wb.save(filename)

    # return response

#Export tickets to XLSX End#

#Export User to XLSX Start#
