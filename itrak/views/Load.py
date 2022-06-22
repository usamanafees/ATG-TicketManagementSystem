from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from itrak.models import *
import functools
from django.http import HttpResponseRedirect, HttpResponse, Http404, JsonResponse, HttpResponseBadRequest
from django.db.models.query import QuerySet
from itertools import chain
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
import pytz
import json
from django.db.models import Case, F, FloatField, IntegerField, Sum, When, Count
from django.db.models.functions import Cast
from django.core import serializers
from django.conf import settings
from django.core import signing
from django import template
from django.core import signing
from django.db.models import F
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
import json
from html.parser import HTMLParser
from django.core.mail import send_mail
import smtplib
import mimetypes
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.message import Message
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.text import MIMEText
from django.db import connection

register = template.Library()
# Get Menus Start#
def get_sections(request):
    if request.user.user_type == '0' and request.user.admin == True:
        sections = Sections.objects.filter(section_is_active=True)
    else:
        sections = Sections.objects.filter(section_is_active=True).exclude(section_id=2)
    return sections

# Get Menus End#


#Get Menus Start#
def get_menus(request):
    if request.user.user_org_id !=1:
        menus = Menus.objects.filter(menu_is_active=True).order_by('menu_order').exclude(menu_name='Organizations').exclude(menu_name='Client Information').exclude(menu_name='Solutions')
    else:
        menus = Menus.objects.filter(menu_is_active=True).order_by('menu_order').exclude(menu_name='Client Information').exclude(menu_name='Solutions')
    return menus

#Get Menus End#


#Get Sub Menus Start#
def get_sub_menus(request):
    sub_menus = SubMenus.objects.filter(submenu_is_active=True).order_by('submenu_order')
    return sub_menus

#Get Sub Menus End#


#Get Sub Menus Start#
def get_sidebar(request):
    sections = get_sections(request)
    menus = get_menus(request)
    sub_menus = get_sub_menus(request)
    menus_allowed = list(get_sidebar_menus_access(request, request.user.id))
    submenus_allowed = list(get_sidebar_submenus_access(request, request.user.id))
    result = sections
    result.menus = menus
    result.sub_menus = sub_menus
    result.menus_allowed = menus_allowed
    result.submenus_allowed = submenus_allowed
    print(result)
    return result

#Get Sub Menus End#


#Get User Menu Permissions Start#
def get_user_menus_permit(request, userId = id):
    menu_ids = UserMenuPermissions.objects.filter(menu_id__isnull= False).filter(user_id=userId).values_list('menu_id', flat=True)
    return menu_ids

#Get User Menu Permissions End#



#Get User Sub Menu Permissions Start#
def get_user_submenus_permit(request, userId = id):
    submenu_ids = UserMenuPermissions.objects.filter(submenu_id__isnull= False).filter(user_id=userId).values_list('submenu_id', flat=True)
    return submenu_ids

#Get User Sub Menu Permissions End#



#Get User Menu Permissions Start#
def get_sidebar_menus_access(request, userId = id):
    data = User.objects.get(pk=userId)
    u_menu_ids = UserMenuPermissions.objects.filter(menu_id__isnull= False).filter(user_id=userId).values_list('menu_id', flat=True)
    allowed_groups = list(data.userMembership.filter(is_delete=0).values_list('m_group_id', flat=True))
    g_menu_ids = GroupMenuPermissions.objects.filter(menu_id__isnull= False).filter(group_id__in=allowed_groups).values_list('menu_id', flat=True)
    menu_ids = chain(u_menu_ids, g_menu_ids)
    return menu_ids

#Get User Menu Permissions End#



#Get User Sub Menu Permissions Start#
def get_sidebar_submenus_access(request, userId = id):
    data = User.objects.get(pk=userId)
    u_submenu_ids = UserMenuPermissions.objects.filter(submenu_id__isnull= False).filter(user_id=userId).values_list('submenu_id', flat=True)
    allowed_groups = list(data.userMembership.filter(is_delete=0).values_list('m_group_id', flat=True))
    g_submenu_ids = GroupMenuPermissions.objects.filter(submenu_id__isnull=False).filter(group_id__in=allowed_groups).values_list('submenu_id', flat=True)
    submenu_ids = chain(u_submenu_ids, g_submenu_ids)
    return submenu_ids

#Get User Sub Menu Permissions End#



#Get Group Menu Permissions Start#
def get_group_menus_permit(request, groupId = id):
    menu_ids = GroupMenuPermissions.objects.filter(menu_id__isnull= False).filter(group_id=groupId).values_list('menu_id', flat=True)
    return menu_ids

#Get Group Menu Permissions End#



#Get Group Sub Menu Permissions Start#
def get_group_submenus_permit(request, groupId = id):
    submenu_ids = GroupMenuPermissions.objects.filter(submenu_id__isnull= False).filter(group_id=groupId).values_list('submenu_id', flat=True)
    return submenu_ids

#Get Group Sub Menu Permissions End#



#Get Load Sidebar Start#

def append_sidebar(request):
    sections = get_sections(request)
    menus = get_menus(request)
    submenus = get_sub_menus(request)
    result = ''
    if sections:
        for section in sections:
            result = ''.join([result, '<hr class ="separator" />'])
            result = ''.join([result, '<div class ="sidebar-widget widget-tasks" ><div class ="widget-header" ><h6 style = "color: white" >' ])
            result = ''.join([result, section.section_name ])
            result = ''.join([result, '</h6><div class ="widget-toggle" > + </div></div> <div class ="widget-content" >' ])
            result = ''.join([result, '<nav id = "menu" class ="nav-main" role="navigation" ><ul class ="nav nav-main" >' ])

            if menus:
                for menu in menus:
                    if menu.m_section == section.section_id:
                        result = ''.join([result, '<li class ="nav-parent" ><a><i class ="'])
                        result = ''.join([result, menu.icon])
                        result = ''.join([result, '" aria-hidden="true" > </i><span>'])
                        result = ''.join([result, menu.menu_name])
                        result = ''.join([result, '</span ></a><ul class ="nav nav-children" >'])
                        if submenus:
                            for submenu in submenus:
                                if submenu.submenu_menu_id == menu.menu_id:
                                    result = ''.join([result, '<li><a href = "&#123 % url &#39' ])
                                    result = ''.join([result, submenu.submenu_link])
                                    result = ''.join([result, '&#39 % &#125" >'])
                                    result = ''.join([result, submenu.submenu_name])
                                    result = ''.join([result, '</a></li>'])
                        result = ''.join([result, '</ul></li>'])
            result = ''.join([result, '</ul></nav></div></div>'])

    return result

#Get Load SideBar End#


#Get Group List against User Type For Group Membership Start#
def get_group_list(request, userType = type):
    # id = request.GET.get('UserID')
    # print(id)
    # user_id = signing.loads(id,salt=settings.SALT_KEY)
    # data = User.objects.get(pk=user_id)
    # print(data.user_org_id)
    # print(user_id)
    user_id = request.user.id
    org_id = request.user.user_org_id
    # if user_id != 3108:
    if int(userType) == 0:          #0 indicate for User Type Agent #
        group_lists = Group.objects.filter(gp_is_delete=0).filter(gp_is_active= 1).filter(group_org=org_id).values_list('group_id', 'group_display_name')
    else: #Case: User Type is Normal User#
        group_lists = Group.objects.filter(gp_is_delete=0).filter(gp_is_active=1).filter(membership_type=1).filter(group_org=org_id).values_list('group_id', 'group_display_name')
    # else:
    #     if int(userType) == 0:          #0 indicate for User Type Agent #
    #         group_lists = Group.objects.filter(group_org=data.user_org_id).filter(gp_is_delete=0).filter(gp_is_active= 1).values_list('group_id', 'group_display_name')
    #     else: #Case: User Type is Normal User#
    #         group_lists = Group.objects.filter(gp_is_delete=0).filter(gp_is_active=1).filter(membership_type=1).values_list('group_id', 'group_display_name')

    return group_lists

#Get Group List against User Type For Group Membership End#


@csrf_exempt
#Get Client Information Against ID Start#
def getClientInformationById(request):
    if request.is_ajax() and request.method == 'POST':
        clientinfo = request.POST.get('clientinfo')
        # probably you want to add a regex check if the username value is valid here
        if clientinfo:
            result = ClientInformation.objects.filter(clientinfo_id= clientinfo).filter(clientinfo_is_delete=0)
            response_data = {}
            try:
                response_data['response'] = serializers.serialize('json', result)
            except:
                response_data['response'] = 'No Record Found'
            return JsonResponse(response_data)


#Get Client Information Against ID End#



@csrf_exempt
#Get Ticket Type Child Against ID Start#
def getTicketTypeChildById(request):
    if request.is_ajax() and request.method == 'POST':
        ttype_id = request.POST.get('ttype_id')
        # probably you want to add a regex check if the username value is valid here
        if ttype_id:
            if request.user.id != 3108:
                org_id = request.user.user_org_id
                result = TicketType.objects.filter(parent_id= ttype_id).filter(ttype_is_delete=0,ttype_is_active=1).filter(user_org_id=org_id)
            else:
                result = TicketType.objects.filter(parent_id= ttype_id).filter(ttype_is_delete=0,ttype_is_active=1)
            response_data = {}
            try:
                response_data['response'] = serializers.serialize('json', result)
            except:
                response_data['response'] = 'No Record Found'
            return JsonResponse(response_data)


#Get Ticket Type Child Against ID End#


#Validate Username for Uniqueness Start#

@csrf_exempt
def getClientId(request):
    if request.is_ajax() and request.method == 'POST':
        caller_id = request.POST.get('caller_id')
        # probably you want to add a regex check if the username value is valid here
        if caller_id:
            result = User.objects.get(pk=caller_id)
            response_data = { 'response': result.user_client_id}
            return JsonResponse(response_data)
    else:
        return HttpResponse('fail')


#Validate Username for Uniqueness End#

@csrf_exempt
#Get User Against Org ID Start#

def getUsersByOrgId(request):
    if request.is_ajax() and request.method == 'POST':
        org_id = request.POST.get('org_id')
        # probably you want to add a regex check if the username value is valid here
        if org_id:
            # Action Permission check
            if check_action_permission("Ag_perm_Can_be_assigned_tickets",request.user.id):
                result = User.objects.filter(user_org_id= org_id).filter(is_delete=0).filter(user_type=0)
            else:
                result = User.objects.filter(user_org_id= org_id).filter(is_delete=0).filter(user_type=0).exclude(id=request.user.id)
            response_data = {}
            try:
                response_data['response'] = serializers.serialize('json', result)
            except:
                response_data['response'] = 'No Record Found'
            return JsonResponse(response_data)


#Get User Against Org ID End#



#Get Ticket Events Role Log Start#
def get_ticket_event_logs(request, ticketId = id, event = ''):
    user_ids = TicketUserRoleLog.objects.filter(urlog_ticket_id=ticketId).filter(urlog_event=event).values_list('urlog_user_id', 'urlog_created_at', 'urlog_user__display_name').order_by('-urlog_created_at')
    return user_ids

#Get Ticket Events Role Log End#



#Get Ticket Attachments Start#
def get_ticket_attachments(request, ticketId = id):
    attach_ids = TicketAttachments.objects.filter(attach_ticket_id=ticketId).filter(attach_is_delete=0).values_list('attach_id','file_path', 'file_name', 'file_size')
    return attach_ids

#Get Ticket Attachments End#




#Get Ticket List of Task Manager Start#
def get_task_mgr_ticket_list(userId):
    print("im here")
    print(type(userId))
    ticket_ids = TaskManager.objects.filter(tmgr_ticket_id__isnull=False).filter(task_assigned_to_id=userId).filter(tmgr_is_complete=0).values_list('tmgr_ticket_id', flat=True)
    return ticket_ids

#Get Ticket List of Task Manager End#

#Get Ticket List of Task Manager Start#
def get_task_mgr_ticket_list_by_userID_list(accountIdList):
    print("im here")
    print(type(accountIdList))
    ticket_ids = TaskManager.objects.filter(tmgr_ticket_id__isnull=False).filter(account_id__in=userIdList).filter(tmgr_is_complete=0).values_list('tmgr_ticket_id', flat=True)
    return ticket_ids

#Get Ticket List of Task Manager End#



#Get Ticket Excel Export Start#


#Export ticket to XLSX Start#

class MLStripper(HTMLParser):
    def __init__(self):
        self.reset()
        self.strict = False
        self.convert_charrefs= True
        self.fed = []
    def handle_data(self, d):
        self.fed.append(d)
    def get_data(self):
        return ''.join(self.fed)

@csrf_exempt
def get_xls_from_tickets(request, tickets, type):
    """
    Downloads all movies as Excel file with a worksheet for each movie category
    """
    tickets = tickets
    xlsType = type
    # category_queryset = Ticket.objects.all()
    ticket_queryset = Ticket.objects.filter(ticket_id__in=json.loads(str(tickets))).filter(ticket_is_delete=0).filter(ticket_is_active=1)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-TicketSearchList.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Delete the default worksheet
    workbook.remove(workbook.active)

    # Define some styles and formatting that will be later used for cells
    header_font = Font(size=7, name='Segoe UI', bold=True, color='FFFFFF')
    centered_alignment = Alignment(horizontal='left')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='21316f'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    if xlsType == '0':
        # Define the column titles and widths
        columns = [
            ('Ticket #', 7),
            ('Status', 7),
            ('SubStatus', 15),
            ('Subject', 40),
            ('Ticket Type', 15),
            ('Subtype', 20),
            ('Priority', 7),
            ('Submitted Date', 15),
            ('Submitted By Name', 15),
            ('Entered By Name', 15),
            ('Assigned To Name', 15),
            ('Assigned Date', 15),
            ('Next Action Name', 15),
            ('Closed Date', 15),
            ('Last Activity', 15),
            ('Total Time Open', 15),
            ('Time Open (Days)', 15),
            ('Adj Time Open', 15),
            ('Adj Time Open (Days)', 15),
        ]
    else:
        # Define the column titles and widths
        columns = [
            ('Ticket #', 7),
            ('Status', 7),
            ('SubStatus', 15),
            ('Subject', 40),
            ('Full Description', 40),
            ('Ticket Type', 15),
            ('Subtype', 20),
            ('Subtype 2', 20),
            ('Subtype 3', 20),
            ('Subtype 4', 20),
            ('Priority', 7),
            ('Organization', 10),
            ('Account ID', 7),
            ('Account', 15),
            ('Record Locator', 10),
            ('Caller Name', 15),
            ('Caller Phone', 15),
            ('Caller Email', 20),
            ('Passenger Name', 15),
            ('Address', 40),
            ('Agent Error or Goodwill', 20),
            ('Agent Responsible (if applic)', 20),
            ('Airline Ticket #', 15),
            ('Amount of Payout', 15),
            ('Amount Saved', 10),
            ('Attention', 15),
            ('Check Number', 15),
            ('Company', 15),
            ('Correction / Containment Actions', 40),
            ('Corrective Action', 40),
            ('Is Traveler a VIP?', 10),
            ('Notes to print on check', 40),
            ('Pay to the order of', 30),
            ('Payout Required?', 15),
            ('Resp Vendor\'s City (if applic)', 40),
            ('Root Cause', 40),
            ('Vendor Name Resp (if applic)', 30),
            ('Who Approved', 20),
            ('Submitted Date', 15),
            ('Submitted By Name', 15),
            ('Submitted By Dept', 20),
            ('Submitted By Phone', 20),
            ('Assigned To', 15),
            ('Assigned To Name', 15),
            ('Assigned Date', 15),
            ('Assigned To Phone', 15),
            ('Next Action', 15),
            ('Next Action Name', 15),
            ('Next Action Date', 15),
            ('Next Action Phone', 15),
            ('Closed By', 15),
            ('Closed By Name', 15),
            ('Closed Date', 15),
            ('Closed By Phone', 15),
            ('ReOpened By', 15),
            ('ReOpened By Name', 15),
            ('ReOpened Date', 15),
            ('ReOpened By Phone', 15),
            ('Last Activity', 15),
            ('Total Time Open', 15),
            ('Time Open (Days)', 15),
            ('Adj Time Open', 15),
            ('Adj Time Open (Days)', 15),
        ]

    # Create a worksheet/tab with the title of the category
    worksheet = workbook.create_sheet(
        title='TicketSearchEcxel',
        index=0,
    )
    # Define the background color of the header cells
    fill = PatternFill(
        start_color='21316f',
        end_color='21316f',
        fill_type='solid',
    )
    row_num = 1

    # Assign values, styles, and formatting for each cell in the header
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = border_bottom
        cell.alignment = centered_alignment
        cell.fill = fill
        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width

    # Iterate through all movies of a category
    for ticket in ticket_queryset:
        row_num += 1

        if ticket.ticket_status == 0:
            delta = datetime.now(timezone.utc) - ticket.ticket_created_at
        else:
            delta = ticket.ticket_closed_at - ticket.ticket_created_at

        days, seconds = delta.days, delta.seconds
        hours = days * 24 + seconds
        minutes = (seconds % 3600)
        seconds = seconds % 60

        if ticket.description:
            s = MLStripper()
            s.feed(ticket.description)
            plain_description = s.get_data()
        else:
            plain_description = ''

        if xlsType == '0':
            # Define data and formats for each cell in the row
            row = [
                (ticket.ticket_id, 'Normal'),
                ('Open' if ticket.ticket_status == 0 else 'Close', 'Normal'),
                (ticket.ticket_sub_status.sub_status_text if ticket.ticket_sub_status_id else '', 'Normal'),
                (ticket.subject, 'Normal'),
                # (timedelta(minutes=movie.ticket_created_at), 'Normal'),
                (ticket.ticket_type.ttype_name if ticket.ticket_type_id else '', 'Normal'),
                (ticket.ticket_subtype1.ttype_name if ticket.ticket_subtype1_id else '', 'Normal'),
                (ticket.priority.priority_name if ticket.priority_id else '', 'Normal'),
                (datetime.strptime(str(ticket.submitted_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S%z').strftime('%d/%m/%Y %I:%M %p') if ticket.submitted_at else '', 'Normal'),
                # (datetime.strptime(str(
                    # datetime.strptime(str(ticket.submitted_date) + ' ' + str(ticket.submitted_time) + '.000001+00:00','%Y-%m-%d %H:%M:%S.%f%z').astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %H:%M %p') if ticket.submitted_date else '', 'Normal'),
                # (datetime.strptime(str(ticket.submitted_date), '%Y-%m-%d').strftime('%d/%m/%Y') if ticket.submitted_date else '', 'Normal'),
                (ticket.ticket_caller.display_name if ticket.ticket_caller_id else '', 'Normal'),
                (ticket.ticket_created_by.display_name if ticket.ticket_created_by_id else '', 'Normal'),
                (ticket.ticket_assign_to.display_name if ticket.ticket_assign_to_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_assign_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_assign_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_next_action.display_name if ticket.ticket_next_action_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_closed_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_closed_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_closed_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_closed_at else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_modified_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_modified_at else '', 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),

            ]
        else:

            #Data Building of Condition Fields Start#
            if ticket.agent_error_goodwill:
                if ticket.agent_error_goodwill == '0':
                    agent_error_goodwill = 'Agent Error'
                else:
                    agent_error_goodwill = 'Goodwill'
            else:
                agent_error_goodwill = ''

            if ticket.is_traveler_vip:
                if ticket.is_traveler_vip == '0':
                    is_traveler_vip = 'Yes'
                else:
                    is_traveler_vip = 'No'
            else:
                is_traveler_vip = ''

            if ticket.is_payout_required:
                if ticket.is_payout_required == '0':
                    is_payout_required = 'Yes'
                else:
                    is_payout_required = 'No'
            else:
                is_payout_required = ''

            #Data Building of Condition Fields End#


            # Define data and formats for each cell in the row
            row = [
                (ticket.ticket_id, 'Normal'),
                ('Open' if ticket.ticket_status == 0 else 'Close', 'Normal'),
                (ticket.ticket_sub_status.sub_status_text if ticket.ticket_sub_status_id else '', 'Normal'),
                (ticket.subject, 'Normal'),
                (plain_description, 'Normal'),
                (ticket.ticket_type.ttype_name if ticket.ticket_type_id else '', 'Normal'),
                (ticket.ticket_subtype1.ttype_name if ticket.ticket_subtype1_id else '', 'Normal'),
                (ticket.ticket_subtype2.ttype_name if ticket.ticket_subtype2_id else '', 'Normal'),
                (ticket.ticket_subtype3.ttype_name if ticket.ticket_subtype3_id else '', 'Normal'),
                (ticket.ticket_subtype4.ttype_name if ticket.ticket_subtype4_id else '', 'Normal'),
                (ticket.priority.priority_name if ticket.priority_id else '', 'Normal'),
                (ticket.ticket_org.org_name if ticket.ticket_org_id else '', 'Normal'),
                (ticket.account.id if ticket.account_id else '', 'Normal'),
                (ticket.account.acc_name if ticket.account_id else '', 'Normal'),
                (ticket.ticket_record_locator if ticket.ticket_record_locator else '', 'Normal'),
                (ticket.ticket_caller_name if ticket.ticket_caller_name else '', 'Normal'),
                (ticket.ticket_caller_phone if ticket.ticket_caller_phone else '', 'Normal'),
                (ticket.ticket_caller_email if ticket.ticket_caller_email else '', 'Normal'),
                (ticket.ticket_passenger_name if ticket.ticket_passenger_name else '', 'Normal'),
                (ticket.ticket_address if ticket.ticket_address else '', 'Normal'),
                (agent_error_goodwill, 'Normal'),
                (ticket.agent_responsible if ticket.agent_responsible else '', 'Normal'),
                (ticket.airline_ticket_no if ticket.airline_ticket_no else '', 'Normal'),
                (ticket.ticket_payout_amount if ticket.ticket_payout_amount else '', 'Normal'),
                (ticket.amount_saved if ticket.amount_saved else '', 'Normal'),
                (ticket.ticket_attention if ticket.ticket_attention else '', 'Normal'),
                (ticket.check_number if ticket.check_number else '', 'Normal'),
                (ticket.ticket_company if ticket.ticket_company else '', 'Normal'),
                (ticket.corr_cont_actions if ticket.corr_cont_actions else '', 'Normal'),
                (ticket.corrective_action if ticket.corrective_action else '', 'Normal'),
                (is_traveler_vip, 'Normal'),
                (ticket.notes_on_check if ticket.notes_on_check else '', 'Normal'),
                (ticket.ticket_order_of_pay if ticket.ticket_order_of_pay else '', 'Normal'),
                (is_payout_required, 'Normal'),
                (ticket.vresponsible_city if ticket.vresponsible_city else '', 'Normal'),
                (ticket.ticket_root_cause if ticket.ticket_root_cause else '', 'Normal'),
                (ticket.vendor_responsible if ticket.vendor_responsible else '', 'Normal'),
                (ticket.check_approved_by if ticket.check_approved_by else '', 'Normal'),
                (datetime.strptime(str(ticket.submitted_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S%z').strftime('%d/%m/%Y %I:%M %p') if ticket.submitted_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.submitted_date), '%Y-%m-%d').strftime('%d/%m/%Y') if ticket.submitted_date else '', 'Normal'),
                (ticket.ticket_caller.display_name if ticket.ticket_caller_id else '', 'Normal'),
                (ticket.ticket_caller.user_dep.dep_name if ticket.ticket_caller_id and ticket.ticket_caller.user_dep_id else '', 'Normal'),
                (ticket.ticket_caller.phone_no if ticket.ticket_caller_id and ticket.ticket_caller_id else '', 'Normal'),
                (ticket.ticket_assign_to.display_name if ticket.ticket_assign_to_id else '', 'Normal'),
                (ticket.ticket_assign_to.first_name if ticket.ticket_assign_to_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_assign_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_assign_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_assign_to.phone_no if ticket.ticket_assign_to_id else '', 'Normal'),
                (ticket.ticket_next_action.display_name if ticket.ticket_next_action_id else '', 'Normal'),
                (ticket.ticket_next_action.first_name if ticket.ticket_next_action_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_next_action_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_next_action_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_next_action.phone_no if ticket.ticket_next_action_id else '', 'Normal'),
                (ticket.ticket_closed_by.display_name if ticket.ticket_closed_by_id else '', 'Normal'),
                (ticket.ticket_closed_by.first_name if ticket.ticket_closed_by_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_closed_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_closed_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_closed_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_closed_at else '', 'Normal'),
                (ticket.ticket_closed_by.phone_no if ticket.ticket_closed_by_id else '', 'Normal'),
                (ticket.ticket_is_reopen_by.display_name if ticket.ticket_is_reopen_by_id else '', 'Normal'),
                (ticket.ticket_is_reopen_by.first_name if ticket.ticket_is_reopen_by_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_is_reopen_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_is_reopen_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_is_reopen_by.phone_no if ticket.ticket_is_reopen_by_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_modified_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=request.user.id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_modified_at else '', 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),

            ]

        if int(row_num) % 2 == 0:
            row_fill = PatternFill(
                start_color='FFEFD5',
                end_color='FFEFD5',
                fill_type='solid',
            )
        else:
            row_fill = PatternFill(
                start_color='FFFFFF',
                end_color='FFFFFF',
                fill_type='solid',
            )
        print(row_num)

        row_font = Font(size=8, name='Segoe UI', bold=False, color='000000')
        # Assign values, styles, and formatting for each cell in the row
        for col_num, (cell_value, cell_format) in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.style = cell_format
            # if cell_format == 'Currency':
            #     cell.number_format = '#,##0.00 €'
            # if col_num == 8:
            #     cell.number_format = '[h]:mm;@'
            cell.alignment = wrapped_alignment
            cell.fill = row_fill
            cell.font = row_font

    # freeze the first row
    worksheet.freeze_panes = worksheet['A2']

    # set tab color
    worksheet.sheet_properties.tabColor = 'FFFFFF'

    workbook.save(response)

    return response

#Export tickets to XLSX End#

#Export User to XLSX Start#

def get_xls_from_user(request,users):
    """
       Downloads all movies as Excel file with a worksheet for each movie category
       """
    users = users
    print(users)

    xlsType = type
    # category_queryset = Ticket.objects.all()
    # user_queryset = User.objects.filter(id__in=json.loads(str(users))).filter(is_delete=0).filter(
    #     is_active=1)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-UserSearchList.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Delete the default worksheet
    workbook.remove(workbook.active)

    # Define some styles and formatting that will be later used for cells
    header_font = Font(size=7, name='Segoe UI', bold=True, color='FFFFFF')
    centered_alignment = Alignment(horizontal='left')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='21316f'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Define the column titles and widths
    columns = [
        ('User ID', 15),
        ('Type', 15),
        ('Last', 15),
        ('First', 15),
        ('Department', 15),
        ('Email', 15),
        ('Phone Number', 15),
        ('Status', 15),
    ]

    worksheet = workbook.create_sheet(
        title='UserSearchEcxel',
        index=0,
    )
    # Define the background color of the header cells
    fill = PatternFill(
        start_color='21316f',
        end_color='21316f',
        fill_type='solid',
    )
    row_num = 1

    # Assign values, styles, and formatting for each cell in the header
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = border_bottom
        cell.alignment = centered_alignment
        cell.fill = fill
        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width

    # Iterate through all movies of a category
    for user in users:
        row_num += 1

        # if user.ticket_status == 0:
        #     delta = datetime.now(timezone.utc) - ticket.ticket_created_at
        # else:
        #     delta = ticket.ticket_closed_at - ticket.ticket_created_at
        #
        # days, seconds = delta.days, delta.seconds
        # hours = days * 24 + seconds
        # minutes = (seconds % 3600)
        # seconds = seconds % 60
        #
        # if ticket.description:
        #     s = MLStripper()
        #     s.feed(ticket.description)
        #     plain_description = s.get_data()
        # else:
        #     plain_description = ''


        # Define data and formats for each cell in the row
        if user.user_type == '0':
            u_type = 'Agent'
        else:
            u_type = 'End User'
        row = [
            (user.username, 'Normal'),
            # ('Agent' if user.user_type == 0 else 'EndUser', 'Normal'),
            (u_type, 'Normal'),
            (user.last_name, 'Normal'),
            (user.first_name, 'Normal'),
            (user.user_dep.dep_name if user.user_dep_id else '', 'Normal'),
            (user.email, 'Normal'),
            (user.phone_no, 'Normal'),
            (user.is_active, 'Normal'),

        ]


        if int(row_num) % 2 == 0:
            row_fill = PatternFill(
                start_color='FFEFD5',
                end_color='FFEFD5',
                fill_type='solid',
            )
        else:
            row_fill = PatternFill(
                start_color='FFFFFF',
                end_color='FFFFFF',
                fill_type='solid',
            )

        row_font = Font(size=8, name='Segoe UI', bold=False, color='000000')
        # Assign values, styles, and formatting for each cell in the row
        for col_num, (cell_value, cell_format) in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.style = cell_format
            # if cell_format == 'Currency':
            #     cell.number_format = '#,##0.00 €'
            # if col_num == 8:
            #     cell.number_format = '[h]:mm;@'
            cell.alignment = wrapped_alignment
            cell.fill = row_fill
            cell.font = row_font

    # freeze the first row
    worksheet.freeze_panes = worksheet['A2']

    # set tab color
    worksheet.sheet_properties.tabColor = 'FFFFFF'

    workbook.save(response)

    return response


#Export User to XLSX End#

#Get Org Detail on Modal Through ID Start#


def getModalOrgDetailById(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        orgID = request.POST.get('org_id')
        organization = Organization.objects.get(pk=orgID) # So we send the company instance
        context['organization'] = organization
    return render(request,'itrak/Organization/org_modal_detail.html',context)

#Get Org Detail on Modal Through ID End#



#Get Org History on Modal Through ID Start#


def getModalOrgHistoryById(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        orgID = request.POST.get('org_id')
        orgN = request.POST.get('orgname')
        orgname = orgN
        print(orgname)
        tickets = Ticket.objects.filter(ticket_org_id=orgID).filter(ticket_is_delete=0) # So we send the company instance
        context['tickets'] = tickets
        context['orgname'] = orgname
    return render(request,'itrak/Organization/org_modal_history.html',context)

#Get Org History on Modal Through ID End#



#Get Caller Detail on Modal Through ID Start#


def getModalCallerDetailById(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        callerID = request.POST.get('caller_id')
        caller = User.objects.get(pk=callerID) # So we send the company instance
        context['caller'] = caller
    return render(request,'itrak/User/caller_modal_detail.html',context)

#Get Caller Detail on Modal Through ID End#



#Get Caller History on Modal Through ID Start#


def getModalCallerHistoryById(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        callerID = request.POST.get('caller_id')
        callername = request.POST.get('callername')
        callerN = callername
        tickets = Ticket.objects.filter(ticket_caller_id=callerID).filter(ticket_is_delete=0) # So we send the company instance
        context['tickets'] = tickets
        context['callerN'] = callerN
    return render(request,'itrak/User/caller_modal_history.html',context)

#Get Caller History on Modal Through ID End#



#Get Client Detail on Modal Through ID Start#


def getModalAccountDetailById(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        accountID = request.POST.get('account_id')
        sql = "SELECT id, AIAN_DK, acc_number, acc_name, agency, login_id, agency_email, country, client, company, currency FROM GlobalACCTS WHERE id = "+accountID
        cursor = connection.cursor()
        cursor.execute(sql)
        accountsTuples = cursor.fetchall()

        accountInformation = {}
        for index, tuple in enumerate(accountsTuples):
            accountInformation['id'] = tuple[0]
            accountInformation['AIAN_DK'] = tuple[1]
            accountInformation['acc_number'] = tuple[2]
            accountInformation['acc_name'] = tuple[3]
            accountInformation['agency'] = tuple[4]
            accountInformation['login_id'] = tuple[5]
            accountInformation['agency_email'] = tuple[6]
            accountInformation['country'] = tuple[7]
            accountInformation['client'] = tuple[8]
            accountInformation['company'] = tuple[9]
            accountInformation['currency'] = tuple[10]

        context['accountInformation'] = accountInformation
    return render(request,'itrak/Client/client_modal_detail.html',context)

#Get Client Detail on Modal Through ID End#



#Get Account History on Modal Through ID Start#


def getModalAccountHistoryById(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        accountID = request.POST.get('account_id')
        accountname = request.POST.get('accountname')
        clientN = accountname
        tickets = Ticket.objects.filter(account_id=accountID).filter(ticket_is_delete=0) # So we send the company instance
        context['tickets'] = tickets
        context['clientN'] = clientN
    return render(request,'itrak/Client/client_modal_history.html',context)

#Get Account History on Modal Through ID End#



#Get Client History on Modal Through ID Start#

def getModalGroupTasksById(request):
    if request.is_ajax() and request.method == 'POST':
        group_id = request.POST.get('group_id')
        # probably you want to add a regex check if the username value is valid here
        if group_id:
            result = TaskGroupManager.objects.filter(tmgrgp_group_id= group_id).order_by('tmgrgp_display_order')
            response_data = {}
            try:
                response_data['response'] = serializers.serialize('json', result)
            except:
                response_data['response'] = 'No Record Found'
            return JsonResponse(response_data)

#Get Client History on Modal Through ID End#


#Get Panel FOr DashBoard Setting Starts#

@csrf_exempt
def getPanelsForDashboardSettings(request):
    # graphPanels = PanelGraph.objects.values_list('panelGraph_text', 'panelGraph_url')
    global_user = isGlobalUser(request)
    if request.user.user_type_slug != global_user:
        graphPanels = PanelGraph.objects.exclude(pk__in=
        DashboardSettings.objects.filter(d_user_id=request.user.id).values_list('d_panel_id', flat=True)).values_list('panelGraph_text', 'panelGraph_url').exclude(panelGraph_url='openTicketsByOrganization')
    else:
        graphPanels = PanelGraph.objects.exclude(pk__in=
        DashboardSettings.objects.filter(d_user_id=request.user.id).values_list('d_panel_id', flat=True)).values_list('panelGraph_text', 'panelGraph_url')
    return graphPanels

#Get Panel FOr DashBoard Setting End#


#Get Client History on Modal Through ID Start#

def getTaskTypeById(request):
    if request.is_ajax() and request.method == 'POST':
        task_id = request.POST.get('task_id')
        # probably you want to add a regex check if the username value is valid here
        if task_id:
            task_type = Task.objects.only('task_type').get(task_id=task_id).task_type
            response_data = {}
            try:
                response_data['task_type'] = task_type
            except:
                response_data['task_type'] = 'No Record Found'
            return JsonResponse(response_data)

#Get Client History on Modal Through ID End#


#Update Ticket Substatus through Ticket ID Start#

def updateTicketSubstatusById(request):
    if request.is_ajax() and request.method == 'POST':
        ticket_id = request.POST.get('ticket_id')
        substatus = request.POST.get('substatus')
        # probably you want to add a regex check if the username value is valid here
        if ticket_id and substatus:
            t_obj = Ticket.objects.get(pk=ticket_id)
            t_obj.ticket_sub_status_id = substatus
            t_obj.save()
            response_data = {}
            try:
                response_data['response'] = 'success'
            except:
                response_data['response'] = 'fail'
            return JsonResponse(response_data)

#Update Ticket Substatus through Ticket ID End#


#Update Ticket Status through Ticket ID Start#

def updateTicketStatusById(request):
    if request.is_ajax() and request.method == 'POST':
        ticket_id = request.POST.get('ticket_id')
        # probably you want to add a regex check if the username value is valid here
        if ticket_id:
            t_obj = Ticket.objects.get(pk=ticket_id)
            t_obj.ticket_status = 1
            t_obj.ticket_is_close = 1
            t_obj.ticket_closed_by_id = request.user.id
            t_obj.ticket_closed_at = datetime.now()
            t_obj.save()
            response_data = {}
            try:
                response_data['response'] = 'success'
            except:
                response_data['response'] = 'fail'
            return JsonResponse(response_data)

#Update Ticket Status through Ticket ID End#


#Get Org Users on Modal Through ID Start#

def getModalOrgUsersById(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        orgID = request.POST.get('org_id')
        users = User.objects.filter(is_delete=0,is_active=1).filter(user_org_id=orgID) # So we send the company instance
        context['users'] = users
        context['orgID'] = orgID

        return render(request,'itrak/Organization/org_modal_users.html',context)

#Get Org Users on Modal Through ID End#


#Get Org View on Modal Through ID Start#

def getModalOrgViewById(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        orgID = request.POST.get('org_id')
        data = Organization.objects.get(pk=orgID)
        # orgID = request.POST.get('org_id')
        # users = User.objects.filter(is_delete=0).filter(user_org_id=orgID) # So we send the company instance
        context['data'] = data
        return render(request,'itrak/Organization/org_modal_orgnization.html',context)

#Get Org View on Modal Through ID End#


#Get Org Tickets on Modal Through ID Start#
def getModalOrgTicketsById(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        orgID = request.POST.get('org_id')
        tickets = Ticket.objects.filter(ticket_is_delete=0).filter(ticket_org_id=orgID) # So we send the company instance
        context['tickets'] = tickets
        return render(request,'itrak/Organization/org_modal_tickets.html',context)

#Get Org Tickets on Modal Through ID End#


#Get Lates Attachment By Ticket ID Start#
def getLatesAttachementByTicketId(request):
    context = {}
    response_data = {}
    if request.method == 'POST' and request.is_ajax():
        ticketId =  request.POST.get('ticketId', '')
        if ticketId:
            try:
                # Get the Latest Attachment By Ticket ID
                attachment = TicketAttachments.objects.filter(attach_ticket_id=ticketId).latest('attach_id')
            except:
                response_data['response'] = ''

            if attachment:
                response_data['response'] = attachment.attach_id
                response_data['path'] = str(attachment.file_path)
            else:
                response_data['response'] = ''
    else:
        response_data['response'] = ''

    return HttpResponse(json.dumps(response_data), content_type="application/json")

#Get Lates Attachment By Ticket ID End#



#Delete Attachment By Attach ID Start#
def deleteAttachementByAttachId(request):
    context = {}
    response_data = {}
    if request.method == 'POST' and request.is_ajax():
        attachId =  request.POST.get('attachId', '')
        if attachId:
            try:
                # Get the Latest Attachment By Ticket ID
                attachment = TicketAttachments.objects.get(pk=attachId)
            except:
                response_data['response'] = 'Fail'

            if attachment:
                attachment.attach_is_delete = 1
                attachment.save()
                response_data['response'] = 'Success'
            else:
                response_data['response'] = 'Fail'
    else:
        response_data['response'] = 'Fail'

    return HttpResponse(json.dumps(response_data), content_type="application/json")

#Delete Attachment By Attach ID End#

# Service Contract By Org ID Start#
def getserviceContractByOrgId(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        org_id = request.POST.get('org_id')
        records = OrginaztionContract.objects.values().filter(oc_org_id=org_id).filter(org_contract_is_delete=0)
        context['records'] = records
    return render(request, 'itrak/Organization/get_contractservices_table_records.html', context)
# Service Contract By Org ID Start#

# Service Contract By Org View Start#
def getserviceContractByOrgViewId(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        org_id = request.POST.get('org_id')
        records = OrginaztionContract.objects.values().filter(oc_org_id=org_id).filter(org_contract_is_delete=0)
        context['records'] = records
    return render(request, 'itrak/Organization/get_contractservice_table_records_org.html', context)
# Service Contract By Org View Start#

#Service Contract By Org Contract ID Start
def getserviceContractByOrgContractId(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        org_contract_id = request.POST.get('org_contract_id')
        records = list(OrginaztionContract.objects.values().filter(org_contract_id=org_contract_id))
    return JsonResponse(records,safe=False)
#Service Contract By Org Contract ID End
# Schedule  Request By sch ID Start#
def getscheduleReportRecpBySchId(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        sch_rep_id = request.POST.get('sch_rep_id')
        records = ScheduleReportResp.objects.filter(sch_rep_id_id=sch_rep_id)
        context['records'] = records
    return render(request, 'itrak/Reports/get_scheduleReportResp_table_records.html', context)
# Schedule  Request By sch ID#

# first Note By ID
def firstNoteAdded(user_id,ticket_id,ticket_created_at,ticket_caller_id):
    if ticket_caller_id != user_id:
        try:
            ticketsNotes= TicketNote.objects.filter(note_ticket=ticket_id).filter(note_is_delete=0 )
            ticketcreated_at = ticket_created_at
            if ticketsNotes != None:
                tic = TicketNote.objects.values('note_created_at').filter(note_ticket=ticket_id).filter(note_is_delete=0 )[0]['note_created_at']
                if tic and ticketcreated_at:
                    delta = tic - ticketcreated_at
                    days, seconds = delta.days, delta.seconds
                    hours = days * 24 + seconds // 3600
                    minutes = (seconds % 3600) // 60
                    avg_time_submit_close = "{:02d}".format(hours + int(minutes / 60))+':'+"{:02d}".format(minutes % 60)
                    return avg_time_submit_close
                else:
                    return '00:00'    
        except:
            return '00:00'
    else:
        return '00:00'

#  first Note By ID
#  first Note  By Assigne ID
def firstNoteAddedByAssignee(ticket_id,ticket_created_at,ticket_assign_to):
    try:
        ticketsNotes= TicketNote.objects.filter(note_is_delete=0 ).filter(note_ticket=ticket_id).filter(note_created_by=ticket_assign_to)
        print(ticketsNotes)
        ticketcreated_at = ticket_created_at
        if ticketsNotes != None:
            tic = TicketNote.objects.values('note_created_at').filter(note_is_delete=0 ).filter(note_ticket=ticket_id)[0]['note_created_at']
            if tic and ticketcreated_at:
                delta = tic - ticketcreated_at
                days, seconds = delta.days, delta.seconds
                hours = days * 24 + seconds // 3600
                minutes = (seconds % 3600) // 60
                avg_time_submit_close = "{:02d}".format(hours + int(minutes / 60))+':'+"{:02d}".format(minutes % 60)
                return avg_time_submit_close
            else:
                return '00:00'     
    except:
        return '00:00' 
#  first Note  By Assigne ID
# get labour hour By ticket
def getLabourHoursByTicketId(ticket_id):
    result = TicketNote.objects.filter(note_ticket=ticket_id).filter(note_is_delete=0).aggregate(
                thours=Sum(Cast('tnote_laborhour_hours', IntegerField())) + Sum(Cast('tnote_laborhour_minutes', IntegerField())) / 60,
                tminutes=Sum(Cast('tnote_laborhour_minutes', IntegerField())) % 60
            )
    if result:
        if result['thours'] is None or str(result['thours']) == '0':
            hours = '00'
        else:
            hours = "{:02d}".format(result['thours'])
        if result['tminutes'] is None or str(result['tminutes']) == '0':
            minutes = '00';
        else:
            minutes = "{:02d}".format(result['tminutes'])
        labour_hours = str(hours)+':'+str(minutes)
        return labour_hours
    else:
        return '00:00'




# get labour Hour by ticket end
# hours open to assign start
def gettickectStat(ticket):
    t_open = 0
    t_close = 0
    t_reopen = 0
    t_assigned = 0
    t_unassigned = 0
    sub_assign_thours = 0
    sub_assign_tminutes = 0
    assign_close_thours = 0
    assign_close_tminutes = 0
    sub_close_thours = 0
    sub_close_tminutes = 0
    time_submit_assign = '00:00'
    time_assign_close = '00:00'
    time_submit_close = '00:00'
    tt_time_entered = '00:00'

    try:
        result = TicketNote.objects.filter(note_ticket=ticket).filter(note_is_delete=0).aggregate(
            thours=Sum(Cast('tnote_laborhour_hours', IntegerField())) + Sum(Cast('tnote_laborhour_minutes', IntegerField())) / 60,
            tminutes=Sum(Cast('tnote_laborhour_minutes', IntegerField())) % 60
        )
    except TicketNote.DoesNotExist:
        result = None
    if result:
        if result['thours'] is None or str(result['thours']) == '0':
            hours = '00'
        else:
            hours = "{:02d}".format(result['thours'])
        if result['tminutes'] is None or str(result['tminutes']) == '0':
            minutes = '00';
        else:
            minutes = "{:02d}".format(result['tminutes'])
        tt_time_entered = str(hours)+':'+str(minutes)+' hours'

    if ticket.ticket_status == 0:
        t_open+=1
    else:
        t_close+=1
    if ticket.ticket_is_reopen == 1:
        t_reopen+=1
    if ticket.ticket_assign_to:
        t_assigned+=1
    else:
        t_unassigned+=1

    if ticket.ticket_assign_to and ticket.ticket_created_at and ticket.ticket_assign_at:
        delta = ticket.ticket_assign_at - ticket.ticket_created_at
        days, seconds = delta.days, delta.seconds
        hours = days * 24 + seconds // 3600
        minutes = (seconds % 3600) // 60
        time_submit_assign = "{:02d}".format(hours + int(minutes / 60))+':'+"{:02d}".format(minutes % 60)

    if ticket.ticket_assign_to and ticket.ticket_closed_at and ticket.ticket_assign_at:
        delta = ticket.ticket_closed_at - ticket.ticket_assign_at
        days, seconds = delta.days, delta.seconds
        hours = days * 24 + seconds // 3600
        minutes = (seconds % 3600) // 60
        time_assign_close = "{:02d}".format(hours + int(minutes / 60))+':'+"{:02d}".format(minutes % 60)

    if ticket.ticket_created_at and ticket.ticket_closed_at:
        delta = ticket.ticket_closed_at - ticket.ticket_created_at
        days, seconds = delta.days, delta.seconds
        hours = days * 24 + seconds // 3600
        minutes = (seconds % 3600) // 60
        time_submit_close = "{:02d}".format(hours + int(minutes / 60))+':'+"{:02d}".format(minutes % 60)

    response = {
        "time_submit_assign": time_submit_assign,
        "time_assign_close": time_assign_close,
        "time_submit_close": time_submit_close,

    }
    return response



# hours open to assign end

# User Summary List start
def getSummaryList(request):
    context = {}
    if 'type' in request.POST and request.method == 'POST' and request.is_ajax():
        type = request.POST.get('type')
        if type == 'Agents':
            # results = User.objects.filter(user_type=0).filter(is_active=1).filter(is_delete=0).distinct()
            # context['results'] = results
            context['summary_type'] = 'Agents'
            return render(request, 'itrak/User/user_modal.html', context)
        if type == 'endUser':
            # results = User.objects.filter(user_type=1).filter(is_active=1).filter(is_delete=0).distinct()
            # context['results'] = results
            context['summary_type'] = 'endUser'
            return render(request, 'itrak/User/user_modal.html', context)
        if type == 'sysAdmin':
            # results = User.objects.filter(is_delete=0).filter(admin=1).distinct()
            # context['results'] = results
            context['summary_type'] = 'sysAdmin'
            return render(request, 'itrak/User/user_modal.html', context)
        if type == 'withlogin':
            # results = User.objects.filter(is_delete=0).filter(is_active=1).filter(login_permit=1).distinct()
            # context['results'] = results
            context['summary_type'] = 'withlogin'
            return render(request, 'itrak/User/user_modal.html', context)
        if type == 'withoutlogin':
            # results = User.objects.filter(is_delete=0).filter(is_active=1).filter(login_permit=0).distinct()
            # context['results'] = results
            context['summary_type'] = 'withoutlogin'
            return render(request, 'itrak/User/user_modal.html', context)
        if type == 'totalActive':
            # results = User.objects.filter(is_active=1).filter(is_delete=0).distinct()
            # context['results'] = results
            context['summary_type'] = 'totalActive'
            return render(request, 'itrak/User/user_modal.html', context)
        if type == 'totalInActive':
            # results = User.objects.filter(is_active=0).filter(is_delete=0).distinct()
            # context['results'] = results
            context['summary_type'] = 'totalInActive'
            return render(request, 'itrak/User/user_modal.html', context)
        if type == 'totalUsers':
            # results = User.objects.filter(is_delete=0).distinct()
            # context['results'] = results
            context['summary_type'] = 'totalUsers'
            return render(request, 'itrak/User/user_modal.html', context)
        if type == 'can_assign_tickets':    
            # user_ids = UserActionPermission.objects.filter(perm_act_id=1).filter(is_delete=0).values_list('user_id').distinct()
            # results=User.objects.filter(pk__in=user_ids)
            # context['results'] = results
            context['summary_type'] = 'can_assign_tickets'
            return render(request, 'itrak/User/user_modal.html', context)
        if type == 'can_be_assigned_tickets':    
            # user_ids = UserActionPermission.objects.filter(perm_act_id=2).filter(is_delete=0).values_list('user_id').distinct()
            # results=User.objects.filter(pk__in=user_ids)
            # context['results'] = results
            context['summary_type'] = 'can_be_assigned_tickets'
            return render(request, 'itrak/User/user_modal.html', context)
        if type == 'can_submit_tickets':    
            # user_ids = UserActionPermission.objects.filter(perm_act_id=3).filter(is_delete=0).values_list('user_id').distinct()
            # results=User.objects.filter(pk__in=user_ids)
            # context['results'] = results
            context['summary_type'] = 'can_submit_tickets'
            return render(request, 'itrak/User/user_modal.html', context)
        if type == 'can_access_and_maintain_admin':    
            # user_ids = UserActionPermission.objects.filter(perm_act_id=4).filter(is_delete=0).values_list('user_id').distinct()
            # results=User.objects.filter(pk__in=user_ids)
            # context['results'] = results
            context['summary_type'] = 'can_access_and_maintain_admin'
            return render(request, 'itrak/User/user_modal.html', context)

# User Summary List end

#Get Permissions List

def user_action_permissions(request, user):
    return list(UserActionPermission.objects.filter(user_id=user,perm_act_id__isnull=False).values_list('perm_act_id',flat=True))

def user_sub_action_permissions(request, user):
    return list(UserSubActionPermission.objects.filter(user_id=user,sub_act_id__isnull=False).values_list('sub_act_id',flat=True))

def permissions_not_allowed(request, user):
    if(user.user_type == str(1)):
        return list(PermissionAction.objects.filter(permit_end_user=False,is_delete=False).values_list('perm_act_id',flat=True))
    else:
        return list(PermissionAction.objects.filter(permit_agent=False,is_delete=False).values_list('perm_act_id',flat=True))
        
def sub_permissions_not_allowed(request, user):
    if(user.user_type == str(1)):
        return list(PermissionSubAction.objects.filter(permit_end_user=False,is_delete=False).values_list('sub_act_id',flat=True))
    else:
        return list(PermissionSubAction.objects.filter(permit_agent=False,is_delete=False).values_list('sub_act_id',flat=True))


# Get Ticket Type Recods  start
def getTicketTypeRecordsByParentID(type_id):
    data = TicketType.objects.filter(parent_id = type_id).filter(ttype_is_delete=0).filter(ttype_is_active=1)
    return data
# Get Ticket Type Recods  ends

# Get ticketTypeExportXLS starts
def ticketTypeExportXLS():
    getAll = TicketType.objects.filter(parent_id = 0, ttype_is_delete=0)
    mainArray = []
    for getAll in getAll:
        levelZeroInnerArray = {}
        dataZeroLevel = getTicketTypeRecordsByParentID(getAll.ttype_id)
        if dataZeroLevel:
            for dataZeroLevel in dataZeroLevel:
                levelOneInnerArray = {}
                dataOneLevel = getTicketTypeRecordsByParentID(dataZeroLevel.ttype_id)
                if dataOneLevel:
                    for dataOneLevel in dataOneLevel:
                        levelTwoInnerArray = {}
                        dataTwoLevel = getTicketTypeRecordsByParentID(dataOneLevel.ttype_id)
                        if dataTwoLevel:
                            for dataTwoLevel in dataTwoLevel:
                                levelThreeInnerArray = {}
                                dataThreeLevel = getTicketTypeRecordsByParentID(dataTwoLevel.ttype_id)
                                if dataThreeLevel:
                                    for dataThreeLevel in dataThreeLevel:
                                        levelFourInnerArray = {}
                                        dataFourLevel = getTicketTypeRecordsByParentID(dataThreeLevel.ttype_id)
                                        levelFourInnerArray['levelZeroIssueTypeID'] = getAll.ttype_id
                                        levelFourInnerArray['levelZeroIssueTypeName'] = getAll.ttype_name
                                        levelFourInnerArray['levelOneIssueTypeID'] = dataZeroLevel.ttype_id
                                        levelFourInnerArray['levelOneIssueTypeName'] = dataZeroLevel.ttype_name
                                        levelFourInnerArray['levelTwoIssueTypeID'] = dataOneLevel.ttype_id
                                        levelFourInnerArray['levelTwoIssueTypeName'] = dataOneLevel.ttype_name
                                        levelFourInnerArray['levelThreeIssueTypeID'] = dataTwoLevel.ttype_id
                                        levelFourInnerArray['levelThreeIssueTypeName'] = dataTwoLevel.ttype_name
                                        levelFourInnerArray['levelFourIssueTypeID'] = dataThreeLevel.ttype_id
                                        levelFourInnerArray['levelFourIssueTypeName'] = dataThreeLevel.ttype_name
                                        levelFourInnerArray['is_active'] = getAll.ttype_is_active
                                        mainArray.append(levelFourInnerArray)
                                else:
                                    levelThreeInnerArray['levelZeroIssueTypeID'] = getAll.ttype_id
                                    levelThreeInnerArray['levelZeroIssueTypeName'] = getAll.ttype_name
                                    levelThreeInnerArray['levelOneIssueTypeID'] = dataZeroLevel.ttype_id
                                    levelThreeInnerArray['levelOneIssueTypeName'] = dataZeroLevel.ttype_name
                                    levelThreeInnerArray['levelTwoIssueTypeID'] = dataOneLevel.ttype_id
                                    levelThreeInnerArray['levelTwoIssueTypeName'] = dataOneLevel.ttype_name
                                    levelThreeInnerArray['levelThreeIssueTypeID'] = dataTwoLevel.ttype_id
                                    levelThreeInnerArray['levelThreeIssueTypeName'] = dataTwoLevel.ttype_name
                                    levelThreeInnerArray['levelFourIssueTypeID'] = 0
                                    levelThreeInnerArray['levelFourIssueTypeName'] = '-'
                                    levelThreeInnerArray['is_active'] = getAll.ttype_is_active
                                    mainArray.append(levelThreeInnerArray)
                        else:
                            levelTwoInnerArray['levelZeroIssueTypeID'] = getAll.ttype_id
                            levelTwoInnerArray['levelZeroIssueTypeName'] = getAll.ttype_name
                            levelTwoInnerArray['levelOneIssueTypeID'] = dataZeroLevel.ttype_id
                            levelTwoInnerArray['levelOneIssueTypeName'] = dataZeroLevel.ttype_name
                            levelTwoInnerArray['levelTwoIssueTypeID'] = dataOneLevel.ttype_id
                            levelTwoInnerArray['levelTwoIssueTypeName'] = dataOneLevel.ttype_name
                            levelTwoInnerArray['levelThreeIssueTypeID'] = 0
                            levelTwoInnerArray['levelThreeIssueTypeName'] = '-'
                            levelTwoInnerArray['levelFourIssueTypeID'] = 0
                            levelTwoInnerArray['levelFourIssueTypeName'] = '-'
                            levelTwoInnerArray['is_active'] = getAll.ttype_is_active
                            mainArray.append(levelTwoInnerArray)
                else:
                    levelOneInnerArray['levelZeroIssueTypeID'] = getAll.ttype_id
                    levelOneInnerArray['levelZeroIssueTypeName'] = getAll.ttype_name
                    levelOneInnerArray['levelOneIssueTypeID'] = dataZeroLevel.ttype_id
                    levelOneInnerArray['levelOneIssueTypeName'] = dataZeroLevel.ttype_name
                    levelOneInnerArray['levelTwoIssueTypeID'] = 0
                    levelOneInnerArray['levelTwoIssueTypeName'] = '-'
                    levelOneInnerArray['levelThreeIssueTypeID'] = 0
                    levelOneInnerArray['levelThreeIssueTypeName'] = '-'
                    levelOneInnerArray['levelFourIssueTypeID'] = 0
                    levelOneInnerArray['levelFourIssueTypeName'] = '-'
                    levelOneInnerArray['is_active'] = getAll.ttype_is_active
                    mainArray.append(levelOneInnerArray)
        else:
            levelZeroInnerArray['levelZeroIssueTypeID'] = getAll.ttype_id
            levelZeroInnerArray['levelZeroIssueTypeName'] = getAll.ttype_name
            levelZeroInnerArray['levelOneIssueTypeID'] = 0
            levelZeroInnerArray['levelOneIssueTypeName'] = '-'
            levelZeroInnerArray['levelTwoIssueTypeID'] = 0
            levelZeroInnerArray['levelTwoIssueTypeName'] = '-'
            levelZeroInnerArray['levelThreeIssueTypeID'] = 0
            levelZeroInnerArray['levelThreeIssueTypeName'] = '-'
            levelZeroInnerArray['levelFourIssueTypeID'] = 0
            levelZeroInnerArray['levelFourIssueTypeName'] = '-'
            levelZeroInnerArray['is_active'] = getAll.ttype_is_active
            mainArray.append(levelZeroInnerArray)
    
    return mainArray
# Get ticketTypeExportXLS ends

#Get Org Users on Modal Through ID Start#

def getModalDepUsersById(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        depID = request.POST.get('dep_id')
        users = User.objects.filter(is_delete=0).filter(user_dep_id=depID) # So we send the company instance
        context['users'] = users
        context['depID'] = depID

        return render(request,'itrak/Department/dep_modal_users.html',context)

#Get Org Users on Modal Through ID End#

#Get Client Users on Modal Through ID Start#

def getModalUsersByClientId(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        client_id = request.POST.get('client_id')
        users = User.objects.filter(is_delete=0).filter(user_client_id = client_id) # So we send the company instance
        context['users'] = users
        context['client_id'] = client_id

        return render(request,'itrak/Client/client_modal_users.html',context)

#Get Client Users on Modal Through ID End#

#check User Action permission. 
def check_action_permission(slug, user_id):
    check_slug_id = PermissionAction.objects.filter(perm_act_slug=slug).values_list('perm_act_id',flat=True)    
    if(check_slug_id.exists()):
        check_slug_id = check_slug_id[0]
        hasUserPermissions = UserActionPermission.objects.filter(user_id=user_id, perm_act_id=check_slug_id).exists()
        user_groups = UserGroupMembership.objects.filter(m_user_id= user_id,is_delete = 0)
        hasGroupPermissions = False
        for group in user_groups:
            groupPermissions = GroupActionPermission.objects.filter(group_id=group.m_group_id, perm_act_id=check_slug_id).exists()
            if groupPermissions:
                hasGroupPermissions = True
        if hasUserPermissions or hasGroupPermissions:
            return True
        else:
            return False
    return False

#check User SubAction permission.
def check_sub_action_permission(slug, user_id):
    check_slug_id = PermissionSubAction.objects.filter(perm_sub_act_slug=slug).values_list('sub_act_id',flat=True)    
    if(check_slug_id.exists()):
        check_slug_id = check_slug_id[0]
        hasUserPermissions = UserSubActionPermission.objects.filter(user_id=user_id, sub_act_id=check_slug_id).exists()
        user_groups = UserGroupMembership.objects.filter(m_user_id= user_id,is_delete = 0)
        hasGroupPermissions = False
        for group in user_groups:
            groupPermissions = GroupSubActionPermission.objects.filter(group_id=group.m_group_id, sub_act_id=check_slug_id).exists()
            if groupPermissions:
                hasGroupPermissions = True
        if hasUserPermissions or hasGroupPermissions:
            return True
        else:
            return False
    return False

#Get Permissions List

def group_action_permissions(request, group_id):
    return list(GroupActionPermission.objects.filter(group_id=group_id,perm_act_id__isnull=False).values_list('perm_act_id',flat=True))

def group_sub_action_permissions(request, group_id):
    return list(GroupSubActionPermission.objects.filter(group_id=group_id,sub_act_id__isnull=False).values_list('sub_act_id',flat=True))

def group_permissions_not_allowed(request):
    if(request.user.id == str(1)):
        return list(PermissionAction.objects.filter(permit_end_user=False,is_delete=False).values_list('perm_act_id',flat=True))
    else:
        return list(PermissionAction.objects.filter(permit_agent=False,is_delete=False).values_list('perm_act_id',flat=True))
        
def group_sub_permissions_not_allowed(request):
    if(request.user.id == str(1)):
        return list(PermissionSubAction.objects.filter(permit_end_user=False,is_delete=False).values_list('sub_act_id',flat=True))
    else:
        return list(PermissionSubAction.objects.filter(permit_agent=False,is_delete=False).values_list('sub_act_id',flat=True))

# Get allUsersByOrganization starts
def getArrayOfallUsersByOrganization():
    getAll = Organization.objects.filter(org_is_delete=0,org_is_active=1)
    mainArray = []
    for getAll in getAll:
        organizatoinData = {}
        organizatoinData['Organization Name'] = getAll.org_name
        organizatoinData['User ID'] = ''
        organizatoinData['Display Name'] = ''
        organizatoinData['First Name'] = ''
        organizatoinData['Last Name'] = ''
        organizatoinData['Email'] = ''
        organizatoinData['Phone'] = ''
        organizatoinData['Address1'] = ''
        organizatoinData['Address2'] = ''
        organizatoinData['City'] = ''
        organizatoinData['State'] = ''
        organizatoinData['Zip'] = ''

        mainArray.append(organizatoinData)
        getAllUsers = User.objects.filter(user_org_id = getAll.org_id,is_active = 1, is_delete=0)
        for getAllUsers in getAllUsers:
            usersData = {}
            usersData['Organization Name'] = ''
            usersData['User ID'] = getAllUsers.username
            usersData['Display Name'] = getAllUsers.last_name+', '+getAllUsers.first_name
            usersData['First Name'] = getAllUsers.first_name
            usersData['Last Name'] = getAllUsers.last_name
            usersData['Email'] = getAllUsers.email
            usersData['Phone'] = getAllUsers.phone_no
            usersData['Address1'] = getAllUsers.address1
            usersData['Address2'] = getAllUsers.address2
            usersData['City'] = getAllUsers.user_city
            usersData['State'] = getAllUsers.user_state
            usersData['Zip'] = getAllUsers.user_zip_code
            mainArray.append(usersData)
    return mainArray
# Get allUsersByOrganization ends

# Get allUsersByDepartment starts
def getArrayOfallUsersByDepartment():
    getAll = Department.objects.filter(d_is_delete=0,d_is_active=1)
    mainArray = []
    for getAll in getAll:
        departmentData = {}
        departmentData['Department Name'] = getAll.dep_name
        departmentData['User ID'] = ''
        departmentData['Display Name'] = ''
        departmentData['First Name'] = ''
        departmentData['Last Name'] = ''
        departmentData['Email'] = ''
        departmentData['Phone'] = ''
        departmentData['Address1'] = ''
        departmentData['Address2'] = ''
        departmentData['City'] = ''
        departmentData['State'] = ''
        departmentData['Zip'] = ''

        mainArray.append(departmentData)
        getAllUsers = User.objects.filter(is_delete=0,is_active = 1,user_dep_id = getAll.dep_id)
        for getAllUsers in getAllUsers:
            usersData = {}
            usersData['Department Name'] = ''
            usersData['User ID'] = getAllUsers.username
            usersData['Display Name'] = getAllUsers.last_name+', '+getAllUsers.first_name
            usersData['First Name'] = getAllUsers.first_name
            usersData['Last Name'] = getAllUsers.last_name
            usersData['Email'] = getAllUsers.email
            usersData['Phone'] = getAllUsers.phone_no
            usersData['Address1'] = getAllUsers.address1
            usersData['Address2'] = getAllUsers.address2
            usersData['City'] = getAllUsers.user_city
            usersData['State'] = getAllUsers.user_state
            usersData['Zip'] = getAllUsers.user_zip_code
            mainArray.append(usersData)
    return mainArray
# Get allUsersByDepartment ends


# Get allUsersByClient starts
def getArrayOfallUsersByclient():
    getAll = Client.objects.filter(cl_is_delete=0,cl_is_active=1)
    mainArray = []
    for getAll in getAll:
        clienttData = {}
        clienttData['Client Name'] = getAll.client_name
        clienttData['User ID'] = ''
        clienttData['Display Name'] = ''
        clienttData['First Name'] = ''
        clienttData['Last Name'] = ''
        clienttData['Email'] = ''
        clienttData['Phone'] = ''
        clienttData['Address1'] = ''
        clienttData['Address2'] = ''
        clienttData['City'] = ''
        clienttData['State'] = ''
        clienttData['Zip'] = ''

        mainArray.append(clienttData)
        getAllUsers = User.objects.filter(is_delete=0,is_active = 1,user_client_id = getAll.client_id)
        for getAllUsers in getAllUsers:
            usersData = {}
            usersData['Client Name'] = ''
            usersData['User ID'] = getAllUsers.username
            usersData['Display Name'] = getAllUsers.last_name+', '+getAllUsers.first_name
            usersData['First Name'] = getAllUsers.first_name
            usersData['Last Name'] = getAllUsers.last_name
            usersData['Email'] = getAllUsers.email
            usersData['Phone'] = getAllUsers.phone_no
            usersData['Address1'] = getAllUsers.address1
            usersData['Address2'] = getAllUsers.address2
            usersData['City'] = getAllUsers.user_city
            usersData['State'] = getAllUsers.user_state
            usersData['Zip'] = getAllUsers.user_zip_code
            mainArray.append(usersData)
    return mainArray
# Get allUsersByOrganization ends




#Get User Attachments Start#
def get_user_attachments(request, UserId = id):
    attach_ids = UserAttachment.objects.filter(ua_user_id=UserId).filter(ua_is_delete=0).values_list('ua_id','ua_file_path', 'ua_file_name', 'ua_file_size')
    return attach_ids

#Get User Attachments End#   

#Get Subject Token Start#
def get_subject_tokens(request, Id = id):
    subject_tokens = CustomMessagesToken.objects.filter(cmt_is_subject=Id).filter(cmt_is_delete=0).values_list('cmt_id','cmt_name','cmt_slug')
    return subject_tokens
#Get Subject Token End#  

#Get Message Token Start#
def get_message_tokens(request, Id = id):
    message_tokens = CustomMessagesToken.objects.filter(cmt_is_subject=Id).filter(cmt_is_delete=0).values_list('cmt_id','cmt_name','cmt_slug')
    return message_tokens
#Get Message Token End#  

#Get Task Group Restrict Start#
def get_taskgrouprestricts(request, Id = id):
    taskgrouprestricts = TaskGroupRestrict.objects.filter(tgr_task_group_id=Id).filter(tgr_is_delete=0).values_list('tgr_id','tgr_group_or_org_name','tgr_type_is_group','tgr_type_is_org')
    return taskgrouprestricts
#Get Task Group Restrict End#

def getDateTimeByTimezone(dateTime, userID):
    if dateTime is None:
        return ''
    else:
        uTimeZone = MySettings.objects.filter(m_user_id=9).first().m_time_zone
        local_dt = dateTime.astimezone(pytz.timezone(uTimeZone))
        result = datetime.strptime(str(local_dt), '%Y-%m-%d').strftime('%m/%d/%Y')
        return result



#Conversion of 12-Hour Format to 24-Format
def get24HourFormattedTime(str1): 
    # Checking if last two elements of time 
    # is AM and first two elements are 12 
    if str1[-2:] == "AM" and str1[:2] == "12": 
        return "00" + str1[2:-2] 
          
    # remove the AM     
    elif str1[-2:] == "AM": 
        return str1[:-2] 
      
    # Checking if last two elements of time 
    # is PM and first two elements are 12    
    elif str1[-2:] == "PM" and str1[:2] == "12": 
        return str1[:-2] 
          
    else: 
        # add 12 to hours and remove PM 
        return str(int(str1[:2]) + 12) + str1[2:8] 

# Saving Value in History Hours of Operation
def CreateHistoryLog(old,new,message,recalculation,user_id):
    HistoryHoursOfOperation.objects.create(
    hhop_recalculation= recalculation ,
    hhop_previous_value = old,
    hhop_current_value = new,
    hhop_type = message ,
    hhop_modified_by = user_id
    )

#Get User ID with permission Token Start#
def get_userids_with_permission(request, Id):
    current_user_id = request.user.id
    org_id = request.user.user_org_id
    global_user = isGlobalUser(request)
    if request.user.user_type_slug != global_user:
        user_ids = UserActionPermission.objects.filter(user_org_id=org_id).filter(perm_act_id=Id).filter(is_delete=0).values_list('user_id')
    else:
        user_ids = UserActionPermission.objects.filter(perm_act_id=Id).filter(is_delete=0).values_list('user_id')
    return user_ids
#Get User ID with permission Token End#     


#Login Permit Start#
def user_login_permit(request, user_name):
    login_permit = User.objects.filter(username=user_name).filter(is_delete=0).values_list('login_permit')
    return login_permit
#Login Permit End#    
  

#Get Task Restrict Start#
def get_taskrestricts(request, Id = id):
    taskrestricts = TaskRestrict.objects.filter(tr_task_id=Id).filter(tr_is_delete=0).values_list('tr_id','tr_group_or_org_name','tr_type_is_group','tr_type_is_org')
    return taskrestricts
#Get Task Restrict End#   

#Get Tickettype Parent Data Start#
def get_tickettype_data(request):
    org_id = request.user.user_org_id
    if request.user.id != 3108:
        tickettype_data = TicketType.objects.raw('''
        select * from AT_TicketType
        where ttype_is_delete = 0 and ttype_is_active=1 
        and has_parent=0 and parent_id=0 and user_org_id = '''+"'"+str(org_id)+"'"+'''
        order by 
        CASE WHEN [t_type_display_order] = '' THEN 1 ELSE 0 END,
        [t_type_display_order]
        ''')
    else:
        tickettype_data = TicketType.objects.raw('''
            select * from AT_TicketType
            where ttype_is_delete = 0 and ttype_is_active=1 
            and has_parent=0 and parent_id=0
            order by 
            CASE WHEN [t_type_display_order] = '' THEN 1 ELSE 0 END,
            [t_type_display_order]
        ''')
    
    return tickettype_data
#Get Tickettype Parent Data End#   


# convert tuple return from cursor to dict start
def dictfetchall(cursor): 
    "Returns all rows from a cursor as a dict" 
    desc = cursor.description 
    return [
        dict(zip([col[0] for col in desc], row)) 
        for row in cursor.fetchall() 
    ]
# convert tuple return from cursor to dict end

# return User Type start
def userType(request): 
    userType = ''
    if(request.user.admin == True and request.user.user_type == '0' ): #Super Admin
        userType = 'superadmin'
    elif(request.user.admin == False and request.user.user_type == '0' ):#Agent 
        userType = 'agent'
    elif(request.user.user_type == '1' and check_action_permission("da_Can_view_Tickets_submitted_by_other_users",request.user.id) == True ):#End User - Manager
        userType = 'manager'
    elif(request.user.user_type == '1' and check_action_permission("da_Can_view_Tickets_submitted_by_other_users",request.user.id) == False ):#End User
        userType = 'enduser'
    return userType
# return User Type End

# GET ACCOUNTS OF CURRENT LOGIN USER STARTS
def accountsOfCurrentUser(request): 
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM AT_UserAccountRelation WHERE user_id = %s", [request.user.id])
    accountsSelectedValues = cursor.fetchall()
    accountsSelectedValuesList = []
    for index, tuple in enumerate(accountsSelectedValues):
        account_id = tuple[2]
        accountsSelectedValuesList.append(str(account_id))
    accountsSelectedValuesList = ', '.join( repr(e) for e in accountsSelectedValuesList) 
    sql = "SELECT * FROM GlobalACCTS WHERE id in ("+accountsSelectedValuesList+")"
    cursor.execute(sql)
    accountsTuples = cursor.fetchall()
    
    accountsList = []
    for index, tuple in enumerate(accountsTuples):
        innerPortion = {}
        innerPortion['account_id'] = tuple[0]
        innerPortion['account_name'] = tuple[3]
        accountsList.append(innerPortion)
    return accountsList
# GET ACCOUNTS OF CURRENT LOGIN USER ENDS

#Mapped AccuntID's to Current Login User - Starts
def getAccountIDsOfCurrentUser(request):
    #User-Account Mapping Check Start
    user_id = request.user.id 
    user_type = userType(request)
    if user_type == 'superadmin' or user_type == 'agent':
        SQL  = """
            select distinct account_id
            from AT_UserAccountRelation a
            where 1=1
            AND a.user_id = '"""+str(user_id)+"""'
        """
    elif user_type == 'manager' or user_type == 'enduser':
        SQL  = """
            select distinct account_id
            from AT_UserAccountRelation a
            where 1=1
            AND a.user_id = '"""+str(user_id)+"""'
            AND (
                SELECT COUNT(*)
                FROM AT_Users AU
                WHERE AU.ID = A.user_id 
                AND AU.user_type = 1 -- END USER
            )> 0 
        """
    
    cursor = connection.cursor()
    cursor.execute(SQL)
    accounts = dictfetchall(cursor) 

    accountsList = []
    for account in accounts:
        accountsList.append(account['account_id'])
    return accountsList
#Mapped AccuntID's to Current Login User - Ends


#Mapped AccuntID's to Current Login User - Starts
def getUsersOfCurrentOrg(request):
    #User-Account Mapping Check Start
    org_id = request.user.user_org_id 
    user_id = request.user.id 
    user_type = userType(request)
    if user_type == 'superadmin' or user_type == 'agent':
            SQL  = """
                select distinct id
                from At_USERS a
                where 1=1
                AND a.id = '"""+str(user_id)+"""'
            """
    elif user_type == 'manager' or user_type == 'enduser':
        SQL  = """
            select distinct id
            from AT_USERS a
            where 1=1
            AND a.id = '"""+str(user_id)+"""'
            AND (
                SELECT COUNT(*)
                FROM AT_Users AU
                WHERE AU.ID = A.user_id 
                AND AU.user_type = 1 -- END USER
            )> 0 
        """
        
    cursor = connection.cursor()
    cursor.execute(SQL)
    users = dictfetchall(cursor)
    usersList = []
    for account in users:
        usersList.append(account['id'])
    
    if len(usersList) > 0:
        usersFilter = {'ticket_created_by_id__in':usersList}
    else: 
        usersFilter = {}
#Mapped AccuntID's to Current Login User - Ends



#Get Mapped Users with currents login User - Starts
def getMappedUserIDsWithCurrentUer(request):
    #User-Account Mapping Check Start
    user_id = request.user.id 
    SQL  = """
        select distinct user_id
        from AT_UserAccountRelation a
        where 1=1
        AND (
            SELECT COUNT(*)
            FROM AT_Users AU
            WHERE AU.id = a.user_id
            AND AU.user_type = 1 --END USER
        )>0
        AND a.account_id in (
            select account_id
            from AT_UserAccountRelation b
            where b.user_id = '"""+str(user_id)+"""'
        )
    """
    
    cursor = connection.cursor()
    cursor.execute(SQL)
    users = dictfetchall(cursor) 

    usersList = []
    for user in users:
        usersList.append(user['user_id'])
    return usersList
#Get Mapped Users with currents login User - Ends

def isGlobalUser(request):

    global_user = 'global_user'
    return global_user
@csrf_exempt
def orgUserExists(request):
    # if request.user.user_type_slug != 'global_user':
    #     orgs_id = request.user.user_org_id
    #     user_org_id = User.objects.filter(user_org_id = orgs_id,admin=1,user_type=0,)
    #     if len(user_org_id) > 0:
    #         res = 2
    #         return JsonResponse(res , safe=False)
    #     else:
    #         res = 0
    #         return JsonResponse(res, safe=False)
    # if 'sys_admin' in request.POST and request.POST['sys_admin']:
    if request.method == 'POST' and request.is_ajax()  :
        id = request.GET.get('UserID')
        print(id)
        org_id = request.POST.get('org_id')
        if 'sys_admin' in request.POST:
            sys_admin = request.POST.get('sys_admin')
        print(sys_admin)
        print(type(sys_admin))
        user_org_id = User.objects.filter(user_org_id = org_id,admin=1,user_type=0, is_delete=0)
        if len(user_org_id) > 0 and sys_admin == 'on':
            res = 1
            return JsonResponse(res , safe=False)
        else:
            res = 0
            return JsonResponse(res, safe=False)
    else:
        return render(request, 'itrak/User/user_edit.html')
    # else:
    #     return render(request, 'itrak/User/user_edit.html')

@csrf_exempt
def getOrg(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        query = 'select user_org_id from AT_USERS where id = '+user_id+''
        cursor = connection.cursor()
        cursor.execute(query)
        org_id = cursor.fetchone()
        print(type(org_id))
        org_id = functools.reduce(lambda sub, ele: sub * 10 + ele, org_id)
        print(org_id)
        # super_admin_user = User.objects.get(user_org_id=org_id , admin=1 , is_delete=0).id
        # print(type(super_admin_user))
        print(type(user_id))
        exists = User.objects.filter(user_org_id=org_id,admin=1,user_type=0, is_delete=0)
        if len(exists) > 0:
            # if str(super_admin_user) == user_id:
            #     res = 2
            #     return JsonResponse(res, safe=False)
            # else:
            res = 1
            return JsonResponse(res, safe=False)
        else:
            res = 0
            return JsonResponse(res, safe=False)

@csrf_exempt
def getDepartment(request):
    if request.method == 'POST':
        org_id = request.POST.get('org_id')
        query = 'select * from AT_Departments where user_org_id = '+org_id+' '
        
        cursor = connection.cursor()
        cursor.execute(query)
        dept = cursor.fetchall()
        query_group = 'select * from AT_Groups where group_org_id = '+org_id+' '
        cursor = connection.cursor()
        cursor.execute(query_group)
        group = cursor.fetchall()
        data = []
        groups = []
        for i in dept:
            depts = {}
            depts['dep_id']=i[0]
            depts['dep_name']=i[1]
            data.append(depts)
        for j in group:
            group_detail = {}
            group_detail['group_id']=j[0]
            group_detail['group_display_name']=j[2]
            groups.append(group_detail)
        
        main = {}
        main['department'] = data
        main['groups'] = groups
        return JsonResponse(main, safe=False)
