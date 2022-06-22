from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404, render_to_response
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponseRedirect, HttpResponse, Http404, JsonResponse, HttpResponseBadRequest
from itrak.models import *
from django_datatables_view.base_datatable_view import BaseDatatableView
from django.utils.html import escape
from django.db.models import Q, F
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
import pytz
from django.core.mail import send_mail, EmailMultiAlternatives
from django.utils.crypto import get_random_string
import smtplib
import mimetypes
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.message import Message
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.text import MIMEText
from django.conf.urls import url
from django.template.loader import render_to_string, get_template
from itrak.views.Load import *
from django.db.models.query import QuerySet
from django.core.files.storage import FileSystemStorage
from django.core import signing
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
import json
from html.parser import HTMLParser
from django.db import transaction, IntegrityError
from django.urls import reverse
from django.db.models import FloatField, IntegerField
from django.db.models.functions import Cast
from django.db.models.expressions import RawSQL





# Create your views here.

class MLStripper(HTMLParser):
    def __init__(self):
        self.reset()
        self.strict = False
        self.convert_charrefs= True
        self.fed = []
    def handle_data(self, d):
        self.fed.append(d)
    def get_data(self):
        return ''.join(self.fed)


#Custom Decorator Start#

user_login_required = user_passes_test(lambda user: user.is_active, login_url='/') #Here user_passes_test decorator designates the user is active.

def active_user_required(view_func):
    decorated_view_func = login_required(user_login_required(view_func))
    return decorated_view_func

#Custom Decorator End#


# Ticket Submit Request Start#

@active_user_required
def submitTicket(request):
    # vname = Ticket._meta.get_field("subject").verbose_name.title()
    organizations = Organization.objects.filter(org_is_delete=0).filter(org_is_active=1)
    clients = Client.objects.filter(cl_is_delete=0)
    clientInfos = ClientInformation.objects.filter(clientinfo_is_delete=0)
    ticketTypes = TicketType.objects.filter(has_parent=0).filter(ttype_is_delete=0)
    priorities = Priority.objects.filter(prior_is_delete=0)
    substatus = SubStatus.objects.filter(sstatus_is_delete=0)
    users = User.objects.filter(is_delete=0)
    tasks = Task.objects.filter(task_is_active=1).filter(task_is_delete=0).order_by(F('task_display_order').asc(nulls_last=True))
    taskGroups = TaskGroup.objects.filter(taskgroup_is_active=1).filter(taskgroup_is_delete=0).order_by(F('taskgroup_display_order').asc(nulls_last=True))

    load_sidebar = get_sidebar(request)
    context = {
        'sidebar': load_sidebar,
        'organizations': organizations,
        'clients': clients,
        'clientinformations': clientInfos,
        'ticketTypes': ticketTypes,
        'priorities': priorities,
        'substatus': substatus,
        'users': users,
        'tasks': tasks,
        'taskGroups': taskGroups,
    }

    return render(request, 'itrak/Ticket/ticket_add.html', context)

# Ticket Submit Request End#


# Ticket Save Request Start#

@active_user_required
def saveTicket(request):
    with transaction.atomic():
        if request.method == 'POST':
            if 'submit_date' in request.POST and request.POST['submit_date']:
                submit_date = datetime.strptime(request.POST.get('submit_date'), '%m-%d-%Y').strftime('%Y-%m-%d')
            if 'submit_time' in request.POST:
                submit_time = request.POST.get('submit_time')
            if 'org_id' in request.POST:
                org_id = request.POST.get('org_id')
            if 'caller_id' in request.POST:
                caller_id = request.POST.get('caller_id')
            if 'client_id' in request.POST and request.POST['client_id']:
                client_id = request.POST.get('client_id')
            if 'clientinfo_id' in request.POST:
                clientinfo_id = request.POST.get('clientinfo_id')
            if 'record_locator' in request.POST:
                record_locator = request.POST.get('record_locator')
            if 'caller_name' in request.POST:
                caller_name = request.POST.get('caller_name')
            if 'caller_phone' in request.POST:
                caller_phone = request.POST.get('caller_phone')
            if 'caller_email' in request.POST:
                caller_email = request.POST.get('caller_email')
            if 'passenger_name' in request.POST:
                passenger_name = request.POST.get('passenger_name')
            if 'subject' in request.POST and request.POST['subject']:
                subject = request.POST.get('subject')
            if 'description' in request.POST and request.POST['description']:
                description = request.POST.get('description')
            else:
                description = ''
            if 'ticket_type' in request.POST:
                ticket_type = request.POST.get('ticket_type')
            if 'subtype1' in request.POST:
                subtype1 = request.POST.get('subtype1')
            if 'subtype2' in request.POST:
                subtype2 = request.POST.get('subtype2')
            if 'subtype3' in request.POST:
                subtype3 = request.POST.get('subtype3')
            if 'subtype4' in request.POST:
                subtype4 = request.POST.get('subtype4')
            if 'priority' in request.POST:
                priority = request.POST.get('priority')
            if 'ticket_substatus' in request.POST:
                ticket_substatus = request.POST.get('ticket_substatus')
            if 'traveler_vip' in request.POST:
                traveler_vip = request.POST.get('traveler_vip')
            if 'payout_required' in request.POST:
                payout_required = request.POST.get('payout_required')
            if 'error_goodwill' in request.POST:
                error_goodwill = request.POST.get('error_goodwill')
            if 'amount_saved' in request.POST:
                amount_saved = request.POST.get('amount_saved')
            if 'airline_ticket' in request.POST:
                airline_ticket = request.POST.get('airline_ticket')
            if 'agent_responsible' in request.POST:
                agent_responsible = request.POST.get('agent_responsible')
            if 'vendor_responsible' in request.POST:
                vendor_responsible = request.POST.get('vendor_responsible')
            if 'vresponsible_city' in request.POST:
                vresponsible_city = request.POST.get('vresponsible_city')
            if 'ticket_payout_amount' in request.POST:
                ticket_payout_amount = request.POST.get('ticket_payout_amount')
            if 'ticket_order_of_pay' in request.POST:
                ticket_order_of_pay = request.POST.get('ticket_order_of_pay')
            if 'ticket_attention' in request.POST:
                ticket_attention = request.POST.get('ticket_attention')
            if 'ticket_company' in request.POST:
                ticket_company = request.POST.get('ticket_company')
            if 'ticket_address' in request.POST:
                ticket_address = request.POST.get('ticket_address')
            if 'notes_on_check' in request.POST:
                notes_on_check = request.POST.get('notes_on_check')
            if 'check_number' in request.POST:
                check_number = request.POST.get('check_number')
            if 'check_approved_by' in request.POST:
                check_approved_by = request.POST.get('check_approved_by')
            if 'corr_cont_actions' in request.POST:
                corr_cont_actions = request.POST.get('corr_cont_actions')
            if 'ticket_root_cause' in request.POST:
                ticket_root_cause = request.POST.get('ticket_root_cause')
            if 'corrective_action' in request.POST:
                corrective_action = request.POST.get('corrective_action')
            if 'labour_hours' in request.POST:
                labour_hours = request.POST.get('labour_hours')
            if 'assign_to' in request.POST and request.POST['assign_to']:
                assign_to = request.POST.get('assign_to')

            obj = Ticket(
                submitted_date= submit_date,
                submitted_time=submit_time,
                ticket_org_id=org_id,
                ticket_caller_id=caller_id,
                ticket_client_id=client_id,
                ticket_clientinformation_id=clientinfo_id,
                ticket_record_locator=record_locator,
                ticket_caller_name=caller_name,
                ticket_caller_phone=caller_phone,
                ticket_caller_email=caller_email,
                ticket_passenger_name=passenger_name,
                subject=subject,
                description=description,
                ticket_type_id=ticket_type,
                ticket_subtype1_id=subtype1,
                ticket_subtype2_id=subtype2,
                ticket_subtype3_id=subtype3,
                ticket_subtype4_id=subtype4,
                priority_id=priority,
                ticket_status=0,
                ticket_sub_status_id=ticket_substatus,
                is_traveler_vip=traveler_vip,
                is_payout_required=payout_required,
                agent_error_goodwill=error_goodwill,
                amount_saved=amount_saved,
                airline_ticket_no=airline_ticket,
                agent_responsible=agent_responsible,
                vendor_responsible=vendor_responsible,
                vresponsible_city=vresponsible_city,
                ticket_payout_amount=ticket_payout_amount,
                ticket_order_of_pay=ticket_order_of_pay,
                ticket_attention=ticket_attention,
                ticket_company=ticket_company,
                ticket_address=ticket_address,
                notes_on_check=notes_on_check,
                check_number=check_number,
                check_approved_by=check_approved_by,
                corr_cont_actions=corr_cont_actions,
                ticket_root_cause=ticket_root_cause,
                corrective_action=corrective_action,
                total_labour_hours=labour_hours,
                ticket_created_by_id=request.user.id,
                ticket_is_open_by_id=request.user.id,
            )
            obj.save()
            insert_id = Ticket.objects.latest('pk').ticket_id
            TObj = Ticket.objects.get(pk=insert_id)

            # Assigned Object Insertion
            if 'assign_to' in request.POST and request.POST['assign_to']:
                try:
                    assign_to = request.POST.get('assign_to')
                    TObj.ticket_assign_to_id = assign_to
                    TObj.ticket_assign_by_id = request.user.id
                    TObj.ticket_assign_at = datetime.now(timezone.utc)
                    TObj.save()
                except IntegrityError:
                    transaction.rollback()
                    messages.error(request, 'Request Failed! Ticket cannot be submitted.Please try again.')
                    return redirect("/Home_SubmitTicket")

            # Next Action Object Insertion
            if 'next_action' in request.POST and request.POST['next_action']:
                try:
                    next_action = request.POST.get('next_action')
                    TObj.ticket_next_action_id = next_action
                    TObj.ticket_next_action_by_id = request.user.id
                    TObj.ticket_next_action_at = datetime.now(timezone.utc)
                    TObj.save()
                except IntegrityError:
                    transaction.rollback()
                    messages.error(request, 'Request Failed! Ticket cannot be submitted.Please try again.')
                    return redirect("/Home_SubmitTicket")

            # return HttpResponse(TObj.ticket_id)
            if 'ticket_note' in request.POST and request.POST.get('ticket_note'):
                try:
                    ticket_note = request.POST.get('ticket_note')
                    if 'is_private' in request.POST:
                        is_private = 'True'
                    else:
                        is_private = 'False'

                    obj1 = TicketNote(
                        note_detail= ticket_note,
                        note_ticket_id= TObj.ticket_id,
                        labour_hours=labour_hours,
                        is_private = is_private,
                        note_created_by_id =request.user.id
                    )
                    obj1.save()
                except IntegrityError:
                    transaction.rollback()
                    messages.error(request, 'Request Failed! Ticket cannot be submitted.Please try again.')
                    return redirect("/Home_SubmitTicket")

            if 'assign_to' in request.POST and request.POST['assign_to']:
                try:
                    assign_to = request.POST.get('assign_to')
                    obj1 = TicketUserRoleLog(
                        urlog_ticket_id=TObj.ticket_id,
                        urlog_user_id=assign_to,
                        urlog_event=1,
                        urlog_created_by_id=request.user.id
                    )
                    obj1.save()
                except IntegrityError:
                    transaction.rollback()
                    messages.error(request, 'Request Failed! Ticket cannot be submitted.Please try again.')
                    return redirect("/Home_SubmitTicket")

            if 'next_action' in request.POST and request.POST['next_action']:
                try:
                    next_action = request.POST.get('next_action')
                    obj1 = TicketUserRoleLog(
                        urlog_ticket_id=TObj.ticket_id,
                        urlog_user_id=next_action,
                        urlog_event=2,
                        urlog_created_by_id=request.user.id
                    )
                    obj1.save()
                except IntegrityError:
                    transaction.rollback()
                    messages.error(request, 'Request Failed! Ticket cannot be submitted.Please try again.')
                    return redirect("/Home_SubmitTicket")


            if 'tgroup_id' in request.POST and request.POST['tgroup_id']:
                try:
                    tgroup_id = request.POST.get('tgroup_id')
                    groupTasks = TaskManager.objects.filter(tmgr_group_id=tgroup_id)
                    for task in groupTasks:
                        TMobj = TaskManager(tmgr_ticket_id=TObj.ticket_id, tmgr_task_id=task.tmgr_task_id, task_note=task.task_note,
                                            task_assigned_to_id=task.task_assigned_to_id, task_due_date=task.task_due_date,
                                            task_dependency=task.task_dependency, task_depend_order=task.task_depend_order, task_created_by_id=request.user.id)
                        TMobj.save()
                except IntegrityError:
                    transaction.rollback()
                    messages.error(request, 'Request Failed! Ticket cannot be submitted.Please try again.')
                    return redirect("/Home_SubmitTicket")

            if 'tmanager_task_id[]' in request.POST and request.POST['tmanager_task_id[]']:
                try:
                    task_ids = request.POST.getlist('tmanager_task_id[]')
                    task_types = request.POST.getlist('tmanager_task_type[]')
                    task_notes = request.POST.getlist('tmanager_note[]')
                    task_assign_tos = request.POST.getlist('tmanager_assign_to[]')
                    task_due_dates = request.POST.getlist('tmanager_due_date[]')
                    task_dependencies = request.POST.getlist('tmanager_task_dependency[]')
                    task_orders = request.POST.getlist('tmanager_task_order[]')
                    task_dependency_orders = request.POST.getlist('tmanager_dependency_order[]')
                    for i,task_id in enumerate(task_ids):
                        if task_notes and task_notes[i]:
                            task_note = task_notes[i]
                        else:
                            task_note = None
                        if task_assign_tos and task_assign_tos[i]:
                            task_assigned_to_id = task_assign_tos[i]
                        else:
                            task_assigned_to_id = None
                        if task_due_dates and task_due_dates[i]:
                            task_due_date = datetime.strptime(task_due_dates[i], '%d/%m/%Y').strftime('%Y-%m-%d')
                        else:
                            task_due_date = None
                        if task_dependencies and task_dependencies[i]:
                            task_dependency = task_dependencies[i]
                        else:
                            task_dependency = None
                        if task_types and task_types[i]:
                            task_type = task_types[i]
                        else:
                            task_type = None
                        if task_orders and task_orders[i]:
                            task_order = task_orders[i]
                        else:
                            task_order = None
                        if task_dependency_orders and task_dependency_orders[i]:
                            task_depend_order = task_dependency_orders[i]
                        else:
                            task_depend_order = None

                        TMobj = TaskManager(tmgr_ticket_id=TObj.ticket_id, tmgr_task_id=task_id, task_note=task_note, tmgr_display_order=task_order,
                                            task_assigned_to_id=task_assigned_to_id, task_due_date=task_due_date, task_dependency=task_dependency,
                                            task_type=task_type, task_depend_order=task_depend_order, task_created_by_id=request.user.id)
                        TMobj.save()
                except IntegrityError:
                    transaction.rollback()
                    messages.error(request, 'Request Failed! Ticket cannot be submitted.Please try again.')
                    return redirect("/Home_SubmitTicket")


            if 'task_id' in request.POST and request.POST['task_id']:
                try:
                    task_id = request.POST.get('task_id')
                    if 'task_note' in request.POST:
                        task_note = request.POST.get('task_note')
                    if 'assign_to' in request.POST:
                        task_assigned_to_id = request.POST.get('assign_to')
                    if 'due_date' in request.POST:
                        task_due_date = datetime.strptime(request.POST.get('due_date'), '%m/%d/%Y').strftime(
                            '%Y-%m-%d')
                    if 'dependency' in request.POST:
                        task_dependency = request.POST.get('dependency')
                    if 'task_order' in request.POST:
                        task_depend_order = request.POST.get('task_order')

                    TMobj = TaskManager(tmgr_ticket_id=TObj.ticket_id, tmgr_task_id=task_id, task_note=task_note,
                                        task_assigned_to_id=task_assigned_to_id, task_due_date=task_due_date,
                                        task_dependency=task_dependency, task_depend_order=task_depend_order, task_created_by_id=request.user.id)
                    TMobj.save()
                except IntegrityError:
                    transaction.rollback()
                    messages.error(request, 'Request Failed! Ticket cannot be submitted.Please try again.')

            user_redirect = MySettings.objects.filter(m_user_id=request.user.id).first()
            if 'include_attachs' in request.POST:
                user_redirect =  'Home_AttachTicket/' + str(TObj.ticket_id)
            elif user_redirect.m_redirect_to == 'TicketView':
                rid = signing.dumps(TObj.ticket_id, salt=settings.SALT_KEY)
                user_redirect =  'Home_ViewTicket?tickID=' + str(rid)
            else:
                user_redirect = user_redirect.m_redirect_to
            messages.success(request, 'Request Succeed! Ticket added.')
            return redirect("/"+str(user_redirect))
        else:
            # return HttpResponse('Fail')
            messages.error(request, 'Request Failed! Ticket cannot be submitted.Please try again.')
            return redirect("/Home_SubmitTicket")
# Ticket  Save Request Start#




# Ticket Close Request Start#

@active_user_required
def closeTicket(request, id):
    # Instead of having an error on your server,
    # your user will get a 404 meaning that he tries to access a non existing resource.
    # data = get_object_or_404(Organization , pk = id)
    try:
        obj = Ticket.objects.get(pk=id)
    except Ticket.DoesNotExist:
        return render_to_response('itrak/page-404.html')
    # If Object Response is Empty
    if not obj:
        messages.success(request, 'Request Failed! No Record Found.')
        return redirect("/Home_MyTicket")
    else:
        obj.ticket_status = 1
        obj.ticket_is_close = 1
        obj.ticket_closed_by_id = request.user.id
        obj.ticket_closed_at = datetime.now()
        obj.save()
        messages.success(request, 'Request Success! Ticket Closed.')
        return redirect("/Home_MyTicket")

# Ticket Close Request End#



# Ticket Attachments Request Start#

@active_user_required
def attachTicket(request, id):

    # Instead of having an error on your server,
    # your user will get a 404 meaning that he tries to access a non existing resource.
    # data = get_object_or_404(Organization , pk = id)
    try:
        data = Ticket.objects.get(pk=id)
    except Ticket.DoesNotExist:
        return render_to_response('itrak/page-404.html')
    # If Object Response is Empty
    if not data:
        messages.success(request, 'Request Failed! No Record Found.')
        return redirect("/Home_MyTicket")
    else:
        attachments = get_ticket_attachments(request, id)
        load_sidebar = get_sidebar(request)

        context = {
            'sidebar': load_sidebar,
            'data': data,
            'attachments': attachments
        }
        return render(request, 'itrak/Ticket/ticket_attachs.html', context)

# Ticket Attachments Request End#



# Ticket Save Request Start#

@active_user_required
def saveTAttachs(request):
    if request.method == 'POST':
        if 'ticket_id' in request.POST:
            ticket_id = request.POST.get('ticket_id')
        if 'file1' in request.FILES and request.FILES['file1'] or 'file2' in request.FILES and request.FILES['file2'] or 'file3' in request.FILES and request.FILES['file3']:
            if 'file1' in request.FILES and request.FILES['file1']:
                file1 = request.FILES['file1']
                size = file1.size/1000
                if size > 1000:
                    file_size = str(round((size/1000), 2)) + 'MB'
                else:
                    file_size = str(round(size, 2)) + 'KB'
                newattach = TicketAttachments(attach_ticket_id=ticket_id, file_path=file1, file_name=file1, file_size=file_size, attach_created_by_id=request.user.id)
                newattach.save()
            if 'file2' in request.FILES and request.FILES['file2']:
                file2 = request.FILES['file2']
                size = file2.size / 1000
                if size > 1000:
                    file_size = str(round((size/1000), 2)) + 'MB'
                else:
                    file_size = str(round(size, 2)) + 'KB'
                newattach = TicketAttachments(attach_ticket_id=ticket_id, file_path=file2, file_name=file2, file_size=file_size, attach_created_by_id=request.user.id)
                newattach.save()
            if 'file3' in request.FILES and request.FILES['file3']:
                file3 = request.FILES['file3']
                size = file3.size / 1000
                if size > 1000:
                    file_size = str(round((size/1000), 2)) + 'MB'
                else:
                    file_size = str(round(size, 2)) + 'KB'
                newattach = TicketAttachments(attach_ticket_id=ticket_id, file_path=file3, file_name=file3, file_size=file_size, attach_created_by_id=request.user.id)
                newattach.save()
            messages.success(request, 'Request Succeed! Attachment added.')

        # fs = FileSystemStorage()
        # filename = fs.save(myfile.name, myfile)
        # uploaded_file_url = fs.url(filename)

        files = request.FILES.getlist('attach_files')
        if files:
            for file in files:
                size = file.size / 1000
                if size > 1000:
                    file_size = str(round((size / 1000), 2)) + 'MB'
                else:
                    file_size = str(round(size, 2)) + 'KB'
                newattach = TicketAttachments(attach_ticket_id=ticket_id, file_path=file, file_name=file, file_size=file_size, attach_created_by_id=request.user.id)
                newattach.save()
            messages.success(request, 'Request Succeed! Attachments updated.')
        try:
            ticket_id = signing.dumps(ticket_id, salt=settings.SALT_KEY)
        except Ticket.DoesNotExist:
            return render_to_response('itrak/page-404.html')
        return redirect(viewTicket, id=ticket_id)
    else:
        # return HttpResponse('Fail')
        messages.error(request, 'Request Failed! Attachment cannot be submitted.Please try again.')
        return redirect("/Home_MyTicket")
# Ticket  Save Request Start#



# Ticket Attachment Delete Request Start#

@active_user_required
def deleteAttach(request, id):
    # Instead of having an error on your server,
    # your user will get a 404 meaning that he tries to access a non existing resource.
    # data = get_object_or_404(Organization , pk = id)
    try:
        obj = TicketAttachments.objects.get(pk=id)
    except TicketAttachments.DoesNotExist:
        return render_to_response('itrak/page-404.html')
    # If Object Response is Empty
    if not obj:
        messages.success(request, 'Request Failed! No Record Found.')
        return redirect("/Home_AttachTicket/"+id)
    else:
        obj.attach_is_delete = 1
        obj.attach_modified_by_id = request.user.id
        obj.save()
        messages.success(request, 'Request Success! Attachment deleted.')
        return redirect("/Home_AttachTicket/"+str(obj.attach_ticket_id))

# Ticket Attachment Delete Request End#



# Ticket List Request Start#

@active_user_required
def myTickets(request):
    load_sidebar = get_sidebar(request)

    user_id = request.user.id # Get user_id from request
    kwargs = Q(ticket_assign_to_id=user_id) | Q(ticket_caller_id=user_id) | Q(ticket_next_action_id=user_id) | Q(ticket_created_by_id=user_id)
    summary = Ticket.objects.filter(ticket_is_delete=0).filter(kwargs)
    kwargs = Q(ticket_caller_id=user_id)
    submitter = Ticket.objects.filter(ticket_is_delete=0).filter(kwargs)
    kwargs = Q(ticket_created_by_id=user_id)
    enterer = Ticket.objects.filter(ticket_is_delete=0).filter(kwargs)
    kwargs = Q(ticket_assign_to_id=user_id)
    assignee = Ticket.objects.filter(ticket_is_delete=0).filter(kwargs)
    kwargs = Q(ticket_assign_to_id=user_id) | Q(ticket_next_action_id=user_id)
    assignee_next = Ticket.objects.filter(ticket_is_delete=0).filter(kwargs)
    kwargs = Q(ticket_next_action_id=user_id)
    next_action = Ticket.objects.filter(ticket_is_delete=0).filter(kwargs)
    kwargs = Q(ticket_assign_to_id__isnull=True)
    unassign = Ticket.objects.filter(ticket_is_delete=0).filter(kwargs)
    ticket_lists = get_task_mgr_ticket_list(user_id)
    task_assignee = Ticket.objects.filter(ticket_is_delete=0).filter(ticket_id__in=ticket_lists)
    ticket_lists = get_task_mgr_ticket_list(user_id)
    task_available = Ticket.objects.filter(ticket_is_delete=0).filter(ticket_id__in=ticket_lists)
    tickets_epanded = MySettings.objects.filter(m_user_id=request.user.id).first()
    context = {
        'sidebar': load_sidebar,
        'summary': summary,
        'submitter': submitter,
        'enterer': enterer,
        'assignee': assignee,
        'assignee_next': assignee_next,
        'next_action': next_action,
        'unassign': unassign,
        'task_assignee': task_assignee,
        'task_available': task_available,
        'tickets_epanded': tickets_epanded.m_ticket_screen,
    }
    return render(request, 'itrak/Ticket/ticket_list.html', context)

# Ticket List Request End#



# Ticket View Request Start#

@active_user_required
def viewTicket(request, id):

    # Instead of having an error on your server,
    # your user will get a 404 meaning that he tries to access a non existing resource.
    # data = get_object_or_404(Organization , pk = id)
    try:
        ticket_id = signing.loads(id, salt=settings.SALT_KEY)
        data = Ticket.objects.get(pk=ticket_id)
    except Ticket.DoesNotExist:
        return render_to_response('itrak/page-404.html')
    # If Object Response is Empty
    if not data:
        messages.success(request, 'Request Failed! No Record Found.')
        return redirect("/Home_MyTicket")
    else:
        load_sidebar = get_sidebar(request)
        assigned_logs = get_ticket_event_logs(request, ticket_id, 1)
        next_action_logs = get_ticket_event_logs(request, ticket_id, 2)
        ticketNotes = TicketNote.objects.filter(note_ticket_id=ticket_id, note_is_delete=0)
        users = User.objects.filter(is_delete=0)
        tasks = Task.objects.filter(task_is_active=1).filter(task_is_delete=0).order_by(F('task_display_order').asc(nulls_last=True))
        taskGroups = TaskGroup.objects.filter(taskgroup_is_active=1).filter(taskgroup_is_delete=0).order_by(F('taskgroup_display_order').asc(nulls_last=True))
        substatus = SubStatus.objects.filter(sstatus_is_delete=0)
        task_manager = TaskManager.objects.filter(tmgr_ticket_id=ticket_id).order_by('tmgr_task__task_display_order')
        attachments = get_ticket_attachments(request, ticket_id)

        # Calculate Total Labour Hours
        total_hours = 0
        total_minutes = 0
        if ticketNotes:
            for note in ticketNotes:
                if note.labour_hours is not None:
                    total_hours += note.labour_hours.hour
                    total_minutes += note.labour_hours.minute
        total_labour_hours = str(total_hours)+":"+str(total_minutes)

        # Calculate Total Time Open
        total_time_open = 0
        if data.ticket_status == 0:
            delta = datetime.now(timezone.utc) - data.ticket_created_at
            days, seconds = delta.days, delta.seconds
            hours = days * 24 + seconds // 3600
            minutes = (seconds % 3600) // 60
        else:
            delta = data.ticket_closed_at - data.ticket_created_at
            days, seconds = delta.days, delta.seconds
            hours = days * 24 + seconds // 3600
            minutes = (seconds % 3600) // 60

        total_time_open = round((hours) + (minutes / 100), 2)

        context = {
            'sidebar': load_sidebar,
            'data': data,
            'assigned_logs': assigned_logs,
            'next_action_logs': next_action_logs,
            'total_labour_hours': total_labour_hours,
            'total_time_open': total_time_open,
            'ticketNotes': ticketNotes,
            'tasks': tasks,
            'taskGroups': taskGroups,
            'substatus': substatus,
            'users': users,
            'task_manager': task_manager,
            'attachments': attachments
        }
        return render(request, 'itrak/Ticket/ticket_view.html', context)

# Ticket View Request End#


# Ticket Edit Request Start#

@active_user_required
def editTicket(request):
    id = request.GET.get('ticketNbr', -1)
    if 'mode' in request.GET:
        mode = request.GET.get('mode', -1)
    # Instead of having an error on your server,
    # your user will get a 404 meaning that he tries to access a non existing resource.
    # data = get_object_or_404(Organization , pk = id)
    try:
        ticket_id = signing.loads(id, salt=settings.SALT_KEY)
        data = Ticket.objects.get(pk=ticket_id)
    except Ticket.DoesNotExist:
        return render_to_response('itrak/page-404.html')

    # If Object Response is Empty
    if not data:
        messages.success(request, 'Request Failed! No Record Found.')
        return redirect("/Home_MyTicket")
    else:
        organizations = Organization.objects.filter(org_is_delete=0)
        clients = Client.objects.filter(cl_is_delete=0)
        clientInfos = ClientInformation.objects.filter(clientinfo_is_delete=0)
        ticketTypes = TicketType.objects.filter(has_parent=0).filter(ttype_is_delete=0)
        TicketSubTypes1 = TicketType.objects.filter(parent_id=data.ticket_type_id).filter(ttype_is_delete=0)
        TicketSubTypes2 = TicketType.objects.filter(parent_id=data.ticket_subtype1_id).filter(ttype_is_delete=0)
        TicketSubTypes3 = TicketType.objects.filter(parent_id=data.ticket_subtype2_id).filter(ttype_is_delete=0)
        TicketSubTypes4 = TicketType.objects.filter(parent_id=data.ticket_subtype3_id).filter(ttype_is_delete=0)
        priorities = Priority.objects.filter(prior_is_delete=0)
        substatus = SubStatus.objects.filter(sstatus_is_delete=0)
        users = User.objects.filter(is_delete=0)
        ticketNotes = TicketNote.objects.filter(note_ticket_id=ticket_id, note_is_delete=0)
        tasks = Task.objects.filter(task_is_active=1).filter(task_is_delete=0).order_by(F('task_display_order').asc(nulls_last=True))
        taskGroups = TaskGroup.objects.filter(taskgroup_is_active=1).filter(taskgroup_is_delete=0).exclude(taskgroup_id=ticket_id).order_by(F('taskgroup_display_order').asc(nulls_last=True))
        task_manager = TaskManager.objects.filter(tmgr_ticket_id=ticket_id).order_by('tmgr_task__task_display_order')

        # Calculate Total Labour Hours
        total_hours = 0
        total_minutes = 0
        if ticketNotes:
            for note in ticketNotes:
                if note.labour_hours is not None:
                    total_hours += note.labour_hours.hour
                    total_minutes += note.labour_hours.minute
        total_labour_hours = str(total_hours) + ":" + str(total_minutes)

        # Calculate Total Time Open
        total_time_open = 0
        if data.ticket_status == 0:
            delta = datetime.now(timezone.utc) - data.ticket_created_at
            days, seconds = delta.days, delta.seconds
            hours = days * 24 + seconds // 3600
            minutes = (seconds % 3600) // 60
        else:
            delta = data.ticket_closed_at - data.ticket_created_at
            days, seconds = delta.days, delta.seconds
            hours = days * 24 + seconds // 3600
            minutes = (seconds % 3600) // 60

        total_time_open = round((hours) + (minutes / 100), 2)

        load_sidebar = get_sidebar(request)
        context = {
            'sidebar': load_sidebar,
            'organizations': organizations,
            'users': users,
            'clients': clients,
            'clientinformations': clientInfos,
            'ticketTypes': ticketTypes,
            'TicketSubTypes1': TicketSubTypes1,
            'TicketSubTypes2': TicketSubTypes2,
            'TicketSubTypes3': TicketSubTypes3,
            'TicketSubTypes4': TicketSubTypes4,
            'priorities': priorities,
            'substatus': substatus,
            'total_labour_hours': total_labour_hours,
            'total_time_open': total_time_open,
            'ticketNotes': ticketNotes,
            'tasks': tasks,
            'taskGroups': taskGroups,
            'task_manager': task_manager,
            'data': data,
            'mode': mode
        }
        return render(request, 'itrak/Ticket/ticket_edit.html', context)

# Ticket Edit Request End#

# Ticket Clone Request Start#

@active_user_required
def cloneTicket(request):
    id = request.GET.get('ticketNbr', -1)
    if 'mode' in request.GET:
        mode = request.GET.get('mode', -1)
    # Instead of having an error on your server,
    # your user will get a 404 meaning that he tries to access a non existing resource.
    # data = get_object_or_404(Organization , pk = id)
    try:
        ticket_id = signing.loads(id, salt=settings.SALT_KEY)
        data = Ticket.objects.get(pk=ticket_id)
    except Ticket.DoesNotExist:
        return render_to_response('itrak/page-404.html')

    # If Object Response is Empty
    if not data:
        messages.success(request, 'Request Failed! No Record Found.')
        return redirect("/Home_MyTicket")
    else:
        organizations = Organization.objects.filter(org_is_delete=0)
        clients = Client.objects.filter(cl_is_delete=0)
        clientInfos = ClientInformation.objects.filter(clientinfo_is_delete=0)
        ticketTypes = TicketType.objects.filter(has_parent=0).filter(ttype_is_delete=0)
        TicketSubTypes1 = TicketType.objects.filter(parent_id=data.ticket_type_id).filter(ttype_is_delete=0)
        TicketSubTypes2 = TicketType.objects.filter(parent_id=data.ticket_subtype1_id).filter(ttype_is_delete=0)
        TicketSubTypes3 = TicketType.objects.filter(parent_id=data.ticket_subtype2_id).filter(ttype_is_delete=0)
        TicketSubTypes4 = TicketType.objects.filter(parent_id=data.ticket_subtype3_id).filter(ttype_is_delete=0)
        priorities = Priority.objects.filter(prior_is_delete=0)
        substatus = SubStatus.objects.filter(sstatus_is_delete=0)
        users = User.objects.filter(is_delete=0)
        ticketNotes = TicketNote.objects.filter(note_ticket_id=ticket_id, note_is_delete=0)
        tasks = Task.objects.filter(task_is_active=1).filter(task_is_delete=0).order_by(F('task_display_order').asc(nulls_last=True))
        taskGroups = TaskGroup.objects.filter(taskgroup_is_active=1).filter(taskgroup_is_delete=0).exclude(taskgroup_id=ticket_id).order_by(F('taskgroup_display_order').asc(nulls_last=True))
        task_manager = TaskManager.objects.filter(tmgr_ticket_id=ticket_id).order_by('tmgr_task__task_display_order')

        load_sidebar = get_sidebar(request)
        context = {
            'sidebar': load_sidebar,
            'organizations': organizations,
            'users': users,
            'clients': clients,
            'clientinformations': clientInfos,
            'ticketTypes': ticketTypes,
            'TicketSubTypes1': TicketSubTypes1,
            'TicketSubTypes2': TicketSubTypes2,
            'TicketSubTypes3': TicketSubTypes3,
            'TicketSubTypes4': TicketSubTypes4,
            'priorities': priorities,
            'substatus': substatus,
            'ticketNotes': ticketNotes,
            'tasks': tasks,
            'taskGroups': taskGroups,
            'task_manager': task_manager,
            'data': data,
            'mode': mode
        }
        # return HttpResponse(data.ticket_caller_id)
        return render(request, 'itrak/Ticket/clone_ticket.html', context)

# Ticket Clone Request End#


#Ticket Update Request Start
@active_user_required
def updateTicket(request):
    if request.method == 'POST':
        id = request.POST.get('ticket_id')
        encrypted_id = signing.dumps(id, salt=settings.SALT_KEY)
        # Instead of having an error on your server,
        # your user will get a 404 meaning that he tries to access a non existing resource.
        # data = get_object_or_404(Organization , pk = id)
        try:
            obj = Ticket.objects.get(pk=id)
        except Ticket.DoesNotExist:
            return render_to_response('itrak/page-404.html')

        if obj.ticket_assign_to_id:
            #Save Assingee/Next Actions Log Start#
            if 'assign_to' in request.POST and request.POST['assign_to'] and int(obj.ticket_assign_to_id) != int(request.POST.get('assign_to')):
                assign_to = request.POST.get('assign_to')
                # obj.ticket_assign_by = request.user.id
                obj1 = TicketUserRoleLog(
                    urlog_ticket_id=id,
                    urlog_user_id=assign_to,
                    urlog_event=1,
                    urlog_created_by_id=request.user.id
                )
                obj1.save()
        else:
            #Save Assingee/Next Actions Log Start#
            if 'assign_to' in request.POST and request.POST['assign_to']:

                assign_to = request.POST.get('assign_to')
                # obj.ticket_assign_by = request.user.id
                obj1 = TicketUserRoleLog(
                    urlog_ticket_id=id,
                    urlog_user_id=assign_to,
                    urlog_event=1,
                    urlog_created_by_id=request.user.id
                )
                obj1.save()

        if obj.ticket_next_action_id:
            if 'next_action' in request.POST and request.POST['next_action'] and int(obj.ticket_next_action_id) != int(request.POST.get('next_action')):
                next_action = request.POST.get('next_action')
                obj1 = TicketUserRoleLog(
                    urlog_ticket_id=id,
                    urlog_user_id=next_action,
                    urlog_event=2,
                    urlog_created_by_id=request.user.id
                )
                obj1.save()
            #Save Assingee/Next Actions Log End#
        else:
            if 'next_action' in request.POST and request.POST['next_action']:
                next_action = request.POST.get('next_action')
                obj1 = TicketUserRoleLog(
                    urlog_ticket_id=id,
                    urlog_user_id=next_action,
                    urlog_event=2,
                    urlog_created_by_id=request.user.id
                )
                obj1.save()
            #Save Assingee/Next Actions Log End#

        # If Object Response is Empty
        if not obj:
            messages.success(request, 'Request Failed! No Record Found.')
            return redirect("/Home_MyTicket")
        else:
            if 'submit_date' in request.POST and request.POST['submit_date']:
                obj.submitted_date = datetime.strptime(request.POST.get('submit_date'), '%m-%d-%Y').strftime(
                    '%Y-%m-%d')
            if 'submit_time' in request.POST:
                obj.submitted_time = request.POST.get('submit_time')
            if 'org_id' in request.POST:
                obj.ticket_org_id = request.POST.get('org_id')
            if 'caller_id' in request.POST:
                obj.ticket_caller_id = request.POST.get('caller_id')
            if 'client_id' in request.POST and request.POST['client_id']:
                obj.ticket_client_id = request.POST.get('client_id')
            if 'clientinfo_id' in request.POST:
                obj.ticket_clientinformation_id = request.POST.get('clientinfo_id')
            if 'record_locator' in request.POST:
                obj.ticket_record_locator = request.POST.get('record_locator')
            if 'caller_name' in request.POST:
                obj.ticket_caller_name = request.POST.get('caller_name')
            if 'caller_phone' in request.POST:
                obj.ticket_caller_phone = request.POST.get('caller_phone')
            if 'caller_email' in request.POST:
                obj.ticket_caller_email = request.POST.get('caller_email')
            if 'passenger_name' in request.POST:
                obj.ticket_passenger_name = request.POST.get('passenger_name')
            if 'subject' in request.POST and request.POST['subject']:
                obj.subject = request.POST.get('subject')
            if 'description' in request.POST and request.POST['description']:
                obj.description = request.POST.get('description')
            if 'ticket_type' in request.POST:
                obj.ticket_type_id = request.POST.get('ticket_type')
            if 'subtype1' in request.POST:
                obj.ticket_subtype1_id = request.POST.get('subtype1')
            if 'subtype2' in request.POST:
                obj.ticket_subtype2_id = request.POST.get('subtype2')
            if 'subtype3' in request.POST:
                obj.ticket_subtype3_id = request.POST.get('subtype3')
            if 'subtype4' in request.POST:
                obj.ticket_subtype4_id = request.POST.get('subtype4')
            if 'priority' in request.POST:
                obj.priority_id = request.POST.get('priority')
            if 'ticket_status' in request.POST and obj.ticket_status != request.POST.get('ticket_status'):
                if request.POST.get('ticket_status') == '0':
                    obj.ticket_is_reopen = 1
                    obj.ticket_status = request.POST.get('ticket_status')
                    obj.ticket_is_reopen_at = datetime.now()
                    obj.ticket_is_reopen_by_id = request.user.id
                else:
                    obj.ticket_is_close = 1
                    obj.ticket_status = request.POST.get('ticket_status')
                    obj.ticket_is_close_at = datetime.now()
                    obj.ticket_is_close_by_id = request.user.id
            if 'ticket_substatus' in request.POST:
                obj.ticket_sub_status_id = request.POST.get('ticket_substatus')
            if 'traveler_vip' in request.POST:
                obj.is_traveler_vip = request.POST.get('traveler_vip')
            if 'payout_required' in request.POST:
                obj.is_payout_required = request.POST.get('payout_required')
            if 'error_goodwill' in request.POST:
                obj.agent_error_goodwill = request.POST.get('error_goodwill')
            if 'amount_saved' in request.POST:
                obj.amount_saved = request.POST.get('amount_saved')
            if 'airline_ticket' in request.POST:
                obj.airline_ticket_no = request.POST.get('airline_ticket')
            if 'agent_responsible' in request.POST:
                obj.agent_responsible = request.POST.get('agent_responsible')
            if 'vendor_responsible' in request.POST:
                obj.vendor_responsible = request.POST.get('vendor_responsible')
            if 'vresponsible_city' in request.POST:
                obj.vresponsible_city = request.POST.get('vresponsible_city')
            if 'ticket_payout_amount' in request.POST:
                obj.ticket_payout_amount = request.POST.get('ticket_payout_amount')
            if 'ticket_order_of_pay' in request.POST:
                obj.ticket_order_of_pay = request.POST.get('ticket_order_of_pay')
            if 'ticket_attention' in request.POST:
                obj.ticket_attention = request.POST.get('ticket_attention')
            if 'ticket_company' in request.POST:
                obj.ticket_company = request.POST.get('ticket_company')
            if 'ticket_address' in request.POST:
                obj.ticket_address = request.POST.get('ticket_address')
            if 'notes_on_check' in request.POST:
                obj.notes_on_check = request.POST.get('notes_on_check')
            if 'check_number' in request.POST:
                obj.check_number = request.POST.get('check_number')
            if 'check_approved_by' in request.POST:
                obj.check_approved_by = request.POST.get('check_approved_by')
            if 'corr_cont_actions' in request.POST:
                obj.corr_cont_actions = request.POST.get('corr_cont_actions')
            if 'ticket_root_cause' in request.POST:
                obj.ticket_root_cause = request.POST.get('ticket_root_cause')
            if 'corrective_action' in request.POST:
                obj.corrective_action = request.POST.get('corrective_action')
            if obj.ticket_assign_to_id:
                if 'assign_to' in request.POST and request.POST['assign_to'] and int(obj.ticket_assign_to_id) != int(request.POST.get('assign_to')):
                    obj.ticket_assign_to_id = request.POST.get('assign_to')
                    obj.ticket_assign_by_id = request.user.id
                    obj.ticket_assign_at = datetime.now(timezone.utc)
                elif 'assign_to' in request.POST and request.POST.get('assign_to') == '':
                    obj.ticket_assign_to_id = None
                    obj.ticket_assign_by_id = None
                    obj.ticket_assign_at = None
            else:
                if 'assign_to' in request.POST and request.POST['assign_to']:
                    obj.ticket_assign_to_id = request.POST.get('assign_to')
                    obj.ticket_assign_by_id = request.user.id
                    obj.ticket_assign_at = datetime.now(timezone.utc)
                elif 'assign_to' in request.POST and request.POST.get('assign_to') == '':
                    obj.ticket_assign_to_id = None
                    obj.ticket_assign_by_id = None
                    obj.ticket_assign_at = None

            if obj.ticket_next_action_id:
                if 'next_action' in request.POST and request.POST['next_action'] and int(obj.ticket_next_action_id) != int(request.POST.get('next_action')):
                    obj.ticket_next_action_id = request.POST.get('next_action')
                    obj.ticket_next_action_by_id = request.user.id
                    obj.ticket_next_action_at = datetime.now(timezone.utc)
                elif 'next_action' in request.POST and request.POST.get('next_action') == '':
                    obj.ticket_next_action_id = None
                    obj.ticket_next_action_by_id = None
                    obj.ticket_next_action_at = None
            else:
                if 'next_action' in request.POST and request.POST['next_action']:
                    obj.ticket_next_action_id = request.POST.get('next_action')
                    obj.ticket_next_action_by_id = request.user.id
                    obj.ticket_next_action_at = datetime.now(timezone.utc)
                elif 'next_action' in request.POST and request.POST.get('next_action') == '':
                    obj.ticket_next_action_id = None
                    obj.ticket_next_action_by_id = None
                    obj.ticket_next_action_at = None

            if 'labour_hours' in request.POST:
                t1 = obj.total_labour_hours
                if t1:
                    t1 = datetime.strptime(str(t1), '%H:%M:%S')
                    t2 = datetime.strptime(request.POST.get('labour_hours'), '%H:%M')
                    time_zero = datetime.strptime('00:00:00', '%H:%M:%S')
                    obj.total_labour_hours = (t1 - time_zero + t2).time()
                else:
                    t2 = datetime.strptime(request.POST.get('labour_hours'), '%H:%M')
                    obj.total_labour_hours = t2

            obj.save()

            if 'ticket_note' in request.POST and request.POST.get('ticket_note') and request.POST.get('ticket_note') != '<p><br></p>':
                ticket_note = request.POST.get('ticket_note')
                if 'is_private' in request.POST:
                    is_private = 'True'
                else:
                    is_private = 'False'
                if 'labour_hours' in request.POST:
                    labour_hours = request.POST.get('labour_hours')

                obj1 = TicketNote(
                    note_detail=ticket_note,
                    note_ticket_id=id,
                    labour_hours=labour_hours,
                    is_private=is_private,
                    note_created_by_id=request.user.id
                )
                obj1.save()

            if 'tgroup_id' in request.POST and request.POST['tgroup_id']:
                tgroup_id = request.POST.get('tgroup_id')
                groupTasks = TaskManager.objects.filter(tmgr_group_id=tgroup_id)
                for task in groupTasks:
                    TMobj = TaskManager(tmgr_ticket_id=id, tmgr_task_id=task.tmgr_task_id,
                                        task_note=task.task_note,
                                        task_assigned_to_id=task.task_assigned_to_id, task_due_date=task.task_due_date,
                                        task_dependency=task.task_dependency, task_depend_order=task.task_depend_order, task_created_by_id=request.user.id)
                    TMobj.save()

            TaskManager.objects.filter(tmgr_ticket_id=id).delete()
            if 'tmanager_task_id[]' in request.POST and request.POST['tmanager_task_id[]']:
                task_ids = request.POST.getlist('tmanager_task_id[]')
                task_orders = request.POST.getlist('tmanager_task_order[]')
                task_types = request.POST.getlist('tmanager_task_type[]')
                task_notes = request.POST.getlist('tmanager_note[]')
                task_assign_tos = request.POST.getlist('tmanager_assign_to[]')
                task_due_dates = request.POST.getlist('tmanager_due_date[]')
                task_dependencies = request.POST.getlist('tmanager_task_dependency[]')
                task_dependency_orders = request.POST.getlist('tmanager_dependency_order[]')
                for i,task_id in enumerate(task_ids):
                    if task_notes and task_notes[i]:
                        task_note = task_notes[i]
                    else:
                        task_note = None
                    if task_assign_tos and task_assign_tos[i]:
                        task_assigned_to_id = task_assign_tos[i]
                    else:
                        task_assigned_to_id = None
                    if task_due_dates and task_due_dates[i]:
                        task_due_date = datetime.strptime(task_due_dates[i], '%d/%m/%Y').strftime('%Y-%m-%d')
                    else:
                        task_due_date = None
                    if task_dependencies and task_dependencies[i]:
                        task_dependency = task_dependencies[i]
                    else:
                        task_dependency = None
                    if task_types and task_types[i]:
                        task_type = task_types[i]
                    else:
                        task_type = None
                    if task_orders and task_orders[i]:
                        task_order = task_orders[i]
                    else:
                        task_order = None
                    if task_dependency_orders and task_dependency_orders[i]:
                        task_depend_order = task_dependency_orders[i]
                    else:
                        task_depend_order = None

                    TMobj = TaskManager(tmgr_ticket_id=id, tmgr_task_id=task_id, tmgr_display_order=task_order, task_note=task_note,
                                        task_assigned_to_id=task_assigned_to_id, task_due_date=task_due_date, task_dependency=task_dependency,
                                        task_type=task_type, task_depend_order=task_depend_order, task_created_by_id=request.user.id)
                    TMobj.save()

            if 'task_id' in request.POST and request.POST['task_id']:
                task_id = request.POST.get('task_id')
                if 'task_note' in request.POST:
                    task_note = request.POST.get('task_note')
                if 'assign_to' in request.POST:
                    task_assigned_to_id = request.POST.get('assign_to')
                if 'due_date' in request.POST:
                    task_due_date = datetime.strptime(request.POST.get('due_date'), '%m/%d/%Y').strftime('%Y-%m-%d')
                if 'dependency' in request.POST:
                    task_dependency = request.POST.get('dependency')
                if 'task_order' in request.POST:
                    task_depend_order = request.POST.get('task_order')

                TMobj = TaskManager(tmgr_ticket_id=id, tmgr_task_id=task_id, task_note=task_note,
                                    task_assigned_to_id=task_assigned_to_id, task_due_date=task_due_date,
                                    task_dependency=task_dependency, task_depend_order=task_depend_order, task_created_by_id=request.user.id)
                TMobj.save()


        # return HttpReshtponse('Success')
        messages.success(request, 'Request Succeed! Ticket updated.')
        return redirect(viewTicket, id=encrypted_id)
    else:
        messages.error(request, 'Request Failed! Ticket cannot be updated.Please try again.')
        return redirect("/Home_MyTicket")

# Ticket  Update Request End#


#Ticket Update Request Start
@active_user_required
def updateTicketTaskManager(request):
    if request.method == 'POST':
        t_id = request.POST.get('ticket_id')
        # Instead of having an error on your server,
        # your user will get a 404 meaning that he tries to access a non existing resource.
        # data = get_object_or_404(Organization , pk = id)
        try:
            ticket_id = signing.loads(t_id, salt=settings.SALT_KEY)
            obj = Ticket.objects.get(pk=ticket_id)
        except Ticket.DoesNotExist:
            return render_to_response('itrak/page-404.html')

        # If Object Response is Empty
        if not obj:
            messages.success(request, 'Request Failed! No Record Found.')
            return redirect("/Home_MyTicket")
        else:
            TaskManager.objects.filter(tmgr_ticket_id=ticket_id).delete()
            if 'tmanager_task_id[]' in request.POST and request.POST['tmanager_task_id[]']:
                task_ids = request.POST.getlist('tmanager_task_id[]')
                task_orders = request.POST.getlist('tmanager_task_order[]')
                task_types = request.POST.getlist('tmanager_task_type[]')
                task_notes = request.POST.getlist('tmanager_note[]')
                task_assign_tos = request.POST.getlist('tmanager_assign_to[]')
                task_due_dates = request.POST.getlist('tmanager_due_date[]')
                task_dependencies = request.POST.getlist('tmanager_task_dependency[]')
                task_dependency_orders = request.POST.getlist('tmanager_dependency_order[]')
                for i,task_id in enumerate(task_ids):
                    if task_notes and task_notes[i]:
                        task_note = task_notes[i]
                    else:
                        task_note = None
                    if task_assign_tos and task_assign_tos[i]:
                        task_assigned_to_id = task_assign_tos[i]
                    else:
                        task_assigned_to_id = None
                    if task_due_dates and task_due_dates[i]:
                        task_due_date = datetime.strptime(task_due_dates[i], '%d/%m/%Y').strftime('%Y-%m-%d')
                    else:
                        task_due_date = None
                    if task_dependencies and task_dependencies[i]:
                        task_dependency = task_dependencies[i]
                    else:
                        task_dependency = None
                    if task_types and task_types[i]:
                        task_type = task_types[i]
                    else:
                        task_type = None
                    if task_orders and task_orders[i]:
                        task_order = task_orders[i]
                    else:
                        task_order = None
                    if task_dependency_orders and task_dependency_orders[i]:
                        task_depend_order = task_dependency_orders[i]
                    else:
                        task_depend_order = None

                    TMobj = TaskManager(tmgr_ticket_id=ticket_id, tmgr_task_id=task_id, tmgr_display_order=task_order, task_note=task_note,
                                        task_assigned_to_id=task_assigned_to_id, task_due_date=task_due_date, task_dependency=task_dependency,
                                        task_type=task_type, task_depend_order=task_depend_order, task_created_by_id=request.user.id)
                    TMobj.save()

        # return HttpReshtponse('Success')
        messages.success(request, 'Request Succeed! Ticket Manager updated.')
        return redirect(viewTicket, id=t_id)
    else:
        messages.error(request, 'Request Failed! Ticket cannot be updated.Please try again.')
        return redirect("/Home_MyTicket")

# Ticket  Update Request End#



# Ticket Delete Request Start#

@active_user_required
def deleteTicket(request):

    id = request.GET.get('ticketNbr', -1)
    # Instead of having an error on your server,
    # your user will get a 404 meaning that he tries to access a non existing resource.
    # data = get_object_or_404(Organization , pk = id)
    try:
        ticket_id = signing.loads(id, salt=settings.SALT_KEY)
        obj = Ticket.objects.get(pk=ticket_id)
    except Ticket.DoesNotExist:
        return render_to_response('itrak/page-404.html')
    # If Object Response is Empty

    if not obj:
        messages.success(request, 'Request Failed! No Record Found.')
        return redirect("/Home_MyTicket")
    else:
        obj.ticket_is_delete = 1
        obj.save()
        messages.success(request, 'Request Success! Ticket deleted.')
        return redirect("/Home_MyTicket")

# Ticket Delete Request End#



# Ticket Note Edit Request Start#

@active_user_required
def editNoteTicket(request):
    if 'noteID' in request.GET and 'ticketNbr' in request.GET:
        note_id = request.GET.get('noteID', -1)
        ticket_id = request.GET.get('ticketNbr', -1)
        # Instead of having an error on your server,
        # your user will get a 404 meaning that he tries to access a non existing resource.
        # data = get_object_or_404(Organization , pk = id)
        try:
            note_id = signing.loads(note_id, salt=settings.SALT_KEY)
            ticket_id = signing.loads(ticket_id, salt=settings.SALT_KEY)
            data = TicketNote.objects.filter(note_ticket_id=ticket_id).filter(note_id=note_id)
        except TicketNote.DoesNotExist:
            return render_to_response('itrak/page-404.html')
        # If Object Response is Empty
        if not data:
            messages.success(request, 'Request Failed! No Record Found.')
            return redirect(viewTicket, id=ticket_id)
        else:
            substatus = SubStatus.objects.filter(sstatus_is_delete=0)

            load_sidebar = get_sidebar(request)
            context = {
                'sidebar': load_sidebar,
                'substatus': substatus,
                'data': data,
            }
            return render(request, 'itrak/Ticket/ticket_note_edit.html', context)
    else:
        return render_to_response('itrak/page-404.html')

# Task Edit Request End#



#Ticket Note Update Request Start
@active_user_required
def updateNoteTicket(request):
    if request.method == 'POST':
        t_id = request.POST.get('ticket_id')
        n_id = request.POST.get('note_id')
        # Instead of having an error on your server,
        # your user will get a 404 meaning that he tries to access a non existing resource.
        # data = get_object_or_404(Organization , pk = id)
        try:
            note_id = signing.loads(n_id, salt=settings.SALT_KEY)
            ticket_id = signing.loads(t_id, salt=settings.SALT_KEY)
            obj = TicketNote.objects.get(pk=note_id)
        except TicketNote.DoesNotExist:
            return render_to_response('itrak/page-404.html')
        # If Object Response is Empty
        if not obj:
            messages.success(request, 'Request Failed! No Record Found.')
            return redirect(viewTicket, id=t_id)
        else:
            try:
                obj1 = Ticket.objects.get(pk=ticket_id)
            except Ticket.DoesNotExist:
                return render_to_response('itrak/page-404.html')

            if 'ticket_substatus' in request.POST:
                obj1.ticket_sub_status_id = request.POST.get('ticket_substatus')

                if obj.labour_hours is not None and request.POST.get('labour_hours') is not None:
                    t1 = datetime.strptime(str(obj1.total_labour_hours), '%H:%M:%S')
                    t2 = datetime.strptime(request.POST.get('labour_hours')+":00", '%H:%M:%S')
                    time_zero = datetime.strptime(str(obj.labour_hours), '%H:%M:%S')
                    obj1.total_labour_hours = (t1 - time_zero + t2).time()
                elif obj.labour_hours is None and request.POST.get('labour_hours') is not None:
                    t1 = datetime.strptime(str(obj1.total_labour_hours), '%H:%M:%S')
                    t2 = datetime.strptime(request.POST.get('labour_hours') + ":00", '%H:%M:%S')
                    time_zero = datetime.strptime("00:00:00", '%H:%M:%S')
                    obj1.total_labour_hours = (t1 - time_zero + t2).time()
                obj1.save()

            if 'ticket_note' in request.POST and request.POST.get('ticket_note'):
                obj.note_detail = request.POST.get('ticket_note')
                if 'is_private' in request.POST:
                    obj.is_private = 'True'
                else:
                    obj.is_private = 'False'
                if 'labour_hours' in request.POST:
                    obj.labour_hours = request.POST.get('labour_hours')

                obj.note_modified_by_id=request.user.id
                obj.note_modified_at=datetime.now()

                obj.save()

        # return HttpReshtponse('Success')
        messages.success(request, 'Request Succeed! Ticket Note updated.')
        return redirect(viewTicket, id=t_id)
    else:
        messages.error(request, 'Request Failed! Ticket Note cannot be updated.Please try again.')
        return redirect(viewTicket, id=t_id)

# Ticket Note Update Request End#



# Ticket Note Delete Request Start#

@active_user_required
def deleteNoteTicket(request):
    if 'noteID' in request.GET and 'ticketNbr' in request.GET:
        n_id = request.GET.get('noteID', -1)
        t_id = request.GET.get('ticketNbr', -1)
        # Instead of having an error on your server,
        # your user will get a 404 meaning that he tries to access a non existing resource.
        # data = get_object_or_404(Organization , pk = id)
        try:
            note_id = signing.loads(n_id, salt=settings.SALT_KEY)
            obj = TicketNote.objects.get(pk=note_id)
        except TicketNote.DoesNotExist:
            return render_to_response('itrak/page-404.html')
        # If Object Response is Empty

        if not obj:
            messages.success(request, 'Request Failed! No Record Found.')
            return redirect("/Home_MyTicket")
        else:
            obj.note_is_delete = 1
            obj.save()
            messages.success(request, 'Request Success! Ticket Note deleted.')
            return redirect(viewTicket, id=t_id)
    else:
        return render_to_response('itrak/page-404.html')
    # Ticket Note Delete Request End#


# Search Ticket with Filter Request Start#

@active_user_required
def searchTicket(request):
    if 'reportNbr' in request.GET:
        report_id = request.GET.get('reportNbr', -1)
        try:
            report = TicketSavedSearch.objects.get(pk=report_id)
        except TicketSavedSearch.DoesNotExist:
            return render_to_response('itrak/page-404.html')

    # Instead of having an error on your server,
    # your user will get a 404 meaning that he tries to access a non existing resource.

    organizations = Organization.objects.filter(org_is_delete=0)
    departments = Department.objects.filter(d_is_delete=0)
    clients = Client.objects.filter(cl_is_delete=0)
    clientInfos = ClientInformation.objects.filter(clientinfo_is_delete=0)
    ticketTypes = TicketType.objects.filter(has_parent=0).filter(ttype_is_delete=0)
    priorities = Priority.objects.filter(prior_is_delete=0)
    substatus = SubStatus.objects.filter(sstatus_is_delete=0)
    users = User.objects.filter(is_delete=0)
    savedSearches = TicketSavedSearch.objects.filter()
    load_sidebar = get_sidebar(request)
    context = {
        'sidebar': load_sidebar,
        'organizations': organizations,
        'departments': departments,
        'clients': clients,
        'substatus': substatus,
        'priorities': priorities,
        'ticketTypes': ticketTypes,
        'users': users,
        'timezones': pytz.common_timezones,
        'savedSearches': savedSearches
    }
    if 'reportNbr' in request.GET:
        context['report'] = report
        context['TicketSubTypes1'] = TicketType.objects.filter(parent_id=report.ticket_type_id).filter(ttype_is_delete=0)
        context['TicketSubTypes2'] = TicketType.objects.filter(parent_id=report.ticket_subtype1_id).filter(ttype_is_delete=0)
        context['TicketSubTypes3'] = TicketType.objects.filter(parent_id=report.ticket_subtype2_id).filter(ttype_is_delete=0)
        context['TicketSubTypes4'] = TicketType.objects.filter(parent_id=report.ticket_subtype3_id).filter(ttype_is_delete=0)

        return render(request, 'itrak/Ticket/ticket_search_edit.html', context)
    else:
        return render(request, 'itrak/Ticket/ticket_search.html', context)

# Search Ticket with Filter Request End#



# Ticket Search Filter Result Start#

@active_user_required
def ticketSearchResults(request):
    if request.method == 'POST':
        kwargs = {
            '{0}__{1}'.format('ticket_is_delete', 'iexact'): 0,
            '{0}__{1}'.format('ticket_is_active', 'iexact'): 1,
        }
        ticket_status_dict = {"0": "Opened", "1": "Closed"}
        labor_hours_dict = {"0": "Less Than", "1": "More Than", "2": "Equal"}
        fielddict = {"1": ''}
        if 'ticket_status' in request.POST and request.method == 'POST' and request.POST.get('ticket_status') != '':
            ticket_status = request.POST.get('ticket_status')
            kwargs.setdefault('ticket_status__iexact', ticket_status)
            fielddict.update({'Ticket Status': ticket_status_dict[ticket_status]})
        if 'ticket_sub_status' in request.POST and request.method == 'POST' and request.POST.get('ticket_sub_status') != '':
            ticket_sub_status = request.POST.get('ticket_sub_status')
            kwargs.setdefault('ticket_sub_status_id', ticket_sub_status)
            result = SubStatus.objects.only('sub_status_text').get(pk=ticket_sub_status).sub_status_text
            if 'ticket_status' in request.POST and request.method == 'POST' and request.POST.get('ticket_status') != '':
                fielddict.update({'Ticket Status': ticket_status_dict[ticket_status] + ' - ' + result})
            else:
                fielddict.update({'Ticket Status': ' - ' + result})
        if 'priority' in request.POST and request.method == 'POST' and request.POST.get('priority') != '':
            priority = request.POST.get('priority')
            kwargs.setdefault('priority_id', priority)
            result = Priority.objects.only('priority_name').get(pk=priority).priority_name
            fielddict.update({'Priority': result})
        if 'ticket_type' in request.POST and request.method == 'POST' and request.POST.get('ticket_type') != '':
            ticket_type = request.POST.get('ticket_type')
            kwargs.setdefault('ticket_type_id', ticket_type)
            result = TicketType.objects.only('ttype_name').get(pk=ticket_type).ttype_name
            fielddict.update({'Ticket Type': result})
        if 'subtype1' in request.POST and request.method == 'POST' and request.POST.get('subtype1') != '':
            subtype1 = request.POST.get('subtype1')
            kwargs.setdefault('ticket_subtype1_id', subtype1)
            result = TicketType.objects.only('ttype_name').get(pk=subtype1).ttype_name
            fielddict.update({'Subtype 1': result})
        if 'subtype2' in request.POST and request.method == 'POST' and request.POST.get('subtype2') != '':
            subtype2 = request.POST.get('subtype2')
            kwargs.setdefault('ticket_subtype2_id', subtype2)
            result = TicketType.objects.only('ttype_name').get(pk=subtype2).ttype_name
            fielddict.update({'Subtype 2': result})
        if 'subtype3' in request.POST and request.method == 'POST' and request.POST.get('subtype3') != '':
            subtype3 = request.POST.get('subtype3')
            kwargs.setdefault('ticket_subtype3_id', subtype3)
            result = TicketType.objects.only('ttype_name').get(pk=subtype3).ttype_name
            fielddict.update({'Subtype 3': result})
        if 'subtype4' in request.POST and request.method == 'POST' and request.POST.get('subtype4') != '':
            subtype4 = request.POST.get('subtype4')
            kwargs.setdefault('ticket_subtype4_id', subtype4)
            result = TicketType.objects.only('ttype_name').get(pk=subtype4).ttype_name
            fielddict.update({'Subtype 4': result})
        if 'subject' in request.POST and request.method == 'POST' and request.POST.get('subject') != '':
            subject = request.POST.get('subject')
            kwargs.setdefault('subject__icontains', subject)
            fielddict.update({'Subject': subject})
        if 'ticket_note' in request.POST and request.method == 'POST' and request.POST.get('ticket_note') != '':
            ticket_note = request.POST.get('ticket_note')
            kwargs.setdefault('ticketNote__note_detail__icontains', ticket_note)
            fielddict.update({'Notes': ticket_note})
        if 'all_three' in request.POST and request.method == 'POST' and request.POST.get('all_three') != '':
            all_three = request.POST.get('all_three')
            args = Q(subject__icontains=all_three) | Q(ticketNote__note_detail__icontains=all_three)
            fielddict.update({'Search All Three': all_three})
        if 'record_locator' in request.POST and request.method == 'POST' and request.POST.get('record_locator') != '':
            record_locator = request.POST.get('record_locator')
            kwargs.setdefault('ticket_record_locator__startswith', record_locator)
            fielddict.update({'Record Locator': record_locator})
        if 'caller_name' in request.POST and request.method == 'POST' and request.POST.get('caller_name') != '':
            caller_name = request.POST.get('caller_name')
            kwargs.setdefault('ticket_caller_name__startswith', caller_name)
            fielddict.update({'Caller Name': caller_name})
        if 'caller_phone' in request.POST and request.method == 'POST' and request.POST.get('caller_phone') != '':
            caller_phone = request.POST.get('caller_phone')
            kwargs.setdefault('ticket_caller_phone__startswith', caller_phone)
            fielddict.update({'Caller Phone': caller_phone})
        if 'caller_email' in request.POST and request.method == 'POST' and request.POST.get('caller_email') != '':
            caller_email = request.POST.get('caller_email')
            kwargs.setdefault('ticket_caller_email__startswith', caller_email)
            fielddict.update({'Caller Email': caller_email})
        if 'passenger_name' in request.POST and request.method == 'POST' and request.POST.get('passenger_name') != '':
            passenger_name = request.POST.get('passenger_name')
            kwargs.setdefault('ticket_passenger_name__startswith', passenger_name)
            fielddict.update({'Passenger Name': passenger_name})
        if 'note_entered_by' in request.POST and request.method == 'POST' and request.POST.get('note_entered_by') != '':
            note_entered_by = request.POST.get('note_entered_by')
            kwargs.setdefault('ticketNote__note_created_by_id', note_entered_by)
            result = User.objects.only('display_name').get(pk=note_entered_by).display_name
            fielddict.update({'Note Entered By': result})
        if 'submitted_by' in request.POST and request.method == 'POST' and request.POST.get('submitted_by') != '':
            submitted_by = request.POST.get('submitted_by')
            kwargs.setdefault('ticket_caller_id', submitted_by)
            result = User.objects.only('display_name').get(pk=submitted_by).display_name
            fielddict.update({'Submitted By': result})
        if 'entered_by' in request.POST and request.method == 'POST' and request.POST.get('entered_by') != '':
            entered_by = request.POST.get('entered_by')
            kwargs.setdefault('ticket_created_by_id', entered_by)
            result = User.objects.only('display_name').get(pk=entered_by).display_name
            fielddict.update({'Entered By': result})
        if 'assigned_to' in request.POST and request.method == 'POST' and request.POST.get('assigned_to') != '' and 'ever_assign' not in request.POST:
            assigned_to = request.POST.get('assigned_to')
            kwargs.setdefault('ticket_assign_to_id', assigned_to)
            result = User.objects.only('display_name').get(pk=assigned_to).display_name
            fielddict.update({'Assigned To': result})
        if 'assigned_to' in request.POST and request.method == 'POST' and request.POST.get('assigned_to') != '' and 'ever_assign' in request.POST:
            assigned_to = request.POST.get('assigned_to')
            kwargs.setdefault('ticketURLog__urlog_event', 1)
            kwargs.setdefault('ticketURLog__urlog_user_id', assigned_to)
            result = User.objects.only('display_name').get(pk=assigned_to).display_name
            fielddict.update({'Assigned To': result})
        if 'assigned_by' in request.POST and request.method == 'POST' and request.POST.get('assigned_by') != '':
            assigned_by = request.POST.get('assigned_by')
            kwargs.setdefault('ticket_assign_by_id', assigned_by)
            result = User.objects.only('display_name').get(pk=assigned_by).display_name
            fielddict.update({'Assigned By': result})
        if 'next_action' in request.POST and request.method == 'POST' and request.POST.get('next_action') != '' and 'ever_next_action' not in request.POST:
            next_action = request.POST.get('next_action')
            kwargs.setdefault('ticket_next_action_id', next_action)
            result = User.objects.only('display_name').get(pk=next_action).display_name
            fielddict.update({'Next Action': result})
        if 'next_action' in request.POST and request.method == 'POST' and request.POST.get('next_action') != '' and 'ever_next_action' in request.POST:
            next_action = request.POST.get('next_action')
            kwargs.setdefault('ticketURLog__urlog_event', 2)
            kwargs.setdefault('ticketURLog__urlog_user_id', next_action)
            result = User.objects.only('display_name').get(pk=next_action).display_name
            fielddict.update({'Next Action': result})
        if 'closed_by' in request.POST and request.method == 'POST' and request.POST.get('closed_by') != '':
            closed_by = request.POST.get('closed_by')
            kwargs.setdefault('ticket_closed_by_id', closed_by)
            result = User.objects.only('display_name').get(pk=closed_by).display_name
            fielddict.update({'Closed By': result})
        if 'org_id' in request.POST and request.method == 'POST' and request.POST.get('org_id') != '':
            org_id = request.POST.get('org_id')
            kwargs.setdefault('ticket_org_id', org_id)
            result = Organization.objects.only('org_name').get(pk=org_id).org_name
            fielddict.update({'Organization': result})
        if 'client_id' in request.POST and request.method == 'POST' and request.POST.get('client_id') != '':
            client_id = request.POST.get('client_id')
            kwargs.setdefault('ticket_client_id', client_id)
            result = Client.objects.only('client_name').get(pk=client_id).client_name
            fielddict.update({'Client': result})
        if 'date_opened' in request.POST and request.method == 'POST' and request.POST.get('date_opened') != '':
            date_opened = request.POST.get('date_opened')
            dopen_start = datetime.strptime(request.POST.get('date_opened').split(' - ')[0], '%m/%d/%Y').strftime('%Y-%m-%d')
            dopen_end = datetime.strptime(request.POST.get('date_opened').split(' - ')[1], '%m/%d/%Y').strftime('%Y-%m-%d')
            kwargs.setdefault('ticket_created_at__gte', dopen_start)
            kwargs.setdefault('ticket_created_at__lte', dopen_end)
            fielddict.update({'Date Opened': date_opened})
        if 'date_closed' in request.POST and request.method == 'POST' and request.POST.get('date_closed') != '':
            date_closed = request.POST.get('date_closed')
            dclose_start = datetime.strptime(request.POST.get('date_closed').split(' - ')[0], '%m/%d/%Y').strftime('%Y-%m-%d')
            dclose_end = datetime.strptime(request.POST.get('date_closed').split(' - ')[1], '%m/%d/%Y').strftime('%Y-%m-%d')
            kwargs.setdefault('ticket_closed_at__gte', dclose_start)
            kwargs.setdefault('ticket_closed_at__lte', dclose_end)
            fielddict.update({'Date Closed': date_closed})
        if 'labour_hours_val' in request.POST and request.method == 'POST' and request.POST.get('labour_hours_val') != '':
            labour_hours_val = request.POST.get('labour_hours_val')
            labour_hours_val += ':00:00'
            labour_hours = request.POST.get('labour_hours')
            if labour_hours == '0':
                condition = 'lte'
            elif labour_hours == '1':
                condition = 'gte'
            else:
                condition = 'startswith'
            kwargs.setdefault('labour_hours__'+condition, str(labour_hours_val))
            fielddict.update({'Labor Hours - ' + labor_hours_dict[labour_hours]: labour_hours})
        if 'task_description' in request.POST and request.method == 'POST' and request.POST.get('task_description') != '':
            task_description = request.POST.get('task_description')
            kwargs.setdefault('ticketManager__tmgr_task__task_description__icontains', task_description)
            fielddict.update({'Task Description': task_description})
        if 'task_assigned_to' in request.POST and request.method == 'POST' and request.POST.get('task_assigned_to') != '':
            task_assigned_to = request.POST.get('task_assigned_to')
            kwargs.setdefault('ticketManager__task_assigned_to_id', task_assigned_to)
            result = User.objects.only('display_name').get(pk=task_assigned_to).display_name
            fielddict.update({'Task Assigned To': result})
        if 'task_completion_date' in request.POST and request.method == 'POST' and request.POST.get('task_completion_date') != '':
            task_completion_date = request.POST.get('task_completion_date')
            dtask_start = datetime.strptime(request.POST.get('task_completion_date').split(' - ')[0], '%m/%d/%Y').strftime('%Y-%m-%d')
            dtask_end = datetime.strptime(request.POST.get('task_completion_date').split(' - ')[1], '%m/%d/%Y').strftime('%Y-%m-%d')
            kwargs.setdefault('ticketManager__task_due_date__gte', dtask_start)
            kwargs.setdefault('ticketManager__task_due_date__lte', dtask_end)
            fielddict.update({'Task Completion Date': task_completion_date})
        if 'search_title' in request.POST and request.method == 'POST' and request.POST.get('search_title') != '':
            search_title = request.POST.get('search_title')
        else:
            search_title = 'Search Results'

        print(kwargs.items())

        if 'search_title' in request.POST and request.method == 'POST' and request.POST.get('search_title') != '':
            search_title = request.POST.get('search_title')
        if 'output_view' in request.POST and request.method == 'POST':
            output_view = request.POST.get('output_view')
        sortargs = []
        sortresponse = []
        sortdict = {
            "ticket_assign_to__display_name": "Assigned To",
            "ticket_closed_at": "Date Closed",
            "ticket_created_at": "Date Opened",
            "ticket_created_by__display_name": "Entered By",
            "ticket_id": "Ticket Nbr",
            "ticket_subtype1__ttype_name": "Ticket Subtype",
            "ticket_type__ttype_name": "Ticket Type",
            "ticket_modified_at": "Last Activity",
            "ticket_next_action__display_name": "Next Action",
            "ticket_org__org_name": "Organization",
            "priority__p_display_order": "Priority",
            "ticket_status": "Status",
            "ticket_caller__display_name": "Submitted By",
            "ticket_client__client_name": "Client"
        }
        sortorder_dict = {
            "0": "Asc",
            "1": "Desc"
        }
        if 'sort_column1' in request.POST and request.method == 'POST' and request.POST.get('sort_column1') != '':
            sort_column1 = request.POST.get('sort_column1')
            sort_order1 = request.POST.get('sort_order1')
            if sort_order1 == '0':
                sort_value = sort_column1
            else:
                sort_value = '-' + sort_column1

            sortargs.append(sort_value)
            sortresponse.append(sortdict[sort_column1] + " , " + sortorder_dict[sort_order1])

        if 'sort_column2' in request.POST and request.method == 'POST' and request.POST.get('sort_column2') != '':
            sort_column2 = request.POST.get('sort_column2')
            sort_order2 = request.POST.get('sort_order2')
            if sort_order2 == '0':
                sort_value = sort_column2
            else:
                sort_value = '-' + str(sort_column2)

            sortargs.append(sort_value)
            sortresponse.append(sortdict[sort_column2] + " , " + sortorder_dict[sort_order2])

        if 'sort_column3' in request.POST and request.method == 'POST' and request.POST.get('sort_column3') != '':
            sort_column3 = request.POST.get('sort_column3')
            sort_order3 = request.POST.get('sort_order3')
            if sort_order3 == '0':
                sort_value = sort_column3
            else:
                sort_value = '-' + str(sort_column3)

            sortargs.append(sort_value)
            sortresponse.append(sortdict[sort_column3] + " , " + sortorder_dict[sort_order3])

        if not sortargs:
            if 'all_three' in request.POST and request.method == 'POST' and request.POST.get('all_three') != '':
                tickets = Ticket.objects.filter(**kwargs).filter(args).distinct()
                ticketid_list = Ticket.objects.filter(**kwargs).filter(args).distinct().values_list('ticket_id', flat=True)
            else:
                tickets = Ticket.objects.filter(**kwargs).distinct()
                ticketid_list = Ticket.objects.filter(**kwargs).distinct().values_list('ticket_id', flat=True)
        else:
            if 'all_three' in request.POST and request.method == 'POST' and request.POST.get('all_three') != '':
                tickets = Ticket.objects.filter(**kwargs).filter(args).distinct().order_by(*sortargs)
                ticketid_list = Ticket.objects.filter(**kwargs).filter(args).distinct().values_list('ticket_id', flat=True).order_by(*sortargs)
            else:
                tickets = Ticket.objects.filter(**kwargs).distinct().order_by(*sortargs)
                ticketid_list = Ticket.objects.filter(**kwargs).distinct().values_list('ticket_id', flat=True).order_by(*sortargs)

        # print(tickets.query)


        # Total Time Open Case Start #
        if 'total_time_open_val' in request.POST and request.method == 'POST' and request.POST.get('total_time_open_val') != '':
            total_time_open_val = request.POST.get('total_time_open_val')
            total_time_open = request.POST.get('total_time_open')

            ticket_list = []
            if tickets:
                for ticket in tickets:
                    if ticket.ticket_status == 0:
                        delta = datetime.now(timezone.utc) - ticket.ticket_created_at
                    else:
                        delta = ticket.ticket_closed_at - ticket.ticket_created_at
                    if total_time_open == '0':
                        if int(delta.days) <= int(total_time_open_val):
                            ticket_list.append(ticket.ticket_id)
                    elif total_time_open == '1':
                        if int(delta.days) >= int(total_time_open_val):
                            ticket_list.append(ticket.ticket_id)
                    else:
                        if int(delta.days) == int(total_time_open_val):
                            ticket_list.append(ticket.ticket_id)

                if not sortargs and ticket_list:
                   tickets = Ticket.objects.filter(ticket_id__in=ticket_list).filter(ticket_is_delete=0).filter(ticket_is_active=1)
                   ticketid_list = Ticket.objects.filter(ticket_id__in=ticket_list).filter(ticket_is_delete=0).filter(ticket_is_active=1).values_list('ticket_id', flat=True)

                elif sortargs and ticket_list:
                    tickets = Ticket.objects.filter(ticket_id__in=ticket_list).filter(ticket_is_delete=0).filter(ticket_is_active=1).order_by(*sortargs)
                    ticketid_list = Ticket.objects.filter(ticket_id__in=ticket_list).filter(ticket_is_delete=0).filter(ticket_is_active=1).values_list('ticket_id', flat=True).order_by(*sortargs)
                else:
                    tickets = []
            fielddict.update({'Total Time Open - ' + labor_hours_dict[total_time_open]: total_time_open_val})
        # Total Time Open Case End #

        if 'show_criteria' not in request.POST:
            fielddict = {"1": ''}

        # for key, value in kwargs.items():
        #     print("The value of {} is {}".format(key, value))
        # return HttpResponse('OK')
        print(list(ticketid_list))

        context = {
            'tickets': tickets,
            'search_title': search_title,
            'sortresponses': sortresponse,
            'fielddict': fielddict,
            'output_view': output_view,
            'ticketid_list': list(ticketid_list)
        }

        if output_view == 'OutputToExcel' or output_view == 'DetailOutputToExcel':
            if output_view ==  'OutputToExcel':
                return get_xls_from_tickets(list(ticketid_list), '0')
            else:
                return get_xls_from_tickets(list(ticketid_list), '1')
        else:
            return render(request, 'itrak/Ticket/tickets_search_list.html', context)

# Ticket Search Filter Result End#



# Ticket Search Save Start#

@active_user_required
def ticketSaveSearch(request):
    if request.method == 'POST':
        if 'ticket_status' in request.POST and request.method == 'POST' and request.POST.get('ticket_status') != '':
            ticket_status = request.POST.get('ticket_status')
        else:
            ticket_status = None
        if 'ticket_sub_status' in request.POST and request.method == 'POST':
            ticket_sub_status = request.POST.get('ticket_sub_status')
        if 'priority' in request.POST and request.method == 'POST':
            priority = request.POST.get('priority')
        if 'ticket_type' in request.POST and request.method == 'POST':
            ticket_type = request.POST.get('ticket_type')
        if 'subtype1' in request.POST and request.method == 'POST':
            subtype1 = request.POST.get('subtype1')
        if 'subtype2' in request.POST and request.method == 'POST':
            subtype2 = request.POST.get('subtype2')
        if 'subtype3' in request.POST and request.method == 'POST':
            subtype3 = request.POST.get('subtype3')
        if 'subtype4' in request.POST and request.method == 'POST':
            subtype4 = request.POST.get('subtype4')
        if 'subject' in request.POST and request.method == 'POST':
            subject = request.POST.get('subject')
        if 'ticket_note' in request.POST and request.method == 'POST':
            ticket_note = request.POST.get('ticket_note')
        if 'all_three' in request.POST and request.method == 'POST':
            all_three = request.POST.get('all_three')
        if 'record_locator' in request.POST and request.method == 'POST':
            record_locator = request.POST.get('record_locator')
        if 'caller_name' in request.POST and request.method == 'POST':
            caller_name = request.POST.get('caller_name')
        if 'caller_phone' in request.POST and request.method == 'POST':
            caller_phone = request.POST.get('caller_phone')
        if 'caller_email' in request.POST and request.method == 'POST':
            caller_email = request.POST.get('caller_email')
        if 'passenger_name' in request.POST and request.method == 'POST':
            passenger_name = request.POST.get('passenger_name')
        if 'note_entered_by' in request.POST and request.method == 'POST':
            note_entered_by = request.POST.get('note_entered_by')
        if 'submitted_by' in request.POST and request.method == 'POST':
            submitted_by = request.POST.get('submitted_by')
        if 'entered_by' in request.POST and request.method == 'POST':
            entered_by = request.POST.get('entered_by')
        if 'assigned_to' in request.POST and request.method == 'POST' and 'ever_assign' not in request.POST:
            assigned_to = request.POST.get('assigned_to')
            ever_assigned = 0
        if 'assigned_to' in request.POST and request.method == 'POST' and 'ever_assign' in request.POST:
            assigned_to = request.POST.get('assigned_to')
            ever_assigned = 1
        if 'assigned_by' in request.POST and request.method == 'POST':
            assigned_by = request.POST.get('assigned_by')
        if 'next_action' in request.POST and request.method == 'POST' and 'ever_next_action' not in request.POST:
            next_action = request.POST.get('next_action')
            ever_next_action = 0
        if 'next_action' in request.POST and request.method == 'POST' and 'ever_next_action' in request.POST:
            next_action = request.POST.get('next_action')
            ever_next_action = 1
        if 'closed_by' in request.POST and request.method == 'POST':
            closed_by = request.POST.get('closed_by')
        if 'org_id' in request.POST and request.method == 'POST':
            org_id = request.POST.get('org_id')
        if 'client_id' in request.POST and request.method == 'POST':
            client_id = request.POST.get('client_id')
        if 'date_opened' in request.POST and request.method == 'POST':
            date_opened = request.POST.get('date_opened')
        if 'date_closed' in request.POST and request.method == 'POST':
            date_closed = request.POST.get('date_closed')
        if 'labour_hours_val' in request.POST and request.method == 'POST':
            labour_hours_val = request.POST.get('labour_hours_val')
            labour_hours = request.POST.get('labour_hours')
        if 'task_description' in request.POST and request.method == 'POST':
            task_description = request.POST.get('task_description')
        if 'task_assigned_to' in request.POST and request.method == 'POST':
            task_assigned_to = request.POST.get('task_assigned_to')
        if 'task_completion_date' in request.POST and request.method == 'POST':
            task_completion_date = request.POST.get('task_completion_date')
        if 'search_title' in request.POST and request.method == 'POST':
            search_title = request.POST.get('search_title')
        else:
            search_title = 'Search Results'

        if 'search_title' in request.POST and request.method == 'POST':
            search_title = request.POST.get('search_title')
        if 'output_view' in request.POST and request.method == 'POST':
            output_view = request.POST.get('output_view')
        if 'sort_column1' in request.POST and request.method == 'POST':
            sort_column1 = request.POST.get('sort_column1')
            sort_order1 = request.POST.get('sort_order1')

        if 'sort_column2' in request.POST and request.method == 'POST':
            sort_column2 = request.POST.get('sort_column2')
            sort_order2 = request.POST.get('sort_order2')

        if 'sort_column3' in request.POST and request.method == 'POST':
            sort_column3 = request.POST.get('sort_column3')
            sort_order3 = request.POST.get('sort_order3')


        # Total Time Open Case Start #
        if 'total_time_open_val' in request.POST and request.method == 'POST':
            total_time_open_val = request.POST.get('total_time_open_val')
            total_time_open = request.POST.get('total_time_open')

        # Total Time Open Case End #

        if 'show_criteria' not in request.POST:
            show_criteria = 0
        else:
            show_criteria = 1

        obj = TicketSavedSearch(
            ticket_status=ticket_status,
            ticket_sub_status_id=ticket_sub_status,
            priority_id=priority,
            ticket_type_id=ticket_type,
            ticket_subtype1_id=subtype1,
            ticket_subtype2_id=subtype2,
            ticket_subtype3_id=subtype3,
            ticket_subtype4_id=subtype4,
            subject=subject,
            ticket_note=ticket_note,
            all_three=all_three,
            ticket_record_locator=record_locator,
            ticket_caller_name=caller_name,
            ticket_caller_phone=caller_phone,
            ticket_caller_email=caller_email,
            ticket_passenger_name=passenger_name,
            submitted_by_id=submitted_by,
            note_entered_by_id=note_entered_by,
            entered_by_id=entered_by,
            assigned_by_id=assigned_by,
            ticket_assigned_to_id=assigned_to,
            ever_assigned=ever_assigned,
            next_action_id=next_action,
            ever_next_action=ever_next_action,
            closed_by_id=closed_by,
            org_id=org_id,
            client_id=client_id,
            date_opened=date_opened,
            date_closed=date_closed,
            labour_hours_val=labour_hours_val,
            labour_hours=labour_hours,
            task_description=task_description,
            task_assigned_to_id=task_assigned_to,
            task_completion_date=task_completion_date,
            search_title=search_title,
            output_view=output_view,
            sort_column1=sort_column1,
            sort_order1=sort_order1,
            sort_column2=sort_column2,
            sort_order2=sort_order2,
            sort_column3=sort_column3,
            sort_order3=sort_order3,
            total_time_open_val=total_time_open_val,
            total_time_open=total_time_open,
            show_criteria=show_criteria,
            save_created_by_id=request.user.id
        )
        obj.save()
        insert_id = TicketSavedSearch.objects.latest('pk').saved_search_id

        messages.success(request, 'Request Success! Search Saved.')
        return redirect("/Home_SearchTicket?reportNbr=" + str(insert_id))

# Ticket Search Save End#




# Ticket Search Save Updating Start#

@active_user_required
def ticketUpdateSavedSearch(request):
    if request.method == 'POST':
        id = request.POST.get('savedSearchId')
        # Instead of having an error on your server,
        # your user will get a 404 meaning that he tries to access a non existing resource.
        # data = get_object_or_404(Organization , pk = id)
        try:
            obj = TicketSavedSearch.objects.get(pk=id)
        except TicketSavedSearch.DoesNotExist:
            return render_to_response('itrak/page-404.html')


        if 'ticket_status' in request.POST and request.method == 'POST' and request.POST.get('ticket_status') != '':
            obj.ticket_status = request.POST.get('ticket_status')
        else:
            ticket_status = None
        if 'ticket_sub_status' in request.POST and request.method == 'POST':
            obj.ticket_sub_status_id = request.POST.get('ticket_sub_status')
        if 'priority' in request.POST and request.method == 'POST':
            obj.priority_id = request.POST.get('priority')
        if 'ticket_type' in request.POST and request.method == 'POST':
            obj.ticket_type_id = request.POST.get('ticket_type')
        if 'subtype1' in request.POST and request.method == 'POST':
            obj.ticket_subtype1_id = request.POST.get('subtype1')
        if 'subtype2' in request.POST and request.method == 'POST':
            obj.ticket_subtype2_id = request.POST.get('subtype2')
        if 'subtype3' in request.POST and request.method == 'POST':
            obj.ticket_subtype3_id = request.POST.get('subtype3')
        if 'subtype4' in request.POST and request.method == 'POST':
            obj.ticket_subtype4_id = request.POST.get('subtype4')
        if 'subject' in request.POST and request.method == 'POST':
            obj.subject = request.POST.get('subject')
        if 'ticket_note' in request.POST and request.method == 'POST':
            obj.ticket_note = request.POST.get('ticket_note')
        if 'all_three' in request.POST and request.method == 'POST':
            obj.all_three = request.POST.get('all_three')
        if 'record_locator' in request.POST and request.method == 'POST':
            obj.ticket_record_locator = request.POST.get('record_locator')
        if 'caller_name' in request.POST and request.method == 'POST':
            obj.ticket_caller_name = request.POST.get('caller_name')
        if 'caller_phone' in request.POST and request.method == 'POST':
            obj.ticket_caller_phone = request.POST.get('caller_phone')
        if 'caller_email' in request.POST and request.method == 'POST':
            obj.ticket_caller_email = request.POST.get('caller_email')
        if 'passenger_name' in request.POST and request.method == 'POST':
            obj.ticket_passenger_name = request.POST.get('passenger_name')
        if 'note_entered_by' in request.POST and request.method == 'POST':
            obj.note_entered_by_id = request.POST.get('note_entered_by')
        if 'submitted_by' in request.POST and request.method == 'POST':
            obj.submitted_by_id = request.POST.get('submitted_by')
        if 'entered_by' in request.POST and request.method == 'POST':
            obj.entered_by_id = request.POST.get('entered_by')
        if 'assigned_to' in request.POST and request.method == 'POST' and 'ever_assign' not in request.POST:
            obj.ticket_assigned_to_id = request.POST.get('assigned_to')
            obj.ever_assigned = 0
        if 'assigned_to' in request.POST and request.method == 'POST' and 'ever_assign' in request.POST:
            obj.ticket_assigned_to_id = request.POST.get('assigned_to')
            obj.ever_assigned = 1
        if 'assigned_by' in request.POST and request.method == 'POST':
            obj.assigned_by_id = request.POST.get('assigned_by')
        if 'next_action' in request.POST and request.method == 'POST' and 'ever_next_action' not in request.POST:
            obj.next_action_id = request.POST.get('next_action')
            obj.ever_next_action = 0
        if 'next_action' in request.POST and request.method == 'POST' and 'ever_next_action' in request.POST:
            obj.next_action_id = request.POST.get('next_action')
            obj.ever_next_action = 1
        if 'closed_by' in request.POST and request.method == 'POST':
            obj.closed_by_id = request.POST.get('closed_by')
        if 'org_id' in request.POST and request.method == 'POST':
            obj.org_id = request.POST.get('org_id')
        if 'client_id' in request.POST and request.method == 'POST':
            obj.client_id = request.POST.get('client_id')
        if 'date_opened' in request.POST and request.method == 'POST':
            obj.date_opened = request.POST.get('date_opened')
        if 'date_closed' in request.POST and request.method == 'POST':
            obj.date_closed = request.POST.get('date_closed')
        if 'labour_hours_val' in request.POST and request.method == 'POST':
            obj.labour_hours_val = request.POST.get('labour_hours_val')
            obj.labour_hours = request.POST.get('labour_hours')
        if 'task_description' in request.POST and request.method == 'POST':
            obj.task_description = request.POST.get('task_description')
        if 'task_assigned_to' in request.POST and request.method == 'POST':
            obj.task_assigned_to_id = request.POST.get('task_assigned_to')
        if 'task_completion_date' in request.POST and request.method == 'POST':
            obj.task_completion_date = request.POST.get('task_completion_date')
        if 'search_title' in request.POST and request.method == 'POST':
            obj.search_title = request.POST.get('search_title')
        else:
            obj.search_title = 'Search Results'

        if 'search_title' in request.POST and request.method == 'POST':
            obj.search_title = request.POST.get('search_title')
        if 'output_view' in request.POST and request.method == 'POST':
            obj.output_view = request.POST.get('output_view')
        if 'sort_column1' in request.POST and request.method == 'POST':
            obj.sort_column1 = request.POST.get('sort_column1')
            obj.sort_order1 = request.POST.get('sort_order1')

        if 'sort_column2' in request.POST and request.method == 'POST':
            obj.sort_column2 = request.POST.get('sort_column2')
            obj.sort_order2 = request.POST.get('sort_order2')

        if 'sort_column3' in request.POST and request.method == 'POST':
            obj.sort_column3 = request.POST.get('sort_column3')
            obj.sort_order3 = request.POST.get('sort_order3')


        # Total Time Open Case Start #
        if 'total_time_open_val' in request.POST and request.method == 'POST':
            obj.total_time_open_val = request.POST.get('total_time_open_val')
            obj.total_time_open = request.POST.get('total_time_open')

        # Total Time Open Case End #

        if 'show_criteria' not in request.POST:
            obj.show_criteria = 0
        else:
            obj.show_criteria = 1

        obj.save_modified_by_id = request.user.id
        obj.save_modified_at = datetime.now(timezone.utc)
        obj.save()

        messages.success(request, 'Request Success! Search Updated.')
        return redirect("/Home_SearchTicket?reportNbr="+id)

# Ticket Search Save Updating End#


@csrf_exempt
#Delete Saved Search Against ID Start#
def deleteSavedSearch(request):
    if request.is_ajax() and request.method == 'POST':
        save_id = request.POST.get('save_id')
        # probably you want to add a regex check if the username value is valid here
        if save_id:
            print(save_id)
            result = TicketSavedSearch.objects.filter(saved_search_id= save_id).delete()
            response_data = {}
            try:
                response_data['response'] = 'Success'
            except:
                response_data['response'] = 'No Record Found'
            return JsonResponse(response_data)


#Delete Saved Search Against ID End#



#Datatable Code Start Here#
class MyTicketListJson(BaseDatatableView):

    # The model we're going to show
    model = Ticket

    # define the columns that will be returned
    # columns = ['action', 'org_id', 'org_name', 'is_internal', 'display']

    # define column names that will be used in sorting
    # order is important and should be same as order of columns
    # displayed by datatables. For non sortable columns use empty
    # value like ''
    # order_columns = ['', 'org_id', 'org_name', 'is_internal']

    # set max limit of records returned, this is used to protect our site if someone tries to attack our site
    # and make it return huge amount of data
    max_display_length = 500

    def get_initial_queryset(self):
        # return queryset used as base for futher sorting/filtering
        # these are simply objects displayed in datatable
        # You should not filter data returned here by any filter values entered by user. This is because
        # we need some base queryset to count total number of records.
        user_id = self.request.user.id # Get user_id from request
        tab = self.request.GET.get('tab_name')
        if tab == 'summary':
            kwargs = Q(ticket_assign_to_id=user_id) | Q(ticket_caller_id=user_id) | Q(ticket_next_action_id=user_id) | Q(ticket_created_by_id=user_id)
        elif tab == 'submitter':
            kwargs = Q(ticket_caller_id=user_id)
        elif tab == 'enterer':
            kwargs = Q(ticket_created_by_id=user_id)
        elif tab == 'assignee':
            kwargs = Q(ticket_assign_to_id=user_id)
        elif tab == 'assignee_next':
            kwargs = Q(ticket_assign_to_id=user_id) | Q(ticket_next_action_id=user_id)
        elif tab == 'next_action':
            kwargs = Q(ticket_next_action_id=user_id)
        elif tab == 'unassign':
            kwargs = Q(ticket_assign_to_id__isnull=True)
        elif tab == 'task_assignee':
            ticket_lists = get_task_mgr_ticket_list(user_id)
            return Ticket.objects.filter(ticket_is_delete=0).filter(ticket_id__in=ticket_lists)
        elif tab == 'task_available':
            ticket_lists = get_task_mgr_ticket_list(user_id)
            return Ticket.objects.filter(ticket_is_delete=0).filter(ticket_id__in=ticket_lists)
        else:
            kwargs = Q(ticket_assign_to_id=user_id) | Q(ticket_caller_id=user_id) | Q(ticket_next_action_id=user_id) | Q(ticket_created_by_id=user_id)

        return Ticket.objects.filter(ticket_is_delete=0).filter(ticket_org__org_is_delete=0).filter(kwargs)


        # return Organization.objects.filter(org_is_active=0, org_is_delete=1)

    def render_column(self, row, column):
        rid = signing.dumps(row.ticket_id, salt=settings.SALT_KEY)
        # We want to render user as a custom column
        if column == 'ticket_id':
            # escape HTML for security reasons
            return '<a class="id_link" href="Home_ViewTicket?tickID=' + str(rid) + '">'+str(row.ticket_id)+'</a>'
        elif column == 'ticket_id_chk':
            # escape HTML for security reasons
            return '<input type="checkbox" name="ticket_id_chk[]" class="ticket_id_chk" value="'+str(row.ticket_id)+'">'
        elif column == 'ticket_created_at':
                return datetime.strptime(str(row.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y')
        elif column == 'ticket_modified_at':
                return datetime.strptime(str(row.ticket_modified_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y')
        elif column == 'task_is_active':
            if row.task_is_active == 1:
                return 'Y'
            else:
                return 'N'
        elif column == 'ticket_status':
            if row.ticket_status == 1:
                return 'Open'
            else:
                return 'Close'
        else:
            return super(MyTicketListJson, self).render_column(row, column)

    def filter_queryset(self, qs):
        # use parameters passed in GET request to filter queryset

        # simple example:
        search = self.request.GET.get('search[value]')
        if search:
            # org = user_org.org_name
            qs = qs.filter(Q(ticket_created_at__icontains=search) | Q(subject__icontains=search) | Q(ticket_caller__display_name__icontains=search) | Q(ticket_next_action__display_name__icontains=search) | Q(priority__priority_name__icontains=search) | Q(ticket_modified_at__icontains=search) | Q(ticket_sub_status__sub_status_text__icontains=search))
            # qs = qs.filter(name__istartswith=search)

        # more advanced example using extra parameters
        # filter_customer = self.request.GET.get('customer', None)
        #
        # if filter_customer:
        #     customer_parts = filter_customer.split(' ')
        #     qs_params = None
        #     for part in customer_parts:
        #         q = Q(customer_firstname__istartswith=part)|Q(customer_lastname__istartswith=part)
        #         qs_params = qs_params | q if qs_params else q
        #     qs = qs.filter(qs_params)
        return qs

        # def prepare_results(self, qs):
        #     # prepare list with output column data
        #     # queryset is already paginated here
        #     json_data = []
        #     for item in qs:
        #         json_data.append([
        #             escape("{0}".format('OK')),
        #             escape(item.org_id),  # escape HTML for security reasons
        #             escape(item.org_name),  # escape HTML for security reasons
        #             escape(item.is_internal),  # escape HTML for security reasons
        #             escape("{0}".format('OK'))
        #             # item.get_state_display(),
        #             # item.created.strftime("%Y-%m-%d %H:%M:%S"),
        #
        #         ])
        #     return json_data


#Datatable Code End Here#



#Validate Username for Uniqueness Start#

@csrf_exempt
def validateTaskUnique(request):
    if request.is_ajax() and request.method == 'POST':
        task_description = request.POST.get('task_description')
        # probably you want to add a regex check if the Task value is valid here
        if task_description:
            is_exist = Task.objects.filter(task_description=task_description).exists()
            response_data = { 'response': is_exist}
            return JsonResponse(response_data)
    else:
        return HttpResponse('fail')


#Validate Username for Uniqueness End#



#Export ticket to XLSX Start#

@csrf_exempt
def export_tickets_xls(request):
    """
    Downloads all movies as Excel file with a worksheet for each movie category
    """
    if request.method == 'POST':
        tickets = request.POST.get('tickets')
        xlsType = request.POST.get('xlsType')
    # category_queryset = Ticket.objects.all()
    ticket_queryset = Ticket.objects.filter(ticket_id__in=json.loads(tickets)).filter(ticket_is_delete=0).filter(ticket_is_active=1)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-TicketSearchList.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Delete the default worksheet
    workbook.remove(workbook.active)

    # Define some styles and formatting that will be later used for cells
    header_font = Font(size=7, name='Segoe UI', bold=True, color='FFFFFF')
    centered_alignment = Alignment(horizontal='left')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='21316f'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    if xlsType == '0':
        # Define the column titles and widths
        columns = [
            ('Ticket #', 7),
            ('Status', 7),
            ('SubStatus', 15),
            ('Subject', 40),
            ('Ticket Type', 15),
            ('Subtype', 20),
            ('Priority', 7),
            ('Submitted Date', 15),
            ('Submitted By Name', 15),
            ('Entered By Name', 15),
            ('Assigned To Name', 15),
            ('Assigned Date', 15),
            ('Next Action Name', 15),
            ('Closed Date', 15),
            ('Last Activity', 15),
            ('Total Time Open', 15),
            ('Time Open (Days)', 15),
            ('Adj Time Open', 15),
            ('Adj Time Open (Days)', 15),
        ]
    else:
        # Define the column titles and widths
        columns = [
            ('Ticket #', 7),
            ('Status', 7),
            ('SubStatus', 15),
            ('Subject', 40),
            ('Full Description', 40),
            ('Ticket Type', 15),
            ('Subtype', 20),
            ('Subtype 2', 20),
            ('Subtype 3', 20),
            ('Subtype 4', 20),
            ('Priority', 7),
            ('Organization', 10),
            ('Client ID', 7),
            ('Client', 15),
            ('Record Locator', 10),
            ('Caller Name', 15),
            ('Caller Phone', 15),
            ('Caller Email', 20),
            ('Passenger Name', 15),
            ('Address', 40),
            ('Agent Error or Goodwill', 20),
            ('Agent Responsible (if applic)', 20),
            ('Airline Ticket #', 15),
            ('Amount of Payout', 15),
            ('Amount Saved', 10),
            ('Attention', 15),
            ('Check Number', 15),
            ('Company', 15),
            ('Correction / Containment Actions', 40),
            ('Corrective Action', 40),
            ('Is Traveler a VIP?', 10),
            ('Notes to print on check', 40),
            ('Pay to the order of', 30),
            ('Payout Required?', 15),
            ('Resp Vendor\'s City (if applic)', 40),
            ('Root Cause', 40),
            ('Vendor Name Resp (if applic)', 30),
            ('Who Approved', 20),
            ('Submitted Date', 15),
            ('Submitted By Name', 15),
            ('Submitted By Dept', 20),
            ('Submitted By Phone', 20),
            ('Assigned To', 15),
            ('Assigned To Name', 15),
            ('Assigned Date', 15),
            ('Assigned To Phone', 15),
            ('Next Action Name', 15),
            ('Next Action Date', 15),
            ('Next Action', 15),
            ('Next Action Phone', 15),
            ('Closed Date', 15),
            ('Closed By', 15),
            ('Closed By Name', 15),
            ('Closed By Phone', 15),
            ('ReOpened Date', 15),
            ('ReOpened By', 15),
            ('ReOpened By Name', 15),
            ('ReOpened By Phone', 15),
            ('Last Activity', 15),
            ('Total Time Open', 15),
            ('Time Open (Days)', 15),
            ('Adj Time Open', 15),
            ('Adj Time Open (Days)', 15),
        ]

    # Create a worksheet/tab with the title of the category
    worksheet = workbook.create_sheet(
        title='TicketSearchEcxel',
        index=0,
    )
    # Define the background color of the header cells
    fill = PatternFill(
        start_color='21316f',
        end_color='21316f',
        fill_type='solid',
    )
    row_num = 1

    # Assign values, styles, and formatting for each cell in the header
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = border_bottom
        cell.alignment = centered_alignment
        cell.fill = fill
        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width

    # Iterate through all movies of a category
    for ticket in ticket_queryset:
        row_num += 1

        if ticket.ticket_status == 0:
            delta = datetime.now(timezone.utc) - ticket.ticket_created_at
        else:
            delta = ticket.ticket_closed_at - ticket.ticket_created_at

        days, seconds = delta.days, delta.seconds
        hours = days * 24 + seconds
        minutes = (seconds % 3600)
        seconds = seconds % 60

        if ticket.description:
            s = MLStripper()
            s.feed(ticket.description)
            plain_description = s.get_data()
        else:
            plain_description = ''

        if xlsType == '0':
            # Define data and formats for each cell in the row
            row = [
                (ticket.ticket_id, 'Normal'),
                ('Open' if ticket.ticket_status == 0 else 'Close', 'Normal'),
                (ticket.ticket_sub_status.sub_status_text if ticket.ticket_sub_status_id else '', 'Normal'),
                (ticket.subject, 'Normal'),
                # (timedelta(minutes=movie.ticket_created_at), 'Normal'),
                (ticket.ticket_type.ttype_name if ticket.ticket_type_id else '', 'Normal'),
                (ticket.ticket_subtype1.ttype_name if ticket.ticket_subtype1_id else '', 'Normal'),
                (ticket.priority.priority_name if ticket.priority_id else '', 'Normal'),
                (datetime.strptime(str(ticket.submitted_date), '%Y-%m-%d').strftime('%d/%m/%Y') if ticket.submitted_date else '', 'Normal'),
                (ticket.ticket_caller.display_name if ticket.ticket_caller_id else '', 'Normal'),
                (ticket.ticket_created_by.display_name if ticket.ticket_created_by_id else '', 'Normal'),
                (ticket.ticket_assign_to.display_name if ticket.ticket_assign_to_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_next_action.display_name if ticket.ticket_next_action_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_closed_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_closed_at else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_modified_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_modified_at else '', 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),

            ]
        else:

            #Data Building of Condition Fields Start#
            if ticket.agent_error_goodwill:
                if ticket.agent_error_goodwill == '0':
                    agent_error_goodwill = 'Agent Error'
                else:
                    agent_error_goodwill = 'Goodwill'
            else:
                agent_error_goodwill = ''

            if ticket.is_traveler_vip:
                if ticket.is_traveler_vip == '0':
                    is_traveler_vip = 'Yes'
                else:
                    is_traveler_vip = 'No'
            else:
                is_traveler_vip = ''

            if ticket.is_payout_required:
                if ticket.is_payout_required == '0':
                    is_payout_required = 'Yes'
                else:
                    is_payout_required = 'No'
            else:
                is_payout_required = ''

            #Data Building of Condition Fields End#


            # Define data and formats for each cell in the row
            row = [
                (ticket.ticket_id, 'Normal'),
                ('Open' if ticket.ticket_status == 0 else 'Close', 'Normal'),
                (ticket.ticket_sub_status.sub_status_text if ticket.ticket_sub_status_id else '', 'Normal'),
                (ticket.subject, 'Normal'),
                (plain_description, 'Normal'),
                (ticket.ticket_type.ttype_name if ticket.ticket_type_id else '', 'Normal'),
                (ticket.ticket_subtype1.ttype_name if ticket.ticket_subtype1_id else '', 'Normal'),
                (ticket.ticket_subtype2.ttype_name if ticket.ticket_subtype2_id else '', 'Normal'),
                (ticket.ticket_subtype3.ttype_name if ticket.ticket_subtype3_id else '', 'Normal'),
                (ticket.ticket_subtype4.ttype_name if ticket.ticket_subtype4_id else '', 'Normal'),
                (ticket.priority.priority_name if ticket.priority_id else '', 'Normal'),
                (ticket.ticket_org.org_name if ticket.ticket_org_id else '', 'Normal'),
                (ticket.ticket_client.client_cus_id if ticket.ticket_client_id else '', 'Normal'),
                (ticket.ticket_client.client_name if ticket.ticket_client_id else '', 'Normal'),
                (ticket.ticket_record_locator if ticket.ticket_record_locator else '', 'Normal'),
                (ticket.ticket_caller_name if ticket.ticket_caller_name else '', 'Normal'),
                (ticket.ticket_caller_phone if ticket.ticket_caller_phone else '', 'Normal'),
                (ticket.ticket_caller_email if ticket.ticket_caller_email else '', 'Normal'),
                (ticket.ticket_passenger_name if ticket.ticket_passenger_name else '', 'Normal'),
                (ticket.ticket_address if ticket.ticket_address else '', 'Normal'),
                (agent_error_goodwill, 'Normal'),
                (ticket.agent_responsible if ticket.agent_responsible else '', 'Normal'),
                (ticket.airline_ticket_no if ticket.airline_ticket_no else '', 'Normal'),
                (ticket.ticket_payout_amount if ticket.ticket_payout_amount else '', 'Normal'),
                (ticket.amount_saved if ticket.amount_saved else '', 'Normal'),
                (ticket.ticket_attention if ticket.ticket_attention else '', 'Normal'),
                (ticket.check_number if ticket.check_number else '', 'Normal'),
                (ticket.ticket_company if ticket.ticket_company else '', 'Normal'),
                (ticket.corr_cont_actions if ticket.corr_cont_actions else '', 'Normal'),
                (ticket.corrective_action if ticket.corrective_action else '', 'Normal'),
                (is_traveler_vip, 'Normal'),
                (ticket.notes_on_check if ticket.notes_on_check else '', 'Normal'),
                (ticket.ticket_order_of_pay if ticket.ticket_order_of_pay else '', 'Normal'),
                (is_payout_required, 'Normal'),
                (ticket.vresponsible_city if ticket.vresponsible_city else '', 'Normal'),
                (ticket.ticket_root_cause if ticket.ticket_root_cause else '', 'Normal'),
                (ticket.vendor_responsible if ticket.vendor_responsible else '', 'Normal'),
                (ticket.check_approved_by if ticket.check_approved_by else '', 'Normal'),
                (datetime.strptime(str(ticket.submitted_date), '%Y-%m-%d').strftime('%d/%m/%Y') if ticket.submitted_date else '', 'Normal'),
                (ticket.ticket_caller.display_name if ticket.ticket_caller_id else '', 'Normal'),
                (ticket.ticket_caller.user_dep.dep_name if ticket.ticket_caller_id and ticket.ticket_caller.user_dep_id else '', 'Normal'),
                (ticket.ticket_caller.phone_no if ticket.ticket_caller_id and ticket.ticket_caller_id else '', 'Normal'),
                (ticket.ticket_assign_to.display_name if ticket.ticket_assign_to_id else '', 'Normal'),
                (ticket.ticket_assign_to.first_name if ticket.ticket_assign_to_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_assign_to.phone_no if ticket.ticket_assign_to_id else '', 'Normal'),
                (ticket.ticket_next_action.display_name if ticket.ticket_next_action_id else '', 'Normal'),
                (ticket.ticket_next_action.first_name if ticket.ticket_next_action_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_next_action.phone_no if ticket.ticket_next_action_id else '', 'Normal'),
                (ticket.ticket_next_action.display_name if ticket.ticket_next_action_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_closed_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_closed_at else '', 'Normal'),
                (ticket.ticket_created_by.display_name if ticket.ticket_created_by_id else '', 'Normal'),
                (ticket.ticket_created_by.phone_no if ticket.ticket_created_by_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_created_by.first_name if ticket.ticket_created_by_id else '', 'Normal'),
                (ticket.ticket_created_by.display_name if ticket.ticket_created_by_id else '', 'Normal'),
                (ticket.ticket_created_by.phone_no if ticket.ticket_created_by_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_modified_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_modified_at else '', 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),

            ]

        if int(row_num) % 2 == 0:
            row_fill = PatternFill(
                start_color='FFEFD5',
                end_color='FFEFD5',
                fill_type='solid',
            )
        else:
            row_fill = PatternFill(
                start_color='FFFFFF',
                end_color='FFFFFF',
                fill_type='solid',
            )
        print(row_num)

        row_font = Font(size=8, name='Segoe UI', bold=False, color='000000')
        # Assign values, styles, and formatting for each cell in the row
        for col_num, (cell_value, cell_format) in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.style = cell_format
            # if cell_format == 'Currency':
            #     cell.number_format = '#,##0.00 €'
            # if col_num == 8:
            #     cell.number_format = '[h]:mm;@'
            cell.alignment = wrapped_alignment
            cell.fill = row_fill
            cell.font = row_font

    # freeze the first row
    worksheet.freeze_panes = worksheet['A2']

    # set tab color
    worksheet.sheet_properties.tabColor = 'FFFFFF'

    workbook.save(response)

    return response

#unassignedTickets tickets to XLSX End#

#Unassigned tickets Start#

@active_user_required
def unassignedTickets(request):
    load_sidebar = get_sidebar(request)

    user_id = request.user.id # Get user_id from request
    kwargs = Q(ticket_assign_to_id__isnull=True)
    unassign = Ticket.objects.filter(ticket_is_delete=0).filter(kwargs)
    context = {
        'unassign': unassign,
        'sidebar': load_sidebar
    }
    return render(request, 'itrak/Ticket/unassigned_tickets.html', context)

#Unassigned tickets End#

#Save Unassigned tickets Start#

def save_UnassignedTickets(request):
    if request.method == 'POST':
        if 'ticket_id_chk[]' in request.POST and request.POST['ticket_id_chk[]']:
            tickets_ids = request.POST.getlist('ticket_id_chk[]')
            for t in tickets_ids:
                try:
                    TObj = Ticket.objects.get(pk=t)
                except Ticket.DoesNotExist:
                    return render_to_response('itrak/page-404.html')

                TObj.ticket_assign_to_id = request.user.id
                TObj.ticket_assign_by_id = request.user.id
                TObj.ticket_assign_at = datetime.now(timezone.utc)
                TObj.save()

                obj1 = TicketUserRoleLog(
                    urlog_ticket_id=TObj.ticket_id,
                    urlog_user_id=request.user.id,
                    urlog_event=1,
                    urlog_created_by_id=request.user.id
                )
                obj1.save()

            messages.success(request, 'Tickets successfully assigned to you.')
            return redirect("/Home_UnassignedTickets")
        else:
            messages.error(request, 'Request Failed! Atleast Select One Ticket.')
            return redirect("/Home_UnassignedTickets")

#Save Unassigned tickets End#


#Assign To Me Start#

def assignTicketToSelf(request):
    id = request.GET.get('ticketNbr', -1)
    # Instead of having an error on your server,
    # your user will get a 404 meaning that he tries to access a non existing resource.
    if id:
        try:
            ticket_id = signing.loads(id, salt=settings.SALT_KEY)
            TObj = Ticket.objects.get(pk=ticket_id)
        except Ticket.DoesNotExist:
            return render_to_response('itrak/page-404.html')

        TObj.ticket_assign_to_id = request.user.id
        TObj.ticket_assign_by_id = request.user.id
        TObj.ticket_assign_at = datetime.now(timezone.utc)
        TObj.save()

        obj1 = TicketUserRoleLog(
            urlog_ticket_id=TObj.ticket_id,
            urlog_user_id=request.user.id,
            urlog_event=1,
            urlog_created_by_id=request.user.id
        )
        obj1.save()

        messages.success(request, 'Ticket successfully assigned to you.')
        return redirect(viewTicket, id=id)
    else:
        messages.error(request, 'Request Failed! Ticket failed to assign. Please try again.')
        return redirect(viewTicket, id=id)

#Assign To Me Start End#

#Ticket Look Uptickets Start#

@active_user_required
def ticketLookUp(request):
    load_sidebar = get_sidebar(request)
    if request.method == 'POST':
        ticket_count = Ticket.objects.filter(Q(pk__contains=request.POST.get('ticket_number')) | Q(subject__contains=request.POST.get('ticket_number'))).filter(ticket_is_delete=0).count()
        # return HttpResponse(ticket_count)
        ticket = Ticket.objects.filter(Q(pk__contains=request.POST.get('ticket_number')) | Q(subject__contains=request.POST.get('ticket_number'))).filter(ticket_is_delete=0).first()
        if ticket_count == 1:
            # return HttpResponse(ticket.ticket_id)
            rid = signing.dumps(ticket.ticket_id, salt=settings.SALT_KEY)
            user_redirect =  'Home_ViewTicket?tickID=' + str(rid)
            return redirect("/"+str(user_redirect))
        elif ticket_count > 1:
            kwargs = {
                '{0}__{1}'.format('ticket_is_delete', 'iexact'): 0,
                '{0}__{1}'.format('ticket_is_active', 'iexact'): 1,
            }
            fielddict = {"1": ''}
            if 'ticket_number' in request.POST and request.method == 'POST' and request.POST.get('ticket_number') != '':
                subject = request.POST.get('ticket_number')
                kwargs.setdefault('subject__icontains', subject)
                fielddict.update({'Subject': subject})

            # print(kwargs.items())

            sortargs = []
            sortresponse = ['No sort selected']
            sortdict = {
                "ticket_assign_to__display_name": "Assigned To",
                "ticket_closed_at": "Date Closed",
                "ticket_created_at": "Date Opened",
                "ticket_created_by__display_name": "Entered By",
                "ticket_id": "Ticket Nbr",
                "ticket_subtype1__ttype_name": "Ticket Subtype",
                "ticket_type__ttype_name": "Ticket Type",
                "ticket_modified_at": "Last Activity",
                "ticket_next_action__display_name": "Next Action",
                "ticket_org__org_name": "Organization",
                "priority__p_display_order": "Priority",
                "ticket_status": "Status",
                "ticket_caller__display_name": "Submitted By",
                "ticket_client__client_name": "Client"
            }
            sortorder_dict = {
                "0": "Asc",
                "1": "Desc"
            }

            tickets = Ticket.objects.filter(**kwargs).distinct()
            ticketid_list = Ticket.objects.filter(**kwargs).distinct().values_list('ticket_id', flat=True)

            print(tickets.query)
            output_view = "BriefList"
            search_title = 'Search Results'
            # for key, value in kwargs.items():
            #     print("The value of {} is {}".format(key, value))
            # return HttpResponse('OK')
            print(list(ticketid_list))

            context = {
                'tickets': tickets,
                'search_title': search_title,
                'sortresponses': sortresponse,
                'fielddict': fielddict,
                'output_view': output_view,
                'ticketid_list': list(ticketid_list)
            }

            # if output_view == 'OutputToExcel' or output_view == 'DetailOutputToExcel':
            #     if output_view ==  'OutputToExcel':
            #         return get_xls_from_tickets(list(ticketid_list), '0')
            #     else:
            #         return get_xls_from_tickets(list(ticketid_list), '1')
            # else:
            return render(request, 'itrak/Ticket/tickets_search_list.html', context)
        else:
            messages.error(request, 'No Ticket Found!')
            return redirect(request.META['HTTP_REFERER'])
    else:
        context = {
            'sidebar': load_sidebar
        }
        return render(request, 'itrak/Ticket/ticket_look_up.html', context)

#Ticket Look Uptickets End#


#Get Task Manager on Modal Through ID Start#
def getModalTaskManagerById(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        records = json.loads(request.POST.get('records', ''))
        context['records'] = records
    return render(request, 'itrak/Ticket/get_taskmanager_modal_records.html', context)

#Get Task Manager on Modal Through ID End#


#Get Task Manager on Table Through ID Start#
def getTableTaskManagerById(request):
    context = {}
    if request.method == 'POST' and request.is_ajax():
        records = json.loads(request.POST.get('records', ''))
        context['depend_value'] = request.POST.get('depend_value', '')
        context['records'] = records
    return render(request, 'itrak/Ticket/get_taskmanager_table_records.html', context)

#Get Task Manager on Table Through ID End#



#Save Task Manager on Table Through Group ID Start#
def saveModalTaskManagerByTicketId(request):
    context = {}
    response_data = {}
    if request.method == 'POST' and request.is_ajax():
        records = json.loads(request.POST.get('records', ''))
        ticketId =  request.POST.get('ticketId', '')
        with transaction.atomic():
            TaskManager.objects.filter(tmgr_ticket_id=ticketId).delete()
            print(records)
            if records:
                try:
                    for record in records:
                        task_order_val = record['task_order_val']
                        task_id_val = record['task_id_val']
                        task_assign_to_val = record['task_assign_to_val']
                        if record['task_due_date_val']:
                            task_due_date_val = datetime.strptime(record['task_due_date_val'], '%d/%m/%Y').strftime('%Y-%m-%d')
                        else:
                            task_due_date_val = None
                        task_type_val = record['task_type_val']
                        task_note_val = record['task_note_val']
                        task_dependency_val = record['task_dependency_val']
                        task_dependency_order_val = record['task_dependency_order_val']
                        modal_ttype_group_yes_val = record['modal_ttype_group_yes_val']
                        modal_ttype_copen_yes_val = record['modal_ttype_copen_yes_val']
                        modal_ttype_cticket_yes_val = record['modal_ttype_cticket_yes_val']
                        modal_ttype_substatus_yes_val = record['modal_ttype_substatus_yes_val']
                        modal_ttype_group_no_val = record['modal_ttype_group_no_val']
                        modal_ttype_copen_no_val = record['modal_ttype_copen_no_val']
                        modal_ttype_cticket_no_val = record['modal_ttype_cticket_no_val']
                        modal_ttype_substatus_no_val = record['modal_ttype_substatus_no_val']
                        modal_ttype_group_na_val = record['modal_ttype_group_na_val']
                        modal_ttype_copen_na_val = record['modal_ttype_copen_na_val']
                        modal_ttype_cticket_na_val = record['modal_ttype_cticket_na_val']
                        modal_ttype_substatus_na_val = record['modal_ttype_substatus_na_val']
                        task_is_cancel = record['task_is_cancel']
                        task_is_complete = record['task_is_complete']
                        task_completion_userId = record['task_completion_userId'] if record['task_completion_userId'] else None
                        task_completion_userName = record['task_completion_userName'] if record['task_completion_userName'] else None
                        task_laborhour_hours = record['task_laborhour_hours']
                        task_laborhour_minutes = record['task_laborhour_minutes']
                        task_labor_note = record['task_labor_note']
                        task_response_status = record['task_response_status']

                        if task_is_cancel == 1 or task_is_complete == 1:
                            task_completion_datetime = datetime.strptime(record['task_completion_datetime'], '%m/%d/%Y %I:%M %p').strftime('%Y-%m-%d %H:%M:%S.%f%z')
                        else:
                            task_completion_datetime = None

                        TMobj = TaskManager(tmgr_display_order=task_order_val, tmgr_task_id=task_id_val, tmgr_ticket_id = ticketId,
                                            task_assigned_to_id=task_assign_to_val, task_due_date=task_due_date_val,
                                            task_type=task_type_val, task_note=task_note_val,
                                            task_dependency=task_dependency_val, task_depend_order=task_dependency_order_val,
                                            ttype_group_yes_id=modal_ttype_group_yes_val, ttype_copen_yes=modal_ttype_copen_yes_val,
                                            ttype_cticket_yes=modal_ttype_cticket_yes_val, ttype_substatus_yes_id=modal_ttype_substatus_yes_val,
                                            ttype_group_no_id=modal_ttype_group_no_val, ttype_copen_no=modal_ttype_copen_no_val,
                                            ttype_cticket_no=modal_ttype_cticket_no_val, ttype_substatus_no_id=modal_ttype_substatus_no_val,
                                            ttype_group_na_id=modal_ttype_group_na_val, ttype_copen_na=modal_ttype_copen_na_val,
                                            ttype_cticket_na=modal_ttype_cticket_na_val, ttype_substatus_na_id=modal_ttype_substatus_na_val,
                                            tmgr_is_cancel=task_is_cancel, tmgr_is_complete=task_is_complete,
                                            tmgr_completion_at=task_completion_datetime,
                                            tmgr_completion_userId_id=task_completion_userId, tmgr_completion_userName=task_completion_userName,
                                            tmgr_laborhour_hours=task_laborhour_hours, tmgr_laborhour_minutes=task_laborhour_minutes,
                                            tmgr_labor_note=task_labor_note, tmgr_response_status=task_response_status,
                                            task_created_by_id=request.user.id)
                        TMobj.save()
                except IntegrityError:
                    transaction.rollback()

        response_data['response'] = 'Success'
    else:
        response_data['response'] = 'No Record Found'

    return HttpResponse(response_data)

#Save Task Manager on Table Through Group ID End#


#Get Task Manager Database Records By Ticket ID Start#
@csrf_exempt
def getTaskManagerByTicketId(request):
    if request.is_ajax() and request.method == 'POST':
        ticket_id = request.POST.get('ticket_id')
        if ticket_id:
            result = TaskManager.objects.filter(tmgr_ticket_id=ticket_id)\
                                        .order_by(Cast('tmgr_display_order', IntegerField()))
            response_data = {}
            try:
                response_data['response'] = serializers.serialize('json', result)
            except:
                response_data['response'] = ''
            return JsonResponse(response_data)

#Get Task Manager Database Records By Ticket ID End#



