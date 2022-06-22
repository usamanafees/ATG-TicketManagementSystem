from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404, render_to_response
from django.core import signing
from django import template
from itrak.models import *
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.urls import reverse
from django.db.models import Case, F, FloatField, IntegerField, Sum, When, Count
from django.db.models.functions import Cast
from django.utils.html import conditional_escape
from django.utils.safestring import mark_safe
import pytz, json

register = template.Library()


# Get Encrypted ID For Template Link Start#


@register.simple_tag
def get_encrypted_id(id):
    encrypted_val = signing.dumps(id, salt=settings.SALT_KEY)
    return encrypted_val

@register.simple_tag
def get_encrypted_id1(id):
    encrypted_val = signing.dumps(id,salt=settings.SALT_KEY)
    return encrypted_val

# Get Encrypted ID For Template Link End#


# Get Actual URL Value Start#


@register.simple_tag
def get_actual_url(request, menu):
    url_parts = request.path.split('/')
    flag = ''
    try:
        if '/' + url_parts[1] == reverse(menu):
            try:
                if '/' + url_parts[1] == reverse(submenu):
                    flag = 'nav-active'
            except:
                return flag
    except:
        return flag
    else:
        flag = ''
    return flag


# Get Actual URL Value End#


# Get Expand URL Value Start#


@register.simple_tag
def get_expand_url(request, menu):
    url_parts = request.path.split('/')
    flag = '';
    submenus = SubMenus.objects.values_list('submenu_link', flat=True).filter(submenu_menu_id=menu)
    for submenu in submenus:
        try:
            if '/' + url_parts[1] == reverse(submenu):
                flag = 'nav-expanded nav-active'
        except:
            return flag
    return flag


# Get Expand URL Value End#


# Get Encrypted ID For Template Link Start#


@register.simple_tag
def get_tasks_by_ticket(id):
    response = TaskManager.objects.filter(tmgr_ticket_id=id).order_by('tmgr_task__task_display_order')
    return response


# Get Encrypted ID For Template Link End#


# Get Encrypted ID For Template Link Start#


@register.simple_tag
def tickets_status(tickets):
    t_open = 0
    t_close = 0
    t_reopen = 0
    t_assigned = 0
    t_unassigned = 0
    sub_assign_thours = 0
    sub_assign_tminutes = 0
    assign_close_thours = 0
    assign_close_tminutes = 0
    sub_close_thours = 0
    sub_close_tminutes = 0
    avg_time_submit_assign = '00:00 hours'
    avg_time_assign_close = '00:00 hours'
    avg_time_submit_close = '00:00 hours'
    tt_time_entered = '00:00 hours'

    try:
        result = TicketNote.objects.filter(note_ticket__in=tickets).filter(note_is_delete=0).aggregate(
            thours=Sum(Cast('tnote_laborhour_hours', IntegerField())) + Sum(
                Cast('tnote_laborhour_minutes', IntegerField())) / 60,
            tminutes=Sum(Cast('tnote_laborhour_minutes', IntegerField())) % 60
        )
    except TicketNote.DoesNotExist:
        result = None
    if result:
        if result['thours'] is None or str(result['thours']) == '0':
            hours = '00'
        else:
            hours = "{:02d}".format(result['thours'])
        if result['tminutes'] is None or str(result['tminutes']) == '0':
            minutes = '00';
        else:
            minutes = "{:02d}".format(result['tminutes'])
        tt_time_entered = str(hours) + ':' + str(minutes) + ' hours'

    for ticket in tickets:
        if ticket.ticket_status == 0:
            t_open += 1
        else:
            t_close += 1
        if ticket.ticket_is_reopen == 1:
            t_reopen += 1
        if ticket.ticket_assign_to:
            t_assigned += 1
        else:
            t_unassigned += 1

        if ticket.ticket_assign_to and ticket.ticket_created_at and ticket.ticket_assign_at:
            delta = ticket.ticket_assign_at - ticket.ticket_created_at
            days, seconds = delta.days, delta.seconds
            hours = days * 24 + seconds // 3600
            minutes = (seconds % 3600) // 60
            sub_assign_thours += hours
            sub_assign_tminutes += minutes
            avg_time_submit_assign = "{:02d}".format(
                sub_assign_thours + int(sub_assign_tminutes / 60)) + ':' + "{:02d}".format(
                sub_assign_tminutes % 60) + ' hours'

        if ticket.ticket_assign_to and ticket.ticket_closed_at and ticket.ticket_assign_at:
            delta = ticket.ticket_closed_at - ticket.ticket_assign_at
            days, seconds = delta.days, delta.seconds
            hours = days * 24 + seconds // 3600
            minutes = (seconds % 3600) // 60
            assign_close_thours += hours
            assign_close_tminutes += minutes
            avg_time_assign_close = "{:02d}".format(
                assign_close_thours + int(assign_close_tminutes / 60)) + ':' + "{:02d}".format(
                assign_close_tminutes % 60) + ' hours'

        if ticket.ticket_created_at and ticket.ticket_closed_at:
            delta = ticket.ticket_closed_at - ticket.ticket_created_at
            days, seconds = delta.days, delta.seconds
            hours = days * 24 + seconds // 3600
            minutes = (seconds % 3600) // 60
            sub_close_thours += hours
            sub_close_tminutes += minutes
            avg_time_submit_close = "{:02d}".format(
                sub_close_thours + int(sub_close_tminutes / 60)) + ':' + "{:02d}".format(
                sub_close_tminutes % 60) + ' hours'

    response = {
        "t_open": t_open,
        "t_close": t_close,
        "t_reopen": t_reopen,
        "t_assigned": t_assigned,
        "t_unassigned": t_unassigned,
        "tt_time_extered": tt_time_entered,
        "avg_time_submit_assign": avg_time_submit_assign,
        "avg_time_assign_close": avg_time_assign_close,
        "avg_time_submit_close": avg_time_submit_close,

    }
    return response


# Get Encrypted ID For Template Link End#


# Export ticket to XLSX Start#

@register.simple_tag
def export_tickets_xls(tickets):
    """
    Downloads all movies as Excel file with a worksheet for each movie category
    """
    category_queryset = Ticket.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-TicketSearchList.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Delete the default worksheet
    workbook.remove(workbook.active)

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Define the column titles and widths
    columns = [
        ('ID', 8),
        ('Title', 40),
        ('Description', 80),
        ('Length', 15),
        ('Rating', 15),
        ('Price', 15),
    ]

    # Iterate through movie categories
    for ticket_id, ticket in enumerate(category_queryset):
        # Create a worksheet/tab with the title of the category
        worksheet = workbook.create_sheet(
            title=ticket.priority.priority_name,
            index=ticket_id,
        )
        # Define the background color of the header cells
        fill = PatternFill(
            start_color='21316f',
            end_color='21316f',
            fill_type='solid',
        )
        row_num = 1

        # Assign values, styles, and formatting for each cell in the header
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.border = border_bottom
            cell.alignment = centered_alignment
            cell.fill = fill
            # set column width
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = column_width

        # Iterate through all movies of a category
        for movie in category_queryset:
            row_num += 1

            # Define data and formats for each cell in the row
            row = [
                (movie.ticket_id, 'Normal'),
                (movie.priority.priority_name, 'Normal'),
                (movie.priority.priority_name, 'Normal'),
                # (timedelta(minutes=movie.ticket_created_at), 'Normal'),
                (movie.priority.priority_name, 'Normal'),
                (movie.priority.priority_name, 'Normal'),
                (movie.ticket_type.ttype_name, 'Normal'),
            ]

            # Assign values, styles, and formatting for each cell in the row
            for col_num, (cell_value, cell_format) in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.style = cell_format
                # if cell_format == 'Currency':
                #     cell.number_format = '#,##0.00 €'
                # if col_num == 4:
                #     cell.number_format = '[h]:mm;@'
                cell.alignment = wrapped_alignment

        # freeze the first row
        worksheet.freeze_panes = worksheet['A2']

        # set tab color
        worksheet.sheet_properties.tabColor = '21316f'

    workbook.save(response)

    return response


# Get Correct Filter Expression Start#


@register.simple_tag
def get_correct_filter_expression(expression):
    express = expression.split(':')
    result = '<b>' + express[0] + ':</b>' + express[1] + ''
    return mark_safe(result)


# Get Correct Filter Expression End#

@register.simple_tag
def convert_list_into_json(expression):
    return json.loads(expression)


# Get Encrypted ID For Template Link Start#


@register.simple_tag
def get_data_type_formats(id):
    response = ReportSettingFormat.objects.filter().filter(format_data_type_id=id)
    return response


# Get Encrypted ID For Template Link End#

# Select Report Setting Format By DataType ID Start#
@register.simple_tag
def select_report_setting_format_by_dataType(dtID, formatID , user_id):
    response = SaveReportSetting.objects.filter(rpt_setting_created_by_id = user_id).filter(rpt_setting_dataType_id=dtID).first()
    if response:
        if response.rpt_setting_format_id == formatID:
            return 'selected'


# Select Report Setting Format By DataType ID End#

# Select Report Setting Justification By DataType ID Start#
@register.simple_tag
def select_report_setting_justification_by_dataType(dtID, JustificationID , user_id):
    print(user_id)
    print('user_id')
    response = SaveReportSetting.objects.filter(rpt_setting_created_by_id = user_id).filter(rpt_setting_dataType_id=dtID).first()
    if response:
        if response.rpt_setting_justification_id == JustificationID:
            return 'selected'


# Select Report Setting Justification By DataType ID End#

# Select Report Setting Width By DataType ID Start#
@register.simple_tag
def select_report_setting_width_by_dataType(dtID , user_id):
    response = SaveReportSetting.objects.filter(rpt_setting_created_by_id = user_id).filter(rpt_setting_dataType_id=dtID).first()
    if response:
        return response.rpt_setting_width


# Select Report Setting Width By DataType ID End#

# Get User Display Name Through User ID Start#
@register.simple_tag
def get_user_displayname_by_id(id):
    if id != '' and id != None:
        response = User.objects.values_list('display_name', flat=True).get(pk=id)
    else:
        response = 'No Submitter'
    return response


# Get User Display Name Through User ID End#


# Get Organization Name Through Org ID Start#
@register.simple_tag
def get_org_name_by_id(id):
    if id != '' and id != None:
        response = Organization.objects.values_list('org_name', flat=True).get(pk=id)
    else:
        response = 'No Submitter'
    return response


# Get Organization Name Through Org ID End#


@register.simple_tag
def get_pair_fields_by_pair(id):
    response = DataSetsPairFields.objects.filter(df_pair_id=id).exclude(df_actual_table_name = 'Client').order_by('df_name')
    return response


@register.simple_tag
def get_reopen_ticket_by_user(id):
    if id:
        try:
            response = User.objects.filter(is_delete=0).filter(is_active=1).get(id=id)
        except:
            response = None
    else:
        response = None
    return response


@register.simple_tag
def get_tickets_dateTime_by_timezone(dateTime, userID):
    if dateTime:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = dateTime.astimezone(pytz.timezone(uTimeZone))
        # print(datetime.strptime(str(local_dt), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%m/%d/%Y %I:%M %p'))
        return datetime.strptime(str(local_dt), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%m/%d/%Y %I:%M %p')


@register.simple_tag
def get_tickets_dateTime_by_timezone1(dateTime, userID):
    if dateTime is None:
        return ''
    else:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = dateTime.astimezone(pytz.timezone(uTimeZone))
        result = datetime.strptime(str(local_dt), '%Y-%m-%d %H:%M:%S%z').strftime('%m/%d/%Y %I:%M %p')
        print(result)
        return result
        # uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        # tempdateTime = str(date)+' '+str(time)+'.000001+00:00'
        # dateTime = datetime.strptime(tempdateTime, '%Y-%m-%d %H:%M:%S.%f%z')
        # local_dt = dateTime.astimezone(pytz.timezone(uTimeZone))
        # return datetime.strptime(str(local_dt), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %H:%M %p')


@register.simple_tag
def getLabourHoursByTicketId(ticket_id):
    result = TicketNote.objects.filter(note_ticket=ticket_id).filter(note_is_delete=0).aggregate(
        thours=Sum(Cast('tnote_laborhour_hours', IntegerField())) + Sum(
            Cast('tnote_laborhour_minutes', IntegerField())) / 60,
        tminutes=Sum(Cast('tnote_laborhour_minutes', IntegerField())) % 60
    )
    if result:
        if result['thours'] is None or str(result['thours']) == '0':
            hours = '00'
        else:
            hours = "{:02d}".format(result['thours'])
        if result['tminutes'] is None or str(result['tminutes']) == '0':
            minutes = '00';
        else:
            minutes = "{:02d}".format(result['tminutes'])
        labour_hours = str(hours) + ':' + str(minutes)
        return labour_hours
    else:
        return '00:00'


@register.simple_tag
def getTotalTimeOpenFormat(totaltime):
    if totaltime:
        ttime = str(totaltime).split('.')
        ttime_hours = int(ttime[0]) + int(ttime[1]) / 60
        ttime_minutes = int(ttime[1]) % 60
        response = "{:02d}".format(int(ttime_hours)) + ':' + "{:02d}".format(int(ttime_minutes))
        return response
    else:
        return '00:00'


@register.simple_tag
def getTimeOpenDaysFormat(totaltime):
    if totaltime:
        tdays = totaltime / 24
        response = round(tdays, 2)
        return response
    else:
        return '0'


@register.simple_tag
def getFileTypeIconByFileName(file_name):
    if file_name:
        ext = file_name.split('.').pop();
        ext = ext.lower()
        if ext == "pdf":
            return "itrak/images/icons/dummypdf.jpg"
        elif ext == "doc":
            return "itrak/images/icons/dummyword.jpg"
        elif ext == "xls" or ext == "xlsx":
            return "itrak/images/icons/dummyxls.jpg"
        elif ext == "txt":
            return "itrak/images/icons/dummytext1.jpg"
        elif ext == "pptx":
            return "itrak/images/icons/dummyppt.jpg"
        else:
            return "itrak/images/icons/dummydefault.png"
    else:
        return "itrak/images/icons/dummydefault.png"


@register.simple_tag
def getFullTimezone(timeZone):
    local_dt = datetime.now().astimezone(pytz.timezone(timeZone))
    a = datetime.strptime(str(local_dt), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%Y-%m-%d %H:%M:%S %Z')
    a = a.split(' ')
    return a[2]


@register.simple_tag
def getTaskManagerByTicketId(ticket_id):
    if ticket_id:
        result = TaskManager.objects.filter(tmgr_ticket_id=ticket_id) \
            .order_by(Cast('tmgr_display_order', IntegerField()))
        return (result)


@register.simple_tag
def getReportDataTypeFormatsByFieldName(field_name, query_id):
    if field_name and query_id:
        try:
            datatype = DataSetsPairFields.objects.values('df_rpt_data_type_id').filter(df_name=field_name)[0][
                'df_rpt_data_type_id']
            print(datatype)
            if datatype and (datatype != 5 and datatype != 6):
                formats = ReportSettingFormat.objects.values('rpt_format_id', 'rpt_format_name').filter(
                    format_data_type_id=datatype)
                if len(formats):
                    return formats
                else:
                    return None
            else:
                return None
        except:
            return None


@register.simple_tag
def getReportDataTypeSubtotalByFieldName(field_name, query_id):
    if field_name and query_id:
        try:
            datatype = DataSetsPairFields.objects.values('df_rpt_data_type_id').filter(df_name=field_name)[0][
                'df_rpt_data_type_id']
            if datatype and (datatype != 4 and datatype != 6):
                return datatype
            else:
                return None
        except:
            return None


@register.simple_tag
def getSelectedTypeFormatByFieldName(field_name, query_id, format_id):
    if field_name and query_id:
        try:
            datatype = DataSetsPairFields.objects.values('df_rpt_data_type_id').filter(df_name=field_name)[0][
                'df_rpt_data_type_id']
            if datatype and (datatype != 5 and datatype != 6):
                try:
                    result = SaveReportSetting.objects.filter(rpt_setting_dataType_id=datatype).filter(
                        rpt_setting_format_id=format_id)
                    if len(result) > 0:
                        return 'selected'
                    else:
                        return None
                except:
                    return None
            else:
                return None
        except:
            return None


@register.simple_tag
def getSelectedTypeJustificationByFieldName(field_name, query_id, justify_id):
    if field_name and query_id:
        try:
            datatype = DataSetsPairFields.objects.values('df_rpt_data_type_id').filter(df_name=field_name)[0][
                'df_rpt_data_type_id']
            if datatype:
                try:
                    result = SaveReportSetting.objects.filter(rpt_setting_dataType_id=datatype).filter(
                        rpt_setting_justification_id=justify_id)
                    if len(result) > 0:
                        return 'selected'
                    else:
                        return None
                except:
                    return None
            else:
                return None
        except:
            return None


@register.simple_tag
def getReportDataTypeWidthByFieldName(field_name, query_id):
    if field_name and query_id:
        try:
            datatype = DataSetsPairFields.objects.values('df_rpt_data_type_id').filter(df_name=field_name)[0][
                'df_rpt_data_type_id']
            if datatype:
                try:
                    result = \
                    SaveReportSetting.objects.values('rpt_setting_width').filter(rpt_setting_dataType_id=datatype)[0][
                        'rpt_setting_width']
                    if result:
                        return result
                    else:
                        return 0
                except:
                    return 0
            else:
                return 0
        except:
            return 0


@register.simple_tag
def firstNoteAdded(user_id, ticket_id, ticket_created_at, ticket_caller_id):
    if ticket_caller_id != user_id:
        try:
            ticketsNotes = TicketNote.objects.filter(note_ticket=ticket_id).filter(note_is_delete=0)
            ticketcreated_at = ticket_created_at
            if ticketsNotes != None:
                tic = \
                TicketNote.objects.values('note_created_at').filter(note_ticket=ticket_id).filter(note_is_delete=0)[0][
                    'note_created_at']
                if tic and ticketcreated_at:
                    delta = tic - ticketcreated_at
                    days, seconds = delta.days, delta.seconds
                    hours = days * 24 + seconds // 3600
                    minutes = (seconds % 3600) // 60
                    avg_time_submit_close = "{:02d}".format(hours + int(minutes / 60)) + ':' + "{:02d}".format(
                        minutes % 60)
                    return avg_time_submit_close
                else:
                    return '00:00'
        except:
            return '00:00'
    else:
        return '00:00'


@register.simple_tag
def firstNoteAddedByAssignee(ticket_id, ticket_created_at, ticket_assign_to):
    try:
        ticketsNotes = TicketNote.objects.filter(note_is_delete=0).filter(note_ticket=ticket_id).filter(
            note_created_by=ticket_assign_to)
        print(ticketsNotes)
        ticketcreated_at = ticket_created_at
        if len(ticketsNotes) > 0:
            tic = \
            TicketNote.objects.values('note_created_at').filter(note_is_delete=0).filter(note_ticket=ticket_id)[0][
                'note_created_at']
            if tic and ticketcreated_at:
                delta = tic - ticketcreated_at
                days, seconds = delta.days, delta.seconds
                hours = days * 24 + seconds // 3600
                minutes = (seconds % 3600) // 60
                avg_time_submit_close = "{:02d}".format(hours + int(minutes / 60)) + ':' + "{:02d}".format(minutes % 60)
                return avg_time_submit_close
            else:
                return '00:00'
        else:
            return '00:00'
    except:
        return '00:00'


@register.simple_tag
def getAdjHoursByTicketId(ticket_id):
    ticket = Ticket.objects.get(pk=ticket_id)
    if ticket.submitted_at:
        if ticket.ticket_status == 0:
            delta = datetime.now(timezone.utc) - ticket.submitted_at
            days, seconds = delta.days, delta.seconds
            hours = days * 24 + seconds // 3600
            minutes = (seconds % 3600) // 60
        else:
            delta = ticket.ticket_closed_at - ticket.submitted_at
            days, seconds = delta.days, delta.seconds
            hours = days * 24 + seconds // 3600
            minutes = (seconds % 3600) // 60

        total_hours = round((hours) + (minutes / 100), 2)
        if total_hours:
            print(total_hours)
            ttime = str(total_hours).split('.')
            ttime_hours = int(ttime[0]) + int(ttime[1]) / 60
            ttime_minutes = int(ttime[1]) % 60
            response = "{:02d}".format(int(ttime_hours)) + ':' + "{:02d}".format(int(ttime_minutes))
            return response
        else:
            return '00:00'
    else:
        return '00:00'


@register.simple_tag
def get_tickets_attachments(ticket_id):
    attach_ids = TicketAttachments.objects.filter(attach_ticket_id=ticket_id).filter(attach_is_delete=0)
    response = attach_ids
    return response


@register.simple_tag
def getGroupSortByIndex(groups_sorting, index):
    result = groups_sorting[index]
    return result


@register.simple_tag
def getFieldFormatingByColumnName(field_formatings, index):
    column_name = field_formatings[index]['actual_column_name']
    keyValList = []
    keyValList.append(column_name)
    result = list(filter(lambda d: d['actual_column_name'] in keyValList, field_formatings))[0]
    return result


@register.simple_tag
def getformatingValue(value, formatId):
    if formatId == '11':
        dateformat = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%d %B, %Y')
        return dateformat
    if formatId == '10':
        dateformat = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%A, %B,%d,%Y')
        return dateformat
    if formatId == '12':
        dateformat = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y')
        return dateformat
    if formatId == '13':
        dateformat = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y')
        return dateformat
    if formatId == '8':
        return value
    if formatId == '9':
        boolformat = ("Yes", "No")[value]
        return boolformat
    return value


@register.simple_tag
def getformatingValue1(value, formatId, key):
    if formatId == '1':
        frmt = value
        return frmt
    if formatId == '2':
        frmt = format(int(value), ',d')
        return frmt
    if formatId == '3':
        frmt = value + "%"
        return frmt
    # if formatId == '11':
    #     dateformat = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%d %B, %Y')
    #     return dateformat
    # if formatId == '10':
    #     dateformat = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%A, %B,%d,%Y')
    #     return dateformat
    if formatId == '10':
        stringDate=str(value)
        stringDate=stringDate[:19]
        dateformat = datetime.strptime(stringDate,'%Y-%m-%d %H:%M:%S')
        dateformat= dateformat.strftime('%A, %B,%d,%Y')
        return dateformat
    if formatId == '12':
        dateformat = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y')
        return dateformat
    if key == 'ticketAttach__attach_created_at':
        if value != None:
            value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
            return value1
        else:
            return ''
    if key == 'client_created_at':
        if value != None:
            value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
            return value1
        else:
            return ''
    if key == 'ticket_assign_at':
        if value != None:
            value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
            return value1
        else:
            return ''
    if key == 'ticket_next_action_at':
        if value != None:
            value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
            return value1
        else:
            return ''
    if key == 'submitted_at':
        if value != None:
            value1 = datetime.strptime(str(value), '%Y-%m-%d %H:%M:%S%z').strftime('%m/%d/%Y %I:%M %p')
            return value1
        else:
            return ''
    if key == 'ticket_created_at':
        if value != None:
            value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
            return value1
        else:
            return ''
    return (value)


@register.simple_tag
def getformatingValue2(key, value):
    return 'abc'


@register.simple_tag
def getTimeformating(value, key, userID):
    print(key)
    print(value)
    if key == 'ticket_created_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'ticketOrg__ticket_closed_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'ticket_is_reopen_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'ticket_closed_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'ticketNote__note_modified_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'ticketNote__note_created_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'ticketOrg__ticket_assign_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'ticketOrg__ticket_next_action_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'qb_created_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'qb_modified_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'rbQuerySetsQuery__rb_created_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'rbQuerySetsQuery__rb_modified_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'ticketManager__tmgr_completion_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'ticket_modified_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'created_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'submitted_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt), '%Y-%m-%d %H:%M:%S%z').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'ticketAttach__attach_created_at' and value is not None:
        if value != None:
            value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
            return value1
        else:
            return ''
    if key == 'UserOrgId__created_at' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'UserOrgId__last_login' and value is not None:
        uTimeZone = MySettings.objects.filter(m_user_id=userID).first().m_time_zone
        local_dt = value.astimezone(pytz.timezone(uTimeZone))
        value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
        return value1
    if key == 'client_created_at' and value is not None:
        if value != None:
            value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
            return value1
        else:
            return ''
    if key == 'ticket_assign_at' and value is not None:
        if value != None:
            value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
            return value1
        else:
            return ''
    if key == 'ticket_next_action_at' and value is not None:
        if value != None:
            value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
            return value1
        else:
            return ''
    if key == 'submitted_at' and value is not None:
        if value != None:
            value1 = datetime.strptime(str(value), '%Y-%m-%d %H:%M:%S%z').strftime('%m/%d/%Y %I:%M %p')
            return value1
        else:
            return ''
    if key == 'ticketOrg__submitted_date' and value is not None:
        if value != None:
            value1 = datetime.strptime(str(value), '%Y-%m-%d').strftime('%m/%d/%Y')
            return value1
        else:
            return ''
    if key == 'ticket_created_at' and value is not None:
        if value != None:
            value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y %I:%M %p')
            return value1
        else:
            return ''
    if key == 'ticket_next_action_by__display_name' and value is not None:
        if value != None:
            return value
        else:
            return ''

    return value


@register.simple_tag
def getqbqueryRecordCheckingEmpty(value, key):
    if key and value:
        print(key)
        print(value)
        return value


@register.simple_tag
def getGroupFieldFormatingByColumnName(query_result, group_sorting_array, selected_fields_formating, qb_query_pair_id,
                                       groupFilterArgs):
    keyValList = group_sorting_array['field_name']
    records_index = list(filter(lambda d: d['actual_column_name'] in keyValList, selected_fields_formating))[0]
    records_index = selected_fields_formating.index(records_index)
    records_display = list(filter(lambda d: d['actual_column_name'] not in keyValList, selected_fields_formating))
    kwargs = {}
    for groupFilterArg in groupFilterArgs:
        for key, value in groupFilterArg.items():
            kwargs.setdefault(key, value)
    group_selected_columns = []
    sortargs = []
    try:
        actual_field = \
        DataSetsPairFields.objects.values('df_actual_column_name').filter(df_pair_id=qb_query_pair_id).filter(
            df_name=group_sorting_array['field_name'])[0]['df_actual_column_name']
    except:
        return render_to_response('itrak/page-404.html')
    group_selected_columns.append(actual_field)

    grouped_records = []

    # Setting Sort Order Asc and Desc For Group Sort
    if group_sorting_array['sort_order'] == '0':
        sort_value = actual_field
    else:
        sort_value = '-' + actual_field
    sortargs.append(sort_value)

    grouped_records = query_result.values(*tuple(group_selected_columns)).filter(**kwargs).annotate(
        tcount=Count(actual_field)).order_by(*sortargs)

    # Setting Justification, Column Width, Format
    keyValList = group_sorting_array['field_name']
    group_format_array = list(filter(lambda d: d['actual_column_name'] in keyValList, selected_fields_formating))[0]

    # Setting Check For Header/Footer for Display
    for record in grouped_records:
        record['group_record_name'] = list(record.items())[0][1]
        record['group_record_column'] = list(record.items())[0][0]
        record['header'] = group_sorting_array['header']
        record['footer'] = group_sorting_array['footer']
        record['colspan'] = len(records_display) - records_index
        if group_format_array:
            record['justification'] = group_format_array['justification']
            record['column_width'] = group_format_array['column_width']
            record['format'] = group_format_array['format']

    return grouped_records


@register.simple_tag
def getGroupFieldRecordsDictByIndex(record):
    return record


@register.simple_tag
def getReportRecordsWithinGroupField(records):
    return records


@register.simple_tag
def getGroupChildFields(group_record, query_result, selected_columns, group_index, groupFilterArgs):
    # print(group_index)
    # print(group_record)
    # print(groupFilterArgs)
    kwargs = {}
    for groupFilterArg in groupFilterArgs:
        for key, value in groupFilterArg.items():
            kwargs.setdefault(key, value)

    # kwargs = {
    #     '{0}__{1}'.format(group_record['group_record_column'], 'iexact'): group_record['group_record_name'],
    # }
    # group_record_index = selected_columns.index(group_record['group_record_column']) + 1
    # selected_columns = selected_columns[group_record_index:]
    finalquery = query_result.values(*tuple(selected_columns)).filter(**kwargs)
    group_records = finalquery
    # print(query_result)
    return group_records


@register.simple_tag
def getorgidByTicketlist(records):
    totalrecord = records.values('note_ticket__ticket_org_id', 'note_ticket__ticket_org__org_name', 'hours',
                                 'minutes').distinct()

    totalrecord = list({v['note_ticket__ticket_org_id']: v for v in totalrecord}.values())
    return totalrecord


@register.simple_tag
def labourgetTicketByID(tickets, orgID):
    totallabourrecord = tickets.filter(note_ticket__ticket_org_id=str(orgID['note_ticket__ticket_org_id']))
    return totallabourrecord


@register.simple_tag
def SumlabourgetTicketByID(tickets, orgID):
    result = tickets.filter(note_ticket__ticket_org_id=str(orgID['note_ticket__ticket_org_id'])).aggregate(
        thours=Sum(Cast('hours', IntegerField())) + Sum(Cast('minutes', IntegerField())) / 60,
        tminutes=Sum(Cast('minutes', IntegerField())) % 60
    )
    if result:
        if result['thours'] is None or str(result['thours']) == '0':
            hours = '00'
        else:
            hours = "{:02d}".format(result['thours'])
        if result['tminutes'] is None or str(result['tminutes']) == '0':
            minutes = '00';
        else:
            minutes = "{:02d}".format(result['tminutes'])
        labour_hours = str(hours) + ':' + str(minutes)
        return labour_hours
    else:
        return '00:00'


# Site appearances functions
@register.filter(name='getSiteAppearanceLeftLogoFile')
@register.simple_tag
def getSiteAppearanceLeftLogoFile(param , param2):
    print(param)
    print(param2)
    data = SiteAppearance.objects.filter(site_org_id=param).values_list('upload_left_logo', flat=True).first()
    print(data)
    return data     

@register.filter(name='getSiteAppearanceLeftLogoURL')
@register.simple_tag
def getSiteAppearanceLeftLogoURL(param,param2):
    print(param)
    data = SiteAppearance.objects.filter(site_org_id=param).values_list('left_logo_url', flat=True).first()
    # print(data)
    return data      

@register.filter(name='getSiteAppearanceRightLogoFile')
@register.simple_tag
def getSiteAppearanceRightLogoFile(param,param2):
    data = SiteAppearance.objects.filter(site_org_id=param).values_list('upload_right_logo', flat=True).first()

    return data     

@register.filter(name='getSiteAppearanceRightLogoURL')
@register.simple_tag
def getSiteAppearanceRightLogoURL(param,param2):
    data = SiteAppearance.objects.filter(site_org_id=param).values_list('right_logo_url', flat=True).first()

    return data   

@register.filter(name='getSiteAppearanceSiteTitle')
@register.simple_tag
def getSiteAppearanceSiteTitle(param,param2):
    data = SiteAppearance.objects.filter(site_org_id=param).values_list('site_title', flat=True).first()

    if data:
        return data
    else:
        return 'ATG | Extra'

@register.filter(name='getSiteAppearanceFavicon')
@register.simple_tag
def getSiteAppearanceFavicon(param,param2):
    data = SiteAppearance.objects.filter(site_org_id=param).values_list('upload_favicon', flat=True).first()

    return data   

@register.filter(name='getSiteAppearanceLoginScreen')
@register.simple_tag
def getSiteAppearanceLoginScreen(param,param2):
    data = SiteAppearance.objects.filter(site_org_id=param).values_list('login_screen', flat=True).first()
    if data:
        return data
    else:
        return ''
      
@register.filter(name='getSiteAppearanceHomeScreen')
@register.simple_tag
def getSiteAppearanceLoginScreen(param,param2):
    data = SiteAppearance.objects.filter(site_org_id=param).values_list('home_screen', flat=True).first()

    if data:
        return data
    else:
        return ''

@register.filter(name='getSiteAppearanceHomeAgent')
@register.simple_tag
def getSiteAppearanceHomeAgent(param,param2):
    data = SiteAppearance.objects.filter(site_org_id=param).values_list('home_agent', flat=True).first()

    if data:
        return data
    else:
        return ''

@register.filter(name='checkUserHasAlreadyPermisisonsFromGroups')
@register.simple_tag
def checkUserHasAlreadyPermisisonsFromGroups(user_id,perm_act):
    print(user_id)
    print(perm_act)
    allGroupUsersPermissions = GroupActionPermission.objects.raw('''
        SELECT 
            1 AS group_act_per_id
            ,COUNT(*) as total
        FROM AT_GroupActionPermissions GRP
        WHERE GRP.perm_act_id = %s
        AND GRP.group_id in 
        (
            Select JOG.m_group_id
            From  AT_UserGroupMembership JOG with(nolock)
            Where JOG.m_user_id = %s
        )
        '''
        ,[perm_act,user_id]
    )
    isPermissionsAlreadyGivenByGroup = False
    if allGroupUsersPermissions:
        for allGroupUsersPermissions in allGroupUsersPermissions:
            if allGroupUsersPermissions.total > 0:
                print("basit")
                isPermissionsAlreadyGivenByGroup = True
    return isPermissionsAlreadyGivenByGroup
      
      
