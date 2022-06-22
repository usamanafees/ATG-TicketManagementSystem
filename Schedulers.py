import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE','ATG_itrak.settings')
django.setup()
# from django.conf import settings
from django.core.exceptions import PermissionDenied
from django.shortcuts import render,reverse, redirect, get_object_or_404, render_to_response
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponseRedirect, HttpResponse, Http404, JsonResponse, HttpResponseBadRequest, request
from itrak.models import *
from django_datatables_view.base_datatable_view import BaseDatatableView
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
import pytz
from django.core.mail import EmailMessage
from django.core.mail import EmailMultiAlternatives

import time
import smtplib
import mimetypes
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.message import Message
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.text import MIMEText
from django.utils.crypto import get_random_string
from django.conf.urls import url
from django.template.loader import render_to_string, get_template
from itrak.views.Load import *
from itrak.views.Email import *
from django.db.models.query import QuerySet
from django.core import signing
from datetime import datetime, timezone, timedelta
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
import json
import operator
from functools import reduce
from django.db.models import Q, Count, F, Func, Sum
from django.apps import apps
from django.db import transaction, IntegrityError
from copy import deepcopy
from django.db.models import Case, F, FloatField, IntegerField, Sum, When
from django.db.models.functions import Cast
from django.db.models.expressions import RawSQL
from operator import itemgetter
import collections
from django.utils.html import escape
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter


def qbQueryProcess(request, queryId=None, call_reportProcessFun=False , user_id=0 ):
    # user_type = 'superadmin'
    user_types = User.objects.get(pk=user_id)
    # for user in user_types:
    print(user_types.admin)
    userType = ''
    if(user_types.admin == True and user_types.user_type == '0' ): #Super Admin
        userType = 'superadmin'
    elif(user_types.admin == False and user_types.user_type == '0' ):#Agent 
        userType = 'agent'
    elif(user_types.user_type == '1' and check_action_permission("da_Can_view_Tickets_submitted_by_other_users",user_id) == True ):#End User - Manager
        userType = 'manager'
    elif(user_types.user_type == '1' and check_action_permission("da_Can_view_Tickets_submitted_by_other_users",user_id) == False ):#End User
        userType = 'enduser'
    if queryId:
        query_id = queryId
    else:
        query_id = request.GET.get('qbQuery')

    # Get List Of Associated Accounts
    # accountsList = getAccountIDsOfCurrentUser(request)
    print(user_id)
    if userType == 'superadmin' or userType == 'agent':
        SQL  = """
            select distinct account_id
            from AT_UserAccountRelation a
            where 1=1
            AND a.user_id = '"""+str(user_id)+"""'
        """
    elif userType == 'manager' or userType == 'enduser':
        SQL  = """
            select distinct account_id
            from AT_UserAccountRelation a
            where 1=1
            AND a.user_id = '"""+str(user_id)+"""'
            AND (
                SELECT COUNT(*)
                FROM AT_Users AU
                WHERE AU.ID = A.user_id 
                AND AU.user_type = 1 -- END USER
            )> 0 
        """
    
    cursor = connection.cursor()
    cursor.execute(SQL)
    accounts = dictfetchall(cursor) 

    accountsList = []
    for account in accounts:
        accountsList.append(account['account_id'])
    if len(accountsList) > 0:
        accountFilter = {'account_id__in':accountsList}
    else: 
        accountFilter = {}
    
    # user_id = user_id
    # user_type = userType(request) 
    # accountsList = getMappedUserIDsWithCurrentUer(request)
    SQL  = """
        select distinct user_id
        from AT_UserAccountRelation a
        where 1=1
        AND (
            SELECT COUNT(*)
            FROM AT_Users AU
            WHERE AU.id = a.user_id
            AND AU.user_type = 1 --END USER
        )>0
        AND a.account_id in (
            select account_id
            from AT_UserAccountRelation b
            where b.user_id = '"""+str(user_id)+"""'
        )
    """
    
    cursor = connection.cursor()
    cursor.execute(SQL)
    users = dictfetchall(cursor) 

    usersList = []
    for user in users:
        usersList.append(user['user_id'])
    print(usersList)
    if userType == 'manager':
        userFilter = Q(ticket_assign_to_id__in = usersList) | Q(ticket_caller_id__in = usersList) | Q(ticket_next_action_id__in=usersList) | Q(ticket_created_by_id__in=usersList)
    elif userType == 'enduser':
        userFilter = Q(ticket_assign_to_id=user_id) | Q(ticket_caller_id=user_id) | Q(ticket_next_action_id=user_id) | Q(ticket_created_by_id=user_id)
    else:
        userFilter = Q(pk__isnull=False)

    context = {}
    obj = SavedQBQuries.objects.get(pk=query_id)
    
    query_name = obj.qb_query_name
    query_description = obj.qb_query_description
    qb_query_pair_id = obj.qb_query_pair_id
    

    filter_expression_array = json.loads(obj.qb_filter_expression_array)
    filter_statement = (obj.qb_filter_statement).split(" ") if obj.qb_filter_statement else ''
    selected_fields = json.loads(obj.qb_selected_fields)
    
    selected_columns = []
    argument_list = []
    conditiondict = {
        "=": "exact",
        "<>": "exact",
        "In": "in",
        "Not In": "in",
        "Like": "contains",
        "Not Like": "contains",
        "Is Blank": "exact",
        "Is Not Blank": "exact",
        "Is Null": "isnull",
        "Is Not Null": "isnull",
        "Is True": "exact",
        "Is False": "exact",
        ">": "gt",
        "<": "lt",
        ">=": "gte",
        "<=": "lte",
        "Starts with": "startswith",
        "Starts Not with": "startswith",
        "Today": "range",
        "Yesterday": "range",
        "Tomorrow": "range",
        "This Week": "range",
        "Last Week": "range",
        "Next Week": "range",
        "This Month": "range",
        "Last Month": "range",
        "Next Month": "range",
        "This Year": "range",
        "Last Year": "range",
        "Next Year": "range",
        "Custom Date Range": "range"
    }

    for field in selected_fields:
        try:
            actual_field = \
            DataSetsPairFields.objects.values('df_actual_column_name').filter(df_pair_id=qb_query_pair_id).filter(
                df_name=field)[0]['df_actual_column_name']
        except:
            return render_to_response('itrak/page-404.html')
        selected_columns.append(actual_field)
    
    counter = 0
    filter_operator = 'and'
    query_commit = 0
    query_not_commit = 0
    compound_filter = ''
    if filter_statement:
        if len(filter_statement) == 1:
            for key in filter_statement:
                # try:
                if len(key) == 1:
                    expression = filter_expression_array[key]
                    modelName = expression['modelName']
                    modelAttr = expression['modelAttr']
                    condition = expression['condition']
                    fieldValue = expression['fieldValue']
                    fieldLabel = expression['fieldLabel']

                    # print(expression)
                    # print(modelName)
                    # print(modelAttr)
                    # print(condition)
                    # print(fieldValue)
                    # print(fieldLabel)

                    modelInstance = apps.get_model('itrak', modelName)
                    print(modelInstance)
                    # print(selected_columns)
                    if modelName == 'Ticket':
                        finalquery = ModelInstance.objects.values(*tuple(selected_columns)).filter().filter(**accountFilter).filter(userFilter)
                    else:
                        finalquery = ModelInstance.objects.values(*tuple(selected_columns)).filter()
                    # print(modelInstance)
                    if condition == '<>' or condition == 'Not In' or condition == 'Not Like' or condition == 'Is Not Blank' or condition == 'Is Not Null' or condition == 'Is False' or condition == 'Starts Not with':
                        argument_list.append(~Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                    else:
                        argument_list.append(Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                    tinalquery = Finalquery.filter(reduce(operator.and_, argument_list))
                else:
                    return render_to_response('itrak/page-404.html')
            # except:
            #     return render_to_response('itrak/page-404.html')
        else:
            if filter_statement.__contains__('(') or filter_statement.__contains__(')'):
                try:
                    modelName = DataSetsPairFields.objects.only('df_primary_table_name').get(
                        df_actual_column_name=selected_columns[0]).df_primary_table_name
                except:
                    return render_to_response('itrak/page-404.html')
                modelInstance = apps.get_model('itrak', modelName)
                if modelName == 'Ticket':
                    finalquery = modelInstance.objects.values(*tuple(selected_columns)).filter().filter(**accountFilter).filter(userFilter)
                else:
                    finalquery = modelInstance.objects.values(*tuple(selected_columns)).filter()
                compound_filter = getCompoundFilterForQuery(filter_statement, filter_expression_array, conditiondict)
                print(compound_filter)
            else:
                for key in filter_statement:
                    # print(key)
                    # print(filter_operator)
                    try:
                        if len(key) == 1:
                            key = key.upper()
                            # print(key)
                            # print(filter_expression_array[key])
                            expression = filter_expression_array[key]
                            modelName = expression['modelName']
                            modelAttr = expression['modelAttr']
                            condition = expression['condition']
                            fieldValue = expression['fieldValue']
                            fieldLabel = expression['fieldLabel']

                            # print(expression)
                            # print(modelName)
                            # print(modelAttr)
                            # print(condition)
                            # print(fieldValue)
                            # print(fieldLabel)

                            modelInstance = apps.get_model('itrak', modelName)
                            # print(modelInstance)
                            # print(selected_columns)

                            # Condition Implementation For NOT Case AND Condition BASED
                            if query_not_commit == 0:
                                if condition == '<>' or condition == 'Not In' or condition == 'Not Like' or condition == 'Is Not Blank' or condition == 'Is Not Null' or condition == 'Is False' or condition == 'Starts Not with':
                                    argument_list.append(
                                        ~Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                                else:
                                    argument_list.append(Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                            elif query_not_commit == 1:
                                query_not_commit = 0
                                if condition == '<>' or condition == 'Not In' or condition == 'Not Like' or condition == 'Is Not Blank' or condition == 'Is Not Null' or condition == 'Is False' or condition == 'Starts Not with':
                                    argument_list.append(Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                                else:
                                    argument_list.append(
                                        ~Q(**{modelAttr + '__' + conditiondict[condition]: fieldValue}))
                                    print(argument_list)

                            # First Iteration of Filter Expression
                            if counter == 0:
                                if modelName == 'Ticket':
                                    finalquery = modelInstance.objects.values(*tuple(selected_columns)).filter().filter(**accountFilter).filter(userFilter)
                                else:
                                    finalquery = modelInstance.objects.values(*tuple(selected_columns)).filter()
                                compound_filter = argument_list[0]
                                argument_list.clear()

                            counter = counter + 1
                            # First Increment in Filter Expression
                            if filter_operator == 'or' and query_commit == 1:
                                argument_list.insert(0, compound_filter)
                                compound_filter = argument_list[0].__or__(argument_list[1])
                                # finalquery = finalquery.filter(reduce(operator.or_, argument_list))
                                argument_list.clear()
                            elif filter_operator == 'and' and query_commit == 1:
                                argument_list.insert(0, compound_filter)
                                compound_filter = argument_list[0].__and__(argument_list[1])
                                print(compound_filter)
                                # finalquery = finalquery.filter(reduce(operator.and_, argument_list))
                                argument_list.clear()
                        elif key.lower() == 'and' or key.lower() == 'or':
                            filter_operator = key.lower()
                            query_commit = 1
                        elif key.lower() == 'not':
                            query_not_commit = 1
                        else:
                            return render_to_response('itrak/page-404.html')
                    except:
                        return render_to_response('itrak/page-404.html')
            if modelName == 'Ticket':
                finalquery = finalquery.filter(compound_filter).filter(**accountFilter).filter(userFilter)
            else:
                finalquery = finalquery.filter(compound_filter)
    else:
        try:
            modelName = \
            DataSetsPairFields.objects.values('df_primary_table_name').filter(df_pair_id=qb_query_pair_id).filter(
                df_actual_column_name=selected_columns[0])[0]['df_primary_table_name']
        except:
            return render_to_response('itrak/page-404.html')
        modelInstance = apps.get_model('itrak', modelName)
        if modelName == 'Ticket':
            finalquery = modelInstance.objects.values(*tuple(selected_columns)).filter().filter(**accountFilter).filter(userFilter)
        else:
            finalquery = modelInstance.objects.values(*tuple(selected_columns)).filter()
        
    if call_reportProcessFun:
        return finalquery
    
    timezone = MySettings.objects.filter(m_user_id=user_id).first()
    eastern = pytz.timezone(timezone.m_time_zone)
    time = datetime.now(eastern)
    time = time.strftime('%m/%d/%Y %I:%M %p')
    context = {
        'query_name': query_name,
        'query_description': query_description,
        'selected_fields': selected_fields,
        'time':time,
        'query_id':query_id,
        'records': finalquery
    }

    return render(request, 'itrak/Reports/qb_query_result_list.html', context)
def runDailyScheduleJob(request):
    scheduledReports = ScheduledReport.objects.filter(sch_rpt_is_delete= 0).filter(is_active= 1).filter(schedule = 1)
    res = 1
    if scheduledReports:
        print(scheduledReports)
        print('here1')
        for scheduledReport in scheduledReports:
            print('here2') , print(scheduledReport.sch_rpt_saved_search_id)
            # For Saved Search
            print(scheduledReport.sch_rpt_saved_search_id)
            if scheduledReport.sch_rpt_saved_search_id:
                print(scheduledReport.sch_rpt_saved_search_id)
                obj = TicketSavedSearch.objects.get(pk=scheduledReport.sch_rpt_saved_search_id)
                print('here3')
                kwargs = {
                    '{0}__{1}'.format('ticket_is_delete', 'iexact'): 0,
                    '{0}__{1}'.format('ticket_is_active', 'iexact'): 1,
                }
                ticket_status_dict = {0: "Opened", 1: "Closed"}
                labor_hours_dict = {"0": "Less Than", "1": "More Than", "2": "Equal"}
                fielddict = {}
                # fielddict = {"1": ''}
                if obj.ticket_status != None:
                    ticket_status = obj.ticket_status
                    kwargs.setdefault('ticket_status__iexact', ticket_status)
                    fielddict.update({'Ticket Status': ticket_status_dict[ticket_status]})
                if obj.ticket_sub_status_id != None:
                    ticket_sub_status = obj.ticket_sub_status_id
                    kwargs.setdefault('ticket_sub_status_id', ticket_sub_status)
                    result = SubStatus.objects.only('sub_status_text').get(pk=ticket_sub_status).sub_status_text
                    if obj.ticket_status != None:
                        fielddict.update({'Ticket Status': ticket_status_dict[ticket_status] + ' - ' + result})
                    else:
                        fielddict.update({'Ticket Status': ' - ' + result})
                if obj.priority_id != None:
                    priority = obj.priority_id
                    kwargs.setdefault('priority_id', priority)
                    result = Priority.objects.only('priority_name').get(pk=priority).priority_name
                    fielddict.update({'Priority': result})
                if obj.ticket_type_id != None:
                    ticket_type = obj.ticket_type_id
                    kwargs.setdefault('ticket_type_id', ticket_type)
                    result = TicketType.objects.only('ttype_name').get(pk=ticket_type).ttype_name
                    fielddict.update({'Ticket Type': result})
                if obj.ticket_subtype1_id != None:
                    subtype1 = obj.ticket_subtype1_id
                    kwargs.setdefault('ticket_subtype1_id', subtype1)
                    result = TicketType.objects.only('ttype_name').get(pk=subtype1).ttype_name
                    fielddict.update({'Subtype 1': result})
                if obj.ticket_subtype2_id != None:
                    subtype2 = obj.ticket_subtype2_id
                    kwargs.setdefault('ticket_subtype2_id', subtype2)
                    result = TicketType.objects.only('ttype_name').get(pk=subtype2).ttype_name
                    fielddict.update({'Subtype 2': result})
                if obj.ticket_subtype3_id != None:
                    subtype3 = obj.ticket_subtype3_id
                    kwargs.setdefault('ticket_subtype3_id', subtype3)
                    result = TicketType.objects.only('ttype_name').get(pk=subtype3).ttype_name
                    fielddict.update({'Subtype 3': result})
                if obj.ticket_subtype4_id != None:
                    subtype4 = obj.ticket_subtype4_id
                    kwargs.setdefault('ticket_subtype4_id', subtype4)
                    result = TicketType.objects.only('ttype_name').get(pk=subtype4).ttype_name
                    fielddict.update({'Subtype 4': result})
                if obj.subject != '':
                    subject = obj.subject
                    kwargs.setdefault('subject__icontains', subject)
                    fielddict.update({'Subject': subject})
                if obj.ticket_note != '':
                    ticket_note = request.POST.get('ticket_note')
                    kwargs.setdefault('ticketNote__note_detail__icontains', ticket_note)
                    fielddict.update({'Notes': ticket_note})
                if obj.all_three != '':
                    all_three = request.POST.get('all_three')
                    args = Q(subject__icontains=all_three) | Q(ticketNote__note_detail__icontains=all_three)
                    fielddict.update({'Search All Three': all_three})
                if obj.ticket_record_locator != '':
                    record_locator = obj.ticket_record_locator
                    kwargs.setdefault('ticket_record_locator__startswith', record_locator)
                    fielddict.update({'Record Locator': record_locator})
                if obj.ticket_caller_name != '':
                    caller_name = obj.ticket_caller_name
                    kwargs.setdefault('ticket_caller_name__startswith', caller_name)
                    fielddict.update({'Caller Name': caller_name})
                if obj.ticket_caller_phone != '':
                    caller_phone = obj.ticket_caller_phone
                    kwargs.setdefault('ticket_caller_phone__startswith', caller_phone)
                    fielddict.update({'Caller Phone': caller_phone})
                if obj.ticket_caller_email != '':
                    caller_email = obj.ticket_caller_email
                    kwargs.setdefault('ticket_caller_email__startswith', caller_email)
                    fielddict.update({'Caller Email': caller_email})
                if obj.ticket_passenger_name != '':
                    passenger_name = request.POST.get('passenger_name')
                    kwargs.setdefault('ticket_passenger_name__startswith', passenger_name)
                    fielddict.update({'Passenger Name': passenger_name})
                if obj.note_entered_by_id != None:
                    note_entered_by = request.POST.get('note_entered_by')
                    kwargs.setdefault('ticketNote__note_created_by_id', note_entered_by)
                    result = User.objects.only('display_name').get(pk=note_entered_by).display_name
                    fielddict.update({'Note Entered By': result})
                if obj.submitted_by_id != None:
                    submitted_by = request.POST.get('submitted_by')
                    kwargs.setdefault('ticket_caller_id', submitted_by)
                    result = User.objects.only('display_name').get(pk=submitted_by).display_name
                    fielddict.update({'Submitted By': result})
                if obj.entered_by_id != None:
                    entered_by = request.POST.get('entered_by')
                    kwargs.setdefault('ticket_created_by_id', entered_by)
                    result = User.objects.only('display_name').get(pk=entered_by).display_name
                    fielddict.update({'Entered By': result})
                if obj.task_assigned_to_id != None and obj.ever_assigned == 0:
                    assigned_to = obj.task_assigned_to_id
                    kwargs.setdefault('ticket_assign_to_id', assigned_to)
                    result = User.objects.only('display_name').get(pk=assigned_to).display_name
                    fielddict.update({'Assigned To': result})
                if obj.task_assigned_to_id != None and obj.ever_assigned == 1:
                    assigned_to = obj.task_assigned_to_id
                    kwargs.setdefault('ticketURLog__urlog_event', 1)
                    kwargs.setdefault('ticketURLog__urlog_user_id', assigned_to)
                    result = User.objects.only('display_name').get(pk=assigned_to).display_name
                    fielddict.update({'Assigned To': result})
                if obj.assigned_by_id != None:
                    assigned_by = obj.assigned_by_id
                    kwargs.setdefault('ticket_assign_by_id', assigned_by)
                    result = User.objects.only('display_name').get(pk=assigned_by).display_name
                    fielddict.update({'Assigned By': result})
                if obj.next_action_id != None and obj.ever_next_action == 0:
                    next_action = obj.next_action_id
                    kwargs.setdefault('ticket_next_action_id', next_action)
                    result = User.objects.only('display_name').get(pk=next_action).display_name
                    fielddict.update({'Next Action': result})
                if obj.next_action_id != None and obj.ever_next_action == 1:
                    next_action = obj.next_action_id
                    kwargs.setdefault('ticketURLog__urlog_event', 2)
                    kwargs.setdefault('ticketURLog__urlog_user_id', next_action)
                    result = User.objects.only('display_name').get(pk=next_action).display_name
                    fielddict.update({'Next Action': result})
                if obj.closed_by_id != None:
                    closed_by = obj.closed_by_id
                    kwargs.setdefault('ticket_closed_by_id', closed_by)
                    result = User.objects.only('display_name').get(pk=closed_by).display_name
                    fielddict.update({'Closed By': result})
                if obj.org_id != None:
                    org_id = obj.org_id
                    kwargs.setdefault('ticket_org_id', org_id)
                    result = Organization.objects.only('org_name').get(pk=org_id).org_name
                    fielddict.update({'Organization': result})
                # if obj.client_id != None:
                #     client_id = obj.client_id
                #     kwargs.setdefault('ticket_client_id', client_id)
                #     result = Client.objects.only('client_name').get(pk=client_id).client_name
                #     fielddict.update({'Client': result})
                if obj.date_opened != '':
                    date_opened = obj.date_opened
                    dopen_start = datetime.strptime(obj.date_opened.split(' - ')[0], '%m/%d/%Y').strftime('%Y-%m-%d')
                    dopen_end = datetime.strptime(obj.date_opened.split(' - ')[1], '%m/%d/%Y').strftime('%Y-%m-%d')
                    kwargs.setdefault('ticket_created_at__gte', dopen_start)
                    kwargs.setdefault('ticket_created_at__lte', dopen_end)
                    fielddict.update({'Date Opened': date_opened})
                if obj.date_closed != '':
                    date_closed = obj.date_closed
                    dclose_start = datetime.strptime(obj.date_closed.split(' - ')[0], '%m/%d/%Y').strftime('%Y-%m-%d')
                    dclose_end = datetime.strptime(obj.date_closed.split(' - ')[1], '%m/%d/%Y').strftime('%Y-%m-%d')
                    kwargs.setdefault('ticket_closed_at__gte', dclose_start)
                    kwargs.setdefault('ticket_closed_at__lte', dclose_end)
                    fielddict.update({'Date Closed': date_closed})
                if obj.labour_hours_val != '':
                    labour_hours_val = obj.labour_hours_val
                    labour_hours_val += ':00:00'
                    labour_hours = obj.labour_hours
                    if labour_hours == '0':
                        condition = 'lte'
                    elif labour_hours == '1':
                        condition = 'gte'
                    else:
                        condition = 'startswith'
                    kwargs.setdefault('labour_hours__' + condition, str(labour_hours_val))
                    fielddict.update({'Labor Hours - ' + labor_hours_dict[labour_hours]: labour_hours})
                if obj.task_description != '':
                    task_description = obj.task_description
                    kwargs.setdefault('ticketManager__tmgr_task__task_description__icontains', task_description)
                    fielddict.update({'Task Description': task_description})
                if obj.task_assigned_to_id != None:
                    task_assigned_to = obj.task_assigned_to_id
                    kwargs.setdefault('ticketManager__task_assigned_to_id', task_assigned_to)
                    result = User.objects.only('display_name').get(pk=task_assigned_to).display_name
                    fielddict.update({'Task Assigned To': result})
                if obj.task_completion_date != '':
                    task_completion_date = obj.task_completion_date
                    dtask_start = datetime.strptime(obj.task_completion_date.split(' - ')[0], '%m/%d/%Y').strftime('%Y-%m-%d')
                    dtask_end = datetime.strptime(obj.task_completion_date.split(' - ')[1], '%m/%d/%Y').strftime('%Y-%m-%d')
                    kwargs.setdefault('ticketManager__task_due_date__gte', dtask_start)
                    kwargs.setdefault('ticketManager__task_due_date__lte', dtask_end)
                    fielddict.update({'Task Completion Date': task_completion_date})
                if obj.search_title != '':
                    search_title = obj.search_title
                else:
                    search_title = 'Search Results'

                if obj.search_title != '':
                    search_title = obj.search_title
                if obj.output_view != '':
                    output_view = obj.output_view 
                # Query To get all ticket ids
                if obj.all_three != '':
                    tickets = Ticket.objects.filter(**kwargs).filter(args).distinct()
                    ticketid_list = Ticket.objects.filter(**kwargs).filter(args).distinct().values_list('ticket_id', flat=True)
                else:
                    tickets = Ticket.objects.filter(**kwargs).distinct()
                    ticketid_list = Ticket.objects.filter(**kwargs).distinct().values_list('ticket_id', flat=True)

                print(list(ticketid_list))

                # Saving Excel in Specific Directory
                get_xls_from_scheduled_saved_search(request,list(ticketid_list), '0', scheduledReport.sch_rpt_created_by_id)

                # Send Email Request with Attachment Start#
                smtp = getSMTPSettings(1)
                smtpserver = smtplib.SMTP(smtp['email_server'], smtp['port'])
                # print(smtpserver.ehlo())
                smtpserver.starttls()
                # print(smtpserver.ehlo)
                smtpserver.login(smtp["user_name"], smtp["password"]) 
                print(smtp)
                emails_to = []
                if scheduledReport.notify_error_id:
                    to = User.objects.values_list('email',flat=True).get(pk=scheduledReport.notify_error_id)
                    emails_to.append(to)
                print(emails_to)
                # # recipients = ScheduleReportResp.objects.values_list('sr_resp_recipt_user_id',flat=True).filter(sch_rep_id_id = 9)
                # # print(recipients)
                # to_recipients = User.objects.values_list('email',flat=True).filter(pk=2)
                # print(to_recipients)
                # for to_recipient in to_recipients:
                #     emails_to.append(to_recipient) 
                print(emails_to)
                # for email_to in emails_to:
                #     msg = EmailMessage('Scheduled Report', 'Report For Saved Search', smtp["email_sender_name"], [email_to])
                #     msg.content_subtype = "html"  
                #     dirname = os.path.dirname(__file__)
                #     filename = os.path.join(dirname, 'Attachments\Report.xlsx')
                #     msg.attach_file(filename)
                #     msg.send()
                # Send Email Request with Attachment End#

                # return HttpResponse(request,ticketid_list)
            # time.sleep(1)
            # # For Report Writer Search
            elif scheduledReport.sch_rpt_report_writer_id:
                saveRBProcessheader = SavedRBReports.objects.get(pk=scheduledReport.sch_rpt_report_writer_id)
                print(saveRBProcessheader)
                # saveQBProcess = SavedQBQuries.objects.get(pk=reportheader)
                selected_fields = json.loads(saveRBProcessheader.rb_selected_query_fields_array)
                # # print('abc')
                # # print(query_result)
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = 'attachment; filename={date}-ReportResult.xlsx'.format(
                    date=datetime.now().strftime('%Y-%m-%d'),
                )
                # workbook = Workbook()
                wb = Workbook()
                ws1 = wb.active

                # Delete the default worksheet
                # workbook.remove(workbook.active)
                #
                # Define some styles and formatting that will be later used for cells
                header_font = Font(size=7, name='Segoe UI', bold=True, color='FFFFFF')
                centered_alignment = Alignment(horizontal='left')
                border_bottom = Border(
                    bottom=Side(border_style='medium', color='21316f'),
                )
                wrapped_alignment = Alignment(
                    vertical='top',
                    wrap_text=True
                )
                #
                # Define the column titles and widths
                columns = []
                #
                # Create a worksheet/tab with the title of the category
                worksheet = wb.create_sheet(
                    title='ReportResultEcxel',
                    index=0,
                )
                # Define the background color of the header cells
                fill = PatternFill(
                    start_color='21316f',
                    end_color='21316f',
                    fill_type='solid',
                )
                row_num = 1
                #
                # # Assign values, styles, and formatting for each cell in the header
                for col_num, (column_title) in enumerate(selected_fields, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title
                    cell.font = header_font
                    cell.border = border_bottom
                    cell.alignment = centered_alignment
                    cell.fill = fill
                    # set column width
                    column_letter = get_column_letter(col_num)
                    column_dimensions = worksheet.column_dimensions[column_letter]
                    column_dimensions.width = 20
                #
                obj = SavedRBReports.objects.get(pk=scheduledReport.sch_rpt_report_writer_id)

                report_name = obj.rb_report_name
                report_title = obj.rb_report_title
                report_description = obj.rb_report_description
                rb_report_query_id = obj.rb_report_query_id

                selected_fields = selected_query_fields = json.loads(obj.rb_selected_query_fields_array)
                unselected_query_fields = json.loads(obj.rb_unselected_query_fields_array)
                selected_group_fields = json.loads(obj.rb_selected_group_fields_array)
                selected_group_fields = list(filter(lambda x: x != '', selected_group_fields))
                # test = filter(None,selected_group_fields)
                unselected_group_fields = json.loads(obj.rb_unselected_group_fields_array)
                selected_order_fields = json.loads(obj.rb_selected_order_fields_array)
                selected_group_sorting = json.loads(obj.rb_selected_group_sorting)
                selected_group_sorting = list(filter(lambda x: x['field_name'] != '', selected_group_sorting))
                selected_sort_expressions = json.loads(obj.rb_selected_sort_expressions)
                selected_sort_expressions = list(filter(lambda x: x != '', selected_sort_expressions))
                report_headers = selected_fields_formating = json.loads(obj.rb_selected_fields_formating)
                selected_format_fields = json.loads(obj.rb_selected_format_fields_array)
                print('before query_result')
                query_result = qbQueryProcess(request, rb_report_query_id, True, obj.rb_created_by_id)
                print('after_query')
                queryObj = SavedQBQuries.objects.get(pk=rb_report_query_id)
                print(queryObj)
                qb_query_pair_id = queryObj.qb_query_pair_id
                print(qb_query_pair_id)
                if selected_group_fields:
                    print('debugger0')
                    group_field_formating = []
                    for ele in selected_group_fields:
                        keyValList = ele
                        selected_fields.remove(ele)
                        # Setting Group Fields At Start of Report Header
                        group_index = list(filter(lambda d: d['actual_column_name'] in keyValList, report_headers))[0]
                        report_headers.remove(group_index)
                        group_field_formating.append(group_index)
                    selected_fields[0:0] = selected_group_fields
                    report_headers[0:0] = group_field_formating
                print('debugger1')
                selected_columns = []
                sorted_columns = []

                for field in selected_fields:
                    print('debugger2')
                    try:
                        print('debugger3')
                        actual_field = DataSetsPairFields.objects.values('df_actual_column_name').filter(df_pair_id=qb_query_pair_id).filter(df_name=field)[0]['df_actual_column_name']
                        print(actual_field)
                        print('debugger4')
                    except:
                        return render_to_response('itrak/page-404.html')
                    selected_columns.append(actual_field)
                    print(selected_columns)
                # Set the Sorting If Sorting Expression Exist
                print('dubugger1qa5')
                for sort_expression in selected_sort_expressions:
                    sort_string = sort_expression
                    if sort_string.find("(Asc)") == -1:
                        column_name = sort_string.split(' (Desc)')[0]
                    else:
                        column_name = sort_string.split(' (Asc)')[0]
                    try:
                        actual_field = \
                        DataSetsPairFields.objects.values('df_actual_column_name').filter(df_pair_id=qb_query_pair_id).filter(
                            df_name=column_name)[0]['df_actual_column_name']
                    except:
                        return render_to_response('itrak/page-404.html')
                    if sort_string.find("(Asc)") == -1:
                        sort_value = '-' + actual_field
                    else:
                        sort_value = actual_field
                    sorted_columns.append(sort_value)

                if selected_sort_expressions:
                    finalquery = query_result.values(*tuple(selected_columns)).order_by(*sorted_columns)
                    print(finalquery)
                else:
                    finalquery = query_result.values(*tuple(selected_columns))
                    print(finalquery)

                # # Iterate through all movies of a category
                for record in finalquery:
                    row_num += 1
                    if int(row_num) % 2 == 0:
                        row_fill = PatternFill(
                            start_color='FFEFD5',
                            end_color='FFEFD5',
                            fill_type='solid',
                        )
                    else:
                        row_fill = PatternFill(
                            start_color='FFFFFF',
                            end_color='FFFFFF',
                            fill_type='solid',
                        )

                    row_font = Font(size=8, name='Segoe UI', bold=False, color='000000')
                #     # Assign values, styles, and formatting for each cell in the row
                #
                    recordList = []
                    for key, value in record.items():
                        print(key)
                        if key == 'created_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'client_created_at' and value is not None:
                            value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                    '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticket_created_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticket_is_reopen_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticket_closed_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticketNote__note_modified_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticketNote__note_created_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticketOrg__ticket_assign_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticketOrg__ticket_next_action_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'qb_created_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'qb_modified_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'rbQuerySetsQuery__rb_created_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'rbQuerySetsQuery__rb_modified_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticketManager__tmgr_completion_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticket_modified_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                '%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'submitted_at' and value is not None:
                            uTimeZone = MySettings.objects.filter(m_user_id=obj.rb_created_by_id).first().m_time_zone
                            local_dt = value.astimezone(pytz.timezone(uTimeZone))
                            value1 = datetime.strptime(str(local_dt), '%Y-%m-%d %H:%M:%S%z').strftime('%m/%d/%Y %I:%M %p')
                            recordList.append(value1)
                        elif key == 'ticketAttach__attach_created_at' and value is not None:
                            if value != None:
                                value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                    '%m/%d/%Y %I:%M %p')
                                recordList.append(value1)
                        elif key == 'client_created_at' and value is not None:
                            if value != None:
                                value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                    '%m/%d/%Y %I:%M %p')
                                recordList.append(value1)
                        elif key == 'ticket_assign_at' and value is not None:
                            if value != None:
                                value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                    '%m/%d/%Y %I:%M %p')
                                recordList.append(value1)
                        elif key == 'ticket_next_action_at' and value is not None:
                            if value != None:
                                value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                    '%m/%d/%Y %I:%M %p')
                                recordList.append(value1)
                        elif key == 'submitted_at' and value is not None:
                            if value != None:
                                value1 = datetime.strptime(str(value), '%Y-%m-%d %H:%M:%S%z').strftime('%m/%d/%Y %I:%M %p')
                                recordList.append(value1)
                        elif key == 'ticket_created_at' and value is not None:
                            if value != None:
                                value1 = datetime.strptime(str(value).split(".")[0], '%Y-%m-%d %H:%M:%S').strftime(
                                    '%m/%d/%Y %I:%M %p')
                                recordList.append(value1)
                        else:
                            recordList.append(value)
                    for col_num, (cell_value) in enumerate(recordList, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.style = 'Normal'
                        # if cell_format == 'Currency':
                        #     cell.number_format = '#,##0.00 €'
                        # if col_num == 8:
                        #     cell.number_format = '[h]:mm;@'
                        cell.alignment = wrapped_alignment
                        cell.fill = row_fill
                        cell.font = row_font
                #
                # # freeze the first row
                worksheet.freeze_panes = worksheet['A2']
                #
                # # set tab color
                worksheet.sheet_properties.tabColor = 'FFFFFF'
                #
                # workbook.save(response)
                # return response
                dirname = os.path.dirname(__file__)
                # filename = os.path.join(dirname, 'Attachments\Report.xlsx')
                filename = os.path.join(dirname, 'Attachments\Report_writer.'+format(int(time.time()))+'.xlsx')
                # return HttpResponse(dirname)
                wb.save(filename)

                # Send Email Request with Attachment Start#
                smtp = getSMTPSettings(1)
                smtpserver = smtplib.SMTP(smtp['email_server'], smtp['port'])
                smtpserver.ehlo()
                smtpserver.starttls()
                smtpserver.ehlo
                smtpserver.login(smtp["user_name"], smtp["password"]) 

                emails_to = []
                if scheduledReport.notify_error_id:
                    to = User.objects.values_list('email',flat=True).get(pk=scheduledReport.notify_error_id)
                    emails_to.append(to)
                
                recipients = ScheduleReportResp.objects.values_list('sr_resp_recipt_user_id',flat=True).filter(sch_rep_id_id = 2)
                to_recipients = User.objects.values_list('email',flat=True).filter(pk__in=recipients)
                
                for to_recipient in to_recipients:
                    emails_to.append(to_recipient) 
                print(emails_to)
                # return HttpResponse(emails_to)
                
                # for email_to in emails_to:
                #     msg = EmailMessage('Scheduled Report', 'Report For Report Writer', smtp["email_sender_name"], [email_to])
                #     msg.content_subtype = "html"  
                #     dirname = os.path.dirname(__file__)
                #     filename = os.path.join(dirname, 'Attachments\Report.xlsx')
                #     msg.attach_file(filename)
                #     msg.send()
    else:
        return HttpResponse('failed')



@csrf_exempt
def get_xls_from_scheduled_saved_search(request, tickets, type, user_id):
    print('export')
    """
    Downloads all movies as Excel file with a worksheet for each movie category
    """
    tickets = tickets
    xlsType = type
    # category_queryset = Ticket.objects.all()
    ticket_queryset = Ticket.objects.filter(ticket_id__in=json.loads(str(tickets))).filter(ticket_is_delete=0).filter(ticket_is_active=1)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    dirname = os.path.dirname(__file__)
    response['Content-Disposition'] = 'attachment; filename={date}-TicketSearchList.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    # workbook = Workbook()
    wb = Workbook()
    ws1 = wb.active
    # ws1.title = "1st Hour"
    # wb.save('D:/atg-extra/FileName11.xlsx')

    # workbook = xlsxwriter.workbook('demo.xlsx')
    # workbook= workbook.create_sheet('TicketSearchList.xlsx', idx)
    
    # save_path = os.path.join(BASE_DIR, 'backup')
    # writer = ExcelWriter('D:/atg-extra/media/PythonExportt.xlsx'.format(save_path))
    # writer.save()
    # workbook= openpyxl.load_workbook('myfile.xlsx')
    # path='D:/atg-extra/workbook.xlsx'
    # workbook=openpyxl.load_workbook(path)

    # Define some styles and formatting that will be later used for cells
    header_font = Font(size=7, name='Segoe UI', bold=True, color='FFFFFF')
    centered_alignment = Alignment(horizontal='left')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='21316f'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    if xlsType == '0':
        # Define the column titles and widths
        columns = [
            ('Ticket #', 7),
            ('Status', 7),
            ('SubStatus', 15),
            ('Subject', 40),
            ('Ticket Type', 15),
            ('Subtype', 20),
            ('Priority', 7),
            ('Submitted Date', 15),
            ('Submitted By Name', 15),
            ('Entered By Name', 15),
            ('Assigned To Name', 15),
            ('Assigned Date', 15),
            ('Next Action Name', 15),
            ('Closed Date', 15),
            ('Last Activity', 15),
            ('Total Time Open', 15),
            ('Time Open (Days)', 15),
            ('Adj Time Open', 15),
            ('Adj Time Open (Days)', 15),
        ]
    else:
        # Define the column titles and widths
        columns = [
            ('Ticket #', 7),
            ('Status', 7),
            ('SubStatus', 15),
            ('Subject', 40),
            ('Full Description', 40),
            ('Ticket Type', 15),
            ('Subtype', 20),
            ('Subtype 2', 20),
            ('Subtype 3', 20),
            ('Subtype 4', 20),
            ('Priority', 7),
            ('Organization', 10),
            ('Client ID', 7),
            ('Client', 15),
            ('Record Locator', 10),
            ('Caller Name', 15),
            ('Caller Phone', 15),
            ('Caller Email', 20),
            ('Passenger Name', 15),
            ('Address', 40),
            ('Agent Error or Goodwill', 20),
            ('Agent Responsible (if applic)', 20),
            ('Airline Ticket #', 15),
            ('Amount of Payout', 15),
            ('Amount Saved', 10),
            ('Attention', 15),
            ('Check Number', 15),
            ('Company', 15),
            ('Correction / Containment Actions', 40),
            ('Corrective Action', 40),
            ('Is Traveler a VIP?', 10),
            ('Notes to print on check', 40),
            ('Pay to the order of', 30),
            ('Payout Required?', 15),
            ('Resp Vendor\'s City (if applic)', 40),
            ('Root Cause', 40),
            ('Vendor Name Resp (if applic)', 30),
            ('Who Approved', 20),
            ('Submitted Date', 15),
            ('Submitted By Name', 15),
            ('Submitted By Dept', 20),
            ('Submitted By Phone', 20),
            ('Assigned To', 15),
            ('Assigned To Name', 15),
            ('Assigned Date', 15),
            ('Assigned To Phone', 15),
            ('Next Action', 15),
            ('Next Action Name', 15),
            ('Next Action Date', 15),
            ('Next Action Phone', 15),
            ('Closed By', 15),
            ('Closed By Name', 15),
            ('Closed Date', 15),
            ('Closed By Phone', 15),
            ('ReOpened By', 15),
            ('ReOpened By Name', 15),
            ('ReOpened Date', 15),
            ('ReOpened By Phone', 15),
            ('Last Activity', 15),
            ('Total Time Open', 15),
            ('Time Open (Days)', 15),
            ('Adj Time Open', 15),
            ('Adj Time Open (Days)', 15),
        ]

    # Create a worksheet/tab with the title of the category
    worksheet = wb.create_sheet(
        title='TicketSearchEcxel',
        index=0,
        
    )
    # Define the background color of the header cells
    fill = PatternFill(
        start_color='21316f',
        end_color='21316f',
        fill_type='solid',
    )
    row_num = 1

    # Assign values, styles, and formatting for each cell in the header
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = border_bottom
        cell.alignment = centered_alignment
        cell.fill = fill
        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width

    # Iterate through all movies of a category
    for ticket in ticket_queryset:
        row_num += 1

        if ticket.ticket_status == 0:
            delta = datetime.now(timezone.utc) - ticket.ticket_created_at
        else:
            delta = ticket.ticket_closed_at - ticket.ticket_created_at

        days, seconds = delta.days, delta.seconds
        hours = days * 24 + seconds
        minutes = (seconds % 3600)
        seconds = seconds % 60

        if ticket.description:
            s = MLStripper()
            s.feed(ticket.description)
            plain_description = s.get_data()
        else:
            plain_description = ''

        if xlsType == '0':
            # Define data and formats for each cell in the row
            row = [
                (ticket.ticket_id, 'Normal'),
                ('Open' if ticket.ticket_status == 0 else 'Close', 'Normal'),
                (ticket.ticket_sub_status.sub_status_text if ticket.ticket_sub_status_id else '', 'Normal'),
                (ticket.subject, 'Normal'),
                # (timedelta(minutes=movie.ticket_created_at), 'Normal'),
                (ticket.ticket_type.ttype_name if ticket.ticket_type_id else '', 'Normal'),
                (ticket.ticket_subtype1.ttype_name if ticket.ticket_subtype1_id else '', 'Normal'),
                (ticket.priority.priority_name if ticket.priority_id else '', 'Normal'),
                (datetime.strptime(str(ticket.submitted_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S%z').strftime('%d/%m/%Y %I:%M %p') if ticket.submitted_at else '', 'Normal'),
                # (datetime.strptime(str(
                    # datetime.strptime(str(ticket.submitted_date) + ' ' + str(ticket.submitted_time) + '.000001+00:00','%Y-%m-%d %H:%M:%S.%f%z').astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %H:%M %p') if ticket.submitted_date else '', 'Normal'),
                # (datetime.strptime(str(ticket.submitted_date), '%Y-%m-%d').strftime('%d/%m/%Y') if ticket.submitted_date else '', 'Normal'),
                (ticket.ticket_caller.display_name if ticket.ticket_caller_id else '', 'Normal'),
                (ticket.ticket_created_by.display_name if ticket.ticket_created_by_id else '', 'Normal'),
                (ticket.ticket_assign_to.display_name if ticket.ticket_assign_to_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_assign_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_assign_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_next_action.display_name if ticket.ticket_next_action_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_closed_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_closed_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_closed_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_closed_at else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_modified_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_modified_at else '', 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),

            ]
        else:

            #Data Building of Condition Fields Start#
            if ticket.agent_error_goodwill:
                if ticket.agent_error_goodwill == '0':
                    agent_error_goodwill = 'Agent Error'
                else:
                    agent_error_goodwill = 'Goodwill'
            else:
                agent_error_goodwill = ''

            if ticket.is_traveler_vip:
                if ticket.is_traveler_vip == '0':
                    is_traveler_vip = 'Yes'
                else:
                    is_traveler_vip = 'No'
            else:
                is_traveler_vip = ''

            if ticket.is_payout_required:
                if ticket.is_payout_required == '0':
                    is_payout_required = 'Yes'
                else:
                    is_payout_required = 'No'
            else:
                is_payout_required = ''

            #Data Building of Condition Fields End#


            # Define data and formats for each cell in the row
            row = [
                (ticket.ticket_id, 'Normal'),
                ('Open' if ticket.ticket_status == 0 else 'Close', 'Normal'),
                (ticket.ticket_sub_status.sub_status_text if ticket.ticket_sub_status_id else '', 'Normal'),
                (ticket.subject, 'Normal'),
                (plain_description, 'Normal'),
                (ticket.ticket_type.ttype_name if ticket.ticket_type_id else '', 'Normal'),
                (ticket.ticket_subtype1.ttype_name if ticket.ticket_subtype1_id else '', 'Normal'),
                (ticket.ticket_subtype2.ttype_name if ticket.ticket_subtype2_id else '', 'Normal'),
                (ticket.ticket_subtype3.ttype_name if ticket.ticket_subtype3_id else '', 'Normal'),
                (ticket.ticket_subtype4.ttype_name if ticket.ticket_subtype4_id else '', 'Normal'),
                (ticket.priority.priority_name if ticket.priority_id else '', 'Normal'),
                (ticket.ticket_org.org_name if ticket.ticket_org_id else '', 'Normal'),
                (ticket.ticket_client.client_cus_id if ticket.ticket_client_id else '', 'Normal'),
                (ticket.ticket_client.client_name if ticket.ticket_client_id else '', 'Normal'),
                (ticket.ticket_record_locator if ticket.ticket_record_locator else '', 'Normal'),
                (ticket.ticket_caller_name if ticket.ticket_caller_name else '', 'Normal'),
                (ticket.ticket_caller_phone if ticket.ticket_caller_phone else '', 'Normal'),
                (ticket.ticket_caller_email if ticket.ticket_caller_email else '', 'Normal'),
                (ticket.ticket_passenger_name if ticket.ticket_passenger_name else '', 'Normal'),
                (ticket.ticket_address if ticket.ticket_address else '', 'Normal'),
                (agent_error_goodwill, 'Normal'),
                (ticket.agent_responsible if ticket.agent_responsible else '', 'Normal'),
                (ticket.airline_ticket_no if ticket.airline_ticket_no else '', 'Normal'),
                (ticket.ticket_payout_amount if ticket.ticket_payout_amount else '', 'Normal'),
                (ticket.amount_saved if ticket.amount_saved else '', 'Normal'),
                (ticket.ticket_attention if ticket.ticket_attention else '', 'Normal'),
                (ticket.check_number if ticket.check_number else '', 'Normal'),
                (ticket.ticket_company if ticket.ticket_company else '', 'Normal'),
                (ticket.corr_cont_actions if ticket.corr_cont_actions else '', 'Normal'),
                (ticket.corrective_action if ticket.corrective_action else '', 'Normal'),
                (is_traveler_vip, 'Normal'),
                (ticket.notes_on_check if ticket.notes_on_check else '', 'Normal'),
                (ticket.ticket_order_of_pay if ticket.ticket_order_of_pay else '', 'Normal'),
                (is_payout_required, 'Normal'),
                (ticket.vresponsible_city if ticket.vresponsible_city else '', 'Normal'),
                (ticket.ticket_root_cause if ticket.ticket_root_cause else '', 'Normal'),
                (ticket.vendor_responsible if ticket.vendor_responsible else '', 'Normal'),
                (ticket.check_approved_by if ticket.check_approved_by else '', 'Normal'),
                (datetime.strptime(str(ticket.submitted_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S%z').strftime('%d/%m/%Y %I:%M %p') if ticket.submitted_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.submitted_date), '%Y-%m-%d').strftime('%d/%m/%Y') if ticket.submitted_date else '', 'Normal'),
                (ticket.ticket_caller.display_name if ticket.ticket_caller_id else '', 'Normal'),
                (ticket.ticket_caller.user_dep.dep_name if ticket.ticket_caller_id and ticket.ticket_caller.user_dep_id else '', 'Normal'),
                (ticket.ticket_caller.phone_no if ticket.ticket_caller_id and ticket.ticket_caller_id else '', 'Normal'),
                (ticket.ticket_assign_to.display_name if ticket.ticket_assign_to_id else '', 'Normal'),
                (ticket.ticket_assign_to.first_name if ticket.ticket_assign_to_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_assign_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_assign_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_assign_to.phone_no if ticket.ticket_assign_to_id else '', 'Normal'),
                (ticket.ticket_next_action.display_name if ticket.ticket_next_action_id else '', 'Normal'),
                (ticket.ticket_next_action.first_name if ticket.ticket_next_action_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_next_action_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_next_action_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_next_action.phone_no if ticket.ticket_next_action_id else '', 'Normal'),
                (ticket.ticket_closed_by.display_name if ticket.ticket_closed_by_id else '', 'Normal'),
                (ticket.ticket_closed_by.first_name if ticket.ticket_closed_by_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_closed_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_closed_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_closed_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_closed_at else '', 'Normal'),
                (ticket.ticket_closed_by.phone_no if ticket.ticket_closed_by_id else '', 'Normal'),
                (ticket.ticket_is_reopen_by.display_name if ticket.ticket_is_reopen_by_id else '', 'Normal'),
                (ticket.ticket_is_reopen_by.first_name if ticket.ticket_is_reopen_by_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_is_reopen_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_is_reopen_at else '', 'Normal'),
                # (datetime.strptime(str(ticket.ticket_created_at), '%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M%p') if ticket.ticket_created_at else '', 'Normal'),
                (ticket.ticket_is_reopen_by.phone_no if ticket.ticket_is_reopen_by_id else '', 'Normal'),
                (datetime.strptime(str(ticket.ticket_modified_at.astimezone(pytz.timezone(MySettings.objects.filter(m_user_id=user_id).first().m_time_zone))),'%Y-%m-%d %H:%M:%S.%f%z').strftime('%d/%m/%Y %I:%M %p') if ticket.ticket_modified_at else '', 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),
                (hours, 'Normal'),
                (days, 'Normal'),

            ]

        if int(row_num) % 2 == 0:
            row_fill = PatternFill(
                start_color='FFEFD5',
                end_color='FFEFD5',
                fill_type='solid',
            )
        else:
            row_fill = PatternFill(
                start_color='FFFFFF',
                end_color='FFFFFF',
                fill_type='solid',
            )
        print(row_num)

        row_font = Font(size=8, name='Segoe UI', bold=False, color='000000')
        # Assign values, styles, and formatting for each cell in the row
        for col_num, (cell_value, cell_format) in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.style = cell_format
            # if cell_format == 'Currency':
            #     cell.number_format = '#,##0.00 €'
            # if col_num == 8:
            #     cell.number_format = '[h]:mm;@'
            cell.alignment = wrapped_alignment
            cell.fill = row_fill
            cell.font = row_font

    # freeze the first row
    worksheet.freeze_panes = worksheet['A2']

    # set tab color
    worksheet.sheet_properties.tabColor = 'FFFFFF'
    dirname = os.path.dirname(__file__)
    filename = os.path.join(dirname, 'Attachments\Report.'+format(int(time.time()))+'.xlsx')
    # return HttpResponse(dirname)
    wb.save(filename)

    # return response

#Export tickets to XLSX End#

#Export User to XLSX Start#
runDailyScheduleJob(request)